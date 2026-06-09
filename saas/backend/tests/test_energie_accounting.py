"""Tests de la matrice comptable ENGIE + fiche de liaison finances."""
import io

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.core.db import Base
from app.models.invoice import (
    EnergyAccountingNatureRule,
    EnergyAccountingSiteMapping,
    EnergyInvoice,
    EnergyInvoiceImport,
    EnergyInvoiceLine,
    EnergyInvoicePeriod,
    EnergyInvoiceSite,
)
from app.services import energie_accounting as svc


@pytest.fixture()
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


def _codification_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sites vers codes"
    ws1.append(["PRM", "Nom du site", "Service", "Libellé service", "Fonction", "Antenne"])
    ws1.append(["12345678901234", "Mairie", "S100", "Bâtiments", "F20", "ANT-A"])
    ws2 = wb.create_sheet("Poste facturé vers Nature ctpab")
    ws2.append(["Poste facturé", "Nature proposée", "Libellé nature"])
    ws2.append(["Abonnement", "6061", "Énergie"])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _make_invoice(db, *, prm="12345678901234", poste="Abonnement"):
    imp = EnergyInvoiceImport(
        city_id=1, uploaded_by_user_id=1, original_filename="f.pdf", stored_filename="s.pdf",
        storage_path="/tmp/s.pdf", file_size_bytes=1, sha256="x" * 64, invoice_number="INV-1",
    )
    db.add(imp)
    db.flush()
    inv = EnergyInvoice(city_id=1, import_id=imp.id, supplier="ENGIE", invoice_number="INV-1", total_ht=100.0)
    db.add(inv)
    db.flush()
    site = EnergyInvoiceSite(invoice_id=inv.id, prm_id=prm, site_name="Mairie")
    db.add(site)
    db.flush()
    period = EnergyInvoicePeriod(invoice_site_id=site.id)
    db.add(period)
    db.flush()
    db.add(EnergyInvoiceLine(invoice_period_id=period.id, poste=poste, normalized_code=poste, amount_ht=42.0))
    db.commit()
    db.refresh(imp)
    return imp


def test_import_codification_upserts(db):
    res = svc.import_codification_workbook(db, _codification_xlsx(), filename="codif.xlsx", city_id=1)
    assert res.errors == []
    assert res.site_mappings_created == 1
    assert res.nature_rules_created == 1
    # ré-import = update, pas de doublon
    res2 = svc.import_codification_workbook(db, _codification_xlsx(), filename="codif.xlsx", city_id=1)
    assert res2.site_mappings_updated == 1
    assert res2.nature_rules_updated == 1
    assert len(svc.list_site_mappings(db, 1)) == 1
    assert len(svc.list_nature_rules(db, 1)) == 1


def test_resolve_codification_ok_and_blocked(db):
    svc.import_codification_workbook(db, _codification_xlsx(), filename="codif.xlsx", city_id=1)
    imp = _make_invoice(db)
    rows = svc.resolve_invoice_codification(db, imp)
    assert len(rows) == 1
    assert rows[0].status == "ok"
    assert rows[0].service_code == "S100"
    assert rows[0].accounting_nature == "6061"

    # facture avec un PRM non codifié -> blocked
    imp2 = _make_invoice(db, prm="99999999999999")
    rows2 = svc.resolve_invoice_codification(db, imp2)
    assert rows2[0].status == "blocked"
    assert rows2[0].service_code is None


def test_bootstrap_from_invoices(db):
    _make_invoice(db, prm="11111111111111")
    out = svc.bootstrap_site_mappings_from_invoices(db, 1)
    assert out["created"] == 1
    mappings = svc.list_site_mappings(db, 1)
    assert mappings[0].prm_id == "11111111111111"
    assert mappings[0].site_name == "Mairie"


def test_build_liaison_workbook(db):
    svc.import_codification_workbook(db, _codification_xlsx(), filename="codif.xlsx", city_id=1)
    imp = _make_invoice(db)
    content = svc.build_energy_liaison_workbook(db, imp)
    assert content[:2] == b"PK"  # xlsx = zip
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    assert ws["A1"].value == "Fiche de liaison finance ENGIE"
    # ligne de données présente avec la nature résolue
    found = any(
        ws.cell(row=r, column=12).value == "6061"
        for r in range(14, ws.max_row + 1)
    )
    assert found
