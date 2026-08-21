"""Réexport ASTECH — le fichier rendu à la collectivité.

Ce que ces tests protègent, dans l'ordre d'importance :

1. **Les en-têtes sont recopiés à l'octet près** depuis le fichier importé. C'est la
   condition de réinjection : un `COD_COMPTABLE` devenu `CODE_COMPTABLE`, un `0#SURF`
   reformaté, et ASTECH refuse l'import. Le gabarit est dérivé du fichier source, il
   n'est jamais retapé dans le code.
2. **Le `CODE_BIEN` n'est jamais réécrit** — clé de mise à jour d'ASTECH.
3. **On n'écrit jamais une valeur qu'on n'a pas su produire proprement.** Le doute
   envoie la ligne en feuille « à vérifier », il ne remplit pas la case au jugé.
"""
from __future__ import annotations

import io
import json

import openpyxl
import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.core.db import Base
from app.models.building import Building
from app.models.city import City
from app.models.local import Local  # noqa: F401 (enregistre la table)
from app.models.patrimoine_legacy import (
    STATUS_GONE,
    STATUS_LINKED,
    STATUS_PROPOSED,
    STATUS_TODO,
    PatrimoineLegacyAsset,
    PatrimoineLegacyImport,
)
from app.models.site import Site  # noqa: F401
from app.services.patrimoine_legacy import compute_candidates, set_asset_gone
from app.services.patrimoine_legacy_export import (
    build_astech_workbook,
    build_refcad,
    format_coordinate,
    normalize_street,
    split_house_number,
)

# En-têtes tels qu'ils apparaissent dans l'export réel, orthographes bizarres comprises.
HEADERS = [
    "CODE_BIEN", "DESIGNATION", "NOMCOURT", "GENRE", "COD_COMPTABLE", "0#SURF",
    "NORUE", "BISTER", "LIBELVOIE", "CODPOST", "VILLE", "COMMUNE", "REFCAD",
    "LATITUDE", "LONGITUDE",
]


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(City(id=1, nom_commune="Sete", code_commune="34301"))
        session.add(
            PatrimoineLegacyImport(
                city_id=1,
                batch="astech_test",
                filename="export.xlsx",
                sheet_name="Feuil1",
                header_row=2,
                headers_json=json.dumps(HEADERS),
                total_rows=1,
            )
        )
        session.commit()
        yield session


def _asset(db: Session, **kwargs) -> PatrimoineLegacyAsset:
    defaults = dict(
        city_id=1,
        code_bien="ADMICIMET02",
        designation="CIMETIERE LE PY",
        nomcourt="CIMETIERE LE PY",
        status=STATUS_LINKED,
        building_id=1,
        resolved_housenumber="2",
        resolved_street="AV DU MARECHAL JUIN",
        resolved_city="Sète",
        resolved_citycode="34301",
        resolved_refcad="AK149",
        resolved_source="building",
        latitude=43.404512,
        longitude=3.693845,
    )
    defaults.update(kwargs)
    asset = PatrimoineLegacyAsset(**defaults)
    db.add(asset)
    db.commit()
    return asset


def _sheets(db: Session) -> dict[str, list[list]]:
    result = build_astech_workbook(db, 1)
    workbook = openpyxl.load_workbook(io.BytesIO(result["content"]))
    return {
        name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
        for name in workbook.sheetnames
    }, result


# --- La contrainte dure : les en-têtes ---------------------------------------

def test_les_entetes_sont_recopies_a_l_octet_pres(db_session: Session):
    db_session.add(Building(id=1, city_id=1, nom_batiment="CIMETIERE LE PY", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session)

    sheets, result = _sheets(db_session)
    rows = sheets["Feuil1"]

    # Ligne 1 vide puis en-têtes en ligne 2 : la mise en page du fichier source.
    assert rows[0] == [None] * len(result["columns"])
    assert result["header_row"] == 2
    header_row = rows[1]
    # Orthographe exacte, reprise du gabarit — pas une constante du code.
    assert header_row[0] == "CODE_BIEN"
    assert set(header_row) <= set(HEADERS)
    assert "CODEBIEN" not in header_row  # l'orthographe de la feuille BAT ferait échouer


def test_le_code_bien_n_est_jamais_reecrit(db_session: Session):
    db_session.add(Building(id=1, city_id=1, nom_batiment="X", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session, code_bien="ADMIANMAI02")

    sheets, _ = _sheets(db_session)
    data = sheets["Feuil1"][2]
    assert data[0] == "ADMIANMAI02"


def test_un_bien_a_creer_sort_avec_une_cle_vide(db_session: Session):
    """Décision Q13 : c'est ASTECH qui attribuera le code."""
    db_session.add(Building(id=1, city_id=1, nom_batiment="X", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session, code_bien="NOUVEAU_1", status="a_creer")

    sheets, _ = _sheets(db_session)
    # Cellule vide : openpyxl relit une chaîne vide comme `None`. Les deux disent la
    # même chose à ASTECH — « pas de code, attribue-lui-en un ».
    assert sheets["Feuil1"][2][0] in (None, "")


# --- La règle de sûreté : ne rien inventer -----------------------------------

def test_un_rattachement_non_confirme_ne_part_pas(db_session: Session):
    """Le moteur PROPOSE, il ne valide pas : personne n'a confirmé ce rattachement."""
    db_session.add(Building(id=1, city_id=1, nom_batiment="X", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session, status=STATUS_PROPOSED)

    sheets, result = _sheets(db_session)
    assert result["exported_rows"] == 0
    assert result["review_rows"] == 1
    motif = sheets["À vérifier"][1][2]
    assert "jamais confirmé" in motif


@pytest.mark.parametrize(
    ("source", "attendu"),
    [
        ("0002", ("2", None)),
        ("2", ("2", None)),
        ("0005B", ("5", "BIS")),
        ("15 T", ("15", "TER")),
    ],
)
def test_numero_de_voirie_et_bis_ter(source: str, attendu: tuple):
    number, bister, issue = split_house_number(source)
    assert (number, bister) == attendu
    assert issue is None


def test_un_indice_de_repetition_inconnu_n_est_pas_ecrit():
    """Mieux vaut une case vide qu'une valeur inventée dans le fichier de la collectivité."""
    number, bister, issue = split_house_number("12Z")
    assert number == "12"
    assert bister is None
    assert issue is not None


@pytest.mark.parametrize(
    ("source", "attendu"),
    [
        # Expansion des abréviations : la source DGFIP mélange BD/BOULEVARD, AV/Avenue.
        ("AV DU MARECHAL JUIN", "AVENUE DU MARECHAL JUIN"),
        ("BD CAMILLE BLANC", "BOULEVARD CAMILLE BLANC"),
        ("IMP DE LA BORDIGUE", "IMPASSE DE LA BORDIGUE"),
        ("Rue Lacan", "RUE LACAN"),
        # Déjà en toutes lettres : inchangé.
        ("BOULEVARD JOLIOT CURIE", "BOULEVARD JOLIOT CURIE"),
        # Pas de type reconnu : recopié tel quel, en majuscules sans accent.
        ("CORNICHE DE NEUBURG", "CORNICHE DE NEUBURG"),
    ],
)
def test_normalisation_du_libelle_de_voie(source: str, attendu: str):
    value, _postcode, issue = normalize_street(source)
    assert value == attendu
    assert issue is None


@pytest.mark.parametrize("source", ["4674", "RUE"])
def test_une_voie_non_analysable_n_est_pas_ecrite(source: str):
    """Relevé en base : `4674` (un nombre là où le type devrait être) et un type seul."""
    value, _postcode, issue = normalize_street(source)
    assert value is None
    assert issue is not None


def test_le_code_postal_ne_fuit_pas_dans_le_libelle_de_voie():
    """Constaté sur les vraies données du 2026-08-20.

    Le géocodage inverse rend un libellé complet ; après découpage, le code postal et la
    commune restaient collés au nom de voie. `LIBELVOIE` serait parti dans ASTECH avec
    « QUAI DE LA RESISTANCE 34200 SETE ». Le code postal est coupé — et récupéré, car il
    est juste.
    """
    value, postcode, issue = normalize_street("Quai de la Résistance 34200 Sète")
    assert value == "QUAI DE LA RESISTANCE"
    assert postcode == "34200"
    assert issue is None


def test_les_coordonnees_sortent_a_virgule_decimale():
    """Format constaté dans le fichier : `'43,404512'`. Un point serait mal relu."""
    assert format_coordinate(43.404512) == "43,404512"
    assert format_coordinate(None) is None


def test_un_numero_de_plan_hors_format_ne_produit_pas_de_refcad():
    """Au-delà de 999, une référence tronquée désignerait une AUTRE parcelle."""
    asset = PatrimoineLegacyAsset(
        city_id=1, code_bien="X", resolved_section="AK", resolved_numero_plan="1234"
    )
    value, issue = build_refcad(asset)
    assert value is None
    assert issue is not None


def test_la_feuille_de_tracabilite_dit_l_ancienne_et_la_nouvelle_valeur(db_session: Session):
    db_session.add(Building(id=1, city_id=1, nom_batiment="X", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session, source_libelvoie="MARECHAL JUIN", source_norue="0")

    sheets, _ = _sheets(db_session)
    trace = sheets["Traçabilité"]
    voie = [row for row in trace[1:] if row[2] == "LIBELVOIE"]
    assert len(voie) == 1
    assert voie[0][3] == "MARECHAL JUIN"           # valeur ASTECH d'origine
    assert voie[0][4] == "AVENUE DU MARECHAL JUIN"  # valeur renvoyée


def test_sans_import_le_gabarit_ne_peut_pas_etre_invente(db_session: Session):
    """Les en-têtes viennent du fichier source : sans lui, on ne produit rien."""
    for row in db_session.scalars(select(PatrimoineLegacyImport)):
        db_session.delete(row)
    db_session.commit()

    with pytest.raises(ValueError, match="gabarit"):
        build_astech_workbook(db_session, 1)


# --- Q23 : les biens à supprimer de AS-TECH -----------------------------------

def test_un_bien_a_supprimer_ne_repart_pas_dans_le_fichier(db_session: Session):
    """Un CODE_BIEN que Po2 demande de supprimer ne doit pas etre remis a jour.

    Il ne figure donc pas dans la feuille réinjectable — mais il est nommé en
    « À vérifier » avec son motif, pour que la référente sache ce que Po2 lui demande
    de retirer.
    """
    db_session.add(Building(id=1, city_id=1, nom_batiment="CIMETIERE LE PY", nom_commune="Sete"))
    db_session.commit()
    _asset(db_session, code_bien="ADMICIMET02", status=STATUS_GONE)

    sheets, result = _sheets(db_session)

    assert result["exported_rows"] == 0
    assert sheets["Feuil1"][2:] == []
    review = sheets["À vérifier"][1:]
    assert [row[0] for row in review] == ["ADMICIMET02"]
    assert "à supprimer de AS-TECH" in review[0][2]


def test_annuler_la_consigne_rend_le_statut_reel(db_session: Session):
    """Annuler ne restaure pas un statut mémorisé : il se déduit de l'état du bien."""
    db_session.add(Building(id=1, city_id=1, nom_batiment="X", nom_commune="Sete"))
    db_session.commit()

    rattache = _asset(db_session, code_bien="ADMICIMET02", building_id=1)
    set_asset_gone(db_session, rattache, True)
    assert rattache.status == STATUS_GONE
    set_asset_gone(db_session, rattache, False)
    assert rattache.status == STATUS_LINKED

    orphelin = _asset(db_session, code_bien="ADMICIMET03", building_id=None, status=STATUS_TODO)
    set_asset_gone(db_session, orphelin, True)
    assert orphelin.status == STATUS_GONE
    set_asset_gone(db_session, orphelin, False)
    assert orphelin.status == STATUS_TODO


def test_un_bien_a_supprimer_n_est_jamais_repropose_par_le_moteur(db_session: Session):
    """`compute_candidates` ne balaie que les biens « à traiter » : celui-ci dort."""
    db_session.add(Building(id=1, city_id=1, nom_batiment="CIMETIERE LE PY", nom_commune="Sete"))
    db_session.commit()
    asset = _asset(db_session, code_bien="ADMICIMET02", building_id=None, status=STATUS_GONE)

    compute_candidates(db_session, 1)

    db_session.refresh(asset)
    assert asset.status == STATUS_GONE
    assert asset.candidate_building_id is None
