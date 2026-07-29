from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.models.cpe import CpeAccountingSiteMapping, CpeFinanceControl, CpeFinanceInvoice, CpeFinanceLine
from app.models.cpe_dpgf_p1 import CpeDpgfP1Line
from app.models.invoice import EnergyInvoice, EnergyInvoiceImport, EnergyInvoiceSite
from app.services import comptable_report
from app.services.comptable_report import (
    PlatformInvoice,
    WorklistInvoice,
    build_comptable_control_workbook,
    extract_supplier_invoice_number,
    parse_comptable_worklist,
)


COMPTA_DIR = Path(__file__).resolve().parents[2] / "energie" / "COMPTA"


def test_line_ttc_applies_vat_rate() -> None:
    assert comptable_report._line_ttc(100.0, 20.0) == 120.0
    assert comptable_report._line_ttc(100.0, None) == 100.0  # taux absent -> 0 %
    assert comptable_report._line_ttc(None, 20.0) is None


def test_invoice_ttc_ratio_from_totals() -> None:
    assert comptable_report._invoice_ttc_ratio(120.0, 100.0) == 1.2
    assert comptable_report._invoice_ttc_ratio(120.0, None) is None
    assert comptable_report._invoice_ttc_ratio(None, 100.0) is None
    assert comptable_report._invoice_ttc_ratio(120.0, 0) is None


@pytest.mark.parametrize(
    ("filename", "expected_count", "first_number"),
    [
        ("FACTURES ENGIE.xlsx", 25, "150000071294"),
        ("FACTURES DALKIA 2IEME TRIMESTRE.xlsx", 50, "0001E2607QRY8"),
    ],
)
def test_parse_comptable_worklist_real_files(filename: str, expected_count: int, first_number: str) -> None:
    workbook = COMPTA_DIR / filename
    if not workbook.exists():
        pytest.skip(f"Worklist comptable reelle absente: {workbook}")

    parsed = parse_comptable_worklist(workbook)

    assert parsed.sheet_name.startswith("_ShowList-")
    assert len(parsed.rows) == expected_count
    assert parsed.rows[0].supplier_invoice_number == first_number
    assert all(row.supplier_invoice_number for row in parsed.rows)
    assert all(row.total_ttc is not None for row in parsed.rows)


def test_extract_supplier_invoice_number_strict_token() -> None:
    assert extract_supplier_invoice_number("FAC. 150000071294 DU 07/07/2026") == "150000071294"
    assert extract_supplier_invoice_number(" FAC. 0001E2607QRY8 DU 30/06/2026 ") == "0001E2607QRY8"
    assert extract_supplier_invoice_number("TOTAL") is None


def test_build_report_creates_one_sheet_per_market_without_uploads() -> None:
    content = build_comptable_control_workbook(db=None, city_id=303, files_by_market={})
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)

    assert wb.sheetnames == ["Synthèse", "DALKIA", "ENGIE", "EDF", "TotalEnergies"]
    assert wb["Synthèse"]["A1"].value == "Synthèse - rapport de contrôle comptable"
    for sheet_name in ["DALKIA", "ENGIE", "EDF", "TotalEnergies"]:
        assert wb[sheet_name]["A1"].value == "Aucune facture à analyser"


def test_cpe_accounting_summary_writes_invoice_level_prices_and_codes() -> None:
    class FakeScalarResult:
        def __init__(self, rows):
            self.rows = rows

        def all(self):
            return self.rows

    class FakeScalarDb:
        def __init__(self, sequences):
            self.sequences = list(sequences)

        def scalars(self, _stmt):
            return FakeScalarResult(self.sequences.pop(0))

    invoice = CpeFinanceInvoice(id=10, invoice_number="D-1", contract_code="C001", total_ht=120.0)
    worklist = WorklistInvoice(
        row_number=2,
        accounting_number="202600001",
        supplier_invoice_number="D-1",
        label="FAC. D-1 DU 30/06/2026",
        total_ttc=144.0,
        invoice_date="30/06/2026",
        arrival_date=None,
        supplier_code=None,
        supplier_name="DALKIA",
        invoice_status=None,
        liquidation_status=None,
        market_code=None,
        raw={},
    )
    platform = PlatformInvoice(
        id=10,
        invoice_number="D-1",
        total_ttc=144.0,
        control_status="blocked",
        decision_status="a_controler",
        problem_summary="Imputation comptable absente",
        raw=invoice,
    )
    lines = [
        CpeFinanceLine(
            id=1,
            invoice_id=10,
            row_number=1,
            market="P2",
            billed_item="P2",
            accounting_site_id=100,
            site_code_detected="S1",
            base_price=1000.0,
            revised_price=1030.0,
            accounting_nature="6156",
            accounting_label="Maintenance",
        ),
        CpeFinanceLine(
            id=2,
            invoice_id=10,
            row_number=2,
            market="P3",
            billed_item="P3.4",
            accounting_site_id=100,
            site_code_detected="S1",
            base_price=500.0,
            revised_price=515.0,
            accounting_nature=None,
        ),
    ]
    controls = [
        CpeFinanceControl(
            invoice_id=10,
            line_id=1,
            control_type="revision_p2",
            index_year=2026,
            index_quarter=2,
            delta_abs=0.01,
        )
    ]
    sites = [
        CpeAccountingSiteMapping(
            id=100,
            code_site="S1",
            site_name="Site 1",
            service_code="020",
            function_code="F01",
            antenna_code="A01",
            operation_code="98004",
        )
    ]
    db = FakeScalarDb([lines, controls, sites])

    summary = comptable_report._cpe_line_enrichments(db, [(worklist, platform)])[10]

    assert summary["base_price"] == 1500.0
    assert summary["revised_price"] == 1545.0
    assert summary["revision_amount"] == 45.0
    assert summary["revision_delta"] == 0.01
    assert summary["revision_control"] == "2026 T2"
    assert "service: 020" in summary["accounting"]
    assert "operation: 98004" in summary["accounting"]
    assert "nature: 6156 - Maintenance" in summary["accounting"]
    assert summary["issue"] == "A completer : 1 nature(s) a completer"



def test_dalkia_sheet_matches_accountant_model_for_p3_invoice() -> None:
    class FakeScalarResult:
        def __init__(self, rows):
            self.rows = rows

        def all(self):
            return self.rows

    class FakeScalarDb:
        def __init__(self, sequences):
            self.sequences = list(sequences)

        def scalars(self, _stmt):
            return FakeScalarResult(self.sequences.pop(0))

    invoice = CpeFinanceInvoice(
        id=10,
        invoice_number="0001E2607QRY8",
        contract_code="C00190116O",
        invoice_date=date(2026, 6, 30),
        total_ht=457.86,
        status="valide",
    )
    worklist = WorklistInvoice(
        row_number=2,
        accounting_number="202605379",
        supplier_invoice_number="0001E2607QRY8",
        label="FAC. 0001E2607QRY8 DU 30/06/2026",
        total_ttc=549.43,
        invoice_date=date(2026, 6, 30),
        arrival_date=None,
        supplier_code=None,
        supplier_name="DALKIA",
        invoice_status=None,
        liquidation_status=None,
        market_code=None,
        raw={},
    )
    platform = {
        "0001E2607QRY8": PlatformInvoice(
            id=10,
            invoice_number="0001E2607QRY8",
            total_ttc=549.43,
            control_status="valid",
            decision_status="valide",
            problem_summary=None,
            raw=invoice,
        )
    }
    line = CpeFinanceLine(
        id=1,
        invoice_id=10,
        row_number=1,
        contract_code="C00190116O",
        market="P3",
        billed_item="P3",
        service_sold="P3 - GARANTIE TOTALE",
        vat_rate=20,
        amount_ht=457.86,
        base_price=1775.0,
        revised_price=1831.42,
        detail="ENTRETIEN VDS-ENS 19 - CENTRE DE LOISIR LE VALLON-ALSH P3 - GARANTIE TOTALE",
        accounting_site_id=100,
        site_code_detected="VDS-ENS 19",
        accounting_nature="6156",
        accounting_label="Maintenance",
    )
    site = CpeAccountingSiteMapping(
        id=100,
        code_site="VDS-ENS 19",
        site_name="Centre de loisir Le Vallon",
        manager="BATI",
        service_code="XSCO",
        function_code="331",
        antenna_code="ALSH",
    )
    db = FakeScalarDb([[line], [site]])
    workbook = openpyxl.Workbook()
    ws = workbook.active
    parsed = comptable_report.WorklistParseResult(sheet_name="_ShowList-001", rows=[worklist])

    comptable_report._write_market_sheet(db, 303, ws, comptable_report.MARKETS[0], parsed, platform)

    # Layout TTC additif au niveau ligne : BASE + REVISION = MONTANT, LC en col 10,
    # plus de colonnes HT / valeur de base / forfait revise trimestriel.
    assert ws.cell(row=4, column=1).value == "CODE CONTRAT"
    assert ws.cell(row=4, column=6).value == "MONTANT BASE TTC"
    assert ws.cell(row=4, column=7).value == "MONTANT REVISION TTC"
    assert ws.cell(row=4, column=8).value == "MONTANT TTC"
    assert ws.cell(row=4, column=10).value == "LC"
    assert ws.cell(row=5, column=4).value == "P3"
    assert ws.cell(row=5, column=5).value == 20
    # dont revision = 457.86 x (1831.42-1775)/1831.42 = 14.11 HT -> 16.93 TTC
    # montant facture (au prix revise) = 457.86 x 1.20 = 549.43 ; base = 549.43 - 16.93
    assert ws.cell(row=5, column=6).value == 532.5   # base
    assert ws.cell(row=5, column=7).value == 16.93   # revision
    assert ws.cell(row=5, column=8).value == 549.43  # montant = base + revision
    assert round(ws.cell(row=5, column=6).value + ws.cell(row=5, column=7).value, 2) == ws.cell(row=5, column=8).value
    # LC avec fonction (331) : gestionnaire-fonction-nature-operation(P3)-service-antenne.
    assert ws.cell(row=5, column=10).value == "BATI-331-21351-98003-XSCO-ALSH"
    assert ws.cell(row=5, column=13).value is None


def test_dalkia_p1_gaz_uses_os3_variable_base_and_positive_revision() -> None:
    class FakeScalarResult:
        def __init__(self, rows):
            self.rows = rows

        def all(self):
            return self.rows

    class FakeScalarDb:
        def __init__(self, sequences):
            self.sequences = list(sequences)

        def scalars(self, _stmt):
            return FakeScalarResult(self.sequences.pop(0))

    invoice = CpeFinanceInvoice(
        id=12,
        invoice_number="0001E2607QRY6",
        contract_code="C00190116O",
        invoice_date=date(2026, 6, 30),
        total_ht=835.37,
        status="valide",
    )
    worklist = WorklistInvoice(
        row_number=2,
        accounting_number="202605378",
        supplier_invoice_number="0001E2607QRY6",
        label="FAC. 0001E2607QRY6 DU 30/06/2026",
        total_ttc=1002.44,
        invoice_date=date(2026, 6, 30),
        arrival_date=None,
        supplier_code=None,
        supplier_name="DALKIA",
        invoice_status=None,
        liquidation_status=None,
        market_code=None,
        raw={},
    )
    platform = {
        "0001E2607QRY6": PlatformInvoice(
            id=12,
            invoice_number="0001E2607QRY6",
            total_ttc=1002.44,
            control_status="valid",
            decision_status="valide",
            problem_summary=None,
            raw=invoice,
        )
    }
    line = CpeFinanceLine(
        id=12,
        invoice_id=12,
        row_number=1,
        contract_code="C00190116O",
        market="P1",
        billed_item="P1",
        service_sold="CHAUFFAGE",
        vat_rate=20,
        amount_ht=835.37,
        base_price=74.17,
        revised_price=3341.49,
        detail="VDS-ENS 19 - CENTRE DE LOISIR LE VALLON",
        accounting_site_id=102,
        site_code_detected="VDS-ENS 19",
        period_start=date(2026, 4, 1),
        period_end=date(2026, 6, 30),
        accounting_nature="60621",
        accounting_label="Chauffage",
    )
    site = CpeAccountingSiteMapping(
        id=102,
        code_site="VDS-ENS 19",
        site_name="Centre de loisir Le Vallon",
        manager="BATI",
        service_code="XSCO",
        function_code="331",
        antenna_code="ALSH",
    )
    os3_reference = CpeDpgfP1Line(
        id=2,
        import_id=2,
        city_id=303,
        lot=1,
        level="rev_temp_prix",
        code_site="VDS-ENS 19",
        period_idx=2,
        period_label="2026",
        period_year=2026,
        prix_unitaire_ht=74.17,
        qt_mwhpcs=42.8,
        p10_var_ht=3174.23,
        p10_total_ht=3842.34,
    )
    db = FakeScalarDb([[line], [site], [os3_reference]])
    ws = openpyxl.Workbook().active
    parsed = comptable_report.WorklistParseResult(sheet_name="_ShowList-001", rows=[worklist])

    comptable_report._write_market_sheet(db, 303, ws, comptable_report.MARKETS[0], parsed, platform)

    assert ws.cell(row=5, column=4).value == "P1"
    # Valeur de base annuelle combustible = part variable OS3 (p10_var_ht) = 3174.23 EUR HT
    # (la part fixe 3842.34 - 3174.23 est facturee sur des lignes REFAC distinctes).
    # Base TTC proratisee = 835.37 x 3174.23 / 3341.49 x 1.20 = 952.27 EUR.
    assert ws.cell(row=5, column=6).value == 952.27
    assert ws.cell(row=5, column=7).value == 50.17
    assert ws.cell(row=5, column=8).value == 1002.44
    assert round(ws.cell(row=5, column=6).value + ws.cell(row=5, column=7).value, 2) == 1002.44
    assert ws.cell(row=5, column=10).value == "BATI-331-60621-XSCO-ALSH"
    assert ws.cell(row=5, column=13).value is None


def test_dalkia_p1_electricity_is_not_revised() -> None:
    line = CpeFinanceLine(
        billed_item="P1.EL", service_sold="ELECTRICITE", amount_ht=2718.36, vat_rate=20
    )

    assert comptable_report._cpe_report_amounts_ttc(line) == (3262.03, 0.0, 3262.03, None)


def test_dalkia_p1_gaz_without_contractual_reference_is_not_invented() -> None:
    line = CpeFinanceLine(
        billed_item="P1", service_sold="CHAUFFAGE", amount_ht=835.37,
        vat_rate=20, base_price=74.17, revised_price=3341.49,
    )

    base, revision, total, issue = comptable_report._cpe_report_amounts_ttc(line)

    assert (base, revision, total) == (None, None, 1002.44)
    assert issue and "Référence P1 gaz OS n°3 absente" in issue


def test_dalkia_p2_line_excludes_operation_from_lc() -> None:
    """Sur une ligne P2 (maintenance 6156), le numero d'operation ne doit pas
    fuir dans la LC, meme si le site porte un operation_code. Reproduit le cas
    remonte par la comptable : `BATI-28-6156-98004-ATBA-CTM` -> sans 98004."""

    class FakeScalarResult:
        def __init__(self, rows):
            self.rows = rows

        def all(self):
            return self.rows

    class FakeScalarDb:
        def __init__(self, sequences):
            self.sequences = list(sequences)

        def scalars(self, _stmt):
            return FakeScalarResult(self.sequences.pop(0))

    invoice = CpeFinanceInvoice(
        id=11, invoice_number="INV-P2", contract_code="C00025812G",
        invoice_date=date(2026, 6, 30), total_ht=100.0, status="valide",
    )
    worklist = WorklistInvoice(
        row_number=2, accounting_number="202605400", supplier_invoice_number="INV-P2",
        label="FAC. INV-P2 DU 30/06/2026", total_ttc=120.0, invoice_date=date(2026, 6, 30),
        arrival_date=None, supplier_code=None, supplier_name="DALKIA",
        invoice_status=None, liquidation_status=None, market_code=None, raw={},
    )
    platform = {
        "INV-P2": PlatformInvoice(
            id=11, invoice_number="INV-P2", total_ttc=120.0, control_status="valid",
            decision_status="valide", problem_summary=None, raw=invoice,
        )
    }
    line = CpeFinanceLine(
        id=1, invoice_id=11, row_number=1, contract_code="C00025812G",
        market="P2", billed_item="P2-11", service_sold="Maintenance",
        vat_rate=20, amount_ht=100.0, base_price=None, revised_price=None,
        detail="ENTRETIEN CTM", accounting_site_id=101, site_code_detected="VDS-STE 01",
        accounting_nature="6156", accounting_label="Maintenance",
    )
    site = CpeAccountingSiteMapping(
        id=101, code_site="VDS-STE 01", site_name="Centre technique municipal",
        manager="BATI", service_code="ATBA", function_code="28",
        antenna_code="CTM", operation_code="98004",
    )
    db = FakeScalarDb([[line], [site]])
    ws = openpyxl.Workbook().active
    parsed = comptable_report.WorklistParseResult(sheet_name="_ShowList-001", rows=[worklist])

    comptable_report._write_market_sheet(db, 303, ws, comptable_report.MARKETS[0], parsed, platform)

    lc = ws.cell(row=5, column=10).value
    # Avec fonction (28), sans operation (98004) : gestionnaire-fonction-nature-service-antenne.
    assert lc == "BATI-28-6156-ATBA-CTM"
    assert "98004" not in lc
    assert "-28-" in lc


def test_report_translates_decisions_and_writes_problem_summary(monkeypatch) -> None:
    monkeypatch.setattr(
        comptable_report,
        "_market_line_enrichments",
        lambda *args: {1: {"issue": None, "revision_control": "PRM-1: BPU 1.0 / TURPE 1.0"}},
    )
    liaison_rows = [
        comptable_report.energie_accounting.LiaisonRow(
            prm_id="PRM-1",
            site_name="Hotel de ville",
            poste="ABONNEMENT",
            label="Abonnement",
            quantity=1,
            unit_price_ht=100.0,
            amount_ht=100.0,
            service_code="ELEC",
            function_code="020",
            antenna_code="ANT",
            operation_code="OP1",
            accounting_nature="60612",
            accounting_label="Electricite",
            status="ok",
        ),
        comptable_report.energie_accounting.LiaisonRow(
            prm_id="PRM-2",
            site_name="Gymnase",
            poste="ACHEMINEMENT",
            label="Acheminement",
            quantity=1,
            unit_price_ht=20.0,
            amount_ht=20.0,
            service_code="SPORT",
            function_code="411",
            antenna_code="GYMN",
            operation_code=None,
            accounting_nature="60612",
            accounting_label="Electricite",
            status="ok",
        ),
    ]
    monkeypatch.setattr(comptable_report.energie_accounting, "resolve_invoice_codification", lambda *_args: liaison_rows)
    workbook = openpyxl.Workbook()
    ws = workbook.active
    parsed = comptable_report.WorklistParseResult(
        sheet_name="_ShowList-001",
        rows=[
            WorklistInvoice(
                row_number=2,
                accounting_number="202600001",
                supplier_invoice_number="F-ENGIE-1",
                label="FAC. F-ENGIE-1 DU 01/07/2026",
                total_ttc=120.0,
                invoice_date="01/07/2026",
                arrival_date=None,
                supplier_code=None,
                supplier_name="ENGIE",
                invoice_status=None,
                liquidation_status=None,
                market_code=None,
                raw={},
            )
        ],
    )
    platform = {
        "F-ENGIE-1": PlatformInvoice(
            id=1,
            invoice_number="F-ENGIE-1",
            total_ttc=120.0,
            control_status="valid (0 erreur(s), 1 alerte(s))",
            decision_status="to_review",
            problem_summary="Ecart prix BPU",
            raw=EnergyInvoiceImport(invoice_number="F-ENGIE-1"),
        )
    }

    comptable_report._write_market_sheet(None, 303, ws, comptable_report.MARKETS[1], parsed, platform)

    # Layout TTC (retour comptable) : plus de MONTANT HT ni de colonne OPERATION.
    assert ws.cell(row=4, column=1).value == "FOURNISSEUR"
    assert ws.cell(row=4, column=5).value == "SITE / POINT"
    assert ws.cell(row=4, column=9).value == "MONTANT TTC"
    assert ws.cell(row=4, column=16).value == "ANTENNE"
    assert ws.cell(row=4, column=17).value == "LC"
    assert ws.cell(row=4, column=18).value == "REVISION / INDICES"
    assert ws.cell(row=4, column=20).value == "POINT A CORRIGER"
    assert ws.cell(row=5, column=1).value == "ENGIE"
    assert ws.cell(row=5, column=2).value == "F-ENGIE-1"
    assert ws.cell(row=5, column=5).value == "Hotel de ville"
    assert ws.cell(row=5, column=6).value == "PRM-1"
    # Ratio TTC facture indisponible ici (facture sans total_ht) -> HT passthrough.
    assert ws.cell(row=5, column=9).value == 100.0
    # LC energie : jamais d'operation d'investissement (reserve DALKIA P3/P3.4),
    # meme si le site porte un operation_code (ici OP1 est volontairement ignore).
    assert ws.cell(row=5, column=17).value == "60612-ELEC-ANT"
    assert ws.cell(row=5, column=18).value == "PRM-1: BPU 1.0 / TURPE 1.0"
    assert ws.cell(row=5, column=19).value == "Conforme (0 erreur(s), 1 alerte(s))"
    assert ws.cell(row=5, column=20).value == "Ecart prix BPU"
    assert ws.cell(row=6, column=2).value == "F-ENGIE-1"
    assert ws.cell(row=6, column=5).value == "Gymnase"
    assert ws.cell(row=6, column=17).value == "60612-SPORT-GYMN"
    assert ws.cell(row=8, column=1).value is None


def test_energy_lines_aggregated_one_row_per_site(monkeypatch) -> None:
    """La comptable veut UNE ligne comptable par site/point : plusieurs postes
    d'un meme PRM doivent etre regroupes (montant somme, imputation commune)."""
    monkeypatch.setattr(comptable_report, "_market_line_enrichments", lambda *args: {})

    def liaison(poste, amount):
        return comptable_report.energie_accounting.LiaisonRow(
            prm_id="PRM-1", site_name="Hotel de ville", poste=poste, label=poste,
            quantity=1, unit_price_ht=amount, amount_ht=amount,
            service_code="ELEC", function_code="020", antenna_code="ANT",
            operation_code=None, accounting_nature="60612", accounting_label="Electricite",
            status="ok",
        )

    lines = [liaison("ABONNEMENT", 100.0), liaison("CONSOMMATION", 50.0), liaison("ACHEMINEMENT", 30.0)]
    monkeypatch.setattr(comptable_report.energie_accounting, "resolve_invoice_codification", lambda *_a: lines)
    ws = openpyxl.Workbook().active
    parsed = comptable_report.WorklistParseResult(
        sheet_name="_ShowList-001",
        rows=[WorklistInvoice(
            row_number=2, accounting_number="202600001", supplier_invoice_number="F-ENGIE-1",
            label="FAC. F-ENGIE-1 DU 01/07/2026", total_ttc=216.0, invoice_date="01/07/2026",
            arrival_date=None, supplier_code=None, supplier_name="ENGIE",
            invoice_status=None, liquidation_status=None, market_code=None, raw={},
        )],
    )
    platform = {
        "F-ENGIE-1": PlatformInvoice(
            id=1, invoice_number="F-ENGIE-1", total_ttc=216.0, control_status="valid",
            decision_status="approved", problem_summary=None, raw=EnergyInvoiceImport(invoice_number="F-ENGIE-1"),
        )
    }

    comptable_report._write_market_sheet(None, 303, ws, comptable_report.MARKETS[1], parsed, platform)

    assert ws.cell(row=4, column=7).value == "POSTES REGROUPES"
    assert ws.cell(row=4, column=8).value == "NB LIGNES"
    # Une seule ligne pour PRM-1 (3 postes regroupes), pas 3.
    assert ws.cell(row=5, column=6).value == "PRM-1"
    assert ws.cell(row=5, column=8).value == 3
    assert ws.cell(row=5, column=9).value == 180.0  # 100 + 50 + 30 (ratio TVA indispo -> HT)
    assert ws.cell(row=5, column=17).value == "60612-ELEC-ANT"
    assert ws.cell(row=6, column=1).value is None  # pas de 2e ligne


def test_report_forces_review_when_chorus_total_differs_from_po2(monkeypatch) -> None:
    monkeypatch.setattr(comptable_report, "_market_line_enrichments", lambda *args: {})
    monkeypatch.setattr(comptable_report.energie_accounting, "resolve_invoice_codification", lambda *_args: [])
    workbook = openpyxl.Workbook()
    ws = workbook.active
    parsed = comptable_report.WorklistParseResult(
        sheet_name="_ShowList-001",
        rows=[
            WorklistInvoice(
                row_number=2,
                accounting_number="202600001",
                supplier_invoice_number="F-ENGIE-1",
                label="FAC. F-ENGIE-1 DU 01/07/2026",
                total_ttc=5359.09,
                invoice_date="01/07/2026",
                arrival_date=None,
                supplier_code=None,
                supplier_name="ENGIE",
                invoice_status=None,
                liquidation_status=None,
                market_code=None,
                raw={},
            )
        ],
    )
    invoice = EnergyInvoiceImport(invoice_number="F-ENGIE-1")
    invoice.normalized_invoice = EnergyInvoice(
        sites=[EnergyInvoiceSite(prm_id=f"PRM-{index}") for index in range(1, 6)]
    )
    platform = {
        "F-ENGIE-1": PlatformInvoice(
            id=1,
            invoice_number="F-ENGIE-1",
            total_ttc=2084.66,
            control_status="valid",
            decision_status="approved",
            problem_summary=None,
            raw=invoice,
        )
    }

    comptable_report._write_market_sheet(None, 303, ws, comptable_report.MARKETS[1], parsed, platform)

    assert ws.cell(row=5, column=1).value == "ENGIE"
    assert ws.cell(row=5, column=12).value == -3274.43
    assert ws.cell(row=5, column=19).value == "\u00c9cart TTC"
    assert ws.cell(row=5, column=20).value == (
        "\u00c9cart TTC Po2 - Chorus : -3274.43 EUR. "
        "Po2 contient 5 site(s) ; 5 PRM ; 2084.66 EUR TTC import\u00e9s ; "
        "Chorus/compta attend 5359.09 EUR TTC. "
        "Il manque probablement une ou plusieurs FIC/sites dans l'export fournisseur import\u00e9. "
        "Comparer avec le PDF fournisseur Chorus."
    )


def test_row_problem_summary_explains_ttc_gap_before_decision() -> None:
    current = PlatformInvoice(
        id=1,
        invoice_number="F-1",
        total_ttc=130.0,
        control_status="valid",
        decision_status="approved",
        problem_summary=None,
        raw=object(),
    )

    assert comptable_report._row_problem_summary("Écart TTC", 10.0, current) == "Écart TTC Po2 - Chorus : 10.00 EUR."


class _FakeExecuteResult:
    def __init__(self, rows):
        self._rows = rows

    def all(self):
        return self._rows


class _FakeDb:
    def __init__(self, rows):
        self._rows = rows

    def execute(self, _stmt):
        return _FakeExecuteResult(self._rows)


def test_cpe_control_summaries_identify_real_dalkia_control_reasons() -> None:
    db = _FakeDb(
        [
            (1, "error", "p2p3_base_dpgf", "Montant divergent"),
            (1, "blocked", "accounting_site", "Site absent"),
            (2, "ok", "invoice_total_ht", "OK"),
        ]
    )

    summaries = comptable_report._cpe_control_summaries(db, [1, 2, 3])

    assert summaries[1]["control_status"] == "error (1 écart(s), 1 bloqué(s))"
    assert "Montant P2/P3 différent du DPGF" in summaries[1]["problem_summary"]
    assert "Site comptable non rattaché" in summaries[1]["problem_summary"]
    assert summaries[2] == {"control_status": "valid", "problem_summary": None}
    assert summaries[3] == {
        "control_status": "not_checked",
        "problem_summary": "Aucun contrôle CPE disponible pour cette facture.",
    }
