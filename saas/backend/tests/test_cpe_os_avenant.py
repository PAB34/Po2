from __future__ import annotations

import io
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.core.db import Base
from app.models.city import City
from app.models.cpe_dalkia import CpeDalkiaRefImport, CpeDalkiaRefP1Gaz, CpeDalkiaRefP2P3, CpeDalkiaRefSite
from app.schemas.cpe_os_avenant import CpeOsAvenantLineCreate, CpeOsAvenantRequestCreate
from app.services.cpe_os_avenant import build_impact_workbook, create_request, list_site_options


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(City(id=1, nom_commune="Sete", code_commune="34301"))
        session.commit()
        yield session


def _seed_reference(db: Session) -> None:
    imp = CpeDalkiaRefImport(city_id=1, lot=1, filename="L1.xlsx", is_active=True)
    db.add(imp)
    db.flush()
    db.add(CpeDalkiaRefSite(import_id=imp.id, city_id=1, lot=1, code_site="VDS-ENS 01", nom_batiment="Ecole test"))
    db.add(
        CpeDalkiaRefP1Gaz(
            import_id=imp.id,
            city_id=1,
            code_site="VDS-ENS 01",
            period_idx=1,
            period_label="2026",
            period_year=2026,
            pce="24300000000000",
            type_tarif="T2",
            p10_total_ht=8000.0,
        )
    )
    db.add(
        CpeDalkiaRefP2P3(
            import_id=imp.id,
            city_id=1,
            code_site="VDS-ENS 01",
            period_idx=1,
            period_label="2026",
            period_year=2026,
            p2_total_ht=2500.0,
            p3_total_ht=4500.0,
        )
    )
    db.add(
        CpeDalkiaRefP1Gaz(
            import_id=imp.id,
            city_id=1,
            code_site="VDS-ENS 01",
            period_idx=2,
            period_label="2027",
            period_year=2027,
            pce="24300000000000",
            type_tarif="T2",
            p10_total_ht=9000.0,
        )
    )
    db.add(
        CpeDalkiaRefP2P3(
            import_id=imp.id,
            city_id=1,
            code_site="VDS-ENS 01",
            period_idx=2,
            period_label="2027",
            period_year=2027,
            p2_total_ht=3000.0,
            p3_total_ht=5000.0,
        )
    )
    db.commit()


def test_site_options_read_active_dpgf_amounts(db_session: Session):
    _seed_reference(db_session)
    options = list_site_options(db_session, 1, year=2026, lot=1)
    assert len(options) == 1
    assert options[0]["code_site"] == "VDS-ENS 01"
    assert options[0]["site_name"] == "Ecole test"
    assert options[0]["p1_gaz_annual_ht"] == pytest.approx(8000.0)
    assert options[0]["total_annual_ht"] == pytest.approx(15000.0)


def test_remove_request_hydrates_current_amounts_and_negative_impact(db_session: Session):
    _seed_reference(db_session)
    payload = CpeOsAvenantRequestCreate(
        title="Sortie Ecole test",
        change_type="remove",
        lot=1,
        effective_date=date(2026, 1, 1),
        lines=[CpeOsAvenantLineCreate(action="remove", code_site="VDS-ENS 01", site_name=None, lot=1)],
    )
    result = create_request(db_session, 1, 42, payload)
    assert result["lines"][0].site_name == "Ecole test"
    assert result["lines"][0].current_p1_gaz_annual_ht == pytest.approx(8000.0)
    assert result["impact"]["total_annual_ht"] == pytest.approx(-15000.0)
    assert result["impact"]["first_year_prorata_ht"] == pytest.approx(-15000.0)


def test_add_request_uses_target_amounts(db_session: Session):
    payload = CpeOsAvenantRequestCreate(
        title="Entree Nouveau site",
        change_type="add",
        effective_date=date(2026, 1, 1),
        lines=[
            CpeOsAvenantLineCreate(
                action="add",
                site_name="Nouveau site",
                p1_gaz_annual_ht=10000.0,
                p2_annual_ht=3000.0,
                p3_annual_ht=7000.0,
            )
        ],
    )
    result = create_request(db_session, 1, 42, payload)
    assert result["impact"]["p1_annual_ht"] == pytest.approx(10000.0)
    assert result["impact"]["total_annual_ht"] == pytest.approx(20000.0)


def test_remove_request_projects_each_budget_year_with_daily_prorata(db_session: Session):
    _seed_reference(db_session)
    payload = CpeOsAvenantRequestCreate(
        title="Sortie Ecole test mi-annee",
        change_type="remove",
        lot=1,
        effective_date=date(2026, 7, 1),
        lines=[CpeOsAvenantLineCreate(action="remove", code_site="VDS-ENS 01", site_name=None, lot=1)],
    )
    result = create_request(db_session, 1, 42, payload)
    annual = result["impact"]["annual_impacts"]
    assert annual[0]["year"] == 2026
    assert annual[0]["ratio"] == pytest.approx(184 / 365)
    assert annual[0]["total_ht"] == pytest.approx(round(-15000.0 * (184 / 365), 2))
    assert annual[1]["year"] == 2027
    assert annual[1]["total_ht"] == pytest.approx(-17000.0)
    assert result["impact"]["first_year_prorata_ht"] == annual[0]["total_ht"]

def test_impact_workbook_exports_summary_lines_and_projection(db_session: Session):
    _seed_reference(db_session)
    payload = CpeOsAvenantRequestCreate(
        title="Sortie Ecole test",
        change_type="remove",
        lot=1,
        effective_date=date(2026, 7, 1),
        lines=[CpeOsAvenantLineCreate(action="remove", code_site="VDS-ENS 01", site_name=None, lot=1)],
    )
    created = create_request(db_session, 1, 42, payload)
    exported = build_impact_workbook(db_session, 1, created["id"])
    assert exported is not None
    content, filename = exported
    assert filename == f"impact-os-avenant-{created['id']}.xlsx"
    wb = load_workbook(io.BytesIO(content), data_only=True)
    assert wb.sheetnames == ["Synthese", "Lignes", "Projection"]
    assert wb["Synthese"]["A1"].value == "Dossier OS / avenant CPE DALKIA"
    assert wb["Lignes"]["A2"].value == "remove"
    assert wb["Projection"]["A2"].value == 2026
