"""Audit des carences d'inventaire CVC.

Le point à verrouiller : la distinction entre un champ **non livré par le format**
d'export du prestataire (la colonne n'existe pas → faire évoluer l'export) et un
champ **livré mais non renseigné** (→ compléter les lignes). Les confondre
décrédibiliserait la demande adressée au titulaire.
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.core.db import Base
from app.models.building import Building  # noqa: F401 (enregistre la table)
from app.models.city import City
from app.models.cvc import CvcInventoryItem
from app.models.equipment import EquipmentReference  # noqa: F401
from app.models.local import Local  # noqa: F401
from app.models.site import Site  # noqa: F401
from app.services.cvc import (
    PROVIDER_DALKIA,
    PROVIDER_SPIE,
    build_carences_workbook,
    get_cvc_carences,
)


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(City(id=1, nom_commune="Sete", code_commune="34301"))
        session.commit()
        yield session


def _add(db: Session, provider: str, batch: str = "b1", **champs) -> CvcInventoryItem:
    item = CvcInventoryItem(
        city_id=1,
        provider=provider,
        designation=champs.pop("designation", "Chaudiere gaz"),
        famille=champs.pop("famille", "Chaudiere"),
        import_batch=batch,
        **champs,
    )
    db.add(item)
    db.flush()
    return item


def _champ(entries, nom):
    return next((e for e in entries if e.champ == nom), None)


def test_champ_absent_du_format_dalkia_est_signale_comme_non_livre(db_session):
    # Le format DALKIA ne comporte pas de n° de série ni de puissance.
    _add(db_session, PROVIDER_DALKIA, marque="Viessmann", modele="V200")
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)
    dalkia = report.providers[0]

    non_livres = {e.champ for e in dalkia.champs_non_livres}
    assert "numero_serie" in non_livres
    assert "puissance" in non_livres
    # Un champ non livré manque par construction sur tout le parc du prestataire.
    assert _champ(dalkia.champs_non_livres, "numero_serie").manquants_pct == 100.0
    # ... et il n'est pas compté comme « à compléter ligne à ligne ».
    assert "numero_serie" not in {e.champ for e in dalkia.champs_incomplets}


def test_champ_livre_mais_vide_est_signale_comme_incomplet(db_session):
    _add(db_session, PROVIDER_DALKIA, marque="Viessmann", date_mis_en_service=2015)
    _add(db_session, PROVIDER_DALKIA, marque=None, date_mis_en_service=None)
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)
    dalkia = report.providers[0]

    marque = _champ(dalkia.champs_incomplets, "marque")
    assert marque is not None
    assert marque.livre_par_format is True
    assert marque.manquants == 1
    assert marque.manquants_pct == 50.0


def test_spie_livre_le_numero_de_serie_contrairement_a_dalkia(db_session):
    _add(db_session, PROVIDER_SPIE, batch="s1", numero_serie=None)
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)
    spie = next(p for p in report.providers if p.provider == PROVIDER_SPIE)

    # Colonne présente dans l'export SPIE : c'est une carence de remplissage.
    assert "numero_serie" not in {e.champ for e in spie.champs_non_livres}
    assert "numero_serie" in {e.champ for e in spie.champs_incomplets}


def test_rattachement_est_hors_demande_titulaire(db_session):
    """Le rattachement vient de NOTRE rapprochement, pas du prestataire."""
    _add(db_session, PROVIDER_DALKIA, building_id=None)
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)

    assert report.rattachement_manquant == 1
    assert report.rattachement_total == 1
    tous_les_champs = {
        e.champ for p in report.providers for e in (p.champs_non_livres + p.champs_incomplets)
    }
    assert "building_id" not in tous_les_champs


def test_equipements_incomplets_comptes_une_seule_fois(db_session):
    # Un équipement à qui il manque 2 champs ne compte que pour 1 équipement.
    _add(db_session, PROVIDER_DALKIA, marque=None, modele=None, date_mis_en_service=None)
    _add(db_session, PROVIDER_DALKIA, marque="X", modele="Y", date_mis_en_service=2020, niveau="RDC", local_name="Chaufferie")
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)

    assert report.providers[0].equipements_incomplets == 1


def test_seul_le_lot_courant_est_audite(db_session):
    _add(db_session, PROVIDER_DALKIA, batch="ancien", marque=None)
    _add(db_session, PROVIDER_DALKIA, batch="recent", marque="Viessmann")
    db_session.commit()

    report = get_cvc_carences(db_session, city_id=1)

    assert report.providers[0].equipements == 1


def test_export_liste_uniquement_les_equipements_incomplets(db_session):
    complet = dict(
        marque="Viessmann",
        modele="V200",
        numero_serie="SN-1",
        date_mis_en_service=2020,
        niveau="RDC",
        local_name="Chaufferie",
        puissance="24 kW",
        capacite=1.0,
    )
    _add(db_session, PROVIDER_SPIE, batch="s1", designation="Complet", **complet)
    _add(db_session, PROVIDER_SPIE, batch="s1", designation="Incomplet", marque="Daikin")
    db_session.commit()

    content = build_carences_workbook(db_session, city_id=1, provider=PROVIDER_SPIE)
    sheet = openpyxl.load_workbook(io.BytesIO(content)).active
    lignes = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(lignes) == 1
    assert "Incomplet" in [str(v) for v in lignes[0]]
    # L'en-tête expose les colonnes exigibles à remplir.
    entetes = [c.value for c in sheet[1]]
    assert "N° de série" in entetes
    assert "Date de mise en service" in entetes


def test_export_prerenseigne_l_identification(db_session):
    batiment = Building(city_id=1, nom_batiment="Ecole Jean Moulin", nom_commune="Sete")
    db_session.add(batiment)
    db_session.flush()
    _add(
        db_session,
        PROVIDER_SPIE,
        batch="s1",
        designation="Split bureau",
        site_raw="VILLA SALIS",
        building_id=batiment.id,
        quantite_relevee=2,
    )
    db_session.commit()

    content = build_carences_workbook(db_session, city_id=1, provider=PROVIDER_SPIE)
    sheet = openpyxl.load_workbook(io.BytesIO(content)).active
    ligne = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))

    # Le titulaire doit pouvoir retrouver l'équipement sur le terrain.
    assert ligne[0] == "VILLA SALIS"
    assert ligne[1] == "Ecole Jean Moulin"
    assert ligne[2] == "Split bureau"
