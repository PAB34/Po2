"""Tests de l'import DPGF P1 revise (livrable separe DALKIA).

Couvre : parsing generique (offsets de colonnes variables), isolation de la persistance
(ne desactive PAS l'import maitre, ne touche pas l'autre lot), reimport du meme lot, et
exposition additive dans le suivi marche (le prevu P1 reste au niveau contrat).
"""
from __future__ import annotations

import io
from types import SimpleNamespace

import openpyxl
import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from app.core.db import Base
from app.models.city import City
from app.models.cpe import CpeContractReference
from app.models.cpe_dalkia import CpeDalkiaRefImport, CpeDalkiaRefP1Gaz
from app.models.cpe_dpgf_p1 import CpeDpgfP1Import, CpeDpgfP1Line
from app.services.cpe_dpgf_p1 import (
    get_dpgf_p1_levels,
    parse_dpgf_p1_file,
    persist_dpgf_p1_import,
)
from app.services.cpe_market_tracking import build_market_tracking

PERIOD_YEARS = [2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033]

# Nom de feuille -> niveau attendu (accents/casse volontairement varies).
LEVEL_SHEETS = {
    "contrat": "Annexe 6 - P1 GAZ contrat",
    "rev_temp": "Annexe 6 - P1 GAZ Rev Temp",
    "rev_temp_prix": "Annexe 6 - P1GAZ Rev T° & prix",
}


def _build_sheet(ws, site_values: dict[str, dict[int, float]], *, offset: int = 0) -> None:
    """Construit une feuille Annexe 6 plausible.

    ``offset`` decale toutes les colonnes (sauf la structure relative) pour simuler les
    decalages reels entre les 3 feuilles -> verifie que le parsing par libelle est robuste.
    code_site en 'N° PROG', 9 triplets QT/P10/P10-TOTAL ; ligne de periode 2 lignes au-dessus
    des en-tetes.
    """
    base = offset
    prog_col = base + 2
    pce_col = base + 5
    tarif_col = base + 6
    pu_col = base + 7
    first_total = base + 13  # QT=first_total-2, P10=first_total-1, TOTAL=first_total
    total_cols = [first_total + 3 * i for i in range(9)]

    # Ligne de periode (hr-2 = ligne 2) : annee au-dessus de la colonne QT (total-2)
    for i, tc in enumerate(total_cols):
        ws.cell(row=2, column=(tc - 2) + 1, value=str(PERIOD_YEARS[i]))
    # Ligne d'en-tete (hr = ligne 4, 0-indexed 3)
    ws.cell(row=4, column=base + 1, value="LOT")
    ws.cell(row=4, column=prog_col + 1, value="N° PROG")
    ws.cell(row=4, column=pce_col + 1, value="PCE")
    ws.cell(row=4, column=tarif_col + 1, value="TYPE DE TARIF")
    ws.cell(row=4, column=pu_col + 1, value="Prix unitaire HT")
    for tc in total_cols:
        ws.cell(row=4, column=tc + 1, value="P10 - TOTAL")
    # Donnees
    r = 5
    for code, by_year in site_values.items():
        ws.cell(row=r, column=prog_col + 1, value=code)
        ws.cell(row=r, column=pce_col + 1, value="PCE" + code[-2:])
        ws.cell(row=r, column=tarif_col + 1, value="T2")
        ws.cell(row=r, column=pu_col + 1, value=82.13)
        for i, tc in enumerate(total_cols):
            val = by_year.get(PERIOD_YEARS[i])
            if val is not None:
                ws.cell(row=r, column=tc + 1, value=val)
        r += 1


def _build_dpgf_bytes(level_site_values: dict[str, dict[str, dict[int, float]]]) -> bytes:
    """Construit un classeur DPGF P1 avec les 3 feuilles ; offset distinct par feuille."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    offsets = {"contrat": 0, "rev_temp": 1, "rev_temp_prix": 3}
    for level, sheet_name in LEVEL_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name)
        _build_sheet(ws, level_site_values[level], offset=offsets[level])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# Jeu de donnees : 2 sites, valeurs 2026 distinctes par niveau.
SAMPLE = {
    "contrat": {"VDS-ENS 01": {2026: 200000.0}, "VDS-ENS 02": {2026: 117775.0}},
    "rev_temp": {"VDS-ENS 01": {2026: 220000.0}, "VDS-ENS 02": {2026: 132073.0}},
    "rev_temp_prix": {"VDS-ENS 01": {2026: 190000.0}, "VDS-ENS 02": {2026: 122197.0}},
}


@pytest.fixture()
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(City(id=1, nom_commune="Sete", code_commune="34301"))
        # Perimetre Lot 1 + Lot 2 (pour le decoupage par lot du suivi marche)
        session.add_all([
            CpeContractReference(
                city_id=1, contract_code="C00190116O", contract_label="LOT 1",
                reference_kind="cpe_contract_scope", year=2026, market="SCOPE",
                billed_item="CPE_VILLE_LOT_1", active=True,
            ),
            CpeContractReference(
                city_id=1, contract_code="C00190155J", contract_label="LOT 2",
                reference_kind="cpe_contract_scope", year=2026, market="SCOPE",
                billed_item="CPE_VILLE_LOT_2", active=True,
            ),
        ])
        session.commit()
        yield session


@pytest.fixture()
def user():
    return SimpleNamespace(city_id=1)


def _seed_master_import(db: Session, *, lot: int, p1_2026: float) -> CpeDalkiaRefImport:
    """Import MAITRE (cpe_dalkia_ref_*) — celui qui ne doit jamais etre desactive par le DPGF."""
    imp = CpeDalkiaRefImport(city_id=1, lot=lot, filename=f"MAITRE_L{lot}.xlsx", is_active=True)
    db.add(imp)
    db.flush()
    db.add(CpeDalkiaRefP1Gaz(
        import_id=imp.id, city_id=1, code_site=f"S{lot}", period_idx=2,
        period_label="2026", period_year=2026, p10_total_ht=p1_2026,
    ))
    db.commit()
    return imp


# ── Parsing ──────────────────────────────────────────────────────────────────


def test_parse_three_levels_totals():
    raw = _build_dpgf_bytes(SAMPLE)
    result = parse_dpgf_p1_file(raw, "P1 - DPGF LOT 1 2026V2.xlsx", lot=1)
    assert result.totals["contrat"][2026] == 317775.0
    assert result.totals["rev_temp"][2026] == 352073.0
    assert result.totals["rev_temp_prix"][2026] == 312197.0
    # 3 niveaux x 2 sites x 9 periodes
    assert len(result.lines) == 3 * 2 * 9
    assert result.nb_sites == {"contrat": 2, "rev_temp": 2, "rev_temp_prix": 2}
    # un site porte bien son tarif/pce (colonnes reperees par libelle malgre l'offset)
    a_line = next(ln for ln in result.lines if ln.level == "rev_temp_prix" and ln.period_year == 2026)
    assert a_line.type_tarif == "T2"


def test_parse_rejects_non_dpgf():
    wb = openpyxl.Workbook()
    wb.active.title = "synthèse"
    out = io.BytesIO()
    wb.save(out)
    with pytest.raises(ValueError):
        parse_dpgf_p1_file(out.getvalue(), "autre.xlsx", lot=1)


# ── Persistance isolee ──────────────────────────────────────────────────────


def test_persist_does_not_deactivate_master(db_session, user):
    master = _seed_master_import(db_session, lot=1, p1_2026=317775.0)
    raw = _build_dpgf_bytes(SAMPLE)
    result = parse_dpgf_p1_file(raw, "DPGF_L1.xlsx", lot=1)
    persist_dpgf_p1_import(db_session, result, user)

    # L'import maitre Lot 1 reste actif, sa donnee P1 intacte.
    db_session.refresh(master)
    assert master.is_active is True
    p1_master = db_session.scalars(
        select(CpeDalkiaRefP1Gaz).where(CpeDalkiaRefP1Gaz.import_id == master.id)
    ).all()
    assert len(p1_master) == 1 and p1_master[0].p10_total_ht == 317775.0
    # Le DPGF P1 a bien ecrit dans sa lignee propre.
    dpgf = db_session.scalars(select(CpeDpgfP1Import).where(CpeDpgfP1Import.is_active.is_(True))).all()
    assert len(dpgf) == 1 and dpgf[0].lot == 1


def test_persist_lot1_does_not_touch_lot2(db_session, user):
    raw2 = _build_dpgf_bytes(SAMPLE)
    persist_dpgf_p1_import(db_session, parse_dpgf_p1_file(raw2, "DPGF_L2.xlsx", lot=2), user)
    raw1 = _build_dpgf_bytes(SAMPLE)
    persist_dpgf_p1_import(db_session, parse_dpgf_p1_file(raw1, "DPGF_L1.xlsx", lot=1), user)

    active = db_session.scalars(select(CpeDpgfP1Import).where(CpeDpgfP1Import.is_active.is_(True))).all()
    assert {b.lot for b in active} == {1, 2}  # les deux lots restent actifs


def test_reimport_same_lot_deactivates_previous_dpgf_only(db_session, user):
    master = _seed_master_import(db_session, lot=1, p1_2026=317775.0)
    first = persist_dpgf_p1_import(db_session, parse_dpgf_p1_file(_build_dpgf_bytes(SAMPLE), "v1.xlsx", lot=1), user)
    second = persist_dpgf_p1_import(db_session, parse_dpgf_p1_file(_build_dpgf_bytes(SAMPLE), "v2.xlsx", lot=1), user)

    db_session.refresh(first)
    db_session.refresh(second)
    db_session.refresh(master)
    assert first.is_active is False  # ancien DPGF desactive
    assert second.is_active is True
    assert master.is_active is True  # maitre toujours intact
    # plus aucune ligne orpheline active pour le v1 dans les lectures
    levels = get_dpgf_p1_levels(db_session, 1, [2026], lot=1)
    assert levels["contrat"][2026] == 317775.0  # une seule version active


# ── Exposition additive dans le suivi marche ─────────────────────────────────


def test_market_tracking_exposes_dpgf_without_changing_prevu(db_session, user):
    # prevu P1 au niveau contrat (import maitre)
    _seed_master_import(db_session, lot=1, p1_2026=317775.0)
    persist_dpgf_p1_import(db_session, parse_dpgf_p1_file(_build_dpgf_bytes(SAMPLE), "DPGF_L1.xlsx", lot=1), user)

    report = build_market_tracking(db_session, 1, year_from=2026, year_to=2026)
    prevu_p1 = next(p for p in report["postes"] if p["poste"] == "P1")["by_year"][0]["prevu"]
    # Le prevu P1 reste au contrat (317775), PAS au Rev Temp (352073)
    assert prevu_p1 == 317775.0

    dpgf = report["p1_dpgf"]
    assert dpgf["has_data"] is True
    by_level = {lv["level"]: lv["total"] for lv in dpgf["levels"]}
    assert by_level["contrat"] == 317775.0
    assert by_level["rev_temp"] == 352073.0
    assert by_level["rev_temp_prix"] == 312197.0
    # decoupage par lot : le Lot 1 porte aussi son bloc DPGF
    lot1 = next(e for e in report["by_lot"] if e["lot"] == 1)
    assert {lv["level"]: lv["total"] for lv in lot1["p1_dpgf"]["levels"]}["rev_temp"] == 352073.0
