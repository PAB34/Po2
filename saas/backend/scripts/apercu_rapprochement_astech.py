"""Aperçu du rapprochement ASTECH <-> patrimoine Po2, sans rien écrire en base.

Sert à faire relire la QUALITÉ du rapprochement par un humain avant de construire
l'écran : on rejoue l'import et le moteur de reconnaissance dans une base SQLite
jetable, et on produit un classeur Excel à quatre onglets.

Sources de bâtiments possibles :
  - `--database-url` (ou la variable DATABASE_URL) : lit la table `buildings` ;
  - `--buildings-csv` : fichier `id|nom|adresse` (une ligne par bâtiment), utile
    quand on n'a pas d'accès direct à la base.

Exemples
--------
    python scripts/apercu_rapprochement_astech.py "Liste BATIMENTS TRAITES.xlsx" \
        --database-url postgresql://... -o apercu.xlsx

    python scripts/apercu_rapprochement_astech.py "Liste BATIMENTS TRAITES.xlsx" \
        --buildings-csv batiments_prod.txt -o apercu.xlsx
"""
from __future__ import annotations

import argparse
import pathlib
import sys

# Le script vit dans saas/backend/scripts/ : on ajoute la racine backend au path
# pour pouvoir importer `app.*` sans installation.
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from sqlalchemy import create_engine, select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.core.db import Base  # noqa: E402
from app.models.building import Building  # noqa: E402
from app.models.city import City  # noqa: E402
from app.models.local import Local  # noqa: F401,E402  (enregistre la table)
from app.models.patrimoine_legacy import (  # noqa: E402
    STATUS_LINKED,
    STATUS_OUT_OF_SCOPE,
    STATUS_TODO,
    PatrimoineLegacyAsset,
)
from app.models.site import Site  # noqa: F401,E402
from app.services.patrimoine_legacy import (  # noqa: E402
    CANDIDATE_MIN_SCORE,
    compute_candidates,
    import_astech_file,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _load_buildings_from_db(session: Session, database_url: str) -> int:
    source_engine = create_engine(database_url)
    with Session(source_engine) as source:
        rows = source.execute(
            select(Building.id, Building.nom_batiment, Building.adresse_reconstituee)
        ).all()
    count = 0
    for building_id, name, address in rows:
        if not name:
            continue
        session.add(
            Building(
                id=building_id, city_id=1, nom_batiment=name,
                nom_commune="Sete", adresse_reconstituee=address,
            )
        )
        count += 1
    session.commit()
    return count


def _load_buildings_from_csv(session: Session, path: pathlib.Path) -> int:
    count = 0
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        parts = line.split("|") if "|" in line else line.split("~")
        if len(parts) < 2 or not parts[0].strip().isdigit() or not parts[1].strip():
            continue
        session.add(
            Building(
                id=int(parts[0]), city_id=1, nom_batiment=parts[1].strip(),
                nom_commune="Sete",
                adresse_reconstituee=(parts[2].strip() or None) if len(parts) > 2 else None,
            )
        )
        count += 1
    session.commit()
    return count


def _sheet(workbook, title: str, headers: list[str], rows: list[list]):
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        worksheet.append(row)
    widths = [16, 46, 30, 46, 12, 30, 26]
    for index, width in enumerate(widths[: len(headers)], start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width
    worksheet.freeze_panes = "A2"
    return worksheet


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("fichier_astech", help="Classeur exporté d'ASTECH (.xlsx)")
    parser.add_argument("--database-url", default=None, help="Base Po2 où lire les bâtiments")
    parser.add_argument("--buildings-csv", default=None, help="Fichier id|nom|adresse")
    parser.add_argument("-o", "--sortie", default="apercu_rapprochement_astech.xlsx")
    parser.add_argument("--genres", default="BATI", help="Genres ASTECH importés (vide = tous)")
    parser.add_argument("--inclure-hors-parc", action="store_true")
    args = parser.parse_args()

    source_path = pathlib.Path(args.fichier_astech)
    if not source_path.exists():
        print(f"Fichier introuvable : {source_path}")
        return 1

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        session.add(City(id=1, nom_commune="Sete", code_commune="34301"))
        session.commit()

        if args.buildings_csv:
            loaded = _load_buildings_from_csv(session, pathlib.Path(args.buildings_csv))
        elif args.database_url:
            loaded = _load_buildings_from_db(session, args.database_url)
        else:
            print("Indiquer --database-url ou --buildings-csv.")
            return 1
        print(f"Bâtiments Po2 chargés : {loaded}")

        genres = tuple(g.strip().upper() for g in args.genres.split(",") if g.strip())
        result = import_astech_file(
            session, city_id=1, filename=source_path.name,
            raw_bytes=source_path.read_bytes(),
            genres=genres, include_out_of_park=args.inclure_hors_parc,
        )
        print(
            f"Feuille lue : {result['sheet_name']} ({result['columns']} colonnes, "
            f"en-têtes ligne {result['header_row']})"
        )
        print(
            f"{result['total_rows']} lignes lues -> {result['created']} biens importés, "
            f"{result['skipped_scope']} écartés (genre / sortis du parc)"
        )

        candidates = compute_candidates(session, 1, auto_link=True)
        print(
            f"Rapprochement : {candidates['auto_linked']} rattachés automatiquement, "
            f"{candidates['proposed']} avec un candidat proposé"
        )

        names = {
            building.id: building.nom_batiment
            for building in session.scalars(select(Building))
        }
        assets = list(session.scalars(select(PatrimoineLegacyAsset)))

        def line(asset: PatrimoineLegacyAsset, target: str | None) -> list:
            return [
                asset.code_bien, asset.designation, asset.nomcourt, target,
                asset.candidate_score, asset.candidate_reason, asset.source_libelvoie,
            ]

        auto = [
            line(a, names.get(a.building_id))
            for a in assets if a.status == STATUS_LINKED
        ]
        to_review = [
            line(a, a.candidate_label)
            for a in assets
            if a.status == STATUS_TODO and (a.candidate_score or 0) >= CANDIDATE_MIN_SCORE
        ]
        orphans = [
            line(a, None)
            for a in assets
            if a.status == STATUS_TODO and (a.candidate_score or 0) < CANDIDATE_MIN_SCORE
        ]
        out_of_scope = [line(a, None) for a in assets if a.status == STATUS_OUT_OF_SCOPE]

        for rows in (auto, to_review, orphans):
            rows.sort(key=lambda r: (-(r[4] or 0), r[0]))

        headers = [
            "Code bien", "Désignation ASTECH", "Nom court",
            "Bâtiment Po2 proposé", "Score", "Motif", "Voie ASTECH",
        ]
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Synthèse")
        for row in [
            ["Aperçu du rapprochement ASTECH <-> Po2"],
            [],
            ["Fichier source", source_path.name],
            ["Feuille lue", result["sheet_name"]],
            ["Colonnes conservées", result["columns"]],
            [],
            ["Lignes lues", result["total_rows"]],
            ["Biens importés (bâtiments en service)", result["created"]],
            ["Écartés (autre genre / sortis du parc)", result["skipped_scope"]],
            [],
            ["Rattachés automatiquement", len(auto)],
            ["À valider (candidat proposé)", len(to_review)],
            ["Sans candidat (à localiser sur la carte)", len(orphans)],
            ["Hors périmètre (hors Sète)", len(out_of_scope)],
        ]:
            summary.append(row)
        summary["A1"].font = Font(bold=True, size=14)
        summary.column_dimensions["A"].width = 42
        summary.column_dimensions["B"].width = 18

        _sheet(workbook, "1 - Rattachés auto", headers, auto)
        _sheet(workbook, "2 - À valider", headers, to_review)
        _sheet(workbook, "3 - Sans candidat", headers, orphans)
        _sheet(workbook, "4 - Hors périmètre", headers, out_of_scope)

        workbook.save(args.sortie)
        print(f"\nRapport écrit : {args.sortie}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
