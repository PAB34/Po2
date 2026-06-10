"""Parser de l'export XLSX ENGIE « Mes Factures ».

L'export ENGIE est un fichier multi-feuilles (Gaz, C2 sur index, C2 sur courbe,
C3, C4, C5) où chaque ligne représente UN site (FIC) pour UNE période de
facturation. Les factures globales (bordereaux) sont identifiées par le numéro
« N° FMC/FUM/Bordereau » et regroupent plusieurs lignes (= plusieurs sites).

Ce parser :
- Lit les headers en ligne 13 (les 12 premières lignes contiennent un cartouche)
- Pour chaque ligne de données (à partir de la ligne 14) : convertit en dict
  facture-site avec champs Po2 normalisés
- Regroupe les sites par bordereau et retourne une liste de dicts `parsed`
  compatibles avec le format produit par `parse_engie_pdf` → l'analyse aval
  est identique (BPU, TURPE, taxes, périodes, etc.)

Mapping confirmé sur un export 144 bordereaux / 1069 sites / 3 segments.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Feuilles ENGIE à traiter (le gaz est exclu : il a son propre BPU lot 7 et le
# module CPE DALKIA s'en occupe via un autre pipeline).
SHEETS_TO_PARSE: tuple[str, ...] = (
    "C2 - sur index",
    "C2 - sur courbe de charge",
    "C3",
    "C4",
    "C5",
)

# Ligne d'en-tête (1-indexed)
HEADER_ROW = 13
# Première ligne de données
DATA_START_ROW = 14

# Postes d'énergie reconnus par l'export ENGIE (selon la feuille)
ENERGY_POSTES = ("BASE", "HP", "HC", "HPH", "HCH", "HPB", "HCB")

# Conversion ENGIE → convention BPU/PDF utilisée par le moteur d'analyse.
# Note : HPB/HCB (basse saison ENGIE) ↔ hpe/hce (été dans la convention BPU).
POSTE_XLSX_TO_BPU: dict[str, str] = {
    "BASE": "base",
    "HP": "hp",
    "HC": "hc",
    "HPH": "hph",
    "HCH": "hch",
    "HPB": "hpe",
    "HCB": "hce",
}

# Mapping clé Po2 → libellé exact dans le header XLSX ENGIE
COLUMN_NAMES = {
    # Identification document
    "fic_number": "N° Facture ou Avoir",
    "document_type": "Type de document",
    "annulled_document": "N° Document annulé",
    "invoice_number": "N° FMC/FUM/Bordereau",
    "compte_contrat": "Compte de contrat",
    "compte_contrat_collectif": "Compte de contrat collectif",
    "regroupement": "Libellé du CCC",
    "invoice_date": "Date d'édition",
    "due_date": "Date d'échéance",
    # Payeur
    "contract_holder": "Raison Sociale Payeur",
    "payer_address": "Adresse Payeur",
    "payer_postcode": "Code Postal Payeur",
    "payer_city": "Localité Payeur",
    # Site
    "delivery_site_name": "Désignation Site",
    "delivery_address": "Adresse  Site",
    "delivery_postcode": "Code Postal Site",
    "delivery_city": "Localité Site",
    "installation": "Installation",
    "prm_id": "PCE/PDL",
    "frequency": "Fréquence de relève",
    "offer_label": "Libellé Offre",
    "meter_number": "Matricule ou numero compteur",
    "segment": "Segment distributeur",
    "tariff_option_label": "Tarif d'acheminement",
    "tariff_code": "Version d'Utilisation",
    "renewable_pct": "Electricité d'origine renouvelable (%)",
    "termination_date": "Date de la résiliation",
    # Période
    "period_start": "Date de début de période de consommation",
    "period_end": "Date de fin de période de consommation",
    "period_days": "Nombre de jours entre le début et la fin de la période facturée",
    # Énergie origine renouvelable
    "renewable_kwh": "Electricité d'origine renouvelable (kWh)",
    "renewable_unit_price": "Prix unitaire HT Electricité d'origine renouvelable (€)",
    "renewable_amount": "Montant Electricité d'origine renouvelable (€)",
    # Capacité
    "capacity_total": "Montant obligation Capacité (€)",
    # Acheminement
    "delivery_fixed_part": "Total part fixe acheminement (€)",
    "delivery_variable_total": "Composante de Soutirage - total part variable (€)",
    "delivery_techn_total": "Total Prestations Techniques distributeur (€)",
    "delivery_variable_full": "Montant total part variable acheminement (€)",
    "cg_amount": "Composante de gestion  (€)",
    "cc_amount": "Composantes de comptage (€)",
    "soutirage_fixed": "Composante de Soutirage part fixe (€)",
    # Taxes
    "cspe_total": "Total CSPE (€)",
    "ticfe_total": "Total TICFE (€)",
    "cta_elec_total": "Total CTA Elec (€)",
    "tax_communal": "Montant taxes communales",
    "tax_departemental": "Montant taxes départementales",
    # TVA
    "vat_reduced": "Montant Total TVA taux réduit (€)",
    "vat_normal": "Montant Total TVA Normal (€)",
    "vat_zero": "Montant Total TVA taux zéro (€)",
    # Sous-totaux / totaux
    "services_total": "Total Services et prestations (€)",
    "delivery_total_grand": "Total Acheminement (€)",
    "amount_excl_vat": "Montant total HTVA (€)",
    "supply_total": "Total fourniture HTT (€)",
    "amount_htt": "Montant total HTT (€)",
    "vat_total": "Montant Total TVA (€)",
    "total_ttc": "Montant TTC (€)",
    "tax_total": "Total Taxes et contributions (€)",
    # Puissances
    "subscribed_power": "Puissances",  # parfois "PUISSANCES" en section header — voir parser
    # Statut / index
    "reading_status": "Statut relève",
    "total_consumption": "Consommation (kWh)",
    # Frais
    "discount_amount": "Montant Remises/Promotions (€)",
    "interest_diff": "Intérêt pour Paiement Différé (€)",
    "interest_late": "Intérêt pour retard de Paiement (€)",
    "financial_fees": "Montant Frais Financiers (€)",
    "expert_advice": "Conseil Expert Energies (€)",
    "location_post_amount": "Location de poste Montant HT (€)",
    "location_post_qty": "Location de poste Qté",
    # Termination
    "is_terminated": "Contrat résilié",
}

# Modalité spécifique au gaz (réservée pour V2)
# COLUMN_NAMES_GAZ = { ... }


@dataclass(frozen=True)
class ColumnIndex:
    """Index colonne 1-based pour une feuille donnée."""

    sheet: str
    column_by_key: dict[str, int]
    column_by_name: dict[str, int]


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("  ", " ")


def _build_column_index(ws: Worksheet) -> ColumnIndex:
    """Localise chaque colonne attendue par son libellé exact dans la ligne 13."""
    name_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw is None and HEADER_ROW > 1:
            raw = ws.cell(row=HEADER_ROW - 1, column=col).value
        if raw is None:
            continue
        name_to_col[_normalize_header(raw)] = col

    key_to_col: dict[str, int] = {}
    for key, expected_name in COLUMN_NAMES.items():
        normalized = _normalize_header(expected_name)
        if normalized in name_to_col:
            key_to_col[key] = name_to_col[normalized]
    return ColumnIndex(sheet=ws.title, column_by_key=key_to_col, column_by_name=name_to_col)


# ── Helpers de conversion (les cellules ENGIE mélangent str/float/None) ───────


def _coerce_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        cleaned = value.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except Exception:
            return None
    return None


def _coerce_float(value: Any) -> float | None:
    dec = _coerce_decimal(value)
    return float(dec) if dec is not None else None


def _coerce_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        try:
            return int(float(s.replace(",", ".")))
        except Exception:
            return None
    return None


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _coerce_date_iso(value: Any) -> str | None:
    """Retourne une date au format YYYY-MM-DD ou None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        # Format ENGIE typique : DD/MM/YYYY
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _cell(ws: Worksheet, row: int, idx: ColumnIndex, key: str) -> Any:
    """Lit la cellule (row, colonne associée à `key` dans le mapping) ou None."""
    col = idx.column_by_key.get(key)
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def _cell_by_name(ws: Worksheet, row: int, idx: ColumnIndex, name: str) -> Any:
    target = _normalize_header(name)
    col = idx.column_by_name.get(target)
    if col is None:
        target_lower = target.lower()
        for header, candidate_col in idx.column_by_name.items():
            if header.lower() == target_lower:
                col = candidate_col
                break
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def _max_float_by_header_prefix(ws: Worksheet, row: int, idx: ColumnIndex, prefix: str) -> float | None:
    values: list[float] = []
    normalized_prefix = _normalize_header(prefix).lower()
    for header, col in idx.column_by_name.items():
        if not header.lower().startswith(normalized_prefix):
            continue
        value = _coerce_float(ws.cell(row=row, column=col).value)
        if value is not None:
            values.append(value)
    return max(values) if values else None


# ── Reconstitution des invoice_lines (énergie + capacité + acheminement) ──────


def _build_energy_lines(
    ws: Worksheet,
    row: int,
    idx: ColumnIndex,
    period_start_iso: str | None,
    period_end_iso: str | None,
) -> list[dict[str, Any]]:
    """Pour chaque poste énergie connu, génère une ligne facture si la
    consommation ou le montant facturé est renseigné.

    Le `normalized_component` est cohérent avec celui produit par le parser PDF
    (cf. invoice_normalization), pour réutilisation du moteur de contrôle BPU.
    """
    lines: list[dict[str, Any]] = []
    for poste in ENERGY_POSTES:
        qty_name = f"Consommation {poste} (kWh)"
        amount_name = f"Montant facturé {poste} (€)"
        # Le prix unitaire suit la convention ENGIE : Base/HP/HC sans S, HPH/HCH/HPB/HCB → HPHS/HCHS/HPBS/HCBS
        price_suffix = poste if poste in {"BASE", "HP", "HC"} else f"{poste}S"
        price_name = f"Prix unitaire {price_suffix} (€)"

        quantity = _coerce_float(_cell_by_name(ws, row, idx, qty_name))
        amount = _coerce_float(_cell_by_name(ws, row, idx, amount_name))
        unit_price = _coerce_float(_cell_by_name(ws, row, idx, price_name))

        if not any([quantity, amount, unit_price]):
            continue

        poste_bpu = POSTE_XLSX_TO_BPU.get(poste, poste.lower())
        lines.append(
            {
                "family": "electricity",
                "label": f"Consommation {poste}",
                "normalized_component": "supply",
                "poste": poste_bpu,
                "period_start": period_start_iso,
                "period_end": period_end_iso,
                "quantity": quantity,
                "quantity_unit": "kWh",
                "unit_price_ht": unit_price,
                "unit_price_unit": "EUR/kWh",
                "amount_ht": amount,
                "vat_rate": None,
                "raw_line": f"XLSX:supply:{poste}",
            }
        )
    return lines


def _build_capacity_lines(
    ws: Worksheet,
    row: int,
    idx: ColumnIndex,
    period_start_iso: str | None,
    period_end_iso: str | None,
) -> list[dict[str, Any]]:
    """Lignes de capacité (CEE, garantie capacité) par poste."""
    lines: list[dict[str, Any]] = []
    for poste in ENERGY_POSTES:
        poste_lower = poste.lower()
        amount_name = f"Montant capacité {poste} (€)" if poste != "BASE" else "Montant capacité Base (€)"
        qty_name = f"Quantité capacité {poste} (kWh)" if poste != "BASE" else "Quantité capacité Base (kWh)"
        price_suffix = poste if poste in {"BASE", "HP", "HC"} else f"{poste}S"
        price_name = (
            f"Prix capacité {price_suffix} (€)" if poste != "BASE" else "Prix capacité base (€)"
        )
        # ENGIE met les suffixes en minuscule pour Base mais en majuscule HPHS/HCHS...
        # On essaye les deux casses :
        if _cell_by_name(ws, row, idx, amount_name) is None and poste == "BASE":
            amount_name = "Montant capacité Base (€)"

        amount = _coerce_float(_cell_by_name(ws, row, idx, amount_name))
        quantity = _coerce_float(_cell_by_name(ws, row, idx, qty_name))
        unit_price = _coerce_float(_cell_by_name(ws, row, idx, price_name))

        if not any([amount, quantity, unit_price]):
            continue

        poste_bpu = POSTE_XLSX_TO_BPU.get(poste, poste.lower())
        lines.append(
            {
                "family": "electricity",
                "label": f"Capacité {poste}",
                "normalized_component": "capacity",
                "poste": poste_bpu,
                "period_start": period_start_iso,
                "period_end": period_end_iso,
                "quantity": quantity,
                "quantity_unit": "kWh",
                "unit_price_ht": unit_price,
                "unit_price_unit": "EUR/kWh",
                "amount_ht": amount,
                "vat_rate": None,
                "raw_line": f"XLSX:capacity:{poste}",
            }
        )
    return lines


def _build_cee_lines(ws: Worksheet, row: int, idx: ColumnIndex) -> list[dict[str, Any]]:
    """Lignes contribution CEE classique + précarité (composantes énergétiques)."""
    lines: list[dict[str, Any]] = []
    # Toutes les sous-catégories CEE remontent dans le même normalized_component "cee"
    # — c'est ce qu'attend le moteur BPU.
    mapping = [
        ("CEE CLASSIQUES Montant HT (€)", "CEE CLASSIQUES Conso. (kWh)", "CEE CLASSIQUES Prix unitaire HT (€)", "CEE Classiques", "classique"),
        ("CEE PRECARITE Montant HT (€)", "CEE PRECARITE Conso. (kWh)", "CEE PRECARITE Prix unitaire HT (€)", "CEE Précarité", "precarite"),
        ("Contribution CEE Montant HT (€)", "Contribution CEE Conso (kWh)", "Contribution CEE Prix unitaire HT  (€)", "Contribution CEE", "contribution"),
    ]
    for amount_name, qty_name, price_name, label, subkind in mapping:
        amount = _coerce_float(_cell_by_name(ws, row, idx, amount_name))
        quantity = _coerce_float(_cell_by_name(ws, row, idx, qty_name))
        unit_price = _coerce_float(_cell_by_name(ws, row, idx, price_name))
        if not any([amount, quantity, unit_price]):
            continue
        lines.append(
            {
                "family": "electricity",
                "label": label,
                "normalized_component": "cee",
                "poste": None,
                "quantity": quantity,
                "quantity_unit": "kWh",
                "unit_price_ht": unit_price,
                "unit_price_unit": "EUR/kWh",
                "amount_ht": amount,
                "vat_rate": None,
                "raw_line": f"XLSX:cee:{subkind}",
            }
        )
    return lines


def _build_delivery_lines(ws: Worksheet, row: int, idx: ColumnIndex) -> list[dict[str, Any]]:
    """Lignes acheminement : composante de gestion (CG), composante de comptage (CC),
    composante de soutirage part fixe, composantes part variable par poste.

    Format compatible avec le moteur TURPE de Po2 (cf. services/turpe.py).
    """
    lines: list[dict[str, Any]] = []

    # Composantes fixes : family="network", composants compatibles avec turpe.py
    cg = _coerce_float(_cell(ws, row, idx, "cg_amount"))
    if cg is not None:
        lines.append(
            {
                "family": "network",
                "label": "Composante de gestion",
                "normalized_component": "network_management",
                "poste": None,
                "amount_ht": cg,
                "raw_line": "XLSX:network:management",
            }
        )
    cc = _coerce_float(_cell(ws, row, idx, "cc_amount"))
    if cc is not None:
        lines.append(
            {
                "family": "network",
                "label": "Composante de comptage",
                "normalized_component": "network_counting",
                "poste": None,
                "amount_ht": cc,
                "raw_line": "XLSX:network:counting",
            }
        )
    sout_fixed = _coerce_float(_cell(ws, row, idx, "soutirage_fixed"))
    if sout_fixed is not None:
        lines.append(
            {
                "family": "network",
                "label": "Composante de soutirage - part fixe",
                "normalized_component": "network_withdrawal",
                "poste": None,
                "amount_ht": sout_fixed,
                "raw_line": "XLSX:network:withdrawal",
            }
        )

    # Soutirage part variable par poste — family="network", normalized_component="network_variable"
    for poste in ENERGY_POSTES:
        if poste in {"BASE", "HP", "HC"}:
            qty_name = f"Composante de Soutirage - part variable {poste} (kWh)"
            price_name = f"Prix composante de Soutirage part variable {poste} (€)"
        else:
            qty_name = f"Composante de Soutirage - part variable {poste}S (kWh)"
            price_name = f"Prix composante de Soutirage part variable {poste}S (€)"
            # Fallback pour HPB/HCB où la colonne quantité utilise parfois "HPB" sans S
            if _cell_by_name(ws, row, idx, qty_name) is None:
                qty_alt = f"Composante de Soutirage - part variable {poste} (kWh)"
                if _cell_by_name(ws, row, idx, qty_alt) is not None:
                    qty_name = qty_alt
        qty = _coerce_float(_cell_by_name(ws, row, idx, qty_name))
        unit_price = _coerce_float(_cell_by_name(ws, row, idx, price_name))
        if not any([qty, unit_price]):
            continue
        amount = qty * unit_price if (qty is not None and unit_price is not None) else None
        poste_bpu = POSTE_XLSX_TO_BPU.get(poste, poste.lower())
        lines.append(
            {
                "family": "network",
                "label": f"Composante de soutirage - part variable {poste}",
                "normalized_component": "network_variable",
                "poste": poste_bpu,
                "quantity": qty,
                "quantity_unit": "kWh",
                "unit_price_ht": unit_price,
                "unit_price_unit": "EUR/kWh",
                "amount_ht": amount,
                "raw_line": f"XLSX:network:variable:{poste}",
            }
        )

    return lines


def _build_tax_lines(ws: Worksheet, row: int, idx: ColumnIndex) -> list[dict[str, Any]]:
    """Lignes taxes : CSPE, TICFE, CTA Elec, taxes communales/départementales."""
    lines: list[dict[str, Any]] = []
    # family="taxes" + composants cspe/cta côté contrôles taxes (cf. _normalized_component PDF)
    for key, label, comp in [
        ("cspe_total", "CSPE", "cspe"),
        ("ticfe_total", "TICFE", "ticfe"),
        ("cta_elec_total", "CTA Elec", "cta"),
        ("tax_communal", "Taxes communales", "tax_communale"),
        ("tax_departemental", "Taxes départementales", "tax_departementale"),
    ]:
        amount = _coerce_float(_cell(ws, row, idx, key))
        if amount is None:
            continue
        lines.append(
            {
                "family": "taxes",
                "label": label,
                "normalized_component": comp,
                "poste": None,
                "amount_ht": amount,
                "raw_line": f"XLSX:taxes:{comp}",
            }
        )
    return lines


def _build_renewable_line(ws: Worksheet, row: int, idx: ColumnIndex) -> list[dict[str, Any]]:
    """Ligne énergie d'origine renouvelable (option garantie d'origine)."""
    amount = _coerce_float(_cell(ws, row, idx, "renewable_amount"))
    qty = _coerce_float(_cell(ws, row, idx, "renewable_kwh"))
    unit_price = _coerce_float(_cell(ws, row, idx, "renewable_unit_price"))
    if not any([amount, qty, unit_price]):
        return []
    return [
        {
            "family": "electricity",
            "label": "Electricité d'origine renouvelable",
            "normalized_component": "green_energy",
            "poste": None,
            "quantity": qty,
            "quantity_unit": "kWh",
            "unit_price_ht": unit_price,
            "unit_price_unit": "EUR/kWh",
            "amount_ht": amount,
            "raw_line": "XLSX:supply:green_energy",
        }
    ]


# ── Extraction d'une ligne XLSX → structure site complète ─────────────────────


def _parse_row(ws: Worksheet, row: int, idx: ColumnIndex) -> dict[str, Any] | None:
    """Convertit une ligne XLSX en dict site avec invoice_lines reconstitués."""
    invoice_number = _coerce_str(_cell(ws, row, idx, "invoice_number"))
    fic_number = _coerce_str(_cell(ws, row, idx, "fic_number"))
    if not invoice_number and not fic_number:
        return None

    period_start = _coerce_date_iso(_cell(ws, row, idx, "period_start"))
    period_end = _coerce_date_iso(_cell(ws, row, idx, "period_end"))

    # Champs site
    site: dict[str, Any] = {
        "fic_number": fic_number,
        "prm_id": _coerce_str(_cell(ws, row, idx, "prm_id")),
        "site_name": _coerce_str(_cell(ws, row, idx, "delivery_site_name")),
        "delivery_site_name": _coerce_str(_cell(ws, row, idx, "delivery_site_name")),
        "delivery_address": _coerce_str(_cell(ws, row, idx, "delivery_address")),
        "delivery_postcode": _coerce_str(_cell(ws, row, idx, "delivery_postcode")),
        "delivery_city": _coerce_str(_cell(ws, row, idx, "delivery_city")),
        "installation": _coerce_str(_cell(ws, row, idx, "installation")),
        "meter_number": _coerce_str(_cell(ws, row, idx, "meter_number")),
        "tariff_option_label": _coerce_str(_cell(ws, row, idx, "tariff_option_label")),
        "tariff_code": _coerce_str(_cell(ws, row, idx, "tariff_code")),
        "segment": _coerce_str(_cell(ws, row, idx, "segment")),
        "regroupement": _coerce_str(_cell(ws, row, idx, "regroupement")),
        "period_start": period_start,
        "period_end": period_end,
        "period_days": _coerce_int(_cell(ws, row, idx, "period_days")),
        "subscribed_power_kva": _max_float_by_header_prefix(ws, row, idx, "Puissance souscrite")
            or _coerce_float(_cell_by_name(ws, row, idx, "Puissances")),
        "max_reached_power_kva": _max_float_by_header_prefix(ws, row, idx, "Puissance atteinte"),
        "total_consumption_kwh": _coerce_float(_cell(ws, row, idx, "total_consumption")),
        "total_ht": _coerce_float(_cell(ws, row, idx, "amount_excl_vat"))
            or _coerce_float(_cell(ws, row, idx, "amount_htt")),
        "total_vat": _coerce_float(_cell(ws, row, idx, "vat_total")),
        "total_ttc": _coerce_float(_cell(ws, row, idx, "total_ttc")),
    }

    # Reconstitution des lignes facture
    invoice_lines: list[dict[str, Any]] = []
    invoice_lines.extend(_build_energy_lines(ws, row, idx, period_start, period_end))
    invoice_lines.extend(_build_capacity_lines(ws, row, idx, period_start, period_end))
    invoice_lines.extend(_build_renewable_line(ws, row, idx))
    invoice_lines.extend(_build_cee_lines(ws, row, idx))
    invoice_lines.extend(_build_delivery_lines(ws, row, idx))
    invoice_lines.extend(_build_tax_lines(ws, row, idx))
    site["invoice_lines"] = invoice_lines

    return {
        "_invoice_number": invoice_number,
        "_sheet": ws.title,
        "_row": row,
        "_invoice_meta": {
            "invoice_number": invoice_number,
            "invoice_date": _coerce_date_iso(_cell(ws, row, idx, "invoice_date")),
            "regroupement": _coerce_str(_cell(ws, row, idx, "regroupement")),
            "contract_holder": _coerce_str(_cell(ws, row, idx, "contract_holder")),
            "compte_contrat": _coerce_str(_cell(ws, row, idx, "compte_contrat")),
            "compte_contrat_collectif": _coerce_str(_cell(ws, row, idx, "compte_contrat_collectif")),
            "due_date": _coerce_date_iso(_cell(ws, row, idx, "due_date")),
            "total_ttc": _coerce_float(_cell(ws, row, idx, "total_ttc")),
            "total_consumption_kwh": _coerce_float(_cell(ws, row, idx, "total_consumption")),
            "period_start": period_start,
            "period_end": period_end,
            # Référence marché ENGIE (constante pour le marché en cours)
            "market_reference": "2024-FCS-03",
        },
        "site": site,
    }


# ── Point d'entrée : parse_engie_xlsx ────────────────────────────────────────


def parse_engie_xlsx(path: Path) -> list[dict[str, Any]]:
    """Parse l'export XLSX ENGIE et retourne une liste de `parsed` dicts.

    Une entrée par bordereau (= n° de facture global FMC/FUM/Bordereau).
    Chaque entrée a la même structure que `parse_engie_pdf` :
        {
          "supplier": "ENGIE",
          "document_type": "facture",
          "site_count": N,
          "fic_count": N,
          "invoice": {...},
          "sites": [...],
        }

    Le moteur d'analyse (_build_control_report) est alors directement
    réutilisable sans modification.
    """
    # The parser performs many keyed cell lookups per row. openpyxl's read-only
    # worksheets are streaming-oriented and make random ws.cell(...) access
    # extremely slow on wide ENGIE exports, so load the small workbook normally.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    # Accumule les sites par bordereau (cross-sheets car un bordereau peut couvrir
    # plusieurs segments tarifaires dans le même export)
    bordereaux: dict[str, dict[str, Any]] = {}

    for sheet_name in SHEETS_TO_PARSE:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # Vérifie qu'il y a bien des données (sheets parfois vides)
        if ws.max_row < DATA_START_ROW:
            continue
        idx = _build_column_index(ws)
        if not idx.column_by_key.get("invoice_number"):
            # Pas de colonne bordereau → feuille inutilisable
            continue

        for row in range(DATA_START_ROW, ws.max_row + 1):
            parsed_row = _parse_row(ws, row, idx)
            if parsed_row is None:
                continue
            bordereau_id = parsed_row["_invoice_number"] or f"_no_bordereau_{sheet_name}_{row}"
            entry = bordereaux.setdefault(
                bordereau_id,
                {
                    "supplier": "ENGIE",
                    "document_type": "facture",
                    "page_count": 1,  # virtuel pour XLSX
                    "fic_count": 0,
                    "site_count": 0,
                    "invoice": dict(parsed_row["_invoice_meta"]),
                    "sites": [],
                    "parser_warnings": [],
                    "_segments": set(),
                    "_sheets": set(),
                },
            )
            entry["sites"].append(parsed_row["site"])
            entry["fic_count"] += 1
            entry["site_count"] += 1
            entry["_segments"].add(parsed_row["site"].get("segment"))
            entry["_sheets"].add(parsed_row["_sheet"])

            # Si les méta initiales sont None, prendre celles d'une ligne suivante
            for k, v in parsed_row["_invoice_meta"].items():
                if entry["invoice"].get(k) in (None, "") and v not in (None, ""):
                    entry["invoice"][k] = v

    # Finalisation : transforme les sets en listes pour la sérialisation JSON
    result: list[dict[str, Any]] = []
    for bordereau_id, entry in bordereaux.items():
        sheets = sorted(s for s in entry.pop("_sheets") if s)
        segments = sorted(s for s in entry.pop("_segments") if s)
        entry["xlsx_sheets"] = sheets
        entry["xlsx_segments"] = segments
        # L'export ENGIE repete le montant TTC au niveau FIC/site, pas au niveau
        # bordereau. Le total document doit donc etre reconstruit depuis les FIC.
        sums = [_coerce_decimal(s.get("total_ttc")) for s in entry["sites"]]
        sums = [s for s in sums if s is not None]
        if sums:
            entry["invoice"]["total_ttc"] = float(sum(sums, Decimal("0")))

        consumption_sums = [_coerce_decimal(s.get("total_consumption_kwh")) for s in entry["sites"]]
        consumption_sums = [s for s in consumption_sums if s is not None]
        if consumption_sums:
            entry["invoice"]["total_consumption_kwh"] = float(sum(consumption_sums, Decimal("0")))
        # Période globale = min/max des périodes sites
        starts = [s.get("period_start") for s in entry["sites"] if s.get("period_start")]
        ends = [s.get("period_end") for s in entry["sites"] if s.get("period_end")]
        if starts and entry["invoice"].get("period_start") is None:
            entry["invoice"]["period_start"] = min(starts)
        if ends and entry["invoice"].get("period_end") is None:
            entry["invoice"]["period_end"] = max(ends)
        result.append(entry)

    return result
