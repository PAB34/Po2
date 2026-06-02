"""Service de parsing et d'import du fichier contractuel DALKIA CPE.

Lit les fichiers L1 (Lot 1 - ecoles/sport) et L2 (Lot 2 - piscines)
et alimente les tables de reference :
  - cpe_dalkia_ref_sites
  - cpe_dalkia_ref_p2p3
  - cpe_dalkia_ref_cibles (GAZ + ELEC)
  - cpe_dalkia_ref_p1_gaz
  - cpe_dalkia_ref_ape

Structure des colonnes (stable entre L1 et L2, peut evoluer avec les avenants) :

Annexe 3.1 P2 / Annexe 4 P3 (57 col) :
  Cols 1-4  : code_site, nom_batiment, entite, lot_label
  9 periodes, offset 6 : colonnes P2.1/P2.2/P2.3/P2.4/P2tot aux cols 5,11,17,...

Annexe 5.1 / 5.2 Cibles GAZ/ELEC (54 col) :
  Cols 1-9  : code_site, site, compteur, ref_globale, sous-compteur, ref_qt, unite, ref_ecs, dju
  9 periodes, offset 5 : cols QT_GLOBAL/NB/q_ECS/QT_ECS/unite aux cols 10,15,20,...

Annexe 6 P1 GAZ (38 col) :
  Cols 1-11 : lot, entite, code_site, commune, libelle, PCE, type_tarif, prix_unitaire,
              ATRD, CTA, P10_fixe
  9 periodes, offset 3 : cols QT/P10var/P10tot aux cols 12,15,18,...

Annexe 2bis APE (20 col) :
  Row 5 : en-tetes
  Col 1-18 : code_site, nom, sit_init, description, annee, montant, cee_mwh, cee_eur,
             subvention, gain, sit_nouvelle, annee_eng, co2, enr_auto, enr_vendue,
             recette, ratio, commentaires
  Plusieurs lignes par site.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Structures de donnees
# ---------------------------------------------------------------------------

PERIOD_YEARS = [2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033]
N_PERIODS = 9


@dataclass
class DalkiaSiteRow:
    code_site: str
    nom_batiment: str
    entite: str | None
    lot_label: str | None
    lot: int


@dataclass
class DalkiaP2P3Row:
    code_site: str
    period_idx: int          # 1-9
    period_label: str
    period_year: int
    p2_1_ht: float | None
    p2_2_ht: float | None
    p2_3_ht: float | None
    p2_4_ht: float | None
    p2_total_ht: float | None
    p3_1_ht: float | None
    p3_2_ht: float | None
    p3_3_ht: float | None
    p3_4_ht: float | None
    p3_total_ht: float | None


@dataclass
class DalkiaCibleRow:
    code_site: str
    fluid: str               # "GAZ" | "ELEC"
    period_idx: int
    period_label: str
    period_year: int
    ref_globale_mwhpci: float | None
    ref_qt_mwhpci: float | None
    dju_reference: float | None
    qt_global_mwhpci: float | None
    nb_mwhpci: float | None
    q_ecs: float | None
    qt_ecs: float | None


@dataclass
class DalkiaP1GazRow:
    code_site: str
    pce: str | None
    type_tarif: str | None
    prix_unitaire_ht: float | None
    atrd_ht: float | None
    cta_ht: float | None
    p10_fixe_ht: float | None
    period_idx: int
    period_label: str
    period_year: int
    qt_mwhpcs: float | None
    p10_var_ht: float | None
    p10_total_ht: float | None


@dataclass
class DalkiaP1TarifRow:
    """En-tete Annexe 6 : composants de prix + coefficients de revision Pu par tarif (T1-T4).

    Formule contractuelle : Pu_GAZ = Pu_0 x (a + b x PEG/PEG0 + c x TVD/TVD0
                                              + d x CEE/CEE0 + e x TICGN/TICGN0)
    Les composants ci-dessous sont les valeurs de base (periode 0, 13/10/2025).
    """
    type_tarif: str
    p0_fournisseur: float | None        # molecule fournisseur (EUR HT/MWhPCS)
    ref_peg: float | None               # reference PEG
    terme_acheminement: float | None    # terme variable d'acheminement (TVD)
    obligation_cee: float | None
    ticgn: float | None
    marge_exploitant_pct: float | None
    prix_unitaire_ht: float | None      # Pu_0 (EUR HT/MWhPCS)
    coef_a: float | None
    coef_b: float | None
    coef_c: float | None
    coef_d: float | None
    coef_e: float | None


@dataclass
class DalkiaApeRow:
    code_site: str
    nom_batiment: str | None
    situation_initiale_mwhpci: float | None
    description_ape: str | None
    annee_achevement: int | None
    montant_ape_ht: float | None
    cee_mwh_cumac: float | None
    cee_eur: float | None
    subvention_ht: float | None
    gain_energetique_mwhpci: float | None
    situation_nouvelle_mwhpci: float | None
    annee_engagement_nouvelle_cible: int | None
    emission_co2_evitee: float | None
    production_enr_auto_mwh: float | None
    production_enr_vendue_mwh: float | None
    recette_vente_energie_ht: float | None
    ratio_ht_mwhpci: float | None
    commentaires: str | None


@dataclass
class DalkiaRecapRow:
    """Ligne du recapitulatif financier (RECAP MARCHE), format long.

    section : engagement | redevance_p1 | redevance_p2p3 | sensibilisation | travaux | bilan
    category: GAZ | ELEC | PV | GLOBAL | P1 | P2 | P3 | P2_SENS | "Travaux APE" | "Conformite"...
    metric  : qt_reference | qt_cible | pct_economie | p1_total_ht | bilan_ht | montant_hors_cee_ht...
    period_year : annee (None pour les totaux "duree du marche" ou montants globaux)
    """

    section: str
    category: str
    metric: str
    metric_label: str
    period_year: int | None
    period_label: str | None
    value: float | None
    unit: str | None


@dataclass
class DalkiaParseResult:
    lot: int
    filename: str
    period_labels: list[str]
    sites: list[DalkiaSiteRow] = field(default_factory=list)
    p2p3_rows: list[DalkiaP2P3Row] = field(default_factory=list)
    cibles_gaz: list[DalkiaCibleRow] = field(default_factory=list)
    cibles_elec: list[DalkiaCibleRow] = field(default_factory=list)
    p1_gaz: list[DalkiaP1GazRow] = field(default_factory=list)
    p1_tarifs: list[DalkiaP1TarifRow] = field(default_factory=list)
    ape_rows: list[DalkiaApeRow] = field(default_factory=list)
    recap_rows: list[DalkiaRecapRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class DalkiaImportPreview:
    """Preview renvoyee avant confirmation d'import."""
    lot: int
    filename: str
    nb_sites: int
    nb_p2p3_rows: int
    nb_cibles_rows: int
    nb_p1_gaz_rows: int
    nb_ape_rows: int
    nb_recap_rows: int
    recap_summary: dict[str, Any]
    period_labels: list[str]
    sample_sites: list[dict[str, Any]]
    classified: dict[str, Any]
    warnings: list[str]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float(v: Any) -> float | None:
    """Convertit une valeur Excel en float, None si vide/S.O."""
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    if s in ("", "S.O.", "S.O", "SO", "-", "N/A", "n/a"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_int(v: Any) -> int | None:
    f = _to_float(v)
    return int(f) if f is not None else None


def _clean_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip().replace("\n", " ").replace("\r", "")
    return s if s else None


def _is_site_row(code_site_val: Any) -> bool:
    """Verifie si la valeur col 1 correspond a un code site valide (VDS-...)."""
    if code_site_val is None:
        return False
    s = str(code_site_val).strip()
    return bool(s and not s.upper().startswith("TOTAL") and not s.upper().startswith("SOUS"))


def _extract_period_labels_p2p3(rows: list[tuple], period_start_cols: list[int]) -> list[str]:
    """Extrait les labels de periode depuis la ligne de headers (row 8, 0-indexed 7)."""
    header_row = rows[7] if len(rows) > 7 else ()
    labels = []
    for col_idx in period_start_cols:
        val = header_row[col_idx - 1] if col_idx - 1 < len(header_row) else None
        if val:
            s = str(val).strip().replace("\n", " ")
            # Normalise "Montant annuel (13 octobre 2025 / 31 dec 2025)" -> annee extraite
            years = re.findall(r"\d{4}", s)
            if years:
                labels.append(f"{years[0]}" if len(set(years)) == 1 else f"{years[0]}-{years[-1]}")
            else:
                labels.append(s[:40])
        else:
            labels.append(f"Periode {len(labels) + 1}")
    return labels


def _extract_period_labels_cibles(rows: list[tuple], period_start_cols: list[int]) -> list[str]:
    """Extrait les labels de periode depuis la ligne de headers des cibles (row 7, 0-indexed 6)."""
    header_row = rows[6] if len(rows) > 6 else ()
    labels = []
    for col_idx in period_start_cols:
        val = header_row[col_idx - 1] if col_idx - 1 < len(header_row) else None
        if val:
            s = str(val).strip().replace("\n", " ")
            years = re.findall(r"\d{4}", s)
            if years:
                labels.append(f"{years[0]}" if len(set(years)) == 1 else f"{years[0]}-{years[-1]}")
            else:
                labels.append(s[:40])
        else:
            labels.append(f"Periode {len(labels) + 1}")
    return labels


# ---------------------------------------------------------------------------
# Parseurs de feuilles
# ---------------------------------------------------------------------------

def _parse_p2(rows: list[tuple], lot: int) -> tuple[list[DalkiaSiteRow], list[DalkiaP2P3Row], list[str]]:
    """
    Parse la feuille Annexe 3.1 - P2 - A.
    Retourne (sites, p2_rows, warnings).
    Structure: headers ligne 9 (0-indexed 8), data from ligne 10 (0-indexed 9).
    Periodes aux cols (1-indexed): 5, 11, 17, 23, 29, 35, 41, 47, 53.
    """
    PERIOD_STARTS = [5, 11, 17, 23, 29, 35, 41, 47, 53]
    period_labels = _extract_period_labels_p2p3(rows, PERIOD_STARTS)

    sites: list[DalkiaSiteRow] = []
    p2_rows: list[DalkiaP2P3Row] = []
    warnings: list[str] = []
    seen_sites: set[str] = set()

    for row in rows[9:]:  # data from row 10
        if len(row) == 0:
            continue
        code_site_val = row[0] if len(row) > 0 else None
        if not _is_site_row(code_site_val):
            continue

        code_site = str(code_site_val).strip()
        nom_batiment = _clean_str(row[1] if len(row) > 1 else None) or code_site
        entite = _clean_str(row[2] if len(row) > 2 else None)
        lot_label = _clean_str(row[3] if len(row) > 3 else None)

        if code_site not in seen_sites:
            sites.append(DalkiaSiteRow(
                code_site=code_site,
                nom_batiment=nom_batiment,
                entite=entite,
                lot_label=lot_label,
                lot=lot,
            ))
            seen_sites.add(code_site)

        for period_i, col_start in enumerate(PERIOD_STARTS):
            c = col_start - 1  # 0-indexed
            p2_1 = _to_float(row[c] if len(row) > c else None)
            p2_2 = _to_float(row[c + 1] if len(row) > c + 1 else None)
            p2_3 = _to_float(row[c + 2] if len(row) > c + 2 else None)
            p2_4 = _to_float(row[c + 3] if len(row) > c + 3 else None)
            p2_tot = _to_float(row[c + 4] if len(row) > c + 4 else None)
            p2_rows.append(DalkiaP2P3Row(
                code_site=code_site,
                period_idx=period_i + 1,
                period_label=period_labels[period_i] if period_i < len(period_labels) else str(PERIOD_YEARS[period_i]),
                period_year=PERIOD_YEARS[period_i],
                p2_1_ht=p2_1, p2_2_ht=p2_2, p2_3_ht=p2_3, p2_4_ht=p2_4, p2_total_ht=p2_tot,
                p3_1_ht=None, p3_2_ht=None, p3_3_ht=None, p3_4_ht=None, p3_total_ht=None,
            ))

    return sites, p2_rows, warnings


def _parse_p3(rows: list[tuple], p2_rows: list[DalkiaP2P3Row]) -> tuple[list[DalkiaP2P3Row], list[str]]:
    """
    Parse la feuille Annexe 4 - P3 et fusionne avec les donnees P2 existantes.
    Meme structure que P2 (memes cols de periodes).
    """
    PERIOD_STARTS = [5, 11, 17, 23, 29, 35, 41, 47, 53]

    # Index des donnees P2 par (code_site, period_idx)
    p2_index: dict[tuple[str, int], DalkiaP2P3Row] = {
        (r.code_site, r.period_idx): r for r in p2_rows
    }
    warnings: list[str] = []

    for row in rows[9:]:
        if len(row) == 0:
            continue
        code_site_val = row[0] if len(row) > 0 else None
        if not _is_site_row(code_site_val):
            continue
        code_site = str(code_site_val).strip()

        for period_i, col_start in enumerate(PERIOD_STARTS):
            c = col_start - 1
            p3_1 = _to_float(row[c] if len(row) > c else None)
            p3_2 = _to_float(row[c + 1] if len(row) > c + 1 else None)
            p3_3 = _to_float(row[c + 2] if len(row) > c + 2 else None)
            p3_4 = _to_float(row[c + 3] if len(row) > c + 3 else None)
            p3_tot = _to_float(row[c + 4] if len(row) > c + 4 else None)

            key = (code_site, period_i + 1)
            if key in p2_index:
                r = p2_index[key]
                r.p3_1_ht = p3_1
                r.p3_2_ht = p3_2
                r.p3_3_ht = p3_3
                r.p3_4_ht = p3_4
                r.p3_total_ht = p3_tot
            else:
                warnings.append(f"P3 site {code_site} periode {period_i + 1} sans P2 correspondant")

    return p2_rows, warnings


def _parse_cibles(rows: list[tuple], fluid: str, lot: int) -> tuple[list[DalkiaCibleRow], list[str]]:
    """
    Parse Annexe 5.1 (GAZ) ou 5.2 (ELEC).
    Structure: headers ligne 8 (0-indexed 7), data from ligne 9 (0-indexed 8).
    Periodes aux cols (1-indexed): 10, 15, 20, 25, 30, 35, 40, 45, 50.
    Cols fixes: 1=CODE_SITE, 2=SITE_PRINCIPAL, 3=COMPTEUR, 4=REF_GLOBALE, 5=SOUS_COMPTEUR,
                6=REF_QT, 7=UNITE, 8=REF_ECS, 9=DJU
    Par periode (5 cols): QT_GLOBAL, NB, q_ECS, QT_ECS, UNITE
    """
    PERIOD_STARTS = [10, 15, 20, 25, 30, 35, 40, 45, 50]
    period_labels = _extract_period_labels_cibles(rows, PERIOD_STARTS)
    cibles: list[DalkiaCibleRow] = []
    warnings: list[str] = []

    for row in rows[8:]:  # data from row 9
        if len(row) == 0:
            continue
        code_site_val = row[0] if len(row) > 0 else None
        if not _is_site_row(code_site_val):
            continue
        code_site = str(code_site_val).strip()

        ref_globale = _to_float(row[3] if len(row) > 3 else None)
        ref_qt = _to_float(row[5] if len(row) > 5 else None)
        dju = _to_float(row[8] if len(row) > 8 else None)

        for period_i, col_start in enumerate(PERIOD_STARTS):
            c = col_start - 1  # 0-indexed
            qt_global = _to_float(row[c] if len(row) > c else None)
            nb = _to_float(row[c + 1] if len(row) > c + 1 else None)
            q_ecs = _to_float(row[c + 2] if len(row) > c + 2 else None)
            qt_ecs = _to_float(row[c + 3] if len(row) > c + 3 else None)
            cibles.append(DalkiaCibleRow(
                code_site=code_site,
                fluid=fluid,
                period_idx=period_i + 1,
                period_label=period_labels[period_i] if period_i < len(period_labels) else str(PERIOD_YEARS[period_i]),
                period_year=PERIOD_YEARS[period_i],
                ref_globale_mwhpci=ref_globale,
                ref_qt_mwhpci=ref_qt,
                dju_reference=dju,
                qt_global_mwhpci=qt_global,
                nb_mwhpci=nb,
                q_ecs=q_ecs,
                qt_ecs=qt_ecs,
            ))

    return cibles, warnings


def _parse_p1_gaz(rows: list[tuple], lot: int) -> tuple[list[DalkiaP1GazRow], list[str]]:
    """
    Parse Annexe 6 - P1 GAZ.
    Headers ligne 30 (0-indexed 29), data from ligne 31 (0-indexed 30).
    Cols 1-11 : lot, entite, code_site(=N°PROG), commune, libelle, PCE, type_tarif,
                prix_unitaire, ATRD, CTA, P10_fixe
    9 periodes offset 3 : QT/P10var/P10tot aux cols 12,15,18,21,24,27,30,33,36
    """
    PERIOD_STARTS = [12, 15, 18, 21, 24, 27, 30, 33, 36]

    # Trouver la ligne de headers (contient "LOT" ou "ENTITE")
    header_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).upper() for v in row if v is not None]
        if any("LOT" in v for v in vals) and any("ENTITE" in v or "PROG" in v for v in vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], ["Annexe 6 P1 GAZ : ligne de headers non trouvee"]

    # Extraire les labels de periode depuis les lignes precedant les headers
    period_label_row = rows[header_row_idx - 2] if header_row_idx >= 2 else ()
    period_labels = []
    for col_start in PERIOD_STARTS:
        c = col_start - 1
        val = period_label_row[c] if len(period_label_row) > c else None
        if val:
            years = re.findall(r"\d{4}", str(val))
            if years:
                period_labels.append(f"{years[0]}" if len(set(years)) == 1 else f"{years[0]}-{years[-1]}")
            else:
                period_labels.append(str(val)[:40])
        else:
            period_labels.append(str(PERIOD_YEARS[len(period_labels)]) if len(period_labels) < 9 else "?")

    p1_rows: list[DalkiaP1GazRow] = []
    warnings: list[str] = []

    for row in rows[header_row_idx + 1:]:
        if len(row) == 0:
            continue
        # code_site est en col 3 (N° PROG)
        code_site_val = row[2] if len(row) > 2 else None
        if not _is_site_row(code_site_val):
            continue
        code_site = str(code_site_val).strip()

        pce = _clean_str(row[5] if len(row) > 5 else None)
        type_tarif = _clean_str(row[6] if len(row) > 6 else None)
        if type_tarif and type_tarif in ("0", ""):
            type_tarif = None
        prix_unitaire = _to_float(row[7] if len(row) > 7 else None)
        atrd = _to_float(row[8] if len(row) > 8 else None)
        cta = _to_float(row[9] if len(row) > 9 else None)
        p10_fixe = _to_float(row[10] if len(row) > 10 else None)

        for period_i, col_start in enumerate(PERIOD_STARTS):
            c = col_start - 1
            qt = _to_float(row[c] if len(row) > c else None)
            p10_var = _to_float(row[c + 1] if len(row) > c + 1 else None)
            p10_tot = _to_float(row[c + 2] if len(row) > c + 2 else None)
            p1_rows.append(DalkiaP1GazRow(
                code_site=code_site,
                pce=pce,
                type_tarif=type_tarif,
                prix_unitaire_ht=prix_unitaire,
                atrd_ht=atrd,
                cta_ht=cta,
                p10_fixe_ht=p10_fixe,
                period_idx=period_i + 1,
                period_label=period_labels[period_i] if period_i < len(period_labels) else str(PERIOD_YEARS[period_i]),
                period_year=PERIOD_YEARS[period_i],
                qt_mwhpcs=qt,
                p10_var_ht=p10_var,
                p10_total_ht=p10_tot,
            ))

    return p1_rows, warnings


def _parse_p1_gaz_tarifs(rows: list[tuple]) -> tuple[list[DalkiaP1TarifRow], list[str]]:
    """Parse l'en-tete de l'Annexe 6 : composants de prix + coefficients de revision par tarif.

    Deux blocs :
      - tableau prix (header contenant "P0 FOURNISSEUR") : 1 ligne par tarif en col 7 (T1..T4),
        cols 8..14 = p0, peg, acheminement, cee, ticgn, marge, prix_unitaire.
      - bloc coefficients (apres la ligne "Formule") : 5 lignes a,b,c,d,e ; chaque ligne porte
        les 4 tarifs aux triplets (label, nom_coef, valeur) en cols (1-3), (5-7), (9-11), (13-15).
    """
    warnings: list[str] = []
    by_tarif: dict[str, DalkiaP1TarifRow] = {}

    # --- Bloc prix ---
    price_header_idx = None
    for i, row in enumerate(rows[:30]):
        if any("P0 FOURNISSEUR" in str(v).upper() for v in row if v is not None):
            price_header_idx = i
            break
    if price_header_idx is None:
        return [], ["Annexe 6 : tableau de prix par tarif (P0 FOURNISSEUR) non trouve"]

    for row in rows[price_header_idx + 1:price_header_idx + 8]:
        tarif = _clean_str(row[6] if len(row) > 6 else None)
        if not tarif or not re.fullmatch(r"T\d", tarif):
            continue
        by_tarif[tarif] = DalkiaP1TarifRow(
            type_tarif=tarif,
            p0_fournisseur=_to_float(row[7] if len(row) > 7 else None),
            ref_peg=_to_float(row[8] if len(row) > 8 else None),
            terme_acheminement=_to_float(row[9] if len(row) > 9 else None),
            obligation_cee=_to_float(row[10] if len(row) > 10 else None),
            ticgn=_to_float(row[11] if len(row) > 11 else None),
            marge_exploitant_pct=_to_float(row[12] if len(row) > 12 else None),
            prix_unitaire_ht=_to_float(row[13] if len(row) > 13 else None),
            coef_a=None, coef_b=None, coef_c=None, coef_d=None, coef_e=None,
        )

    # --- Bloc coefficients : groupes de 3 cols (label tarif, nom coef, valeur) par tarif ---
    # Seule la ligne "a" porte les libelles tarif ; b,c,d,e n'ont que (nom, valeur).
    # On etablit donc le mapping colonne -> tarif depuis la ligne "a", applique par position.
    GROUPS = [0, 4, 8, 12]  # colonne 0-indexed du libelle tarif de chaque groupe
    coef_start = None
    for i, row in enumerate(rows[:40]):
        name = _clean_str(row[1] if len(row) > 1 else None)
        if name and name.lower() == "a" and _to_float(row[2] if len(row) > 2 else None) is not None:
            coef_start = i
            break
    if coef_start is None:
        warnings.append("Annexe 6 : bloc des coefficients de revision (a,b,c,d,e) non trouve")
    else:
        start_row = rows[coef_start]
        group_tarif: dict[int, str] = {}
        for g in GROUPS:
            t = _clean_str(start_row[g] if len(start_row) > g else None)
            if t and re.fullmatch(r"T\d", t):
                group_tarif[g] = t
        coef_attr = {"a": "coef_a", "b": "coef_b", "c": "coef_c", "d": "coef_d", "e": "coef_e"}
        for row in rows[coef_start:coef_start + 5]:
            for g in GROUPS:
                tarif = group_tarif.get(g)
                if not tarif or tarif not in by_tarif:
                    continue
                name = _clean_str(row[g + 1] if len(row) > g + 1 else None)
                value = _to_float(row[g + 2] if len(row) > g + 2 else None)
                attr = coef_attr.get((name or "").lower())
                if attr:
                    setattr(by_tarif[tarif], attr, value)

    tarifs = [by_tarif[t] for t in sorted(by_tarif)]
    if not tarifs:
        warnings.append("Annexe 6 : aucun tarif de prix gaz extrait")
    return tarifs, warnings


def _parse_ape(rows: list[tuple], lot: int) -> tuple[list[DalkiaApeRow], list[str]]:
    """
    Parse Annexe 2bis - Travaux APE.
    Headers ligne 5 (0-indexed 4), data from ligne 6 (0-indexed 5).
    Cols: 1=code_site, 2=nom, 3=sit_init, 4=description, 5=annee_achevement,
          6=montant, 7=cee_mwh_cumac, 8=cee_eur, 9=subvention,
          10=gain_energetique, 11=situation_nouvelle, 12=annee_engagement,
          13=co2_evitee, 14=enr_auto, 15=enr_vendue, 16=recette, 17=ratio, 18=commentaires
    """
    # Trouver la ligne de headers
    header_row_idx = None
    for i, row in enumerate(rows):
        vals = [str(v).upper() for v in row if v is not None]
        if any("CODE" in v and "SITE" in v for v in vals) or any("SITUATION" in v and "INIT" in v for v in vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], ["Annexe 2bis APE : ligne de headers non trouvee"]

    ape_rows: list[DalkiaApeRow] = []
    warnings: list[str] = []

    for row in rows[header_row_idx + 1:]:
        if len(row) == 0:
            continue
        code_site_val = row[0] if len(row) > 0 else None
        if not _is_site_row(code_site_val):
            continue
        code_site = str(code_site_val).strip()

        descr = _clean_str(row[3] if len(row) > 3 else None)
        # Ignorer les lignes de description vide ou "0"
        if descr in ("0", ""):
            descr = None

        ape_rows.append(DalkiaApeRow(
            code_site=code_site,
            nom_batiment=_clean_str(row[1] if len(row) > 1 else None),
            situation_initiale_mwhpci=_to_float(row[2] if len(row) > 2 else None),
            description_ape=descr,
            annee_achevement=_to_int(row[4] if len(row) > 4 else None),
            montant_ape_ht=_to_float(row[5] if len(row) > 5 else None),
            cee_mwh_cumac=_to_float(row[6] if len(row) > 6 else None),
            cee_eur=_to_float(row[7] if len(row) > 7 else None),
            subvention_ht=_to_float(row[8] if len(row) > 8 else None),
            gain_energetique_mwhpci=_to_float(row[9] if len(row) > 9 else None),
            situation_nouvelle_mwhpci=_to_float(row[10] if len(row) > 10 else None),
            annee_engagement_nouvelle_cible=_to_int(row[11] if len(row) > 11 else None),
            emission_co2_evitee=_to_float(row[12] if len(row) > 12 else None),
            production_enr_auto_mwh=_to_float(row[13] if len(row) > 13 else None),
            production_enr_vendue_mwh=_to_float(row[14] if len(row) > 14 else None),
            recette_vente_energie_ht=_to_float(row[15] if len(row) > 15 else None),
            ratio_ht_mwhpci=_to_float(row[16] if len(row) > 16 else None),
            commentaires=_clean_str(row[17] if len(row) > 17 else None),
        ))

    return ape_rows, warnings


def _recap_period_map(header_row: tuple, start_col: int = 1) -> list[tuple[int, int | None, str]]:
    """Construit la liste (col_0based, annee|None, label) depuis une ligne d'en-tete de periodes.

    - Colonne periode : contient une annee + "janvier"/"octobre" -> (col, annee, label)
    - Colonne resume  : contient "total"/"economie"/"engagement"/"duree" -> (col, None, label)
    - Autres colonnes texte (ex "Prestation") : ignorees
    """
    out: list[tuple[int, int | None, str]] = []
    for c in range(start_col, len(header_row)):
        v = header_row[c]
        if v is None or not str(v).strip():
            continue
        s = str(v).strip().replace("\n", " ")
        low = s.lower()
        years = re.findall(r"\d{4}", s)
        is_period = bool(years) and ("janvier" in low or "octobre" in low)
        is_summary = any(k in low for k in ["total", "economie", "économie", "engagement", "durée", "duree"])
        if is_period:
            out.append((c, int(years[0]), s[:70]))
        elif is_summary:
            out.append((c, None, s[:70]))
    return out


def _parse_recap(rows: list[tuple]) -> tuple[list[DalkiaRecapRow], list[str]]:
    """Parse la feuille RECAP MARCHE (recapitulatif financier global).

    6 sections :
      2.6 Engagement consommations (GAZ/ELEC/PV/GLOBAL)
      2.7 Redevances P1 gaz
      2.8 Redevances P2/P3 + sensibilisation
      2.9 Montant travaux APE + obligatoires
      2.10 Bilan sur la duree du marche
    """
    out: list[DalkiaRecapRow] = []
    warnings: list[str] = []

    # Definition des metriques par libelle (substring, insensible casse/accents partiels)
    # (label_match, section, category, metric, unit)
    METRIC_DEFS = [
        ("qt « reference »", "engagement", None, "qt_reference", "MWhPCI"),
        ("qt «reference»", "engagement", None, "qt_reference", "MWhPCI"),
        ("qt « cible »", "engagement", None, "qt_cible", "MWhPCI"),
        ("production pv", "engagement", "PV", "production_pv", "MWhPCI"),
        ("autoconsommation", "engagement", "PV", "autoconsommation", "MWhPCI"),
        ("p1 – parties fixes", "redevance_p1", "P1", "p1_parties_fixes_ht", "EUR_HT"),
        ("p1 – part proportionnelle", "redevance_p1", "P1", "p1_part_proportionnelle_ht", "EUR_HT"),
        ("p1 – total (€ht)", "redevance_p1", "P1", "p1_total_ht", "EUR_HT"),
        ("p1 – total (€ttc)", "redevance_p1", "P1", "p1_total_ttc", "EUR_TTC"),
        ("prestation p2.1 a p2.4 (€ht)", "redevance_p2p3", "P2", "p2_total_ht", "EUR_HT"),
        ("prestation p2.1 a p2.4 (€ttc)", "redevance_p2p3", "P2", "p2_total_ttc", "EUR_TTC"),
        ("prestation p3.1 a p3.4 (€ht)", "redevance_p2p3", "P3", "p3_total_ht", "EUR_HT"),
        ("prestation p3.1 a p3.4 (€ttc)", "redevance_p2p3", "P3", "p3_total_ttc", "EUR_TTC"),
    ]

    current_period_map: list[tuple[int, int | None, str]] = []
    current_section: str | None = None
    current_category: str | None = None
    sensibilisation_seen = 0

    for row in rows:
        if not row:
            continue
        c1 = (_clean_str(row[0]) or "")
        c1_low = c1.lower()

        # --- Marqueurs de section (par prefixe numerote OU mot-cle, robuste L1/L2) ---
        if c1.startswith("2.6") or "engagement de consomm" in c1_low:
            current_section = "engagement"
            continue
        if c1.startswith("2.7") or ("redevances" in c1_low and "p1" in c1_low):
            current_section = "redevance_p1"
            continue
        if c1.startswith("2.8") or ("redevances" in c1_low and ("p2" in c1_low or "p3" in c1_low)):
            current_section = "redevance_p2p3"
            continue
        if c1.startswith("2.9") or "montant travaux" in c1_low or "travaux ape" in c1_low:
            current_section = "travaux"
            continue
        if c1.startswith("2.10") or "bilan sur la dur" in c1_low:
            current_section = "bilan"
            continue

        # --- Marqueurs de categorie (engagement) + facteur CO2 ---
        if c1_low.startswith("pour le gaz"):
            current_category = "GAZ"
            co2 = _to_float(row[1] if len(row) > 1 else None)
            if co2 is not None:
                out.append(DalkiaRecapRow("engagement", "GAZ", "facteur_co2", "Facteur CO2 gaz",
                                          None, "T CO2/MWh", co2, "T_CO2_MWh"))
            continue
        if c1_low.startswith("pour l'electricite") or c1_low.startswith("pour l'électricité"):
            current_category = "ELEC"
            co2 = _to_float(row[1] if len(row) > 1 else None)
            if co2 is not None:
                out.append(DalkiaRecapRow("engagement", "ELEC", "facteur_co2", "Facteur CO2 elec",
                                          None, "T CO2/MWh", co2, "T_CO2_MWh"))
            continue
        if c1_low.startswith("pour le pv"):
            current_category = "PV"
            continue

        # --- Lignes d'en-tete de periodes ---
        if c1.startswith("Paramètre") or c1.startswith("Parametre"):
            current_period_map = _recap_period_map(row, start_col=1)
            continue
        if c1_low == "energie":  # en-tete section P1
            current_period_map = _recap_period_map(row, start_col=1)
            continue

        # --- Section travaux (2.9) : layout categories x montants ---
        if current_section == "travaux":
            if c1_low in ("travaux ape", "amélioration", "amelioration", "conformité", "conformite", "total"):
                # cols: 2=imputation, 3=hors_cee_ht, 4=hors_cee_ttc, 5=cee, 6=subv, 7=apres_deduction
                def _v(idx: int) -> float | None:
                    return _to_float(row[idx] if len(row) > idx else None)
                cat = c1
                pairs = [
                    ("montant_hors_cee_ht", _v(2), "EUR_HT"),
                    ("montant_hors_cee_ttc", _v(3), "EUR_TTC"),
                    ("cee_eur", _v(4), "EUR"),
                    ("subventions_eur", _v(5), "EUR"),
                    ("montant_apres_deduction_ttc", _v(6), "EUR_TTC"),
                ]
                for metric, val, unit in pairs:
                    if val is not None:
                        out.append(DalkiaRecapRow("travaux", cat, metric, c1, None, "Total marche", val, unit))
            continue

        # --- Section bilan (2.10) : col2=HT, col3=TTC ---
        if current_section == "bilan":
            if c1 and not c1.startswith("Paramètre") and len(row) > 2:
                ht = _to_float(row[1] if len(row) > 1 else None)
                ttc = _to_float(row[2] if len(row) > 2 else None)
                if ht is not None:
                    out.append(DalkiaRecapRow("bilan", c1, "bilan_ht", c1, None, "Total marche", ht, "EUR_HT"))
                if ttc is not None:
                    out.append(DalkiaRecapRow("bilan", c1, "bilan_ttc", c1, None, "Total marche", ttc, "EUR_TTC"))
            continue

        # --- Lignes de metrique par periode ---
        if not current_period_map:
            continue

        # Le libelle de metrique peut etre en col 1 (engagement, P2/P3) ou col 2 (P1 gaz).
        c2 = _clean_str(row[1]) if len(row) > 1 else None
        label_search = c1_low
        if c2:
            label_search = (label_search + " " + c2.lower()).strip()
        metric_label = c1 if c1 else (c2 or "")

        def _emit(section: str, category: str, metric: str, unit: str) -> None:
            for col, year, label in current_period_map:
                val = _to_float(row[col] if len(row) > col else None)
                if val is not None:
                    out.append(DalkiaRecapRow(section, category, metric, metric_label, year, label, val, unit))

        # Sensibilisation P2 (2 lignes : 1ere = HT, 2eme = TTC)
        if "sensibilisation" in label_search:
            sensibilisation_seen += 1
            metric = "p2_sensibilisation_ht" if sensibilisation_seen == 1 else "p2_sensibilisation_ttc"
            unit = "EUR_HT" if sensibilisation_seen == 1 else "EUR_TTC"
            _emit("sensibilisation", "P2_SENS", metric, unit)
            continue

        # Pourcentage d'economie total Global (R18/R19) -> category GLOBAL
        if ("pourcentage" in label_search and "economie" in label_search.replace("é", "e") and "total" in label_search):
            _emit("engagement", "GLOBAL", "pct_economie_global", "pct")
            continue

        # Pourcentage d'economie (par fluide)
        if "pourcentage" in label_search and "economie" in label_search.replace("é", "e"):
            _emit("engagement", current_category or "GAZ", "pct_economie", "pct")
            continue

        # Pourcentage d'autoconsommation
        if "pourcentage" in label_search and "autoconsomma" in label_search:
            _emit("engagement", "PV", "pct_autoconsommation", "pct")
            continue

        # Metriques definies par libelle
        for label_match, section, forced_cat, metric, unit in METRIC_DEFS:
            if label_match in label_search:
                cat = forced_cat or current_category or "GAZ"
                _emit(section, cat, metric, unit)
                break

    if not out:
        warnings.append("RECAP MARCHE : aucune donnee financiere extraite")
    return out, warnings


# ---------------------------------------------------------------------------
# Point d'entree principal
# ---------------------------------------------------------------------------

def parse_dalkia_file(raw_bytes: bytes, filename: str, lot: int) -> DalkiaParseResult:
    """
    Parse un fichier DALKIA (L1 ou L2) et retourne un DalkiaParseResult.

    lot : 1 (L1 — ecoles, sport) ou 2 (L2 — piscines).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel : {exc}") from exc

    def _get_rows(sheet_name: str) -> list[tuple]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))

    all_warnings: list[str] = []

    # --- P2 ---
    p2_rows_raw = _get_rows("Annexe 3.1 - P2 - A")
    sites, p2_rows, w = _parse_p2(p2_rows_raw, lot)
    all_warnings.extend(w)

    # --- P3 (fusionne dans p2_rows) ---
    p3_rows_raw = _get_rows("Annexe 4 - P3")
    p2p3_rows, w = _parse_p3(p3_rows_raw, p2_rows)
    all_warnings.extend(w)

    # --- Cibles GAZ ---
    gaz_raw = _get_rows("Annexe 5.1 - CIBLES GAZ")
    cibles_gaz, w = _parse_cibles(gaz_raw, "GAZ", lot)
    all_warnings.extend(w)

    # --- Cibles ELEC ---
    elec_raw = _get_rows("Annexe 5.2 - CIBLES ELEC")
    cibles_elec, w = _parse_cibles(elec_raw, "ELEC", lot)
    all_warnings.extend(w)

    # --- P1 GAZ ---
    p1_raw = _get_rows("Annexe 6 - P1 GAZ")
    p1_gaz, w = _parse_p1_gaz(p1_raw, lot)
    all_warnings.extend(w)

    # --- P1 GAZ : composants de prix + coefficients de revision (en-tete Annexe 6) ---
    p1_tarifs, w = _parse_p1_gaz_tarifs(p1_raw)
    all_warnings.extend(w)

    # --- APE ---
    ape_raw = _get_rows("Annexe 2bis - Travaux APE")
    ape_rows, w = _parse_ape(ape_raw, lot)
    all_warnings.extend(w)

    # --- RECAP MARCHE (recapitulatif financier global) ---
    recap_raw = _get_rows("RECAP MARCHE")
    recap_rows, w = _parse_recap(recap_raw)
    all_warnings.extend(w)

    # Extraire les labels de periodes (depuis P2 si disponible)
    period_labels: list[str] = []
    if p2p3_rows:
        labels_seen: dict[int, str] = {}
        for r in p2p3_rows:
            if r.period_idx not in labels_seen:
                labels_seen[r.period_idx] = r.period_label
        period_labels = [labels_seen.get(i, str(PERIOD_YEARS[i - 1])) for i in range(1, N_PERIODS + 1)]

    return DalkiaParseResult(
        lot=lot,
        filename=filename,
        period_labels=period_labels,
        sites=sites,
        p2p3_rows=p2p3_rows,
        cibles_gaz=cibles_gaz,
        cibles_elec=cibles_elec,
        p1_gaz=p1_gaz,
        p1_tarifs=p1_tarifs,
        ape_rows=ape_rows,
        recap_rows=recap_rows,
        warnings=all_warnings,
    )


def _build_classified(result: DalkiaParseResult) -> dict[str, Any]:
    """Construit la vue complete classifiee de toutes les donnees parsees.

    Format pivote (1 ligne par site, colonnes par annee) pour les donnees
    periodiques (P2/P3, cibles, P1), liste detaillee pour APE et RECAP.
    """
    years = PERIOD_YEARS

    # --- P2/P3 : 1 ligne/site, by_year {p2_total, p3_total} ---
    p2p3_by_site: dict[str, dict[str, Any]] = {}
    for site in result.sites:
        p2p3_by_site[site.code_site] = {
            "code_site": site.code_site,
            "nom_batiment": site.nom_batiment,
            "by_year": {str(y): {"p2": None, "p3": None} for y in years},
        }
    for r in result.p2p3_rows:
        entry = p2p3_by_site.get(r.code_site)
        if entry is not None:
            entry["by_year"][str(r.period_year)] = {"p2": r.p2_total_ht, "p3": r.p3_total_ht}

    # --- Cibles : 1 ligne/site/fluide, by_year qt_global + nb + dju ---
    def _pivot_cibles(rows: list[DalkiaCibleRow]) -> list[dict[str, Any]]:
        by_site: dict[str, dict[str, Any]] = {}
        for r in rows:
            entry = by_site.setdefault(r.code_site, {
                "code_site": r.code_site,
                "ref_globale": r.ref_globale_mwhpci,
                "dju": r.dju_reference,
                "by_year": {},
            })
            # Si plusieurs lignes meme annee (sous-compteurs PV), on somme
            key = str(r.period_year)
            prev = entry["by_year"].get(key)
            val = r.qt_global_mwhpci or 0.0
            entry["by_year"][key] = (prev or 0.0) + val if prev is not None else val
        return list(by_site.values())

    # --- P1 gaz : 1 ligne/site, tarif/pce/prix + by_year p10_total ---
    p1_by_site: dict[str, dict[str, Any]] = {}
    for r in result.p1_gaz:
        entry = p1_by_site.setdefault(r.code_site, {
            "code_site": r.code_site,
            "pce": r.pce,
            "type_tarif": r.type_tarif,
            "prix_unitaire_ht": r.prix_unitaire_ht,
            "by_year": {},
        })
        entry["by_year"][str(r.period_year)] = r.p10_total_ht

    # --- APE : detail complet ---
    ape_rows = [
        {
            "code_site": r.code_site,
            "nom_batiment": r.nom_batiment,
            "description_ape": r.description_ape,
            "annee_achevement": r.annee_achevement,
            "montant_ape_ht": r.montant_ape_ht,
            "cee_eur": r.cee_eur,
            "gain_energetique_mwhpci": r.gain_energetique_mwhpci,
            "annee_engagement_nouvelle_cible": r.annee_engagement_nouvelle_cible,
            "emission_co2_evitee": r.emission_co2_evitee,
        }
        for r in result.ape_rows
    ]

    # --- RECAP : engagement (conso) + redevances + bilan + travaux ---
    def _recap_pivot(section: str, metric: str, category: str | None = None) -> dict[str, float | None]:
        out: dict[str, float | None] = {}
        for r in result.recap_rows:
            if r.section == section and r.metric == metric and (category is None or r.category == category):
                key = str(r.period_year) if r.period_year is not None else (r.period_label or "total")
                out[key] = r.value
        return out

    recap_engagement = {
        "gaz_qt_reference": _recap_pivot("engagement", "qt_reference", "GAZ"),
        "gaz_qt_cible": _recap_pivot("engagement", "qt_cible", "GAZ"),
        "gaz_pct_economie": _recap_pivot("engagement", "pct_economie", "GAZ"),
        "elec_qt_reference": _recap_pivot("engagement", "qt_reference", "ELEC"),
        "elec_qt_cible": _recap_pivot("engagement", "qt_cible", "ELEC"),
        "elec_pct_economie": _recap_pivot("engagement", "pct_economie", "ELEC"),
        "pv_production": _recap_pivot("engagement", "production_pv", "PV"),
        "pv_autoconsommation": _recap_pivot("engagement", "autoconsommation", "PV"),
    }
    recap_redevances = {
        "p1_total_ht": _recap_pivot("redevance_p1", "p1_total_ht"),
        "p2_total_ht": _recap_pivot("redevance_p2p3", "p2_total_ht"),
        "p3_total_ht": _recap_pivot("redevance_p2p3", "p3_total_ht"),
        "p2_sensibilisation_ht": _recap_pivot("sensibilisation", "p2_sensibilisation_ht"),
    }
    recap_travaux = [
        {"categorie": r.category, "metric": r.metric, "value": r.value, "unit": r.unit}
        for r in result.recap_rows if r.section == "travaux"
    ]
    recap_bilan = [
        {"poste": r.category, "metric": r.metric, "value": r.value, "unit": r.unit}
        for r in result.recap_rows if r.section == "bilan"
    ]

    p1_tarifs = [
        {
            "type_tarif": t.type_tarif,
            "p0_fournisseur": t.p0_fournisseur,
            "ref_peg": t.ref_peg,
            "terme_acheminement": t.terme_acheminement,
            "obligation_cee": t.obligation_cee,
            "ticgn": t.ticgn,
            "marge_exploitant_pct": t.marge_exploitant_pct,
            "prix_unitaire_ht": t.prix_unitaire_ht,
            "coef_a": t.coef_a, "coef_b": t.coef_b, "coef_c": t.coef_c,
            "coef_d": t.coef_d, "coef_e": t.coef_e,
        }
        for t in result.p1_tarifs
    ]

    return {
        "years": [str(y) for y in years],
        "p2p3": list(p2p3_by_site.values()),
        "cibles_gaz": _pivot_cibles(result.cibles_gaz),
        "cibles_elec": _pivot_cibles(result.cibles_elec),
        "p1_gaz": list(p1_by_site.values()),
        "p1_tarifs": p1_tarifs,
        "ape": ape_rows,
        "recap_engagement": recap_engagement,
        "recap_redevances": recap_redevances,
        "recap_travaux": recap_travaux,
        "recap_bilan": recap_bilan,
    }


def build_import_preview(result: DalkiaParseResult) -> DalkiaImportPreview:
    """Construit un apercu de l'import avant confirmation."""
    sample_sites = []
    for site in result.sites[:5]:
        # Trouver les totaux P2/P3 pour la periode 2 (annee pleine 2026)
        p2p3_2026 = next(
            (r for r in result.p2p3_rows if r.code_site == site.code_site and r.period_year == 2026),
            None,
        )
        cible_gaz_2026 = next(
            (r for r in result.cibles_gaz if r.code_site == site.code_site and r.period_year == 2026),
            None,
        )
        sample_sites.append({
            "code_site": site.code_site,
            "nom_batiment": site.nom_batiment,
            "lot": site.lot,
            "p2_total_2026": p2p3_2026.p2_total_ht if p2p3_2026 else None,
            "p3_total_2026": p2p3_2026.p3_total_ht if p2p3_2026 else None,
            "qt_gaz_cible_2026": cible_gaz_2026.qt_global_mwhpci if cible_gaz_2026 else None,
        })

    # --- Resume financier depuis RECAP MARCHE ---
    def _recap_value(metric: str, year: int | None = None, category: str | None = None) -> float | None:
        for r in result.recap_rows:
            if r.metric == metric and (year is None or r.period_year == year) and (category is None or r.category == category):
                return r.value
        return None

    # Bilan total sur la duree du marche (period_year None, section bilan)
    bilan = {}
    for r in result.recap_rows:
        if r.section == "bilan" and r.metric == "bilan_ht":
            bilan[r.category] = r.value

    # P1/P2/P3 par annee (totaux HT)
    by_year: dict[int, dict[str, float | None]] = {}
    for year in PERIOD_YEARS:
        by_year[year] = {
            "p1_total_ht": _recap_value("p1_total_ht", year),
            "p2_total_ht": _recap_value("p2_total_ht", year),
            "p3_total_ht": _recap_value("p3_total_ht", year),
        }

    recap_summary = {
        "bilan_marche_ht": bilan,
        "by_year": by_year,
        "facteur_co2_gaz": _recap_value("facteur_co2", category="GAZ"),
        "facteur_co2_elec": _recap_value("facteur_co2", category="ELEC"),
    }

    return DalkiaImportPreview(
        lot=result.lot,
        filename=result.filename,
        nb_sites=len(result.sites),
        nb_p2p3_rows=len(result.p2p3_rows),
        nb_cibles_rows=len(result.cibles_gaz) + len(result.cibles_elec),
        nb_p1_gaz_rows=len(result.p1_gaz),
        nb_ape_rows=len(result.ape_rows),
        nb_recap_rows=len(result.recap_rows),
        recap_summary=recap_summary,
        period_labels=result.period_labels,
        sample_sites=sample_sites,
        classified=_build_classified(result),
        warnings=result.warnings,
    )
