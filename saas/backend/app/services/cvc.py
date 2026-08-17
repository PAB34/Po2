import io
import json
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.building import Building
from app.models.city import City
from app.models.cvc import CvcInventoryItem, CvcRefrigerantItem, CvcSourceBuildingMapping
from app.models.equipment import EquipmentReference
from app.models.local import Local
from app.models.site import Site
from app.schemas.cvc import (
    BuildingMatchSuggestion,
    CvcApplySiteMappingsResult,
    CvcImportSiteMatchResponse,
    CvcImportSiteMatchResult,
    CvcBuildingMapping,
    CvcEquipmentReferenceRead,
    CvcImportBatchSummary,
    CvcImportResult,
    CvcInventoryItemRead,
    CvcInventoryItemUpdate,
    CvcMatchBuildingsResponse,
    CvcParcBatiment,
    CvcParcBucket,
    CvcParcCompletude,
    CvcParcFamille,
    CvcParcTechniqueReport,
    CvcRecomputeReferencesResult,
    CvcRefrigerantBatchSummary,
    CvcRefrigerantDashboard,
    CvcRefrigerantDashboardKpi,
    CvcRefrigerantActionSummary,
    CvcRefrigerantImportResult,
    CvcRefrigerantItemRead,
    CvcRefrigerantItemUpdate,
    CvcRefrigerantMatchCandidate,
    CvcInventoryItemCompact,
    CvcSourceBuildingMappingRead,
    CvcSourceBuildingMappingUpdate,
    CvcTechnicalCoverageReport,
    CvcSiteMapping,
    CvcPreviewResponse,
    PatrimoineSiteSuggestion,
    SiteMatchResult,
)

CURRENT_YEAR = datetime.now().year
HEALTH_AGE_RATIOS = {
    "bon": 0.25,
    "moyen": 0.6,
    "mauvais": 0.95,
}
ALLOWED_CVC_REFERENCE_DOMAINS = {"A.2.1", "A.2.2", "A.2.3"}
NO_FUZZY_FAMILIES = {
    "analyseur",
    "appareil de mesure",
    "autre a qualifier",
    "compteur",
    "plomberie",
}
REFRIGERANT_NIVEAU_3 = {"Production de froid :", "Pompes à chaleur Air/Air, Air/Eau, Eau/Eau"}


DEFAULT_ACTION_STATUS = "À créer"


def _add_months(value: date, months: int) -> date:
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    leap = year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
    month_lengths = [31, 29 if leap else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    return date(year, month, min(value.day, month_lengths[month - 1]))


def _fgas_status(teqco2: float | None) -> str:
    if teqco2 is None:
        return "Données à compléter"
    if teqco2 < 5:
        return "Hors seuil contrôle périodique"
    if teqco2 < 50:
        return "Contrôle annuel"
    if teqco2 < 500:
        return "Contrôle semestriel"
    return "Suivi renforcé ≥500"


def _fgas_frequency_months(teqco2: float | None, detection_permanente: bool | None) -> int | None:
    if teqco2 is None or teqco2 < 5:
        return None
    if teqco2 < 50:
        return 24 if detection_permanente else 12
    if teqco2 < 500:
        return 12 if detection_permanente else 6
    return 6 if detection_permanente else 3


def _document_required(fgas_status: str) -> str:
    if fgas_status == "Hors seuil contrôle périodique":
        return "Fiche intervention uniquement si manipulation de fluide"
    if fgas_status == "Données à compléter":
        return "Fiche équipement / plaque signalétique"
    return "Rapport contrôle étanchéité + Cerfa 15497*04 si intervention fluide"


def _conformity_status(item: CvcRefrigerantItem, fgas_status: str, frequency_months: int | None) -> str:
    if fgas_status == "Données à compléter":
        return "Données à compléter"
    if fgas_status == "Hors seuil contrôle périodique":
        return "Non prioritaire"
    if item.dernier_controle_etancheite is None and item.prochaine_echeance is None:
        return "Dernier contrôle à demander"
    due_date = item.prochaine_echeance
    if due_date is None and item.dernier_controle_etancheite and frequency_months:
        due_date = _add_months(item.dernier_controle_etancheite, frequency_months)
    if due_date is None:
        return "Dernier contrôle à demander"
    today = date.today()
    if due_date < today:
        return "En retard"
    if due_date <= today + timedelta(days=60):
        return "À programmer < 60 j"
    return "OK"


def _priority(item: CvcRefrigerantItem, conformity_status: str) -> str:
    esp = _normalize(item.esp_status)
    if (
        conformity_status in {"En retard", "Dernier contrôle à demander", "Données à compléter"}
        or (item.teqco2 is not None and item.teqco2 >= 50)
        or esp == "manque de donnee"
    ):
        return "Haute"
    if conformity_status == "À programmer < 60 j":
        return "Moyenne"
    return "Basse"


def _next_action(conformity_status: str) -> str:
    if conformity_status == "Données à compléter":
        return "Compléter fluide / charge kg / GWP"
    if conformity_status == "Dernier contrôle à demander":
        return "Demander au titulaire le dernier contrôle et la prochaine échéance"
    if conformity_status == "En retard":
        return "Programmer un contrôle d’étanchéité"
    if conformity_status == "À programmer < 60 j":
        return "Créer un OT GMAO de contrôle"
    if conformity_status == "Non prioritaire":
        return "Surveiller uniquement en cas d’intervention"
    return "Conserver les preuves"


def _computed_refrigerant_fields(item: CvcRefrigerantItem) -> dict[str, object]:
    fgas_status = _fgas_status(item.teqco2)
    frequency = _fgas_frequency_months(item.teqco2, item.detection_permanente)
    conformity = _conformity_status(item, fgas_status, frequency)
    return {
        "fgas_status": fgas_status,
        "frequence_controle_mois": frequency,
        "statut_conformite": conformity,
        "action_prioritaire": _next_action(conformity),
        "preuve_attendue": _document_required(fgas_status),
        "priorite": _priority(item, conformity),
    }


def _action_summary(
    item: CvcRefrigerantItem,
    theme: str,
    constat: str,
    action: str | None = None,
) -> CvcRefrigerantActionSummary:
    computed = _computed_refrigerant_fields(item)
    return CvcRefrigerantActionSummary(
        item_id=item.id,
        priority=str(computed["priorite"]),
        theme=theme,
        site=item.site_raw,
        equipment=item.designation,
        constat=constat,
        action=action or str(computed["action_prioritaire"]),
        preuve_attendue=str(computed["preuve_attendue"]),
        responsable=item.titulaire or item.responsable_collectivite,
        echeance_cible=item.prochaine_echeance,
        statut_action=item.statut_action or DEFAULT_ACTION_STATUS,
    )


def _normalize(value: str | None) -> str:
    if not value:
        return ""
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_value.lower()).strip()


def _combined_item_text(
    famille: str | None,
    designation: str | None = None,
    marque: str | None = None,
    modele: str | None = None,
) -> str:
    return _normalize(" ".join(part for part in [famille, designation, marque, modele] if part))


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


# ---------------------------------------------------------------------------
# Rapprochement des libellés de site (inventaire CVC ↔ patrimoine)
#
# La similarité de chaîne brute est doublement défaillante sur ce corpus :
# elle rate des synonymes évidents (« EGLISE SAINT JOSEPH » vs « EGLISE CATHOLIQUE
# ST JOSEPH ») et rapproche avec assurance des bâtiments de nature différente
# partageant un patronyme (« STADE LOUIS MICHEL » vs « RESTAURANT SCOLAIRE LOUISE
# MICHEL »). On normalise donc les alias, et on ajoute un garde-fou sémantique.
# ---------------------------------------------------------------------------

# Préfixes de codification interne des inventaires (« VDS-ENS 17.03 GS - », « CCAS 10 »).
_SITE_CODE_PREFIX = re.compile(r"^(?:vds|ccas|lr)[-\s/][a-z]{0,4}\s*\d+(?:\.\d+)?\s*(?:gs|bam|ens)?\s*[-–]?\s*")

_SITE_ALIASES = {
    "st": "saint",
    "ste": "sainte",
    "sts": "saints",
    "gs": "groupe scolaire",
    "elem": "elementaire",
    "elemtaire": "elementaire",
    "mat": "maternelle",
    "ec": "ecole",
    "bat": "batiment",
    "cplx": "complexe",
    "mun": "municipal",
}

# Mots sans valeur discriminante pour l'identification d'un site.
_SITE_STOPWORDS = {
    "de", "du", "des", "la", "le", "les", "l", "d", "et", "au", "aux", "a", "en",
    "nouveau", "nouveaux", "nouvelle", "nouvelles", "ancien", "ancienne", "yc",
    "municipal", "municipale", "ville", "sete", "mairie",
}

# Nature du bâtiment : deux natures différentes ⇒ ce n'est pas le même site,
# même si les noms propres coïncident.
_SITE_TYPE_WORDS = {
    "stade", "ecole", "eglise", "cimetiere", "restaurant", "salle", "villa",
    "chapelle", "theatre", "gymnase", "piscine", "maison", "complexe", "atelier",
    "musee", "bibliotheque", "mediatheque", "creche", "garderie", "halte",
    "conservatoire", "cinema", "parking", "tennis", "boulodrome", "dojo", "foyer",
    "logement", "hotel", "marche", "vestiaire", "tribune", "magasin", "depot",
    "garage", "chaufferie", "serre", "elementaire", "maternelle", "college",
    "lycee", "cantine", "presbytere", "camping", "port",
}


def _normalize_site_label(value: str | None) -> str:
    """Minuscule sans accent, préfixe de codification retiré, alias déployés."""
    txt = _normalize(value)
    if not txt:
        return ""
    txt = _SITE_CODE_PREFIX.sub("", txt)
    txt = re.sub(r"[^a-z0-9]+", " ", txt)
    tokens = [_SITE_ALIASES.get(token, token) for token in txt.split()]
    return " ".join(" ".join(tokens).split())


def _site_tokens(value: str | None) -> set[str]:
    return {t for t in _normalize_site_label(value).split() if len(t) > 1 and t not in _SITE_STOPWORDS}


def _site_type_tokens(value: str | None) -> set[str]:
    return _site_tokens(value) & _SITE_TYPE_WORDS


def _site_key_tokens(value: str | None) -> set[str]:
    """Tokens distinctifs : ce qui identifie le site, hors nature du bâtiment."""
    return _site_tokens(value) - _SITE_TYPE_WORDS


def _site_labels_compatible(source: str | None, target: str | None) -> bool:
    """Garde-fou : rejette les rapprochements sémantiquement incohérents."""
    types_source = _site_type_tokens(source)
    types_target = _site_type_tokens(target)
    # « STADE ... » ne peut pas être « RESTAURANT SCOLAIRE ... ».
    if types_source and types_target and not (types_source & types_target):
        return False
    keys_source = _site_key_tokens(source)
    keys_target = _site_key_tokens(target)
    # « CIMETIERE MARIN » ne peut pas être « CIMETIERE LE PY ».
    if keys_source and keys_target and not (keys_source & keys_target):
        return False
    return True


def _site_similarity(source: str | None, target: str | None) -> float:
    """Score de rapprochement de deux libellés de site (0 si incompatibles)."""
    normalized_source = _normalize_site_label(source)
    normalized_target = _normalize_site_label(target)
    if not normalized_source or not normalized_target:
        return 0.0
    if not _site_labels_compatible(source, target):
        return 0.0
    sequence = SequenceMatcher(None, normalized_source, normalized_target).ratio()
    tokens_source = _site_tokens(source)
    tokens_target = _site_tokens(target)
    jaccard = len(tokens_source & tokens_target) / len(tokens_source | tokens_target) if tokens_source and tokens_target else 0.0
    keys_source = _site_key_tokens(source)
    # Couverture : part des tokens distinctifs de la source retrouvés dans la cible.
    coverage = len(keys_source & _site_key_tokens(target)) / len(keys_source) if keys_source else 0.0
    return max(sequence, 0.4 * jaccard + 0.6 * coverage)


def _build_address(b: Building) -> str | None:
    parts = [b.numero_voirie, b.nature_voie, b.nom_voie]
    addr = " ".join(p for p in parts if p)
    if b.nom_commune:
        addr = f"{addr}, {b.nom_commune}" if addr else b.nom_commune
    return addr or None


def parse_excel_preview(raw_bytes: bytes) -> CvcPreviewResponse:
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return CvcPreviewResponse(columns=[], total_rows=0, unique_sites=[], unique_families=[], sample_rows=[])

    header = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = [r for r in rows[1:] if any(v is not None for v in r)]

    site_col = next((i for i, c in enumerate(header) if c == "SITE"), None)
    famille_col = next((i for i, c in enumerate(header) if c == "FAMILLE"), None)

    unique_sites: list[str] = []
    unique_families: list[str] = []

    if site_col is not None:
        seen: set[str] = set()
        for r in data_rows:
            v = r[site_col] if site_col < len(r) else None
            if v and str(v) not in seen:
                seen.add(str(v))
                unique_sites.append(str(v))

    if famille_col is not None:
        seen = set()
        for r in data_rows:
            v = r[famille_col] if famille_col < len(r) else None
            if v and str(v) not in seen:
                seen.add(str(v))
                unique_families.append(str(v))

    sample_rows = []
    for r in data_rows[:5]:
        row_dict: dict = {}
        for i, col in enumerate(header):
            if col and i < len(r):
                row_dict[col] = str(r[i]) if r[i] is not None else None
        sample_rows.append(row_dict)

    return CvcPreviewResponse(
        columns=[c for c in header if c],
        total_rows=len(data_rows),
        unique_sites=unique_sites,
        unique_families=unique_families,
        sample_rows=sample_rows,
    )


def _normalize_model(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", _normalize(value))


def _parse_float(value) -> float | None:
    if value is None:
        return None
    cleaned = str(value).strip().replace(" ", "").replace(",", ".")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_int(value) -> int | None:
    parsed = _parse_float(value)
    return int(parsed) if parsed is not None else None


def _header_lookup(header: list[str]) -> dict[str, int]:
    return {_normalize(col): idx for idx, col in enumerate(header) if col}


def _get_header_value(row: tuple, header_lookup: dict[str, int], name: str):
    idx = header_lookup.get(_normalize(name))
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _clean_cell(value) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


PROVIDER_DALKIA = "DALKIA"
PROVIDER_SPIE = "SPIE"

# Colonnes caractéristiques de chaque format d'export terrain.
_DALKIA_HEADER_MARKERS = ("site", "designation", "qte qte relevee")
_SPIE_HEADER_MARKERS = ("nom du batiment", "nom complet de l equipement (type + complementaire)")


def detect_inventory_provider(header: list[str]) -> str:
    """Devine le prestataire à partir des en-têtes du fichier."""
    normalized = {_normalize(col) for col in header if col}
    if any(marker in normalized for marker in _SPIE_HEADER_MARKERS):
        return PROVIDER_SPIE
    return PROVIDER_DALKIA


def _extract_building_name(raw: str | None) -> str | None:
    """LR/34/SETE/<NOM>-MAIRIE → <NOM> (idempotent, tolère tab/espaces)."""
    if not raw:
        return None
    name = str(raw).strip()
    name = re.sub(r"^[A-Z]{2}/\d+/[A-Z]+/", "", name)
    name = re.sub(r"\s*-\s*MAIRIE\s*$", "", name)
    return name.strip() or None


def _parse_mes_year(value) -> int | None:
    """Extrait l'année d'une date hétérogène (MM/YYYY, YYYY, DD/MM/YYYY, datetime)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)):
        year = int(value)
        return year if 1900 <= year <= CURRENT_YEAR + 1 else None
    match = re.search(r"(19|20)\d{2}", str(value))
    if not match:
        return None
    year = int(match.group(0))
    return year if 1900 <= year <= CURRENT_YEAR + 1 else None


def _read_refrigerant_rows(raw_bytes: bytes) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    lookup = _header_lookup(header)
    years = [str(year) for year in range(2026, 2043)]
    parsed_rows: list[dict] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        designation = _clean_cell(_get_header_value(row, lookup, "DESIGNATION"))
        if not designation:
            continue
        schedule = {
            year: value
            for year in years
            if (value := _clean_cell(_get_header_value(row, lookup, year)))
        }
        parsed_rows.append(
            {
                "row_number": row_number,
                "site_raw": _clean_cell(_get_header_value(row, lookup, "SITE")),
                "designation": designation,
                "quantite_relevee": _parse_int(_get_header_value(row, lookup, "QTE QTE RELEVEE")),
                "famille": _clean_cell(_get_header_value(row, lookup, "FAMILLE")),
                "marque": _clean_cell(_get_header_value(row, lookup, "MARQUE")),
                "modele": _clean_cell(_get_header_value(row, lookup, "MODELE")),
                "fluide_frigorigene": _clean_cell(_get_header_value(row, lookup, "Fluide frigo")),
                "quantite_fluide_kg": _parse_float(_get_header_value(row, lookup, "Quantite fluide en Kg")),
                "puissance_froid_kw": _parse_float(_get_header_value(row, lookup, "Puissance Froid en Kw")),
                "date_mis_en_service": _parse_int(_get_header_value(row, lookup, "DATE MES")),
                "gwp": _parse_float(_get_header_value(row, lookup, "GWP")),
                "teqco2": _parse_float(_get_header_value(row, lookup, "tEqCO2")),
                "esp_status": _clean_cell(_get_header_value(row, lookup, "ESP")),
                "cout_desp_date_eur": _parse_float(_get_header_value(row, lookup, "COUT DESP a Date")),
                "cumul_5_ans_eur": _parse_float(_get_header_value(row, lookup, "CUMUL SUR 5 ANS")),
                "schedule_json": json.dumps(schedule, ensure_ascii=False) if schedule else None,
            }
        )
    return header, parsed_rows


def _inventory_compact(item: CvcInventoryItem) -> CvcInventoryItemCompact:
    return CvcInventoryItemCompact(
        id=item.id,
        site_raw=item.site_raw,
        designation=item.designation,
        famille=item.famille,
        marque=item.marque,
        modele=item.modele,
        date_mis_en_service=item.date_mis_en_service,
        import_batch=item.import_batch,
    )


def _refrigerant_key(data: dict, parts: tuple[str, ...]) -> tuple[str, ...]:
    values: list[str] = []
    for part in parts:
        if part == "modele":
            values.append(_normalize_model(data.get("modele")))
        elif part == "date":
            values.append(str(data.get("date_mis_en_service") or ""))
        else:
            values.append(_normalize(data.get(part)))
    return tuple(values)


def _inventory_key(item: CvcInventoryItem, parts: tuple[str, ...]) -> tuple[str, ...]:
    values: list[str] = []
    for part in parts:
        if part == "modele":
            values.append(_normalize_model(item.modele))
        elif part == "date":
            values.append(str(item.date_mis_en_service or ""))
        else:
            values.append(_normalize(getattr(item, part)))
    return tuple(values)


def _score_inventory_candidate(data: dict, item: CvcInventoryItem) -> tuple[float, str] | None:
    if _normalize(data.get("site_raw")) != _normalize(item.site_raw):
        return None
    designation_score = _similarity(_normalize(data.get("designation")), _normalize(item.designation))
    model_left = _normalize_model(data.get("modele"))
    model_right = _normalize_model(item.modele)
    model_score = _similarity(model_left, model_right) if model_left and model_right else 0.0
    brand_score = _similarity(_normalize(data.get("marque")), _normalize(item.marque)) if data.get("marque") else 0.0
    score = round(designation_score * 0.65 + model_score * 0.25 + brand_score * 0.1, 3)
    return (score, "fuzzy_same_site") if score >= 0.82 else None


def _find_refrigerant_candidates(
    data: dict, inventory_items: list[CvcInventoryItem], limit: int = 5
) -> list[CvcRefrigerantMatchCandidate]:
    key_defs: list[tuple[str, tuple[str, ...], float]] = [
        ("site+designation+famille+marque+modele+date", ("site_raw", "designation", "famille", "marque", "modele", "date"), 1.0),
        ("site+designation+famille+modele+date", ("site_raw", "designation", "famille", "modele", "date"), 0.97),
        ("site+designation+modele+date", ("site_raw", "designation", "modele", "date"), 0.94),
        ("site+designation+marque+modele", ("site_raw", "designation", "marque", "modele"), 0.93),
    ]

    candidates: list[tuple[float, str, CvcInventoryItem]] = []
    for method, parts, score in key_defs:
        key = _refrigerant_key(data, parts)
        if not all(key):
            continue
        exact = [item for item in inventory_items if _inventory_key(item, parts) == key]
        if exact:
            candidates.extend((score, method, item) for item in exact)
            break

    if not candidates:
        for item in inventory_items:
            scored = _score_inventory_candidate(data, item)
            if scored:
                score, method = scored
                candidates.append((score, method, item))

    dedup: dict[int, tuple[float, str, CvcInventoryItem]] = {}
    for score, method, item in sorted(candidates, key=lambda row: row[0], reverse=True):
        dedup.setdefault(item.id, (score, method, item))

    return [
        CvcRefrigerantMatchCandidate(item=_inventory_compact(item), score=score, method=method)
        for score, method, item in list(dedup.values())[:limit]
    ]


def _select_auto_refrigerant_candidate(
    candidates: list[CvcRefrigerantMatchCandidate],
) -> CvcRefrigerantMatchCandidate | None:
    if not candidates:
        return None
    if candidates[0].score < 0.93:
        return None
    if len(candidates) > 1 and candidates[0].score == candidates[1].score:
        return None
    return candidates[0]


def _clear_current_cvc_inventory(db: Session, city_id: int | None, provider: str | None = None) -> None:
    inventory_stmt = select(CvcInventoryItem)
    mapping_stmt = select(CvcSourceBuildingMapping).where(CvcSourceBuildingMapping.source_type == "inventory")
    refrigerant_stmt = select(CvcRefrigerantItem)
    if provider is not None:
        inventory_stmt = inventory_stmt.where(CvcInventoryItem.provider == provider)
    if city_id is not None:
        inventory_stmt = inventory_stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        mapping_stmt = mapping_stmt.where((CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None)))
        refrigerant_stmt = refrigerant_stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))

    items_to_delete = list(db.scalars(inventory_stmt))
    deleted_item_ids = {item.id for item in items_to_delete}
    deleted_batches = {item.import_batch for item in items_to_delete if item.import_batch}

    for refrigerant in db.scalars(refrigerant_stmt):
        # Ne délier que les fluides rattachés aux items du provider purgé.
        if provider is not None and refrigerant.cvc_inventory_item_id not in deleted_item_ids:
            continue
        refrigerant.cvc_inventory_item_id = None
        refrigerant.match_status = "pending"
        refrigerant.match_method = None
        refrigerant.match_score = None

    for mapping in db.scalars(mapping_stmt):
        # En purge ciblée, ne supprimer que les mappings des lots du provider.
        if provider is not None and mapping.import_batch not in deleted_batches:
            continue
        db.delete(mapping)
    for item in items_to_delete:
        db.delete(item)
    db.flush()


def _refresh_refrigerant_inventory_links(db: Session, city_id: int | None) -> int:
    inventory_stmt = select(CvcInventoryItem)
    refrigerant_stmt = select(CvcRefrigerantItem)
    if city_id is not None:
        inventory_stmt = inventory_stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        refrigerant_stmt = refrigerant_stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))

    inventory_items = list(db.scalars(inventory_stmt))
    inventory_by_id = {item.id: item for item in inventory_items}
    relinked = 0
    for refrigerant in db.scalars(refrigerant_stmt):
        candidates = _find_refrigerant_candidates(
            {
                "site_raw": refrigerant.site_raw,
                "designation": refrigerant.designation,
                "famille": refrigerant.famille,
                "marque": refrigerant.marque,
                "modele": refrigerant.modele,
                "date_mis_en_service": refrigerant.date_mis_en_service,
            },
            inventory_items,
        )
        auto_candidate = _select_auto_refrigerant_candidate(candidates)
        if auto_candidate is None:
            refrigerant.cvc_inventory_item_id = None
            refrigerant.match_status = "ambiguous" if candidates else "pending"
            refrigerant.match_method = None
            refrigerant.match_score = None
            continue
        matched_inventory = inventory_by_id.get(auto_candidate.item.id)
        if matched_inventory is None:
            continue
        refrigerant.cvc_inventory_item_id = matched_inventory.id
        refrigerant.site_id = matched_inventory.site_id
        refrigerant.building_id = matched_inventory.building_id
        refrigerant.match_status = "auto_matched"
        refrigerant.match_method = auto_candidate.method
        refrigerant.match_score = auto_candidate.score
        if refrigerant.quantite_fluide_kg is not None:
            matched_inventory.quantite_fluide_frigorigene = refrigerant.quantite_fluide_kg
        relinked += 1
    return relinked


def _latest_inventory_batch(items: list[CvcInventoryItem]) -> str | None:
    batches: dict[str, float] = {}
    for item in items:
        if item.import_batch:
            created_at = item.created_at.timestamp() if item.created_at else 0.0
            current = batches.get(item.import_batch)
            if current is None or created_at < current:
                batches[item.import_batch] = created_at
    if not batches:
        return None
    return max(batches.items(), key=lambda item: item[1])[0]


def _filter_latest_inventory_batch(items: list[CvcInventoryItem]) -> list[CvcInventoryItem]:
    """Garde le dernier lot importé pour CHAQUE provider (DALKIA et SPIE coexistent)."""
    by_provider: dict[str, list[CvcInventoryItem]] = defaultdict(list)
    for item in items:
        by_provider[item.provider or PROVIDER_DALKIA].append(item)

    kept: list[CvcInventoryItem] = []
    for provider_items in by_provider.values():
        latest_batch = _latest_inventory_batch(provider_items)
        if latest_batch is None:
            continue
        kept.extend(item for item in provider_items if item.import_batch == latest_batch)
    return kept


def match_buildings_for_sites(
    db: Session, sites: list[str], city_id: int | None
) -> CvcMatchBuildingsResponse:
    stmt = select(Building)
    if city_id is not None:
        stmt = stmt.where(Building.city_id == city_id)
    buildings = list(db.scalars(stmt))

    results = []
    for site in sites:
        scored: list[tuple[float, Building]] = []
        for b in buildings:
            name = b.nom_batiment or ""
            score = _similarity(site, name)
            scored.append((score, b))
        scored.sort(key=lambda x: x[0], reverse=True)

        top = scored[:5]
        suggestions = [
            BuildingMatchSuggestion(
                building_id=b.id,
                site_id=b.site_id,
                nom_batiment=b.nom_batiment,
                adresse=_build_address(b),
                score=round(s, 3),
            )
            for s, b in top
            if s > 0.1
        ]
        auto_id = top[0][1].id if top and top[0][0] >= 0.65 else None
        results.append(SiteMatchResult(site_raw=site, suggestions=suggestions, auto_selected_id=auto_id))

    return CvcMatchBuildingsResponse(matches=results)


def _best_current_id(values: list[int | None]) -> int | None:
    counts: dict[int, int] = {}
    for value in values:
        if value is not None:
            counts[value] = counts.get(value, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda item: item[1])[0]


def _dedupe_ids(values: list[int | None]) -> list[int]:
    result: list[int] = []
    seen: set[int] = set()
    for value in values:
        if value is None or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def _read_mapping_building_ids(mapping: CvcSourceBuildingMapping) -> list[int]:
    if mapping.building_ids_json:
        try:
            values = json.loads(mapping.building_ids_json)
        except json.JSONDecodeError:
            values = []
        if isinstance(values, list):
            return _dedupe_ids([value if isinstance(value, int) else None for value in values])
    return [mapping.building_id] if mapping.building_id is not None else []


def _write_mapping_building_ids(mapping: CvcSourceBuildingMapping, building_ids: list[int]) -> None:
    clean_ids = _dedupe_ids(building_ids)
    mapping.building_ids_json = json.dumps(clean_ids) if clean_ids else None
    mapping.building_id = clean_ids[0] if len(clean_ids) == 1 else None


def _create_or_reuse_cvc_placeholder_building(
    db: Session,
    source_name: str,
    city_id: int | None,
) -> Building:
    normalized_source = _normalize(source_name)
    stmt = select(Building).where(Building.source_creation == "CVC_IMPORT", Building.site_id.is_(None))
    if city_id is not None:
        stmt = stmt.where(Building.city_id == city_id)
    for building in db.scalars(stmt):
        if _normalize(building.nom_batiment) == normalized_source:
            return building

    city = db.get(City, city_id) if city_id is not None else None
    building = Building(
        city_id=city_id,
        site_id=None,
        nom_batiment=source_name[:255],
        nom_commune=city.nom_commune if city else "A qualifier",
        adresse_reconstituee=None,
        source_creation="CVC_IMPORT",
        statut_geocodage="NON_FAIT",
    )
    db.add(building)
    db.flush()
    return building


def _suggest_source_building(
    source_site_raw: str,
    sites: list[Site],
    buildings: list[Building],
) -> tuple[int | None, int | None, float | None, str | None]:
    site_scored = sorted(
        ((_site_similarity(source_site_raw, site.nom_site), site) for site in sites),
        key=lambda item: item[0],
        reverse=True,
    )
    building_scored = sorted(
        (
            (
                max(
                    _site_similarity(source_site_raw, building.nom_batiment or ""),
                    _site_similarity(source_site_raw, building.adresse_reconstituee or ""),
                ),
                building,
            )
            for building in buildings
        ),
        key=lambda item: item[0],
        reverse=True,
    )
    if building_scored and building_scored[0][0] >= 0.72:
        building = building_scored[0][1]
        return building.site_id, building.id, round(building_scored[0][0], 3), "auto_building"
    if site_scored and site_scored[0][0] >= 0.72:
        site = site_scored[0][1]
        return site.id, None, round(site_scored[0][0], 3), "auto_site"
    return None, None, None, None


def _mapping_suggestions(
    source_site_raw: str,
    sites: list[Site],
    buildings: list[Building],
) -> tuple[list[PatrimoineSiteSuggestion], list[BuildingMatchSuggestion]]:
    site_scored = sorted(
        ((_site_similarity(source_site_raw, site.nom_site), site) for site in sites),
        key=lambda item: item[0],
        reverse=True,
    )
    building_scored = sorted(
        (
            (
                max(
                    _site_similarity(source_site_raw, building.nom_batiment or ""),
                    _site_similarity(source_site_raw, building.adresse_reconstituee or ""),
                ),
                building,
            )
            for building in buildings
        ),
        key=lambda item: item[0],
        reverse=True,
    )
    site_suggestions = [
        PatrimoineSiteSuggestion(
            site_id=site.id,
            nom_site=site.nom_site,
            adresse=site.adresse,
            score=round(score, 3),
        )
        for score, site in site_scored[:5]
        if score > 0.1
    ]
    building_suggestions = [
        BuildingMatchSuggestion(
            building_id=building.id,
            site_id=building.site_id,
            nom_batiment=building.nom_batiment,
            adresse=_build_address(building),
            score=round(score, 3),
        )
        for score, building in building_scored[:5]
        if score > 0.1
    ]
    return site_suggestions, building_suggestions


def ensure_cvc_source_building_mappings(db: Session, city_id: int | None) -> int:
    sites_stmt = select(Site)
    buildings_stmt = select(Building)
    inventory_stmt = select(CvcInventoryItem)
    refrigerant_stmt = select(CvcRefrigerantItem)
    mapping_stmt = select(CvcSourceBuildingMapping)
    if city_id is not None:
        sites_stmt = sites_stmt.where(Site.city_id == city_id)
        buildings_stmt = buildings_stmt.where(Building.city_id == city_id)
        inventory_stmt = inventory_stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        refrigerant_stmt = refrigerant_stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
        mapping_stmt = mapping_stmt.where((CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None)))

    sites = list(db.scalars(sites_stmt))
    buildings = list(db.scalars(buildings_stmt))
    existing = {
        (mapping.source_type, mapping.import_batch, _normalize(mapping.source_site_raw))
        for mapping in db.scalars(mapping_stmt)
    }
    sources: set[tuple[str, str, str]] = set()
    for item in db.scalars(inventory_stmt):
        if item.import_batch and item.site_raw:
            sources.add(("inventory", item.import_batch, item.site_raw.strip()))
    for item in db.scalars(refrigerant_stmt):
        if item.import_batch and item.site_raw:
            sources.add(("refrigerant", item.import_batch, item.site_raw.strip()))

    created = 0
    created_mappings: list[CvcSourceBuildingMapping] = []
    for source_type, import_batch, source_site_raw in sorted(sources):
        key = (source_type, import_batch, _normalize(source_site_raw))
        if key in existing:
            continue
        site_id, building_id, score, method = _suggest_source_building(source_site_raw, sites, buildings)
        status = "matched" if site_id or building_id else "to_review"
        mapping = CvcSourceBuildingMapping(
            city_id=city_id,
            source_type=source_type,
            import_batch=import_batch,
            source_site_raw=source_site_raw,
            site_id=site_id,
            building_id=building_id,
            status=status,
            match_score=score,
            match_method=method,
        )
        db.add(mapping)
        created_mappings.append(mapping)
        created += 1
    if created:
        db.flush()
        for mapping in created_mappings:
            if mapping.status == "matched":
                _apply_source_mapping_to_rows(db, mapping)
        db.commit()
    return created


def _apply_source_mapping_to_rows(db: Session, mapping: CvcSourceBuildingMapping) -> int:
    updated = 0
    selected_building_ids = _read_mapping_building_ids(mapping)
    buildings = [building for building_id in selected_building_ids if (building := db.get(Building, building_id))]
    # Un libellé source peut couvrir plusieurs bâtiments (ex. « Élémentaire LA RENAISSANCE
    # + restaurant scolaire »). Décision 2026-08-17 : on rattache alors au **bâtiment
    # principal** (le premier déclaré) plutôt que de laisser les équipements orphelins ;
    # le périmètre complet reste tracé dans `building_ids_json`.
    primary_building = buildings[0] if buildings else None
    is_multi_building = len(buildings) > 1
    selected_site_ids = {building.site_id for building in buildings if building.site_id is not None}
    next_site_id = mapping.site_id
    if next_site_id is None and len(selected_site_ids) == 1:
        next_site_id = next(iter(selected_site_ids))
    if next_site_id is None and primary_building is not None:
        next_site_id = primary_building.site_id
    if mapping.source_type == "inventory":
        rows = list(
            db.scalars(
                select(CvcInventoryItem).where(
                    CvcInventoryItem.import_batch == mapping.import_batch,
                    CvcInventoryItem.site_raw == mapping.source_site_raw,
                )
            )
        )
        for item in rows:
            item.site_id = next_site_id
            item.building_id = primary_building.id if primary_building else None
            # Le local ciblé n'a de sens que pour un rattachement à un bâtiment unique.
            if primary_building is None or is_multi_building:
                item.local_id = None
            linked_refrigerants = list(
                db.scalars(select(CvcRefrigerantItem).where(CvcRefrigerantItem.cvc_inventory_item_id == item.id))
            )
            for refrigerant in linked_refrigerants:
                refrigerant.site_id = item.site_id
                refrigerant.building_id = item.building_id
            updated += 1
    elif mapping.source_type == "refrigerant":
        rows = list(
            db.scalars(
                select(CvcRefrigerantItem).where(
                    CvcRefrigerantItem.import_batch == mapping.import_batch,
                    CvcRefrigerantItem.site_raw == mapping.source_site_raw,
                )
            )
        )
        for item in rows:
            item.site_id = next_site_id
            item.building_id = primary_building.id if primary_building else None
            updated += 1
    return updated


def reapply_source_building_mappings(db: Session, city_id: int | None) -> dict:
    """Re-propage tous les rattachements déjà résolus vers les équipements.

    Utile après une évolution de la règle de propagation (ex. prise en compte du
    bâtiment principal pour les libellés multi-bâtiments) : les mappings existants
    sont conservés, seules les lignes d'inventaire/fluides sont remises à jour.
    """
    stmt = select(CvcSourceBuildingMapping)
    if city_id is not None:
        stmt = stmt.where(
            (CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None))
        )
    mappings = list(db.scalars(stmt))
    applied = 0
    rows = 0
    for mapping in mappings:
        if not _read_mapping_building_ids(mapping) and mapping.site_id is None:
            continue  # rien à propager tant que la cible n'est pas choisie
        rows += _apply_source_mapping_to_rows(db, mapping)
        applied += 1
    db.commit()
    return {"mappings_total": len(mappings), "mappings_applied": applied, "rows_updated": rows}


def list_site_matches_for_import(
    db: Session, import_batch: str, city_id: int | None
) -> CvcImportSiteMatchResponse:
    ensure_cvc_source_building_mappings(db, city_id)
    stmt = select(CvcInventoryItem).where(CvcInventoryItem.import_batch == import_batch)
    if city_id is not None:
        stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
    items = _filter_latest_inventory_batch(list(db.scalars(stmt)))
    mapping_stmt = select(CvcSourceBuildingMapping).where(
        CvcSourceBuildingMapping.import_batch == import_batch,
        CvcSourceBuildingMapping.source_type == "inventory",
    )
    if city_id is not None:
        mapping_stmt = mapping_stmt.where(
            (CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None))
        )
    source_mappings = {
        _normalize(mapping.source_site_raw): mapping
        for mapping in db.scalars(mapping_stmt)
    }

    sites_stmt = select(Site)
    buildings_stmt = select(Building)
    if city_id is not None:
        sites_stmt = sites_stmt.where(Site.city_id == city_id)
        buildings_stmt = buildings_stmt.where(Building.city_id == city_id)
    patrimoine_sites = list(db.scalars(sites_stmt))
    buildings = list(db.scalars(buildings_stmt))

    grouped: dict[str, list[CvcInventoryItem]] = {}
    for item in items:
        key = (item.site_raw or "").strip()
        if key:
            grouped.setdefault(key, []).append(item)

    results: list[CvcImportSiteMatchResult] = []
    for site_raw, site_items in grouped.items():
        source_mapping = source_mappings.get(_normalize(site_raw))
        current_building_ids = (
            _read_mapping_building_ids(source_mapping)
            if source_mapping is not None
            else _dedupe_ids([item.building_id for item in site_items])
        )
        site_scored = sorted(
            ((_similarity(site_raw, site.nom_site), site) for site in patrimoine_sites),
            key=lambda item: item[0],
            reverse=True,
        )
        building_scored = sorted(
            (
                (
                    max(
                        _similarity(site_raw, building.nom_batiment or ""),
                        _similarity(site_raw, building.adresse_reconstituee or ""),
                    ),
                    building,
                )
                for building in buildings
            ),
            key=lambda item: item[0],
            reverse=True,
        )

        site_suggestions = [
            PatrimoineSiteSuggestion(
                site_id=site.id,
                nom_site=site.nom_site,
                adresse=site.adresse,
                score=round(score, 3),
            )
            for score, site in site_scored[:5]
            if score > 0.1
        ]
        building_suggestions = [
            BuildingMatchSuggestion(
                building_id=building.id,
                site_id=building.site_id,
                nom_batiment=building.nom_batiment,
                adresse=_build_address(building),
                score=round(score, 3),
            )
            for score, building in building_scored[:5]
            if score > 0.1
        ]

        auto_site_id = site_scored[0][1].id if site_scored and site_scored[0][0] >= 0.72 else None
        auto_building_id = building_scored[0][1].id if building_scored and building_scored[0][0] >= 0.72 else None
        if auto_building_id is not None and auto_site_id is None:
            auto_site_id = building_scored[0][1].site_id

        results.append(
            CvcImportSiteMatchResult(
                site_raw=site_raw,
                item_count=len(site_items),
                current_site_id=source_mapping.site_id if source_mapping is not None else _best_current_id([item.site_id for item in site_items]),
                current_building_id=current_building_ids[0] if len(current_building_ids) == 1 else None,
                current_building_ids=current_building_ids,
                site_suggestions=site_suggestions,
                building_suggestions=building_suggestions,
                auto_site_id=auto_site_id,
                auto_building_id=auto_building_id,
            )
        )

    return CvcImportSiteMatchResponse(matches=sorted(results, key=lambda item: item.site_raw.lower()))


def apply_site_mappings_to_import(
    db: Session, import_batch: str, mappings: list[CvcSiteMapping], city_id: int | None
) -> CvcApplySiteMappingsResult:
    updated = 0
    applied = 0

    for mapping in mappings:
        site_raw = mapping.site_raw.strip()
        if not site_raw:
            continue

        create_building = bool(mapping.create_building)
        requested_building_ids = _dedupe_ids(mapping.building_ids or ([mapping.building_id] if mapping.building_id else []))
        site = None if create_building else (db.get(Site, mapping.site_id) if mapping.site_id is not None else None)
        if not create_building and mapping.site_id is not None and site is None:
            raise ValueError(f"Site introuvable pour {site_raw}.")
        if site and city_id is not None and site.city_id != city_id:
            raise ValueError(f"Site hors perimetre pour {site_raw}.")

        buildings: list[Building] = []
        for building_id in requested_building_ids:
            building = db.get(Building, building_id)
            if building is None:
                raise ValueError(f"Batiment introuvable pour {site_raw}.")
            if city_id is not None and building.city_id != city_id:
                raise ValueError(f"Batiment hors perimetre pour {site_raw}.")
            buildings.append(building)

        create_building_names = [
            name.strip()
            for name in (mapping.create_building_names or ([mapping.create_building_name] if mapping.create_building_name else []))
            if name and name.strip()
        ]
        if create_building and not create_building_names:
            create_building_names = [site_raw]
        for building_name in create_building_names:
            created_building = _create_or_reuse_cvc_placeholder_building(db, building_name, city_id)
            if created_building.id not in requested_building_ids:
                requested_building_ids.append(created_building.id)
                buildings.append(created_building)

        single_building = buildings[0] if len(buildings) == 1 else None
        selected_site_ids = {building.site_id for building in buildings if building.site_id is not None}

        next_site_id = site.id if site else None
        if next_site_id is None and len(selected_site_ids) == 1:
            next_site_id = next(iter(selected_site_ids))
        if next_site_id is None and single_building is not None:
            next_site_id = single_building.site_id
        if site:
            for building in buildings:
                if building.site_id not in (None, site.id):
                    raise ValueError(f"Le batiment choisi n'appartient pas au site choisi pour {site_raw}.")

        source_mapping = db.scalar(
            select(CvcSourceBuildingMapping).where(
                CvcSourceBuildingMapping.source_type == "inventory",
                CvcSourceBuildingMapping.import_batch == import_batch,
                CvcSourceBuildingMapping.source_site_raw == site_raw,
            )
        )
        if source_mapping is None:
            source_mapping = CvcSourceBuildingMapping(
                city_id=city_id,
                source_type="inventory",
                import_batch=import_batch,
                source_site_raw=site_raw,
            )
            db.add(source_mapping)
        source_mapping.site_id = next_site_id
        _write_mapping_building_ids(source_mapping, requested_building_ids)
        source_mapping.status = "matched" if requested_building_ids or next_site_id else "to_review"
        source_mapping.match_method = "manual"
        source_mapping.match_score = 1.0 if requested_building_ids or next_site_id else None

        stmt = select(CvcInventoryItem).where(
            CvcInventoryItem.import_batch == import_batch,
            CvcInventoryItem.site_raw == site_raw,
        )
        if city_id is not None:
            stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        items = list(db.scalars(stmt))
        if not items:
            continue

        for item in items:
            building_changed = item.building_id != (single_building.id if single_building else None)
            item.site_id = next_site_id
            item.building_id = single_building.id if single_building else None
            if building_changed:
                item.local_id = None
            linked_refrigerants = list(
                db.scalars(select(CvcRefrigerantItem).where(CvcRefrigerantItem.cvc_inventory_item_id == item.id))
            )
            for refrigerant in linked_refrigerants:
                refrigerant.site_id = item.site_id
                refrigerant.building_id = item.building_id
            updated += 1
        applied += 1

    db.commit()
    return CvcApplySiteMappingsResult(updated=updated, mappings_applied=applied)


def _find_ref(
    all_refs: list[EquipmentReference],
    *,
    id_ligne: int | None = None,
    equipment_contains: str | None = None,
    code_niveau_2: str | None = None,
) -> EquipmentReference | None:
    normalized_contains = _normalize(equipment_contains)
    for ref in all_refs:
        if id_ligne is not None and ref.id_ligne != id_ligne:
            continue
        if code_niveau_2 is not None and ref.code_niveau_2 != code_niveau_2:
            continue
        if normalized_contains and normalized_contains not in _normalize(ref.equipement):
            continue
        return ref
    return None


def _resolve_alias_reference(
    famille: str | None,
    designation: str | None,
    marque: str | None,
    modele: str | None,
    all_refs: list[EquipmentReference],
) -> EquipmentReference | None:
    family = _normalize(famille)
    text = _combined_item_text(famille, designation, marque, modele)

    if family in {"analyseur", "appareil de mesure", "compteur", "plomberie"}:
        return None

    if "armoire electrique" in family or "tableau electrique" in text or "coffret electrique" in text:
        return _find_ref(all_refs, id_ligne=118, equipment_contains="Armoire électrique")

    if family in {"centrale traitement air"} or "cta" in text or "centrale traitement air" in text:
        return _find_ref(all_refs, id_ligne=221, equipment_contains="CTA")

    if family == "chaudiere":
        if "murale" in text:
            return _find_ref(all_refs, id_ligne=169, equipment_contains="Chaudi")
        return _find_ref(all_refs, id_ligne=163, equipment_contains="Chaudi")

    if family in {"preparateur ecs", "ballon de stockage"}:
        return _find_ref(all_refs, id_ligne=97, equipment_contains="Ballon")

    if family == "vase expansion" or "vase d expansion" in text or "vase expansion" in text:
        return _find_ref(all_refs, id_ligne=188, equipment_contains="Vase")

    if family == "gtb / gtc" or re.search(r"\bgtb\b|\bgtc\b", text):
        return _find_ref(all_refs, id_ligne=162, equipment_contains="GTB")

    if family == "circulateur":
        return _find_ref(all_refs, id_ligne=186, equipment_contains="Circulateur")

    if family == "echangeur":
        if "froid" in text or "glacee" in text:
            return _find_ref(all_refs, id_ligne=213, equipment_contains="Echangeur")
        return _find_ref(all_refs, id_ligne=179, equipment_contains="Echangeur")

    if family == "robinet / vanne" or "vanne" in text or "robinet" in text:
        return _find_ref(all_refs, id_ligne=181, equipment_contains="Vannes")

    if family in {"pompe a chaleur", "groupe thermodynamique"} or (
        family in {"", "autre a qualifier"} and ("pac" in text or "thermodynamique" in text)
    ):
        return _find_ref(all_refs, id_ligne=238, equipment_contains="PAC")

    split_keywords = [
        "mono-split",
        "monosplit",
        "multi-split",
        "multisplit",
        "split",
        "ue clim",
        "ui clim",
        "unite interieure",
        "unite exterieure",
        "climatisation",
        "climatiseur",
        "cassette",
        "vrv",
        "drv",
    ]
    split_like = any(token in text for token in split_keywords)
    if family in {"split system", "vrv", "systeme vrv", "climatiseur", "cassette"} or (
        family in {"", "autre a qualifier"} and split_like
    ):
        if "armoire de climatisation" in text or "roof" in text:
            return _find_ref(all_refs, id_ligne=237, equipment_contains="Armoires autonomes")
        return _find_ref(all_refs, id_ligne=236, equipment_contains="Split")

    if family in {"groupe froid"} or "groupe froid" in text:
        return _find_ref(all_refs, id_ligne=207, equipment_contains="Groupe")

    if family == "aerotherme" or "aerotherme" in text:
        return _find_ref(all_refs, id_ligne=201, equipment_contains="Aérothermes") or _find_ref(
            all_refs, id_ligne=219, equipment_contains="Rideau"
        )

    if family == "bouches de soufflage" or "bouche" in text:
        return _find_ref(all_refs, id_ligne=223, equipment_contains="Gaine")

    if family == "ventilation":
        if "desenfum" in text:
            return _find_ref(all_refs, id_ligne=244, equipment_contains="Ventilateurs")
        if "vmc" in text or "extract" in text or "ventil" in text:
            return _find_ref(all_refs, id_ligne=233, equipment_contains="Ventilateur")
        return None

    if family == "filtre":
        if "sable" in text or "piscine" in text:
            return _find_ref(all_refs, id_ligne=92, equipment_contains="Filtre")
        if "pot a boue" in text or "chauffage" in text or "vanne" in text:
            return _find_ref(all_refs, id_ligne=181, equipment_contains="filtres")
        return None

    if family == "pompe":
        if "doseuse" in text:
            return _find_ref(all_refs, id_ligne=91, equipment_contains="Pompe doseuse")
        if "froid" in text or "glacee" in text:
            return _find_ref(all_refs, id_ligne=214, equipment_contains="Pompes")
        if "chauffage" in text or "circulateur" in text or "radiateur" in text:
            return _find_ref(all_refs, id_ligne=186, equipment_contains="Circulateur")
        return None

    if family == "batterie":
        if "chaude" in text or "froide" in text or "cta" in text:
            return _find_ref(all_refs, id_ligne=221, equipment_contains="CTA")
        return None

    if family == "regulation":
        return _find_ref(all_refs, id_ligne=200, equipment_contains="régulation")

    if family == "tube radiant":
        return _find_ref(all_refs, id_ligne=194, equipment_contains="Radiateur")

    return None


def _resolve_family(
    famille: str | None,
    all_refs: list[EquipmentReference],
    cache: dict,
    designation: str | None = None,
    marque: str | None = None,
    modele: str | None = None,
) -> EquipmentReference | None:
    if not famille:
        return None
    cache_key = (
        _normalize(famille),
        _normalize(designation),
        _normalize(marque),
        _normalize(modele),
    )
    if cache_key in cache:
        return cache[cache_key]

    alias_ref = _resolve_alias_reference(famille, designation, marque, modele, all_refs)
    if alias_ref is not None:
        cache[cache_key] = alias_ref
        return alias_ref

    normalized_family = _normalize(famille)
    if normalized_family in NO_FUZZY_FAMILIES:
        cache[cache_key] = None
        return None

    best_score = 0.0
    best_ref = None
    for ref in all_refs:
        if ref.code_niveau_2 not in ALLOWED_CVC_REFERENCE_DOMAINS:
            continue
        score = max(
            _similarity(normalized_family, _normalize(ref.equipement)),
            _similarity(normalized_family, _normalize(ref.niveau_4)),
        )
        if score > best_score:
            best_score = score
            best_ref = ref
    result = best_ref if best_score >= 0.72 else None
    cache[cache_key] = result
    return result


def _requires_refrigerant_quantity(ref: EquipmentReference | None) -> bool:
    return bool(ref and ref.niveau_3 in REFRIGERANT_NIVEAU_3)


def _health_age_ratio(etat_sante: str | None) -> float | None:
    normalized = _normalize(etat_sante)
    if not normalized:
        return None
    for key, ratio in HEALTH_AGE_RATIOS.items():
        if key in normalized:
            return ratio
    return None


def _compute_lifecycle(
    date_mes: int | None,
    ref: EquipmentReference | None,
    etat_sante: str | None = None,
) -> tuple[float | None, float | None, str, str | None]:
    if not ref or not ref.sypemi_reference_annees:
        return None, None, "missing", "Reference duree de vie absente"

    if date_mes:
        age = max(0, CURRENT_YEAR - date_mes)
        return (
            round(ref.sypemi_reference_annees - age, 1),
            round(age, 1),
            "date_mes",
            "Calcule depuis DATE MES",
        )

    ratio = _health_age_ratio(etat_sante)
    if ratio is not None:
        age = round(ref.sypemi_reference_annees * ratio, 1)
        return (
            round(ref.sypemi_reference_annees - age, 1),
            age,
            "etat_sante",
            f"DATE MES absente, estime depuis ETAT SANTE : {etat_sante}",
        )

    return None, None, "missing", "DATE MES et ETAT SANTE exploitables absents"


def _compute_remaining_life(
    date_mes: int | None,
    ref: EquipmentReference | None,
    etat_sante: str | None = None,
) -> float | None:
    remaining, _, _, _ = _compute_lifecycle(date_mes, ref, etat_sante)
    return remaining


def _read_item(item: CvcInventoryItem, ref: EquipmentReference | None) -> CvcInventoryItemRead:
    sypemi_years = ref.sypemi_reference_annees if ref else None
    remaining_life, lifecycle_age, lifecycle_source, lifecycle_label = _compute_lifecycle(
        item.date_mis_en_service,
        ref,
        item.etat_sante,
    )

    criticite_pct = None
    if lifecycle_age is not None and sypemi_years and sypemi_years > 0:
        criticite_pct = min(100.0, round(lifecycle_age / sypemi_years * 100, 1))

    read = CvcInventoryItemRead.model_validate(item)
    # Décision : on garde la valeur fournie par le prestataire (SPIE) quand le calcul
    # SYPEMI n'aboutit pas, et on expose le calcul indicatif quand il existe.
    if remaining_life is not None:
        read.duree_vie_restante = remaining_life
        read.duree_vie_restante_source = "calcule"
    elif item.duree_vie_restante is not None:
        read.duree_vie_restante = item.duree_vie_restante
        read.duree_vie_restante_source = item.duree_vie_restante_source or "fournie"
        lifecycle_label = lifecycle_label or "Durée de vie fournie par le prestataire"
    else:
        read.duree_vie_restante = None
    read.duree_vie_restante_calculee = remaining_life
    read.lifecycle_age_years = lifecycle_age
    read.lifecycle_age_source = lifecycle_source
    read.lifecycle_age_label = lifecycle_label
    read.criticite_pct = criticite_pct
    read.sypemi_reference_annees = sypemi_years
    read.sypemi_mini_annees = ref.sypemi_mini_annees if ref else None
    read.sypemi_maxi_annees = ref.sypemi_maxi_annees if ref else None
    read.requires_refrigerant_quantity = _requires_refrigerant_quantity(ref)
    read.equipment_ref = CvcEquipmentReferenceRead.model_validate(ref) if ref else None
    return read


def _item_match_inputs(item: CvcInventoryItem) -> tuple[str | None, str]:
    """Reproduit la combinaison type+catégorie/désignation utilisée à l'import (SPIE inclus)."""
    if item.provider == PROVIDER_SPIE:
        match_famille = " ".join(part for part in (item.type_equipement, item.famille) if part) or item.famille
        match_designation = " ".join(part for part in (item.type_equipement, item.designation) if part)
        return match_famille, match_designation
    return item.famille, item.designation


def _normalize_inventory_rows(rows: list[tuple], provider: str) -> list[dict]:
    """Convertit les lignes brutes (format DALKIA ou SPIE) en dictionnaires homogènes."""
    header = [str(c) if c is not None else "" for c in rows[0]]
    lookup = _header_lookup(header)

    def cell(row: tuple, name: str):
        return _get_header_value(row, lookup, name)

    parsed: list[dict] = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue

        if provider == PROVIDER_SPIE:
            designation = _clean_cell(cell(row, "Nom complet de l'equipement (type + complementaire)"))
            if not designation:
                continue
            categorie = _clean_cell(cell(row, "Categorie de l'equipement"))
            type_equipement = _clean_cell(cell(row, "Type d'equipement"))
            famille = categorie or type_equipement
            # Matching SYPEMI : on combine type + catégorie + désignation pour plus de robustesse.
            match_famille = " ".join(part for part in (type_equipement, categorie) if part) or famille
            match_designation = " ".join(part for part in (type_equipement, designation) if part)
            parsed.append(
                {
                    "site_raw": _extract_building_name(cell(row, "Nom du batiment")),
                    "batiment": None,
                    "niveau": _clean_cell(cell(row, "Niveau")),
                    "local_name": _clean_cell(cell(row, "Local")),
                    "designation": designation,
                    "type_equipement": type_equipement,
                    "statut": None,
                    "etat_sante": None,
                    "quantite_relevee": _parse_int(cell(row, "Quantite")),
                    "famille": famille,
                    "marque": _clean_cell(cell(row, "Marque")),
                    "modele": _clean_cell(cell(row, "Modele")),
                    "numero_serie": _clean_cell(cell(row, "N° de serie")),
                    "puissance": _clean_cell(cell(row, "Puissance")),
                    "puissance_frigorifique": _parse_float(cell(row, "Puissance frigorifique")),
                    "puissance_calorifique": _parse_float(cell(row, "Puissance calorifique")),
                    "capacite": _parse_float(cell(row, "Capacite")),
                    "date_mes": _parse_mes_year(cell(row, "Date de mise en service")),
                    "duree_vie_provided": _parse_float(cell(row, "Duree de vie restante")),
                    "match_famille": match_famille,
                    "match_designation": match_designation,
                }
            )
        else:
            designation = _clean_cell(cell(row, "DESIGNATION"))
            if not designation:
                continue
            famille = _clean_cell(cell(row, "FAMILLE"))
            parsed.append(
                {
                    "site_raw": _clean_cell(cell(row, "SITE")),
                    "batiment": _clean_cell(cell(row, "BATIMENT")),
                    "niveau": _clean_cell(cell(row, "NIVEAU")),
                    "local_name": _clean_cell(cell(row, "LOCAL")),
                    "designation": designation,
                    "type_equipement": None,
                    "statut": _clean_cell(cell(row, "STATUT")),
                    "etat_sante": _clean_cell(cell(row, "ETAT SANTE")),
                    "quantite_relevee": _parse_int(cell(row, "QTE QTE RELEVEE")),
                    "famille": famille,
                    "marque": _clean_cell(cell(row, "MARQUE")),
                    "modele": _clean_cell(cell(row, "MODELE")),
                    "numero_serie": None,
                    "puissance": None,
                    "puissance_frigorifique": None,
                    "puissance_calorifique": None,
                    "capacite": None,
                    "date_mes": _parse_mes_year(cell(row, "DATE MES")),
                    "duree_vie_provided": None,
                    "match_famille": famille,
                    "match_designation": designation,
                }
            )
    return parsed


def import_cvc_from_excel(
    db: Session,
    raw_bytes: bytes,
    building_mappings: list[CvcBuildingMapping],
    city_id: int | None,
    import_batch: str | None = None,
    provider: str | None = None,
) -> CvcImportResult:
    mapping_dict = {m.site_raw: m.building_id for m in building_mappings}

    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]] if rows else []
    resolved_provider = provider or detect_inventory_provider(header)
    prefix = "spie" if resolved_provider == PROVIDER_SPIE else "import"
    batch_id = import_batch or f"{prefix}_{uuid.uuid4().hex[:8]}"

    if not rows:
        return CvcImportResult(
            imported=0, skipped=0, errors=[], import_batch=batch_id,
            provider=resolved_provider, sypemi_matched=0, sypemi_unmatched=0,
        )

    parsed_rows = _normalize_inventory_rows(rows, resolved_provider)
    all_refs = list(db.scalars(select(EquipmentReference)))
    family_cache: dict = {}

    _clear_current_cvc_inventory(db, city_id, resolved_provider)

    imported = 0
    skipped = len([r for r in rows[1:] if any(v is not None for v in r)]) - len(parsed_rows)
    errors: list[str] = []
    sypemi_matched = 0
    sypemi_unmatched = 0

    for data in parsed_rows:
        site = data["site_raw"]
        building_id = mapping_dict.get(site) if site else None
        building = db.get(Building, building_id) if building_id is not None else None

        ref = _resolve_family(
            data["match_famille"],
            all_refs,
            family_cache,
            data["match_designation"],
            data["marque"],
            data["modele"],
        )

        computed_remaining = _compute_remaining_life(data["date_mes"], ref, data["etat_sante"])
        if computed_remaining is not None:
            duree_vie_restante = computed_remaining
            duree_vie_source = "calcule"
        elif data["duree_vie_provided"] is not None:
            duree_vie_restante = data["duree_vie_provided"]
            duree_vie_source = "fournie"
        else:
            duree_vie_restante = None
            duree_vie_source = None

        item = CvcInventoryItem(
            city_id=city_id,
            site_id=building.site_id if building else None,
            building_id=building_id,
            local_id=None,
            equipment_ref_id=ref.id if ref else None,
            provider=resolved_provider,
            site_raw=site,
            batiment=data["batiment"],
            niveau=data["niveau"],
            local_name=data["local_name"],
            designation=data["designation"],
            type_equipement=data["type_equipement"],
            statut=data["statut"],
            etat_sante=data["etat_sante"],
            quantite_relevee=data["quantite_relevee"],
            famille=data["famille"],
            marque=data["marque"],
            modele=data["modele"],
            numero_serie=data["numero_serie"],
            puissance=data["puissance"],
            puissance_frigorifique=data["puissance_frigorifique"],
            puissance_calorifique=data["puissance_calorifique"],
            capacite=data["capacite"],
            date_mis_en_service=data["date_mes"],
            duree_vie_restante=duree_vie_restante,
            duree_vie_restante_source=duree_vie_source,
            quantite_fluide_frigorigene=None,
            import_batch=batch_id,
        )
        db.add(item)

        if ref:
            sypemi_matched += 1
        else:
            sypemi_unmatched += 1
        imported += 1

    db.flush()
    _refresh_refrigerant_inventory_links(db, city_id)
    db.commit()
    ensure_cvc_source_building_mappings(db, city_id)
    return CvcImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors,
        import_batch=batch_id,
        provider=resolved_provider,
        sypemi_matched=sypemi_matched,
        sypemi_unmatched=sypemi_unmatched,
    )


def list_cvc_import_batches(db: Session, city_id: int | None) -> list[CvcImportBatchSummary]:
    stmt = select(CvcInventoryItem)
    if city_id is not None:
        stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
    items = list(db.scalars(stmt))
    batches: dict[str, list[CvcInventoryItem]] = {}
    for item in items:
        if item.import_batch:
            batches.setdefault(item.import_batch, []).append(item)

    ref_ids = {item.equipment_ref_id for item in items if item.equipment_ref_id}
    refrigerant_ref_ids: set[int] = set()
    if ref_ids:
        refs = db.scalars(select(EquipmentReference).where(EquipmentReference.id.in_(ref_ids)))
        refrigerant_ref_ids = {ref.id for ref in refs if _requires_refrigerant_quantity(ref)}

    summaries = []
    for batch, batch_items in batches.items():
        summaries.append(
            CvcImportBatchSummary(
                import_batch=batch,
                provider=next((item.provider for item in batch_items if item.provider), PROVIDER_DALKIA),
                imported=len(batch_items),
                mapped_items=sum(1 for item in batch_items if item.building_id is not None),
                reference_mapped_items=sum(1 for item in batch_items if item.equipment_ref_id is not None),
                refrigerant_items=sum(1 for item in batch_items if item.equipment_ref_id in refrigerant_ref_ids),
                created_at=min((item.created_at for item in batch_items), default=None),
            )
        )
    return sorted(summaries, key=lambda b: b.created_at or datetime.min, reverse=True)


def list_cvc_items_for_batch(
    db: Session, import_batch: str, city_id: int | None
) -> list[CvcInventoryItemRead]:
    stmt = select(CvcInventoryItem).where(CvcInventoryItem.import_batch == import_batch)
    if city_id is not None:
        stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
    items = list(db.scalars(stmt.order_by(CvcInventoryItem.site_raw, CvcInventoryItem.designation)))
    return _hydrate_items(db, items)


def recompute_cvc_references_for_batch(
    db: Session, import_batch: str, city_id: int | None
) -> CvcRecomputeReferencesResult:
    stmt = select(CvcInventoryItem).where(CvcInventoryItem.import_batch == import_batch)
    if city_id is not None:
        stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
    items = list(db.scalars(stmt))
    all_refs = list(db.scalars(select(EquipmentReference)))
    family_cache: dict = {}

    updated = 0
    changed = 0
    matched = 0
    unmatched = 0
    for item in items:
        match_famille, match_designation = _item_match_inputs(item)
        next_ref = _resolve_family(
            match_famille,
            all_refs,
            family_cache,
            match_designation,
            item.marque,
            item.modele,
        )
        next_ref_id = next_ref.id if next_ref else None
        if item.equipment_ref_id != next_ref_id:
            changed += 1
        item.equipment_ref_id = next_ref_id
        item.duree_vie_restante = _compute_remaining_life(item.date_mis_en_service, next_ref, item.etat_sante)
        if not _requires_refrigerant_quantity(next_ref):
            item.quantite_fluide_frigorigene = None
        if next_ref:
            matched += 1
        else:
            unmatched += 1
        updated += 1

    db.commit()
    return CvcRecomputeReferencesResult(
        import_batch=import_batch,
        updated=updated,
        matched=matched,
        unmatched=unmatched,
        changed=changed,
    )


def import_cvc_refrigerants_from_excel(
    db: Session,
    raw_bytes: bytes,
    city_id: int | None,
    source_filename: str | None = None,
) -> CvcRefrigerantImportResult:
    _, rows = _read_refrigerant_rows(raw_bytes)
    batch_id = f"esp_{uuid.uuid4().hex[:8]}"
    inventory_stmt = select(CvcInventoryItem)
    if city_id is not None:
        inventory_stmt = inventory_stmt.where(
            (CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None))
        )
    inventory_items = list(db.scalars(inventory_stmt))

    imported = 0
    auto_matched = 0
    ambiguous = 0
    total_fluide_kg = 0.0
    total_teqco2 = 0.0

    for row in rows:
        candidates = _find_refrigerant_candidates(row, inventory_items)
        auto_candidate = _select_auto_refrigerant_candidate(candidates)
        match_status = "pending"
        match_method = None
        match_score = None
        inventory_item_id = None

        if auto_candidate:
            match_status = "auto_matched"
            match_method = auto_candidate.method
            match_score = auto_candidate.score
            inventory_item_id = auto_candidate.item.id
        elif candidates:
            match_status = "ambiguous"
            ambiguous += 1
        matched_inventory = next((i for i in inventory_items if i.id == inventory_item_id), None)

        item = CvcRefrigerantItem(
            city_id=city_id,
            site_id=matched_inventory.site_id if matched_inventory else None,
            building_id=matched_inventory.building_id if matched_inventory else None,
            cvc_inventory_item_id=inventory_item_id,
            import_batch=batch_id,
            source_filename=source_filename,
            row_number=row["row_number"],
            site_raw=row["site_raw"],
            designation=row["designation"],
            quantite_relevee=row["quantite_relevee"],
            famille=row["famille"],
            marque=row["marque"],
            modele=row["modele"],
            fluide_frigorigene=row["fluide_frigorigene"],
            quantite_fluide_kg=row["quantite_fluide_kg"],
            puissance_froid_kw=row["puissance_froid_kw"],
            date_mis_en_service=row["date_mis_en_service"],
            gwp=row["gwp"],
            teqco2=row["teqco2"],
            esp_status=row["esp_status"],
            cout_desp_date_eur=row["cout_desp_date_eur"],
            cumul_5_ans_eur=row["cumul_5_ans_eur"],
            schedule_json=row["schedule_json"],
            match_status=match_status,
            match_method=match_method,
            match_score=match_score,
        )
        db.add(item)

        if inventory_item_id is not None:
            if matched_inventory is not None and row["quantite_fluide_kg"] is not None:
                matched_inventory.quantite_fluide_frigorigene = row["quantite_fluide_kg"]
            auto_matched += 1
        total_fluide_kg += row["quantite_fluide_kg"] or 0.0
        total_teqco2 += row["teqco2"] or 0.0
        imported += 1

    db.commit()
    ensure_cvc_source_building_mappings(db, city_id)
    return CvcRefrigerantImportResult(
        import_batch=batch_id,
        imported=imported,
        auto_matched=auto_matched,
        pending=imported - auto_matched - ambiguous,
        ambiguous=ambiguous,
        total_fluide_kg=round(total_fluide_kg, 3),
        total_teqco2=round(total_teqco2, 3),
    )


def list_cvc_refrigerant_batches(db: Session, city_id: int | None) -> list[CvcRefrigerantBatchSummary]:
    stmt = select(CvcRefrigerantItem)
    if city_id is not None:
        stmt = stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
    items = list(db.scalars(stmt))
    batches: dict[str, list[CvcRefrigerantItem]] = defaultdict(list)
    for item in items:
        batches[item.import_batch].append(item)

    summaries: list[CvcRefrigerantBatchSummary] = []
    for import_batch, batch_items in batches.items():
        summaries.append(
            CvcRefrigerantBatchSummary(
                import_batch=import_batch,
                source_filename=next((item.source_filename for item in batch_items if item.source_filename), None),
                imported=len(batch_items),
                matched_items=sum(1 for item in batch_items if item.cvc_inventory_item_id is not None),
                pending_items=sum(1 for item in batch_items if item.cvc_inventory_item_id is None),
                total_fluide_kg=round(sum(item.quantite_fluide_kg or 0.0 for item in batch_items), 3),
                total_teqco2=round(sum(item.teqco2 or 0.0 for item in batch_items), 3),
                created_at=min((item.created_at for item in batch_items), default=None),
            )
        )
    return sorted(summaries, key=lambda item: item.created_at or datetime.min, reverse=True)


def get_cvc_refrigerant_dashboard(db: Session, city_id: int | None) -> CvcRefrigerantDashboard:
    stmt = select(CvcRefrigerantItem)
    if city_id is not None:
        stmt = stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
    items = list(db.scalars(stmt.order_by(CvcRefrigerantItem.created_at.desc(), CvcRefrigerantItem.site_raw)))
    batches = list_cvc_refrigerant_batches(db, city_id) if items else []
    latest_batch = batches[0] if batches else None

    status_counts: dict[str, int] = {}
    conformity_counts: dict[str, int] = {}
    priority_counts: dict[str, int] = {}
    open_actions: list[CvcRefrigerantActionSummary] = []
    esp_signals: list[CvcRefrigerantActionSummary] = []

    for item in items:
        computed = _computed_refrigerant_fields(item)
        fgas_status = str(computed["fgas_status"])
        conformity = str(computed["statut_conformite"])
        priority = str(computed["priorite"])
        status_counts[fgas_status] = status_counts.get(fgas_status, 0) + 1
        conformity_counts[conformity] = conformity_counts.get(conformity, 0) + 1
        priority_counts[priority] = priority_counts.get(priority, 0) + 1

        if conformity not in {"OK", "Non prioritaire"}:
            open_actions.append(
                _action_summary(
                    item,
                    "F-Gaz",
                    f"{fgas_status} : {conformity}",
                )
            )
        if _normalize(item.esp_status) in {"soumis", "manque de donnee"}:
            esp_signals.append(
                _action_summary(
                    item,
                    "ESP/DESP",
                    f"Statut source ESP : {item.esp_status or 'Non renseigné'}",
                    "Récupérer dossier ESP, PV IP/RP, prochaine échéance et preuve LUNE si applicable",
                )
            )

    priority_order = {"Haute": 0, "Moyenne": 1, "Basse": 2}
    open_actions.sort(key=lambda action: (priority_order.get(action.priority, 9), action.echeance_cible or date.max, action.site or ""))
    esp_signals.sort(key=lambda action: (priority_order.get(action.priority, 9), action.site or "", action.equipment))

    missing_data = sum(1 for item in items if item.teqco2 is None or not item.fluide_frigorigene or item.quantite_fluide_kg is None or item.gwp is None)
    over_5 = sum(1 for item in items if item.teqco2 is not None and item.teqco2 >= 5)
    over_50 = sum(1 for item in items if item.teqco2 is not None and item.teqco2 >= 50)
    overdue_or_due = sum(
        1
        for item in items
        if str(_computed_refrigerant_fields(item)["statut_conformite"]) in {"En retard", "À programmer < 60 j"}
    )
    last_check_missing = sum(
        1 for item in items if str(_computed_refrigerant_fields(item)["statut_conformite"]) == "Dernier contrôle à demander"
    )
    unmapped = sum(1 for item in items if item.cvc_inventory_item_id is None)

    kpis = [
        CvcRefrigerantDashboardKpi(key="items", label="Équipements inventoriés", value=len(items), helper="Lignes fluides ESP importées"),
        CvcRefrigerantDashboardKpi(key="fgas", label="À suivre F-Gaz ≥ 5 t", value=over_5, tone="warning", helper="Contrôle périodique requis"),
        CvcRefrigerantDashboardKpi(key="fgas50", label="≥ 50 t éq. CO2", value=over_50, tone="danger", helper="Priorité renforcée"),
        CvcRefrigerantDashboardKpi(key="missing", label="Données à compléter", value=missing_data, tone="danger", helper="Fluide, charge ou GWP absent"),
        CvcRefrigerantDashboardKpi(key="check_missing", label="Contrôle à demander", value=last_check_missing, tone="warning", helper="Dernier contrôle non renseigné"),
        CvcRefrigerantDashboardKpi(key="due", label="À programmer / retard", value=overdue_or_due, tone="danger", helper="Échéance dépassée ou proche"),
        CvcRefrigerantDashboardKpi(key="esp", label="Signaux ESP", value=len(esp_signals), tone="warning", helper="Dossier ESP/DESP à suivre à part"),
        CvcRefrigerantDashboardKpi(key="unmapped", label="Non rattachés CVC", value=unmapped, tone="warning", helper="Lien équipement à valider"),
    ]

    return CvcRefrigerantDashboard(
        total_items=len(items),
        latest_batch=latest_batch.import_batch if latest_batch else None,
        latest_batch_label=latest_batch.source_filename if latest_batch else None,
        kpis=kpis,
        status_counts=status_counts,
        conformity_counts=conformity_counts,
        priority_counts=priority_counts,
        open_actions=open_actions[:80],
        esp_signals=esp_signals[:80],
    )


def _read_refrigerant_item(
    item: CvcRefrigerantItem,
    inventory_map: dict[int, CvcInventoryItem],
    inventory_items: list[CvcInventoryItem],
) -> CvcRefrigerantItemRead:
    read = CvcRefrigerantItemRead.model_validate(item)
    read.schedule = json.loads(item.schedule_json) if item.schedule_json else {}
    computed = _computed_refrigerant_fields(item)
    read.fgas_status = str(computed["fgas_status"])
    read.frequence_controle_mois = computed["frequence_controle_mois"] if isinstance(computed["frequence_controle_mois"], int) else None
    read.statut_conformite = str(computed["statut_conformite"])
    read.action_prioritaire = str(computed["action_prioritaire"])
    read.preuve_attendue = str(computed["preuve_attendue"])
    read.priorite = str(computed["priorite"])
    if item.cvc_inventory_item_id:
        matched = inventory_map.get(item.cvc_inventory_item_id)
        read.matched_inventory_item = _inventory_compact(matched) if matched else None
    read.candidates = _find_refrigerant_candidates(
        {
            "site_raw": item.site_raw,
            "designation": item.designation,
            "famille": item.famille,
            "marque": item.marque,
            "modele": item.modele,
            "date_mis_en_service": item.date_mis_en_service,
        },
        inventory_items,
    )
    return read


def list_cvc_refrigerant_items_for_batch(
    db: Session, import_batch: str, city_id: int | None
) -> list[CvcRefrigerantItemRead]:
    stmt = select(CvcRefrigerantItem).where(CvcRefrigerantItem.import_batch == import_batch)
    inventory_stmt = select(CvcInventoryItem)
    if city_id is not None:
        stmt = stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
        inventory_stmt = inventory_stmt.where(
            (CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None))
        )
    items = list(db.scalars(stmt.order_by(CvcRefrigerantItem.site_raw, CvcRefrigerantItem.designation)))
    inventory_items = list(db.scalars(inventory_stmt))
    inventory_map = {item.id: item for item in inventory_items}
    return [_read_refrigerant_item(item, inventory_map, inventory_items) for item in items]


def update_cvc_refrigerant_item(
    db: Session,
    item_id: int,
    payload: CvcRefrigerantItemUpdate,
    city_id: int | None,
) -> CvcRefrigerantItemRead | None:
    item = db.scalar(select(CvcRefrigerantItem).where(CvcRefrigerantItem.id == item_id))
    if not item:
        return None
    if city_id is not None and item.city_id not in (None, city_id):
        return None

    updates = payload.model_dump(exclude_unset=True)
    inventory_item = None
    if "cvc_inventory_item_id" in updates and payload.cvc_inventory_item_id is not None:
        inventory_item = db.get(CvcInventoryItem, payload.cvc_inventory_item_id)
        if inventory_item is None:
            raise ValueError("Equipement CVC introuvable.")
        if city_id is not None and inventory_item.city_id not in (None, city_id):
            raise ValueError("Equipement CVC hors perimetre utilisateur.")

    if "cvc_inventory_item_id" in updates:
        item.cvc_inventory_item_id = inventory_item.id if inventory_item else None
        item.site_id = inventory_item.site_id if inventory_item else payload.site_id
        item.building_id = inventory_item.building_id if inventory_item else payload.building_id
        item.match_status = "manual_matched" if inventory_item else "pending"
        item.match_method = "manual" if inventory_item else None
        item.match_score = 1.0 if inventory_item else None
        if inventory_item and item.quantite_fluide_kg is not None:
            inventory_item.quantite_fluide_frigorigene = item.quantite_fluide_kg
    else:
        if "site_id" in updates:
            item.site_id = payload.site_id
        if "building_id" in updates:
            item.building_id = payload.building_id

    for field in (
        "detection_permanente",
        "dernier_controle_etancheite",
        "prochaine_echeance",
        "titulaire",
        "responsable_collectivite",
        "statut_action",
        "commentaire_gmao",
    ):
        if field in updates:
            setattr(item, field, updates[field])

    db.commit()
    db.refresh(item)

    inventory_stmt = select(CvcInventoryItem)
    if city_id is not None:
        inventory_stmt = inventory_stmt.where(
            (CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None))
        )
    inventory_items = list(db.scalars(inventory_stmt))
    inventory_map = {inventory.id: inventory for inventory in inventory_items}
    return _read_refrigerant_item(item, inventory_map, inventory_items)


def list_cvc_source_building_mappings(
    db: Session,
    city_id: int | None,
    source_type: str | None = None,
    import_batch: str | None = None,
) -> list[CvcSourceBuildingMappingRead]:
    ensure_cvc_source_building_mappings(db, city_id)
    stmt = select(CvcSourceBuildingMapping)
    sites_stmt = select(Site)
    buildings_stmt = select(Building)
    if city_id is not None:
        stmt = stmt.where((CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None)))
        sites_stmt = sites_stmt.where(Site.city_id == city_id)
        buildings_stmt = buildings_stmt.where(Building.city_id == city_id)
    if source_type:
        stmt = stmt.where(CvcSourceBuildingMapping.source_type == source_type)
    if import_batch:
        stmt = stmt.where(CvcSourceBuildingMapping.import_batch == import_batch)

    mappings = list(db.scalars(stmt.order_by(CvcSourceBuildingMapping.source_type, CvcSourceBuildingMapping.source_site_raw)))
    sites = list(db.scalars(sites_stmt))
    buildings = list(db.scalars(buildings_stmt))

    inventory_counts: dict[tuple[str, str], int] = {}
    refrigerant_counts: dict[tuple[str, str], int] = {}
    inventory_stmt = select(CvcInventoryItem)
    refrigerant_stmt = select(CvcRefrigerantItem)
    if city_id is not None:
        inventory_stmt = inventory_stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        refrigerant_stmt = refrigerant_stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
    for item in db.scalars(inventory_stmt):
        if item.import_batch and item.site_raw:
            key = (item.import_batch, item.site_raw.strip())
            inventory_counts[key] = inventory_counts.get(key, 0) + 1
    for item in db.scalars(refrigerant_stmt):
        if item.import_batch and item.site_raw:
            key = (item.import_batch, item.site_raw.strip())
            refrigerant_counts[key] = refrigerant_counts.get(key, 0) + 1

    results: list[CvcSourceBuildingMappingRead] = []
    for mapping in mappings:
        site_suggestions, building_suggestions = _mapping_suggestions(mapping.source_site_raw, sites, buildings)
        read = CvcSourceBuildingMappingRead.model_validate(mapping)
        read.building_ids = _read_mapping_building_ids(mapping)
        read.site_suggestions = site_suggestions
        read.building_suggestions = building_suggestions
        key = (mapping.import_batch, mapping.source_site_raw.strip())
        read.item_count = inventory_counts.get(key, 0)
        read.refrigerant_count = refrigerant_counts.get(key, 0)
        results.append(read)
    return results


def update_cvc_source_building_mapping(
    db: Session,
    mapping_id: int,
    payload: CvcSourceBuildingMappingUpdate,
    city_id: int | None,
) -> CvcSourceBuildingMappingRead | None:
    mapping = db.get(CvcSourceBuildingMapping, mapping_id)
    if not mapping:
        return None
    if city_id is not None and mapping.city_id not in (None, city_id):
        return None

    site = db.get(Site, payload.site_id) if payload.site_id is not None else None
    if payload.site_id is not None and site is None:
        raise ValueError("Site introuvable.")
    if site and city_id is not None and site.city_id != city_id:
        raise ValueError("Site hors perimetre utilisateur.")

    requested_building_ids = _dedupe_ids(payload.building_ids or ([payload.building_id] if payload.building_id else []))
    buildings: list[Building] = []
    for building_id in requested_building_ids:
        building = db.get(Building, building_id)
        if building is None:
            raise ValueError("Batiment introuvable.")
        if city_id is not None and building.city_id != city_id:
            raise ValueError("Batiment hors perimetre utilisateur.")
        if site and building.site_id not in (None, site.id):
            raise ValueError("Le batiment choisi n'appartient pas au site choisi.")
        buildings.append(building)
    selected_site_ids = {building.site_id for building in buildings if building.site_id is not None}

    mapping.site_id = site.id if site else (next(iter(selected_site_ids)) if len(selected_site_ids) == 1 else None)
    _write_mapping_building_ids(mapping, requested_building_ids)
    mapping.status = payload.status
    mapping.notes = payload.notes
    mapping.match_method = "manual"
    mapping.match_score = 1.0 if requested_building_ids or site else None
    updated_rows = _apply_source_mapping_to_rows(db, mapping)
    db.commit()
    db.refresh(mapping)

    read = CvcSourceBuildingMappingRead.model_validate(mapping)
    read.building_ids = _read_mapping_building_ids(mapping)
    read.item_count = updated_rows if mapping.source_type == "inventory" else 0
    read.refrigerant_count = updated_rows if mapping.source_type == "refrigerant" else 0
    return read


def get_cvc_technical_coverage_report(db: Session, city_id: int | None) -> CvcTechnicalCoverageReport:
    ensure_cvc_source_building_mappings(db, city_id)
    buildings_stmt = select(Building)
    inventory_stmt = select(CvcInventoryItem)
    refrigerant_stmt = select(CvcRefrigerantItem)
    mapping_stmt = select(CvcSourceBuildingMapping)
    if city_id is not None:
        buildings_stmt = buildings_stmt.where(Building.city_id == city_id)
        inventory_stmt = inventory_stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
        refrigerant_stmt = refrigerant_stmt.where((CvcRefrigerantItem.city_id == city_id) | (CvcRefrigerantItem.city_id.is_(None)))
        mapping_stmt = mapping_stmt.where((CvcSourceBuildingMapping.city_id == city_id) | (CvcSourceBuildingMapping.city_id.is_(None)))

    buildings = list(db.scalars(buildings_stmt))
    inventory_items = list(db.scalars(inventory_stmt))
    inventory_items = _filter_latest_inventory_batch(inventory_items)
    refrigerant_items = list(db.scalars(refrigerant_stmt))
    mappings = list(db.scalars(mapping_stmt))

    inventory_building_ids = {item.building_id for item in inventory_items if item.building_id is not None}
    refrigerant_building_ids = {item.building_id for item in refrigerant_items if item.building_id is not None}
    source_mapping_building_ids = {
        building_id
        for mapping in mappings
        if mapping.status == "matched"
        for building_id in _read_mapping_building_ids(mapping)
    }
    covered_building_ids = inventory_building_ids | refrigerant_building_ids | source_mapping_building_ids
    patrimoine_without_cvc = [
        BuildingMatchSuggestion(
            building_id=building.id,
            site_id=building.site_id,
            nom_batiment=building.nom_batiment,
            adresse=_build_address(building),
            score=1.0,
        )
        for building in buildings
        if building.id not in covered_building_ids
    ]

    def grouped_unmapped(rows, source_type: str) -> list[dict]:
        grouped: dict[str, int] = {}
        for item in rows:
            if item.building_id is None and item.site_raw:
                grouped[item.site_raw] = grouped.get(item.site_raw, 0) + 1
        return [
            {"source_type": source_type, "source_site_raw": source_site_raw, "count": count}
            for source_site_raw, count in sorted(grouped.items())
        ]

    return CvcTechnicalCoverageReport(
        patrimoine_buildings=len(buildings),
        cvc_inventory_items=len(inventory_items),
        cvc_refrigerant_items=len(refrigerant_items),
        inventory_without_building=sum(1 for item in inventory_items if item.building_id is None),
        refrigerants_without_building=sum(1 for item in refrigerant_items if item.building_id is None),
        refrigerants_without_inventory_item=sum(1 for item in refrigerant_items if item.cvc_inventory_item_id is None),
        source_mappings_to_review=sum(1 for mapping in mappings if mapping.status == "to_review"),
        source_mappings_not_found=sum(1 for mapping in mappings if mapping.status == "not_found"),
        patrimoine_buildings_without_cvc=patrimoine_without_cvc[:300],
        inventory_unmapped_by_source=grouped_unmapped(inventory_items, "inventory"),
        refrigerants_unmapped_by_source=grouped_unmapped(refrigerant_items, "refrigerant"),
    )


def _hydrate_items(db: Session, items: list[CvcInventoryItem]) -> list[CvcInventoryItemRead]:
    ref_ids = {item.equipment_ref_id for item in items if item.equipment_ref_id}
    refs_map: dict[int, EquipmentReference] = {}
    if ref_ids:
        for ref in db.scalars(select(EquipmentReference).where(EquipmentReference.id.in_(ref_ids))):
            refs_map[ref.id] = ref

    return [
        _read_item(item, refs_map.get(item.equipment_ref_id) if item.equipment_ref_id else None)
        for item in items
    ]


# ---------------------------------------------------------------------------
# État du parc technique
#
# Le cycle de vie est déjà calculé PAR équipement (`_read_item` : âge, durée de
# vie restante, criticité % vs référence SYPEMI). Ici on ne recalcule rien : on
# agrège ces valeurs pour obtenir une lecture du parc.
# Unité de pilotage retenue (arbitrage 2026-08-17) : le NOMBRE d'équipements.
# ---------------------------------------------------------------------------

_AGE_BUCKETS: tuple[tuple[str, str, float, float], ...] = (
    ("0_5", "0-5 ans", 0.0, 5.0),
    ("6_10", "6-10 ans", 5.0, 10.0),
    ("11_15", "11-15 ans", 10.0, 15.0),
    ("16_20", "16-20 ans", 15.0, 20.0),
    ("21_30", "21-30 ans", 20.0, 30.0),
    ("30_plus", "Plus de 30 ans", 30.0, float("inf")),
)

_CRITICITE_BUCKETS: tuple[tuple[str, str, float, float], ...] = (
    ("faible", "Moins de 50 %", 0.0, 50.0),
    ("moyenne", "50 à 80 %", 50.0, 80.0),
    ("elevee", "80 à 100 %", 80.0, 100.0),
    ("depasse", "Durée de vie dépassée", 100.0, float("inf")),
)

_UNKNOWN_BUCKET = ("inconnu", "Non calculable")


def _bucketize(
    values: list[float | None],
    buckets: tuple[tuple[str, str, float, float], ...],
    total: int,
) -> list[CvcParcBucket]:
    counts = {key: 0 for key, _, _, _ in buckets}
    unknown = 0
    for value in values:
        if value is None:
            unknown += 1
            continue
        for key, _, low, high in buckets:
            # Bornes inclusives à gauche, exclusives à droite (dernière ouverte).
            if low <= value < high or (high == float("inf") and value >= low):
                counts[key] += 1
                break
    share = lambda n: round(100 * n / total, 1) if total else 0.0  # noqa: E731
    result = [
        CvcParcBucket(key=key, label=label, count=counts[key], share_pct=share(counts[key]))
        for key, label, _, _ in buckets
    ]
    if unknown:
        result.append(
            CvcParcBucket(key=_UNKNOWN_BUCKET[0], label=_UNKNOWN_BUCKET[1], count=unknown, share_pct=share(unknown))
        )
    return result


def _mean(values: list[float]) -> float | None:
    return round(sum(values) / len(values), 1) if values else None


def get_cvc_parc_technique(
    db: Session,
    city_id: int | None,
    provider: str | None = None,
    building_id: int | None = None,
    famille: str | None = None,
) -> CvcParcTechniqueReport:
    """État du parc CVC : âges, criticité, fin de vie, complétude de la donnée."""
    stmt = select(CvcInventoryItem)
    if city_id is not None:
        stmt = stmt.where((CvcInventoryItem.city_id == city_id) | (CvcInventoryItem.city_id.is_(None)))
    if provider:
        stmt = stmt.where(CvcInventoryItem.provider == provider)
    if building_id is not None:
        stmt = stmt.where(CvcInventoryItem.building_id == building_id)
    if famille:
        stmt = stmt.where(CvcInventoryItem.famille == famille)

    # Ne compter que le lot courant de chaque prestataire : un réimport ne doit
    # jamais gonfler l'état du parc (cf. les 4 lots DALKIA doublons de juin 2026).
    items = _filter_latest_inventory_batch(list(db.scalars(stmt)))
    reads = _hydrate_items(db, items)
    total = len(reads)

    ages = [read.lifecycle_age_years for read in reads]
    criticites = [read.criticite_pct for read in reads]
    known_ages = [value for value in ages if value is not None]

    def _is_depasse(read: CvcInventoryItemRead) -> bool:
        return read.duree_vie_restante is not None and read.duree_vie_restante <= 0

    def _is_fin_de_vie(read: CvcInventoryItemRead) -> bool:
        return read.duree_vie_restante is not None and 0 < read.duree_vie_restante <= 5

    depasses = sum(1 for read in reads if _is_depasse(read))
    fin_de_vie = sum(1 for read in reads if _is_fin_de_vie(read))

    # Répartition par prestataire
    provider_counts: dict[str, int] = defaultdict(int)
    for read in reads:
        provider_counts[read.provider or PROVIDER_DALKIA] += 1
    par_provider = [
        CvcParcBucket(
            key=name,
            label=name,
            count=count,
            share_pct=round(100 * count / total, 1) if total else 0.0,
        )
        for name, count in sorted(provider_counts.items(), key=lambda kv: -kv[1])
    ]

    # Par famille
    by_famille: dict[str, list[CvcInventoryItemRead]] = defaultdict(list)
    for read in reads:
        by_famille[read.famille or "Non renseignée"].append(read)
    par_famille = sorted(
        (
            CvcParcFamille(
                famille=name,
                count=len(group),
                age_moyen=_mean([r.lifecycle_age_years for r in group if r.lifecycle_age_years is not None]),
                fin_de_vie_5ans=sum(1 for r in group if _is_fin_de_vie(r)),
                depasses=sum(1 for r in group if _is_depasse(r)),
            )
            for name, group in by_famille.items()
        ),
        key=lambda entry: -entry.count,
    )

    # Par bâtiment (les équipements non rattachés sont exclus : pas de maille)
    by_building: dict[int, list[CvcInventoryItemRead]] = defaultdict(list)
    for read in reads:
        if read.building_id is not None:
            by_building[read.building_id].append(read)
    names: dict[int, str | None] = {}
    if by_building:
        for building in db.scalars(select(Building).where(Building.id.in_(by_building.keys()))):
            names[building.id] = building.nom_batiment
    par_batiment = sorted(
        (
            CvcParcBatiment(
                building_id=bid,
                nom_batiment=names.get(bid),
                count=len(group),
                age_moyen=_mean([r.lifecycle_age_years for r in group if r.lifecycle_age_years is not None]),
                criticite_moyenne=_mean([r.criticite_pct for r in group if r.criticite_pct is not None]),
                fin_de_vie_5ans=sum(1 for r in group if _is_fin_de_vie(r)),
                depasses=sum(1 for r in group if _is_depasse(r)),
            )
            for bid, group in by_building.items()
        ),
        # Les bâtiments les plus critiques d'abord : c'est la question métier.
        key=lambda entry: (-(entry.depasses + entry.fin_de_vie_5ans), -entry.count),
    )

    pct = lambda n: round(100 * n / total, 1) if total else 0.0  # noqa: E731
    completude = CvcParcCompletude(
        rattachement_pct=pct(sum(1 for r in reads if r.building_id is not None)),
        date_mes_pct=pct(sum(1 for r in reads if r.date_mis_en_service is not None)),
        reference_pct=pct(sum(1 for r in reads if r.equipment_ref_id is not None)),
        duree_vie_pct=pct(sum(1 for r in reads if r.duree_vie_restante is not None)),
    )

    return CvcParcTechniqueReport(
        equipements_total=total,
        equipements_rattaches=sum(1 for r in reads if r.building_id is not None),
        batiments_couverts=len(by_building),
        age_moyen=_mean(known_ages),
        depasses=depasses,
        fin_de_vie_5ans=fin_de_vie,
        ages=_bucketize(ages, _AGE_BUCKETS, total),
        criticites=_bucketize(criticites, _CRITICITE_BUCKETS, total),
        par_provider=par_provider,
        par_famille=par_famille,
        par_batiment=par_batiment,
        completude=completude,
    )


def list_cvc_items_for_building(db: Session, building_id: int) -> list[CvcInventoryItemRead]:
    items = list(
        db.scalars(
            select(CvcInventoryItem)
            .where(CvcInventoryItem.building_id == building_id)
            .order_by(CvcInventoryItem.famille, CvcInventoryItem.designation)
        )
    )
    items = _filter_latest_inventory_batch(items)

    return _hydrate_items(db, items)


def update_cvc_item(
    db: Session, item_id: int, payload: CvcInventoryItemUpdate, city_id: int | None
) -> CvcInventoryItemRead | None:
    item = db.scalar(select(CvcInventoryItem).where(CvcInventoryItem.id == item_id))
    if not item:
        return None
    if city_id is not None and item.city_id not in (None, city_id):
        return None

    updates = payload.model_dump(exclude_unset=True)

    next_building_id = updates.get("building_id", item.building_id)
    next_site_id = updates.get("site_id", item.site_id)
    next_local_id = updates.get("local_id", item.local_id)
    next_ref_id = updates.get("equipment_ref_id", item.equipment_ref_id)
    next_date_mes = updates.get("date_mis_en_service", item.date_mis_en_service)

    building = db.get(Building, next_building_id) if next_building_id is not None else None
    if next_building_id is not None and building is None:
        raise ValueError("Bâtiment introuvable.")
    if building and city_id is not None and building.city_id != city_id:
        raise ValueError("Bâtiment hors périmètre utilisateur.")

    site = db.get(Site, next_site_id) if next_site_id is not None else None
    if next_site_id is not None and site is None:
        raise ValueError("Site introuvable.")
    if site and city_id is not None and site.city_id != city_id:
        raise ValueError("Site hors périmètre utilisateur.")

    local = db.get(Local, next_local_id) if next_local_id is not None else None
    if next_local_id is not None and local is None:
        raise ValueError("Local introuvable.")
    if local and next_building_id is not None and local.building_id != next_building_id:
        raise ValueError("Le local sélectionné n'appartient pas au bâtiment.")

    ref = db.get(EquipmentReference, next_ref_id) if next_ref_id is not None else None
    if next_ref_id is not None and ref is None:
        raise ValueError("Référence durée de vie introuvable.")

    if "site_id" in updates:
        item.site_id = next_site_id
    if "building_id" in updates:
        item.building_id = next_building_id
    if "local_id" in updates:
        item.local_id = next_local_id
    if "equipment_ref_id" in updates:
        item.equipment_ref_id = next_ref_id
    if "date_mis_en_service" in updates:
        item.date_mis_en_service = next_date_mes
    item.duree_vie_restante = _compute_remaining_life(item.date_mis_en_service, ref, item.etat_sante)
    if _requires_refrigerant_quantity(ref):
        if "quantite_fluide_frigorigene" in updates:
            item.quantite_fluide_frigorigene = payload.quantite_fluide_frigorigene
    else:
        item.quantite_fluide_frigorigene = None

    linked_refrigerants = list(
        db.scalars(select(CvcRefrigerantItem).where(CvcRefrigerantItem.cvc_inventory_item_id == item.id))
    )
    for refrigerant in linked_refrigerants:
        refrigerant.site_id = item.site_id
        refrigerant.building_id = item.building_id

    db.commit()
    db.refresh(item)
    return _read_item(item, ref)


def delete_cvc_items_for_building(db: Session, building_id: int) -> int:
    items = list(db.scalars(select(CvcInventoryItem).where(CvcInventoryItem.building_id == building_id)))
    count = len(items)
    for item in items:
        db.delete(item)
    db.commit()
    return count


def delete_cvc_item(db: Session, item_id: int) -> bool:
    item = db.scalar(select(CvcInventoryItem).where(CvcInventoryItem.id == item_id))
    if not item:
        return False
    db.delete(item)
    db.commit()
    return True
