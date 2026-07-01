"""Référentiel comptable et registre finances DALKIA pour le module CPE."""
from __future__ import annotations

import io
import hashlib
import json
import re
import unicodedata
from calendar import monthrange
from collections import defaultdict
from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.cpe import (
    CpeAccountingNatureRule,
    CpeAccountingSiteMapping,
    CpeContractReference,
    CpeFinanceControl,
    CpeFinanceImportBatch,
    CpeFinanceInvoice,
    CpeFinanceLine,
    CpeInvoiceEvidence,
    CpeInvoiceEvidenceLink,
    CpeResultatAnnuel,
    CpeRevisionIndex,
    CpeSite,
)
from app.services.cpe import PCS_PCI_RATIO, get_prix_gaz
from app.services.cpe_dalkia_db import (
    BPU_P2P3_POSTES,
    normalize_p2p3_poste,
    resolve_dalkia_p2p3_forfait,
    resolve_p1_gaz_tarif,
)
from app.schemas.cpe import (
    CpeAccountingImportResult,
    CpeAccountingNatureRuleCreate,
    CpeAccountingNatureRuleUpdate,
    CpeAccountingSiteMappingCreate,
    CpeAccountingSiteMappingUpdate,
    CpeContractReferenceCreate,
    CpeContractReferenceUpdate,
    CpeFinanceImportResult,
    CpeRevisionIndexCreate,
)

# Le separateur entre la categorie et le numero est un ESPACE pour la plupart des sites
# (VDS-ENS 01, VDS-SPORT 02.01) mais un TIRET pour les piscines du Lot 2 (VDS-PSC-01.01,
# VDS-PSC-02.1). On accepte les deux ([\s-]+) et 0..n niveaux de decimale (.01, .01.01) pour
# detecter ces codes et les aligner sur le referentiel DALKIA L2.
_SITE_CODE_RE = re.compile(r"\b(VDS-[A-Z]+[\s-]+\d+(?:\.\d+)*|CCAS\s+\d+)\b", flags=re.IGNORECASE)
CPE_CONTRACT_SCOPE_KIND = "cpe_contract_scope"
P1_GAZ_ACOMPTE_KIND = "p1_gaz_acompte"
P1_GAZ_ACOMPTE_ITEMS = {"P1", "ABT", "CTA", "CPB", "LOCATION", "STOCKAGE", "TERME FIXE"}
ICHT_IME_BASE = 141.4
FSD2_BASE = 169.8
BT40_BASE = 128.4
P2_REVISION_FORMULA = "P2 = P20 x (0,15 + 0,70 x ICHT-IME/ICHT-IME0 + 0,15 x FSD2/FSD20)"
P3_REVISION_FORMULA = "P3 = P30 x (0,15 + 0,30 x ICHT-IME/ICHT-IME0 + 0,55 x BT40/BT400)"
P2_4_OBJECTIVE_RULE = "P2.4 annuel : 100% si objectifs atteints, 50% si objectifs non atteints"
REFERENCE_INDEX_CODES = ("ICHT_IME0", "FSD20", "BT400")
REFERENCE_INDEX_DEFAULTS: tuple[dict[str, Any], ...] = (
    {
        "index_code": "ICHT_IME0",
        "year": 2025,
        "quarter": 0,
        "value": 141.4,
        "source": "CPE DALKIA - Facturation et indices",
        "notes": "Base contractuelle P2/P3 (01/01/2025).",
    },
    {
        "index_code": "FSD20",
        "year": 2025,
        "quarter": 0,
        "value": 169.8,
        "source": "CPE DALKIA - Facturation et indices",
        "notes": "Base contractuelle P2 (01/01/2025).",
    },
    {
        "index_code": "BT400",
        "year": 2025,
        "quarter": 0,
        "value": 128.4,
        "source": "CPE DALKIA - Facturation et indices",
        "notes": "Base contractuelle P3/P3.4 (01/01/2025).",
    },
    {
        "index_code": "P1_CPB0_T1",
        "year": 2026,
        "quarter": 0,
        "value": 38.62,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante CPB T1.",
    },
    {
        "index_code": "P1_CPB0_T2",
        "year": 2026,
        "quarter": 0,
        "value": 38.62,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante CPB T2.",
    },
    {
        "index_code": "P1_CPB0_T3",
        "year": 2026,
        "quarter": 0,
        "value": 38.62,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante CPB T3.",
    },
    {
        "index_code": "P1_TVD0_T1",
        "year": 2026,
        "quarter": 0,
        "value": 44.94,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante TVD T1.",
    },
    {
        "index_code": "P1_TVD0_T2",
        "year": 2026,
        "quarter": 0,
        "value": 12.08,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante TVD T2.",
    },
    {
        "index_code": "P1_TVD0_T3",
        "year": 2026,
        "quarter": 0,
        "value": 8.69,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante TVD T3.",
    },
    {
        "index_code": "P1_CEE0",
        "year": 2026,
        "quarter": 0,
        "value": 7.63,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante CEE.",
    },
    {
        "index_code": "P1_TICGN0",
        "year": 2026,
        "quarter": 0,
        "value": 15.43,
        "source": "OS3 - Prix fixe gaz 5 ans",
        "notes": "Reference P1 gaz composante TICGN.",
    },
)
INVOICE_TYPE_LABELS = {
    "AC": "Acompte",
    "AJ": "Ajustement / avoir",
    "DE": "Facture definitive",
    "EC": "Echeance / facture courante",
    "RE": "Regularisation",
}


def _norm_header(value: Any) -> str:
    raw = "" if value is None else str(value).strip()
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.lower()).strip("_")


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _upper(value: Any) -> str | None:
    text = _clean(value)
    return text.upper() if text else None


def _float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return None


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _site_code(value: Any) -> str | None:
    text = _clean(value)
    if not text:
        return None
    match = _SITE_CODE_RE.search(text)
    if not match:
        return None
    return re.sub(r"\s+", " ", match.group(1).upper()).strip()


def _compact_text(value: Any) -> str:
    text = _norm_text(value)
    return re.sub(r"[^A-Z0-9]", "", text)


def _rows_from_sheet(ws, header_row: int) -> list[dict[str, Any]]:
    headers = [_norm_header(cell.value) for cell in ws[header_row]]
    rows: list[dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in excel_row):
            continue
        rows.append({headers[index]: value for index, value in enumerate(excel_row) if index < len(headers)})
    return rows


def list_accounting_nature_rules(db: Session, city_id: int | None = None) -> list[CpeAccountingNatureRule]:
    query = select(CpeAccountingNatureRule)
    if city_id is not None:
        query = query.where(CpeAccountingNatureRule.city_id == city_id)
    query = query.order_by(
        CpeAccountingNatureRule.contract_code,
        CpeAccountingNatureRule.market,
        CpeAccountingNatureRule.service_sold,
        CpeAccountingNatureRule.billed_item,
    )
    return list(db.scalars(query).all())


def create_accounting_nature_rule(db: Session, payload: CpeAccountingNatureRuleCreate) -> CpeAccountingNatureRule:
    rule = CpeAccountingNatureRule(**payload.model_dump())
    db.add(rule)
    db.commit()
    db.refresh(rule)
    return rule


def update_accounting_nature_rule(
    db: Session,
    rule: CpeAccountingNatureRule,
    payload: CpeAccountingNatureRuleUpdate,
) -> CpeAccountingNatureRule:
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(rule, field, value)
    db.commit()
    db.refresh(rule)
    return rule


def delete_accounting_nature_rule(db: Session, rule: CpeAccountingNatureRule) -> None:
    db.delete(rule)
    db.commit()


def list_contract_references(db: Session, city_id: int | None = None) -> list[CpeContractReference]:
    query = select(CpeContractReference)
    if city_id is not None:
        query = query.where(CpeContractReference.city_id == city_id)
    query = query.order_by(
        CpeContractReference.contract_code,
        CpeContractReference.year.desc(),
        CpeContractReference.reference_kind,
        CpeContractReference.market,
        CpeContractReference.billed_item,
    )
    return list(db.scalars(query).all())


def create_contract_reference(db: Session, payload: CpeContractReferenceCreate) -> CpeContractReference:
    reference = CpeContractReference(**payload.model_dump())
    reference.contract_code = reference.contract_code.strip().upper()
    reference.reference_kind = reference.reference_kind.strip().lower()
    reference.market = reference.market.strip().upper()
    reference.billed_item = reference.billed_item.strip().upper()
    db.add(reference)
    db.commit()
    db.refresh(reference)
    return reference


def update_contract_reference(
    db: Session,
    reference: CpeContractReference,
    payload: CpeContractReferenceUpdate,
) -> CpeContractReference:
    for field, value in payload.model_dump(exclude_unset=True).items():
        if isinstance(value, str) and field in {"contract_code", "market", "billed_item"}:
            value = value.strip().upper()
        elif isinstance(value, str) and field == "reference_kind":
            value = value.strip().lower()
        setattr(reference, field, value)
    db.commit()
    db.refresh(reference)
    return reference


def delete_contract_reference(db: Session, reference: CpeContractReference) -> None:
    db.delete(reference)
    db.commit()


def list_accounting_site_mappings(db: Session, city_id: int | None = None) -> list[CpeAccountingSiteMapping]:
    query = select(CpeAccountingSiteMapping)
    if city_id is not None:
        query = query.where(CpeAccountingSiteMapping.city_id == city_id)
    query = query.order_by(CpeAccountingSiteMapping.code_site)
    return list(db.scalars(query).all())


def create_accounting_site_mapping(db: Session, payload: CpeAccountingSiteMappingCreate) -> CpeAccountingSiteMapping:
    mapping = CpeAccountingSiteMapping(**payload.model_dump())
    db.add(mapping)
    db.commit()
    db.refresh(mapping)
    return mapping


def update_accounting_site_mapping(
    db: Session,
    mapping: CpeAccountingSiteMapping,
    payload: CpeAccountingSiteMappingUpdate,
) -> CpeAccountingSiteMapping:
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(mapping, field, value)
    db.commit()
    db.refresh(mapping)
    return mapping


def delete_accounting_site_mapping(db: Session, mapping: CpeAccountingSiteMapping) -> None:
    db.delete(mapping)
    db.commit()


def _market_from_billed_item(billed_item: str | None) -> str:
    item = (billed_item or "").upper()
    for prefix in ("P1", "P2", "P3", "R1", "R2"):
        if item.startswith(prefix):
            return prefix
    return item[:30] or "AUTRE"


def get_current_cpe_contract_codes(
    db: Session,
    city_id: int | None = None,
    year: int | None = None,
) -> set[str]:
    """Contrats actifs du perimetre CPE Ville, lus depuis le referentiel editable."""
    query = select(CpeContractReference.contract_code).where(
        CpeContractReference.reference_kind == CPE_CONTRACT_SCOPE_KIND,
        CpeContractReference.active.is_(True),
    )
    if city_id is not None:
        query = query.where(or_(CpeContractReference.city_id == city_id, CpeContractReference.city_id.is_(None)))
    if year is not None:
        query = query.where(CpeContractReference.year <= year)
    return {code.strip().upper() for code in db.scalars(query).all() if code and code.strip()}


def _is_current_cpe_contract(
    db: Session,
    contract_code: str | None,
    city_id: int | None = None,
    year: int | None = None,
) -> bool:
    code = (contract_code or "").strip().upper()
    return bool(code) and code in get_current_cpe_contract_codes(db, city_id=city_id, year=year)


def _split_csv_tokens(value: str | None) -> set[str]:
    if not value:
        return set()
    return {part.strip().upper() for part in value.split(",") if part.strip()}


def _reference_included_items(reference: CpeContractReference) -> set[str]:
    raw = reference.included_billed_items
    if not raw:
        return set()
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return _split_csv_tokens(raw)
    if isinstance(parsed, list):
        return {str(item).strip().upper() for item in parsed if str(item).strip()}
    return _split_csv_tokens(raw)


def _rule_notes_from_contract_row(row: dict[str, Any]) -> str | None:
    parts = []
    for key, label in [
        ("marche_perimetre_a_confirmer", "Périmètre"),
        ("statut_codification", "Statut"),
        ("regle_de_codification", "Règle"),
        ("question_precision_a_dalkia", "Question DALKIA"),
        ("alerte_periode", "Alerte période"),
    ]:
        value = _clean(row.get(key))
        if value:
            parts.append(f"{label}: {value}")
    return " | ".join(parts) or None


def _upsert_nature_rule(
    db: Session,
    *,
    city_id: int | None,
    contract_code: str | None,
    market: str,
    service_sold: str | None,
    billed_item: str,
    frequency: str | None,
    accounting_nature: str,
    accounting_label: str | None,
    notes: str | None,
) -> bool:
    existing = db.scalars(
        select(CpeAccountingNatureRule).where(
            CpeAccountingNatureRule.city_id == city_id,
            func.coalesce(CpeAccountingNatureRule.contract_code, "") == (contract_code or ""),
            CpeAccountingNatureRule.market == market,
            func.coalesce(CpeAccountingNatureRule.service_sold, "") == (service_sold or ""),
            CpeAccountingNatureRule.billed_item == billed_item,
            func.coalesce(CpeAccountingNatureRule.frequency, "") == (frequency or ""),
        )
    ).first()
    payload = {
        "city_id": city_id,
        "contract_code": contract_code,
        "market": market,
        "service_sold": service_sold,
        "billed_item": billed_item,
        "frequency": frequency,
        "accounting_nature": accounting_nature,
        "accounting_label": accounting_label,
        "active": True,
        "notes": notes,
    }
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        return False
    db.add(CpeAccountingNatureRule(**payload))
    return True


def import_codification_workbook(
    db: Session,
    raw_bytes: bytes,
    *,
    filename: str | None,
    city_id: int | None,
) -> CpeAccountingImportResult:
    """Importe la matrice `analyse_codification_dalkia.xlsx` en upsert."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    errors: list[str] = []
    created_rules = updated_rules = created_sites = updated_sites = 0

    if "Postes x contrat x nature" in wb.sheetnames:
        ws = wb["Postes x contrat x nature"]
        for row in _rows_from_sheet(ws, 1):
            contract_code = _upper(row.get("code_contrat"))
            billed_item = _upper(row.get("poste_facture"))
            nature = _clean(row.get("nature_proposee"))
            if not contract_code or not billed_item or not nature:
                continue
            if _upsert_nature_rule(
                db,
                city_id=city_id,
                contract_code=contract_code,
                market=_market_from_billed_item(billed_item),
                service_sold=None,
                billed_item=billed_item,
                frequency=_clean(row.get("nb_lignes")),
                accounting_nature=nature,
                accounting_label=_clean(row.get("libelle_nature")),
                notes=_rule_notes_from_contract_row(row),
            ):
                created_rules += 1
            else:
                updated_rules += 1
    elif "Poste facturé vers Nature ctpab" in wb.sheetnames:
        ws = wb["Poste facturé vers Nature ctpab"]
        for row in _rows_from_sheet(ws, 3):
            market = _upper(row.get("marche"))
            billed_item = _upper(row.get("poste_facture"))
            nature = _clean(row.get("nature_proposee"))
            if not market or not billed_item or not nature:
                continue
            service_sold = _upper(row.get("service_vendu"))
            frequency = _clean(row.get("frequence"))
            if _upsert_nature_rule(
                db,
                city_id=city_id,
                contract_code=None,
                market=market,
                service_sold=service_sold,
                billed_item=billed_item,
                frequency=frequency,
                accounting_nature=nature,
                accounting_label=_clean(row.get("libelle_nature")),
                notes=None,
            ):
                created_rules += 1
            else:
                updated_rules += 1
    else:
        errors.append("Feuille absente : Postes x contrat x nature ou Poste facturé vers Nature ctpab")

    if "Sites vers codes" in wb.sheetnames:
        ws = wb["Sites vers codes"]
        for row in _rows_from_sheet(ws, 1):
            code_site = _upper(row.get("code_site"))
            site_name = _clean(row.get("nom_du_site"))
            if not code_site or not site_name:
                continue
            existing = db.scalars(
                select(CpeAccountingSiteMapping).where(
                    CpeAccountingSiteMapping.city_id == city_id,
                    CpeAccountingSiteMapping.code_site == code_site,
                )
            ).first()
            payload = {
                "city_id": city_id,
                "code_site": code_site,
                "site_name": site_name,
                "family": _clean(row.get("famille")),
                "manager": _clean(row.get("gestionnaire")),
                "alternate_manager": _clean(row.get("gestionnaire_alternatif")),
                "service_code": _clean(row.get("service")),
                "service_label": _clean(row.get("libelle_service")),
                "function_code": _clean(row.get("fonction")),
                "function_label": _clean(row.get("libelle_fonction")),
                "antenna_code": _clean(row.get("antenne")),
                "antenna_label": _clean(row.get("libelle_antenne")),
                "operation_code": _clean(row.get("operation_si_travaux")),
                "operation_label": _clean(row.get("libelle_operation")),
                "active": True,
            }
            if existing:
                for key, value in payload.items():
                    setattr(existing, key, value)
                updated_sites += 1
            else:
                db.add(CpeAccountingSiteMapping(**payload))
                created_sites += 1
    else:
        errors.append("Feuille absente : Sites vers codes")

    db.commit()
    return CpeAccountingImportResult(
        filename=filename,
        nature_rules_created=created_rules,
        nature_rules_updated=updated_rules,
        site_mappings_created=created_sites,
        site_mappings_updated=updated_sites,
        errors=errors,
    )


def _find_accounting_rule(
    rules: list[CpeAccountingNatureRule],
    contract_code: str | None,
    market: str | None,
    service_sold: str | None,
    billed_item: str | None,
) -> CpeAccountingNatureRule | None:
    contract_norm = (contract_code or "").upper()
    market_norm = (market or "").upper()
    service_norm = (service_sold or "").upper()
    item_norm = (billed_item or "").upper()
    contract_matches = [
        rule
        for rule in rules
        if (rule.contract_code or "").upper() == contract_norm and (rule.billed_item or "").upper() == item_norm
    ]
    for rule in contract_matches:
        if not rule.service_sold or rule.service_sold.upper() == service_norm:
            return rule
    for rule in rules:
        if rule.contract_code:
            continue
        if rule.market.upper() == market_norm and (rule.billed_item or "").upper() == item_norm:
            if not rule.service_sold or rule.service_sold.upper() == service_norm:
                return rule
    return None


def _find_site_mapping(
    site_mappings_by_code: dict[str, CpeAccountingSiteMapping],
    site_mappings: list[CpeAccountingSiteMapping],
    detail: str | None,
) -> CpeAccountingSiteMapping | None:
    detected_site = _site_code(detail)
    if detected_site:
        direct = site_mappings_by_code.get(detected_site)
        if direct:
            return direct

    compact_detail = _compact_text(detail)
    if not compact_detail:
        return None

    ranked = sorted(
        site_mappings,
        key=lambda mapping: max(
            len(_compact_text(mapping.code_site)),
            len(_compact_text(mapping.site_name)),
            len(_compact_text(mapping.antenna_label)),
        ),
        reverse=True,
    )
    for mapping in ranked:
        for value in (mapping.code_site, mapping.site_name, mapping.antenna_label):
            token = _compact_text(value)
            if len(token) >= 8 and token in compact_detail:
                return mapping
    return None


def import_finance_workbook(
    db: Session,
    raw_bytes: bytes,
    *,
    filename: str | None,
    city_id: int | None,
) -> CpeFinanceImportResult:
    """Importe un export finances DALKIA XLSX et persiste factures + lignes."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = _rows_from_sheet(ws, 1)
    if not rows:
        raise ValueError("Aucune ligne exploitable dans l'export finances.")

    batch = CpeFinanceImportBatch(city_id=city_id, filename=filename)
    db.add(batch)
    db.flush()

    rules = list_accounting_nature_rules(db, city_id)
    site_mapping_rows = list_accounting_site_mappings(db, city_id)
    site_mappings = {m.code_site.upper(): m for m in site_mapping_rows}
    invoice_by_number: dict[str, CpeFinanceInvoice] = {}
    invoice_lines: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
    total_ht = 0.0
    matched_rules = matched_sites = 0
    warnings: list[str] = []

    for index, row in enumerate(rows, start=2):
        invoice_number = _clean(row.get("numero_de_facture")) or f"sans-numero-ligne-{index}"
        amount_ht = _float(row.get("montant_ht")) or 0.0
        total_ht += amount_ht
        invoice_lines[invoice_number].append((index, row))
        if invoice_number not in invoice_by_number:
            invoice = CpeFinanceInvoice(
                batch_id=batch.id,
                city_id=city_id,
                invoice_number=invoice_number,
                contract_code=_clean(row.get("code_contrat")),
                contract_label=_clean(row.get("libelle_contrat")),
                invoice_type=_clean(row.get("type_de_facture")),
                supplier=_clean(row.get("societe")),
                customer_code=_clean(row.get("code_du_client")),
                customer_name=_clean(row.get("nom_du_client")),
                invoice_date=_date(row.get("date_d_edition")),
                due_date=_date(row.get("date_d_echeance_de_la_facture")),
                period_start=_date(row.get("debut_periode_de_facturation")),
                period_end=_date(row.get("fin_periode_de_facturation")),
                total_ht=0.0,
            )
            db.add(invoice)
            db.flush()
            invoice_by_number[invoice_number] = invoice

        invoice = invoice_by_number[invoice_number]
        invoice.total_ht = round((invoice.total_ht or 0.0) + amount_ht, 2)
        start = _date(row.get("debut_periode_de_facturation"))
        end = _date(row.get("fin_periode_de_facturation"))
        if start and (invoice.period_start is None or start < invoice.period_start):
            invoice.period_start = start
        if end and (invoice.period_end is None or end > invoice.period_end):
            invoice.period_end = end

    for invoice_number, source_rows in invoice_lines.items():
        invoice = invoice_by_number[invoice_number]
        for row_number, source_row in source_rows:
            market = _upper(source_row.get("marche"))
            service_sold = _upper(source_row.get("service_vendu"))
            billed_item = _upper(source_row.get("poste_facture"))
            detail = _clean(source_row.get("lieu_ou_detail_de_la_prestation"))
            detected_site = _site_code(detail)
            site_mapping = _find_site_mapping(site_mappings, site_mapping_rows, detail)
            if site_mapping and not detected_site:
                detected_site = site_mapping.code_site
            rule = _find_accounting_rule(
                rules,
                _clean(source_row.get("code_contrat")),
                market,
                service_sold,
                billed_item,
            )
            if site_mapping:
                matched_sites += 1
            if rule:
                matched_rules += 1
            line = CpeFinanceLine(
                batch_id=batch.id,
                invoice_id=invoice.id,
                city_id=city_id,
                row_number=row_number,
                contract_code=_clean(source_row.get("code_contrat")),
                invoice_number=invoice_number,
                market=market,
                market_type=_clean(source_row.get("type_de_marche")),
                service_sold=service_sold,
                billed_item=billed_item,
                vat_rate=_float(source_row.get("taux_de_tva")),
                amount_ht=_float(source_row.get("montant_ht")) or 0.0,
                consumption=_float(source_row.get("consommation")),
                unit=_clean(source_row.get("unite")),
                base_price=_float(source_row.get("prix_de_base")),
                revised_price=_float(source_row.get("prix_ou_forfait_revise")),
                detail=detail,
                site_code_detected=detected_site,
                accounting_site_id=site_mapping.id if site_mapping else None,
                accounting_rule_id=rule.id if rule else None,
                accounting_nature=rule.accounting_nature if rule else None,
                accounting_label=rule.accounting_label if rule else None,
                period_start=_date(source_row.get("debut_periode_de_facturation")),
                period_end=_date(source_row.get("fin_periode_de_facturation")),
                raw_json=json.dumps({key: str(value) if value is not None else None for key, value in source_row.items()}),
            )
            db.add(line)

    batch.line_count = len(rows)
    batch.invoice_count = len(invoice_by_number)
    batch.total_ht = round(total_ht, 2)
    if matched_rules < len(rows):
        warnings.append(f"{len(rows) - matched_rules} ligne(s) sans nature comptable rattachee.")
    if matched_sites == 0:
        warnings.append("Aucun code site VDS/CCAS detecte dans les lignes : un mapping detail DALKIA -> site sera necessaire.")
    db.commit()
    db.refresh(batch)
    invoices = list(invoice_by_number.values())
    for invoice in invoices:
        db.refresh(invoice)
    return CpeFinanceImportResult(
        batch=batch,
        invoices=sorted(invoices, key=lambda item: item.invoice_number),
        line_count=len(rows),
        matched_accounting_rules=matched_rules,
        matched_site_mappings=matched_sites,
        warnings=warnings,
    )


def list_revision_indices(
    db: Session,
    city_id: int | None = None,
    year: int | None = None,
) -> list[CpeRevisionIndex]:
    _ensure_revision_reference_defaults(db, city_id)
    query = select(CpeRevisionIndex)
    if city_id is not None:
        query = query.where(CpeRevisionIndex.city_id == city_id)
    if year is not None:
        query = query.where(CpeRevisionIndex.year == year)
    query = query.order_by(CpeRevisionIndex.year.desc(), CpeRevisionIndex.quarter.desc(), CpeRevisionIndex.index_code)
    return list(db.scalars(query).all())


def list_revision_observations(db: Session, city_id: int | None = None) -> list[dict[str, Any]]:
    """Liste les coefficients appliques par DALKIA, sans les assimiler a des indices officiels."""
    indices = {
        (item.index_code, item.year, item.quarter): item.value
        for item in list_revision_indices(db, city_id)
        if item.quarter > 0
    }
    reference_values = _reference_index_map(db, city_id)
    icht_ime_base = reference_values.get("ICHT_IME0", ICHT_IME_BASE)
    fsd2_base = reference_values.get("FSD20", FSD2_BASE)
    bt40_base = reference_values.get("BT400", BT40_BASE)

    current_contract_codes = get_current_cpe_contract_codes(db, city_id=city_id)
    query = select(CpeFinanceLine).where(
        CpeFinanceLine.contract_code.in_(current_contract_codes),
        CpeFinanceLine.market.in_(("P2", "P3")),
        CpeFinanceLine.base_price.is_not(None),
        CpeFinanceLine.revised_price.is_not(None),
    )
    if city_id is not None:
        query = query.where(CpeFinanceLine.city_id == city_id)

    grouped: dict[tuple[str, int, int, float], dict[str, Any]] = {}
    for line in db.scalars(query).all():
        year, quarter = _line_index_period(line)
        if year is None or quarter is None or not line.base_price:
            continue
        observed_factor = round(line.revised_price / line.base_price, 6)
        key = (line.market, year, quarter, observed_factor)
        item = grouped.setdefault(
            key,
            {
                "market": line.market,
                "year": year,
                "quarter": quarter,
                "observed_factor": observed_factor,
                "invoice_numbers": set(),
                "line_count": 0,
            },
        )
        item["line_count"] += 1
        if line.invoice_number:
            item["invoice_numbers"].add(line.invoice_number)

    observations: list[dict[str, Any]] = []
    for item in grouped.values():
        market = item["market"]
        year = item["year"]
        quarter = item["quarter"]
        icht = _index_value(indices, "ICHT_IME", year, quarter)
        if market == "P2":
            other = _index_value(indices, "FSD2", year, quarter)
            required_indices = ["ICHT-IME", "FSD2"]
            expected_factor = (
                _p2_factor(icht, other, icht_ime_base=icht_ime_base, fsd2_base=fsd2_base)
                if icht is not None and other is not None
                else None
            )
        else:
            other = _index_value(indices, "BT40", year, quarter)
            required_indices = ["ICHT-IME", "BT40"]
            expected_factor = (
                _p3_factor(icht, other, icht_ime_base=icht_ime_base, bt40_base=bt40_base)
                if icht is not None and other is not None
                else None
            )

        delta_factor = round(item["observed_factor"] - expected_factor, 6) if expected_factor is not None else None
        if expected_factor is None:
            status = "to_verify"
            message = (
                f"Nouveau coefficient DALKIA detecte pour {market} {year} T{quarter}. "
                f"Verifier les indices {', '.join(required_indices)} dans une source officielle ou la facture PDF."
            )
        elif abs(delta_factor or 0.0) <= 0.0001:
            status = "matches_validated"
            message = f"Coefficient DALKIA coherent avec les indices valides pour {market} {year} T{quarter}."
        else:
            status = "conflict"
            message = (
                f"Coefficient DALKIA different du calcul avec les indices valides pour {market} {year} T{quarter}. "
                "Verifier la date d'effet et les valeurs saisies."
            )
        observations.append(
            {
                **item,
                "invoice_numbers": sorted(item["invoice_numbers"]),
                "expected_factor": round(expected_factor, 6) if expected_factor is not None else None,
                "delta_factor": delta_factor,
                "status": status,
                "required_indices": required_indices,
                "message": message,
            }
        )

    return sorted(
        observations,
        key=lambda item: (item["year"], item["quarter"], item["market"], item["observed_factor"]),
        reverse=True,
    )


def _pdf_decimal(line: str) -> float | None:
    matches = re.findall(r"\d{2,4},\d{2,6}", line)
    return _float(matches[0]) if matches else None


def extract_invoice_evidence_pdf(raw_bytes: bytes) -> dict[str, Any]:
    """Extrait les indices declares dans la section revision d'une facture DALKIA."""
    try:
        reader = PdfReader(io.BytesIO(raw_bytes))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:
        raise ValueError("PDF DALKIA illisible.") from exc
    if not text.strip():
        raise ValueError("Le PDF DALKIA ne contient aucun texte exploitable.")

    invoice_match = re.search(r"Facture\s*n[°o]\s*([A-Z0-9]+)", text, flags=re.IGNORECASE)
    revision_match = re.search(r"R[ée]vision\s+au\s+(\d{2}/\d{2}/\d{4})", text, flags=re.IGNORECASE)
    factor_match = re.search(r"Coefficient\s+de\s+r[ée]vision\s+(\d+[,.]\d+)", text, flags=re.IGNORECASE)
    values: dict[str, float | None] = {"declared_icht_ime": None, "declared_fsd2": None, "declared_bt40": None}
    for line in text.splitlines():
        normalized = _norm_text(line)
        if "ICHT" in normalized and values["declared_icht_ime"] is None:
            values["declared_icht_ime"] = _pdf_decimal(line)
        elif "FSD2" in normalized and values["declared_fsd2"] is None:
            values["declared_fsd2"] = _pdf_decimal(line)
        elif "BT40" in normalized and values["declared_bt40"] is None:
            values["declared_bt40"] = _pdf_decimal(line)

    return {
        "declared_invoice_number": invoice_match.group(1) if invoice_match else None,
        "revision_date": datetime.strptime(revision_match.group(1), "%d/%m/%Y").date() if revision_match else None,
        "declared_factor": _float(factor_match.group(1)) if factor_match else None,
        **values,
    }


def _link_evidence_to_invoice(db: Session, evidence: CpeInvoiceEvidence, invoice: CpeFinanceInvoice) -> None:
    existing = db.scalars(
        select(CpeInvoiceEvidenceLink).where(
            CpeInvoiceEvidenceLink.evidence_id == evidence.id,
            CpeInvoiceEvidenceLink.invoice_id == invoice.id,
        )
    ).first()
    if existing is None:
        db.add(CpeInvoiceEvidenceLink(evidence_id=evidence.id, invoice_id=invoice.id))
    if evidence.invoice_id is None:
        evidence.invoice_id = invoice.id


def add_revision_evidence_pdf(
    db: Session,
    raw_bytes: bytes,
    *,
    filename: str,
    uploaded_by_user_id: int,
    city_id: int | None,
    invoice: CpeFinanceInvoice | None = None,
) -> CpeInvoiceEvidence:
    if not filename.lower().endswith(".pdf"):
        raise ValueError("Le justificatif doit etre une facture PDF DALKIA.")
    extracted = extract_invoice_evidence_pdf(raw_bytes)
    declared_number = extracted["declared_invoice_number"]
    if invoice is not None and declared_number and declared_number != invoice.invoice_number:
        raise ValueError(
            f"Le PDF concerne la facture {declared_number}, pas la facture {invoice.invoice_number}."
        )

    sha256 = hashlib.sha256(raw_bytes).hexdigest()
    existing = db.scalars(
        select(CpeInvoiceEvidence).where(
            CpeInvoiceEvidence.city_id == city_id,
            CpeInvoiceEvidence.sha256 == sha256,
        )
    ).first()
    if existing:
        if invoice is not None:
            _link_evidence_to_invoice(db, existing, invoice)
            db.commit()
            db.refresh(existing)
        return existing

    if invoice is None and declared_number:
        invoice = db.scalars(
            select(CpeFinanceInvoice).where(
                CpeFinanceInvoice.city_id == city_id,
                CpeFinanceInvoice.invoice_number == declared_number,
            )
        ).first()

    target_dir = Path(settings.invoice_storage_dir) / "cpe" / str(city_id or "global")
    target_dir.mkdir(parents=True, exist_ok=True)
    document_ref = declared_number or (invoice.invoice_number if invoice else "revision")
    storage_path = target_dir / f"{document_ref}-{sha256[:12]}.pdf"
    storage_path.write_bytes(raw_bytes)
    revision_date = extracted.get("revision_date")
    evidence = CpeInvoiceEvidence(
        city_id=city_id,
        invoice_id=invoice.id if invoice else None,
        uploaded_by_user_id=uploaded_by_user_id,
        original_filename=filename,
        storage_path=str(storage_path),
        sha256=sha256,
        extraction_status="parsed",
        validation_status="declared_to_verify",
        evidence_kind="invoice_pdf",
        year=revision_date.year if revision_date else None,
        quarter=((revision_date.month - 1) // 3) + 1 if revision_date else None,
        effective_date=revision_date,
        **extracted,
    )
    db.add(evidence)
    db.flush()
    if invoice is not None:
        _link_evidence_to_invoice(db, evidence, invoice)
    db.commit()
    db.refresh(evidence)
    return evidence


def add_invoice_evidence_pdf(
    db: Session,
    invoice: CpeFinanceInvoice,
    raw_bytes: bytes,
    *,
    filename: str,
    uploaded_by_user_id: int,
) -> CpeInvoiceEvidence:
    return add_revision_evidence_pdf(
        db,
        raw_bytes,
        filename=filename,
        uploaded_by_user_id=uploaded_by_user_id,
        city_id=invoice.city_id,
        invoice=invoice,
    )


def list_revision_evidences(db: Session, city_id: int | None = None) -> list[CpeInvoiceEvidence]:
    query = select(CpeInvoiceEvidence)
    if city_id is not None:
        query = query.where(CpeInvoiceEvidence.city_id == city_id)
    query = query.order_by(CpeInvoiceEvidence.created_at.desc(), CpeInvoiceEvidence.id.desc())
    return list(db.scalars(query).all())


def apply_invoice_evidence_declared_indices(db: Session, evidence: CpeInvoiceEvidence) -> list[CpeRevisionIndex]:
    """Reporte les valeurs du PDF comme declarations DALKIA a verifier, jamais comme indices officiels."""
    if evidence.revision_date is None:
        raise ValueError("Date de revision absente du PDF : impossible de positionner les indices.")
    year = evidence.revision_date.year
    quarter = ((evidence.revision_date.month - 1) // 3) + 1
    declared_values = {
        "ICHT_IME": evidence.declared_icht_ime,
        "FSD2": evidence.declared_fsd2,
        "BT40": evidence.declared_bt40,
    }
    if not any(value is not None for value in declared_values.values()):
        raise ValueError("Aucun indice de revision exploitable n'a ete extrait du PDF.")

    rows: list[CpeRevisionIndex] = []
    for index_code, value in declared_values.items():
        if value is None:
            continue
        existing = db.scalars(
            select(CpeRevisionIndex).where(
                CpeRevisionIndex.city_id == evidence.city_id,
                CpeRevisionIndex.index_code == index_code,
                CpeRevisionIndex.year == year,
                CpeRevisionIndex.quarter == quarter,
            )
        ).first()
        if existing and existing.verification_status == "official_verified":
            rows.append(existing)
            continue
        if existing is None:
            existing = CpeRevisionIndex(
                city_id=evidence.city_id,
                index_code=index_code,
                year=year,
                quarter=quarter,
                value=value,
            )
            db.add(existing)
        existing.value = value
        existing.source = f"Facture DALKIA {evidence.declared_invoice_number or evidence.invoice_id or evidence.id}"
        existing.verification_status = "declared_to_verify"
        existing.evidence_id = evidence.id
        existing.notes = "Valeur declaree dans le PDF DALKIA ; verification externe requise."
        rows.append(existing)
    db.commit()
    for row in rows:
        db.refresh(row)
    return rows


def upsert_revision_index(db: Session, payload: CpeRevisionIndexCreate) -> CpeRevisionIndex:
    index_code = payload.index_code.strip().upper().replace("-", "_")
    existing = db.scalars(
        select(CpeRevisionIndex).where(
            CpeRevisionIndex.city_id == payload.city_id,
            CpeRevisionIndex.index_code == index_code,
            CpeRevisionIndex.year == payload.year,
            CpeRevisionIndex.quarter == payload.quarter,
        )
    ).first()
    if existing:
        existing.value = payload.value
        existing.source = payload.source
        existing.verification_status = payload.verification_status
        existing.evidence_id = payload.evidence_id
        existing.notes = payload.notes
        db.commit()
        db.refresh(existing)
        return existing
    item = CpeRevisionIndex(**payload.model_dump(exclude={"index_code"}), index_code=index_code)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def _ensure_revision_reference_defaults(db: Session, city_id: int | None) -> None:
    created = False
    for default in REFERENCE_INDEX_DEFAULTS:
        existing = db.scalars(
            select(CpeRevisionIndex).where(
                CpeRevisionIndex.city_id == city_id,
                CpeRevisionIndex.index_code == default["index_code"],
                CpeRevisionIndex.year == default["year"],
                CpeRevisionIndex.quarter == default["quarter"],
            )
        ).first()
        if existing is not None:
            continue
        db.add(
            CpeRevisionIndex(
                city_id=city_id,
                index_code=default["index_code"],
                year=default["year"],
                quarter=default["quarter"],
                value=float(default["value"]),
                source=default["source"],
                notes=default["notes"],
            )
        )
        created = True
    if created:
        db.commit()


def _reference_index_map(db: Session, city_id: int | None) -> dict[str, float]:
    _ensure_revision_reference_defaults(db, city_id)
    rows = list(
        db.scalars(
            select(CpeRevisionIndex).where(
                CpeRevisionIndex.city_id == city_id,
                CpeRevisionIndex.quarter == 0,
            )
        ).all()
    )
    reference_by_code: dict[str, tuple[int, float]] = {}
    for row in rows:
        current = reference_by_code.get(row.index_code)
        if current is None or row.year > current[0]:
            reference_by_code[row.index_code] = (row.year, row.value)
    return {code: value for code, (_year, value) in reference_by_code.items()}


def _line_index_period(line: CpeFinanceLine) -> tuple[int | None, int | None]:
    period_date = line.period_end or line.period_start
    if period_date is None:
        return None, None
    return period_date.year, ((period_date.month - 1) // 3) + 1


def _index_value(indices: dict[tuple[str, int, int], float], code: str, year: int, quarter: int) -> float | None:
    return indices.get((code, year, quarter))


def _p3_factor(icht_ime: float, bt40: float, *, icht_ime_base: float, bt40_base: float) -> float:
    return 0.15 + 0.30 * (icht_ime / icht_ime_base) + 0.55 * (bt40 / bt40_base)


def _p2_factor(icht_ime: float, fsd2: float, *, icht_ime_base: float, fsd2_base: float) -> float:
    return 0.15 + 0.70 * (icht_ime / icht_ime_base) + 0.15 * (fsd2 / fsd2_base)


def _line_raw_float(line: CpeFinanceLine, key: str) -> float | None:
    if not line.raw_json:
        return None
    try:
        raw = json.loads(line.raw_json)
    except json.JSONDecodeError:
        return None
    return _float(raw.get(key))


def _line_raw_str(line: CpeFinanceLine, key: str) -> str | None:
    if not line.raw_json:
        return None
    try:
        raw = json.loads(line.raw_json)
    except json.JSONDecodeError:
        return None
    return _clean(raw.get(key))


def _line_revision_breakdown(line: CpeFinanceLine) -> tuple[float | None, float | None]:
    """Decompose le montant HT facture en (acompte hors revision, montant de la revision).

    DALKIA exporte un `prix_de_base` et un `prix_ou_forfait_revise` qui sont les
    montants annuels du poste (le revise = base x coefficient de revision). Le
    `montant_ht` facture est l'acompte de la periode, le plus souvent 1/4 de l'annuel
    pour un acompte trimestriel. On retrouve la quote-part reellement appliquee par
    DALKIA via le ratio `montant_ht / prix_revise`, puis on l'applique au prix de base
    pour isoler la part hors revision. Ce ratio gere aussi bien le decoupage /4 que les
    prorata partiels ou les lignes facturees a la consommation (prix unitaire x quantite).

    Retourne `(None, None)` si la decomposition n'est pas calculable (base ou revise
    absent, ou revise nul).
    """
    base = line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base")
    revised = line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise")
    amount = line.amount_ht
    if base is None or amount is None or not revised:
        return None, None
    base_share = round(base * (amount / revised), 2)
    revision_share = round(amount - base_share, 2)
    return base_share, revision_share


def _invoice_type_label(value: str | None) -> str | None:
    code = (value or "").strip().upper()
    if not code:
        return None
    return INVOICE_TYPE_LABELS.get(code, code)


def _norm_text(value: Any) -> str:
    text = _clean(value) or ""
    normalized = unicodedata.normalize("NFKD", text)
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", ascii_value.upper()).strip()


def _is_p2_4_line(line: CpeFinanceLine) -> bool:
    """Repere les lignes P2.4 malgre les variantes de libelles DALKIA."""
    if (line.market or "").upper() != "P2":
        return False
    haystack = " ".join(
        _norm_text(value)
        for value in (line.service_sold, line.billed_item, line.detail)
        if value not in (None, "")
    )
    return any(token in haystack for token in ("P2.4", "P2-4", "P2 4")) or (
        "MAITRISE" in haystack and "ENERG" in haystack
    )


def _line_duration_days(line: CpeFinanceLine) -> int | None:
    if not line.period_start or not line.period_end:
        return None
    return (line.period_end - line.period_start).days + 1


def _line_control_year(line: CpeFinanceLine) -> int | None:
    period_date = line.period_end or line.period_start
    return period_date.year if period_date else None


def _expected_p2_4_taux(
    db: Session,
    city_id: int | None,
    year: int,
) -> tuple[float | None, str]:
    query = (
        select(CpeResultatAnnuel)
        .join(CpeSite, CpeSite.id == CpeResultatAnnuel.cpe_site_id)
        .where(CpeResultatAnnuel.annee == year)
    )
    if city_id is None:
        query = query.where(CpeSite.city_id.is_(None))
    else:
        query = query.where(CpeSite.city_id == city_id)
    results = list(db.scalars(query).all())
    if not results:
        return None, f"Aucun resultat annuel CPE calcule pour {year} : impossible de valider le taux P2.4."

    incomplete = [item for item in results if item.statut == "partiel" or item.type_resultat in (None, "insuffisant")]
    penalties = [item for item in results if item.type_resultat == "penalite" or item.p2_4_taux == 0.5]
    if penalties:
        return 0.5, f"Objectifs non atteints sur {len(penalties)} site(s) en {year} : P2.4 attendu a 50%."
    if incomplete:
        return None, f"Resultats annuels CPE incomplets pour {year} ({len(incomplete)} site(s)) : validation P2.4 a confirmer."
    return 1.0, f"Objectifs atteints sur les resultats CPE calcules pour {year} : P2.4 attendu a 100%."


def _control_revision_p3(
    line: CpeFinanceLine,
    indices: dict[tuple[str, int, int], float],
    *,
    icht_ime_base: float,
    bt40_base: float,
) -> CpeFinanceControl:
    year, quarter = _line_index_period(line)
    base_price = line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base")
    actual = line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise")
    message = ""
    status = "blocked"
    severity = "warning"
    icht = bt40 = factor = expected = delta = delta_pct = None

    if year is None or quarter is None:
        message = "Période de ligne absente : impossible de sélectionner le trimestre d'indices."
    elif base_price is None or actual is None:
        message = "Prix de base ou prix révisé absent dans l'export DALKIA."
    else:
        icht = _index_value(indices, "ICHT_IME", year, quarter)
        bt40 = _index_value(indices, "BT40", year, quarter)
        if icht is None or bt40 is None:
            missing = []
            if icht is None:
                missing.append("ICHT-IME")
            if bt40 is None:
                missing.append("BT40")
            message = f"Indice(s) manquant(s) pour {year} T{quarter} : {', '.join(missing)}."
        else:
            factor = _p3_factor(icht, bt40, icht_ime_base=icht_ime_base, bt40_base=bt40_base)
            expected = round(base_price * factor, 4)
            delta = round(actual - expected, 4)
            delta_pct = round(delta / expected, 6) if expected else None
            if abs(delta) <= 0.05 or (delta_pct is not None and abs(delta_pct) <= 0.0015):
                status = "ok"
                severity = "info"
                message = f"Prix révisé cohérent avec la formule P3/P3.4 pour {year} T{quarter}."
            else:
                status = "error"
                severity = "error"
                message = (
                    f"Écart de révision P3/P3.4 sur {year} T{quarter} : attendu {expected:.4f}, "
                    f"facturé {actual:.4f}, écart {delta:.4f}."
                )

    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type="revision_p3",
        status=status,
        severity=severity,
        message=message,
        formula=P3_REVISION_FORMULA,
        index_year=year,
        index_quarter=quarter,
        icht_ime_value=icht,
        bt40_value=bt40,
        fsd2_value=None,
        expected_factor=factor,
        base_price=base_price,
        expected_revised_price=expected,
        actual_revised_price=actual,
        delta_abs=delta,
        delta_pct=delta_pct,
    )


def _control_revision_p2(
    line: CpeFinanceLine,
    indices: dict[tuple[str, int, int], float],
    *,
    icht_ime_base: float,
    fsd2_base: float,
) -> CpeFinanceControl:
    year, quarter = _line_index_period(line)
    base_price = line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base")
    actual = line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise")
    message = ""
    status = "blocked"
    severity = "warning"
    icht = fsd2 = factor = expected = delta = delta_pct = None

    if year is None or quarter is None:
        message = "Période de ligne absente : impossible de sélectionner le trimestre d'indices."
    elif base_price is None or actual is None:
        message = "Prix de base ou prix révisé absent dans l'export DALKIA."
    else:
        icht = _index_value(indices, "ICHT_IME", year, quarter)
        fsd2 = _index_value(indices, "FSD2", year, quarter)
        if icht is None or fsd2 is None:
            missing = []
            if icht is None:
                missing.append("ICHT-IME")
            if fsd2 is None:
                missing.append("FSD2")
            message = f"Indice(s) manquant(s) pour {year} T{quarter} : {', '.join(missing)}."
        else:
            factor = _p2_factor(icht, fsd2, icht_ime_base=icht_ime_base, fsd2_base=fsd2_base)
            expected = round(base_price * factor, 4)
            delta = round(actual - expected, 4)
            delta_pct = round(delta / expected, 6) if expected else None
            if abs(delta) <= 0.05 or (delta_pct is not None and abs(delta_pct) <= 0.0015):
                status = "ok"
                severity = "info"
                message = f"Prix révisé cohérent avec la formule P2 pour {year} T{quarter}."
            else:
                status = "error"
                severity = "error"
                message = (
                    f"Écart de révision P2 sur {year} T{quarter} : attendu {expected:.4f}, "
                    f"facturé {actual:.4f}, écart {delta:.4f}."
                )

    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type="revision_p2",
        status=status,
        severity=severity,
        message=message,
        formula=P2_REVISION_FORMULA,
        index_year=year,
        index_quarter=quarter,
        icht_ime_value=icht,
        bt40_value=None,
        fsd2_value=fsd2,
        expected_factor=factor,
        base_price=base_price,
        expected_revised_price=expected,
        actual_revised_price=actual,
        delta_abs=delta,
        delta_pct=delta_pct,
    )


def _control_p2_4_objectives(
    db: Session,
    line: CpeFinanceLine,
) -> CpeFinanceControl:
    year = _line_control_year(line)
    duration_days = _line_duration_days(line)
    expected_taux = None
    status = "blocked"
    severity = "warning"
    message_parts: list[str] = []

    if year is None:
        message_parts.append("Periode de ligne absente : impossible de rattacher P2.4 a un exercice.")
    else:
        expected_taux, objective_message = _expected_p2_4_taux(db, line.city_id, year)
        message_parts.append(objective_message)

    if duration_days is None:
        message_parts.append("Periode de facturation absente : P2.4 doit etre controle comme poste annuel.")
    elif duration_days < 300:
        status = "error"
        severity = "error"
        message_parts.append(
            f"Periode P2.4 de {duration_days} jours : le marche prevoit une facturation annuelle apres validation."
        )
    else:
        message_parts.append(f"Periode P2.4 annuelle detectee ({duration_days} jours).")

    if status != "error":
        if expected_taux is None:
            status = "blocked"
            severity = "warning"
        else:
            status = "ok"
            severity = "info"

    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type="p2_4_objectives",
        status=status,
        severity=severity,
        message=" ".join(message_parts),
        formula=P2_4_OBJECTIVE_RULE,
        index_year=year,
        index_quarter=None,
        icht_ime_value=None,
        bt40_value=None,
        fsd2_value=None,
        expected_factor=expected_taux,
        base_price=line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base"),
        expected_revised_price=None,
        actual_revised_price=line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise"),
        delta_abs=None,
        delta_pct=None,
    )


def _make_basic_control(
    line: CpeFinanceLine,
    *,
    control_type: str,
    status: str,
    severity: str,
    message: str,
    formula: str | None = None,
    delta_abs: float | None = None,
    delta_pct: float | None = None,
) -> CpeFinanceControl:
    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type=control_type,
        status=status,
        severity=severity,
        message=message,
        formula=formula,
        delta_abs=delta_abs,
        delta_pct=delta_pct,
    )


def _control_accounting_nature(line: CpeFinanceLine) -> CpeFinanceControl:
    if line.accounting_nature:
        return _make_basic_control(
            line,
            control_type="accounting_nature",
            status="ok",
            severity="info",
            message=f"Nature comptable rattachee : {line.accounting_nature}.",
        )
    # Non contrôlable (famille C) : matrice de codification incomplète, ce n'est pas
    # une anomalie de facturation. Classé "blocked" (à compléter), pas "error".
    return _make_basic_control(
        line,
        control_type="accounting_nature",
        status="blocked",
        severity="warning",
        message=(
            "Nature comptable absente : la ligne ne peut pas etre envoyee au service finances "
            "sans regle de codification."
        ),
    )


def _control_accounting_site(db: Session, line: CpeFinanceLine, invoice: CpeFinanceInvoice) -> CpeFinanceControl:
    if line.accounting_site_id:
        return _make_basic_control(
            line,
            control_type="accounting_site",
            status="ok",
            severity="info",
            message=f"Site finance rattache ({line.site_code_detected or 'mapping manuel'}).",
        )
    if line.site_code_detected:
        return _make_basic_control(
            line,
            control_type="accounting_site",
            status="blocked",
            severity="warning",
            message=f"Code site detecte ({line.site_code_detected}) mais non rattache a la matrice de codification.",
        )
    if not _is_current_cpe_contract(
        db,
        invoice.contract_code or line.contract_code,
        city_id=invoice.city_id,
        year=(invoice.period_end.year if invoice.period_end else None),
    ):
        return _make_basic_control(
            line,
            control_type="accounting_site",
            status="ok",
            severity="info",
            message=(
                "Site finance non exige : contrat DALKIA hors perimetre CPE Ville cible "
                f"({invoice.contract_code or line.contract_code or 'sans code contrat'})."
            ),
        )
    return _make_basic_control(
        line,
        control_type="accounting_site",
        status="blocked",
        severity="warning",
        message="Aucun site finance rattache a la ligne : verifier si la ligne doit etre imputee a un site.",
    )


def _control_invoice_type(invoice: CpeFinanceInvoice, anchor: CpeFinanceLine) -> CpeFinanceControl:
    code = (invoice.invoice_type or "").strip().upper()
    if code in INVOICE_TYPE_LABELS:
        return _make_basic_control(
            anchor,
            control_type="invoice_type",
            status="ok",
            severity="info",
            message=f"Type de facture reconnu : {INVOICE_TYPE_LABELS[code]} ({code}).",
        )
    if not code:
        # Non contrôlable (famille C) : donnée manquante, impossible de qualifier → "blocked".
        return _make_basic_control(
            anchor,
            control_type="invoice_type",
            status="blocked",
            severity="warning",
            message="Type de facture absent : impossible de qualifier acompte, avoir, regularisation ou definitive.",
        )
    return _make_basic_control(
        anchor,
        control_type="invoice_type",
        status="error",
        severity="error",
        message=f"Type de facture inconnu ({code}) : codification comptable a verifier avant validation.",
    )


def _control_invoice_total(invoice: CpeFinanceInvoice, lines: list[CpeFinanceLine], anchor: CpeFinanceLine) -> CpeFinanceControl:
    line_total = round(sum(line.amount_ht or 0.0 for line in lines), 2)
    invoice_total = round(invoice.total_ht or 0.0, 2)
    delta = round(invoice_total - line_total, 2)
    if abs(delta) <= 0.01:
        return _make_basic_control(
            anchor,
            control_type="invoice_total_ht",
            status="ok",
            severity="info",
            message=f"Total HT coherent : {invoice_total:.2f} EUR pour {len(lines)} ligne(s).",
            formula="Total facture HT = somme des lignes HT importees",
            delta_abs=delta,
        )
    return _make_basic_control(
        anchor,
        control_type="invoice_total_ht",
        status="error",
        severity="error",
        message=f"Total HT incoherent : facture {invoice_total:.2f} EUR, lignes {line_total:.2f} EUR, ecart {delta:.2f} EUR.",
        formula="Total facture HT = somme des lignes HT importees",
        delta_abs=delta,
        delta_pct=round(delta / line_total, 6) if line_total else None,
    )


def _control_invoice_period(invoice: CpeFinanceInvoice, lines: list[CpeFinanceLine], anchor: CpeFinanceLine) -> CpeFinanceControl:
    inverted = [line for line in lines if line.period_start and line.period_end and line.period_start > line.period_end]
    missing = [line for line in lines if not line.period_start or not line.period_end]
    if inverted:
        return _make_basic_control(
            anchor,
            control_type="invoice_period",
            status="error",
            severity="error",
            message=f"{len(inverted)} ligne(s) avec periode inversee : debut posterieur a la fin.",
        )
    if missing:
        return _make_basic_control(
            anchor,
            control_type="invoice_period",
            status="blocked",
            severity="warning",
            message=f"{len(missing)} ligne(s) sans periode complete : verification de la frequence facture impossible.",
        )
    starts = [line.period_start for line in lines if line.period_start]
    ends = [line.period_end for line in lines if line.period_end]
    if starts and ends:
        expected_start = min(starts)
        expected_end = max(ends)
        if invoice.period_start != expected_start or invoice.period_end != expected_end:
            return _make_basic_control(
                anchor,
                control_type="invoice_period",
                status="blocked",
                severity="warning",
                message=(
                    f"Periode facture {invoice.period_start or '-'} au {invoice.period_end or '-'} differente "
                    f"de l'enveloppe des lignes {expected_start} au {expected_end}."
                ),
            )
    return _make_basic_control(
        anchor,
        control_type="invoice_period",
        status="ok",
        severity="info",
        message=f"Periode facture coherente : {invoice.period_start or '-'} au {invoice.period_end or '-'}.",
    )


def _invoice_timeline_metrics(invoice: CpeFinanceInvoice, today: date | None = None) -> dict[str, Any]:
    today = today or date.today()
    billing_days = (
        (invoice.period_end - invoice.period_start).days + 1
        if invoice.period_start and invoice.period_end
        else None
    )
    issue_delay_days = (
        (invoice.invoice_date - invoice.period_end).days
        if invoice.invoice_date and invoice.period_end
        else None
    )
    due_in_days = (invoice.due_date - today).days if invoice.due_date else None
    if due_in_days is None:
        deadline_status = "echeance_absente"
    elif due_in_days < 0:
        deadline_status = "echeance_depassee"
    elif due_in_days <= 7:
        deadline_status = "urgent"
    elif due_in_days <= 30:
        deadline_status = "a_anticiper"
    else:
        deadline_status = "dans_les_temps"
    return {
        "billing_days": billing_days,
        "issue_delay_days": issue_delay_days,
        "due_in_days": due_in_days,
        "deadline_status": deadline_status,
    }


def _control_invoice_timeline(invoice: CpeFinanceInvoice, anchor: CpeFinanceLine) -> CpeFinanceControl:
    if not invoice.invoice_date or not invoice.due_date or not invoice.period_start or not invoice.period_end:
        missing = [
            label
            for label, value in (
                ("date edition", invoice.invoice_date),
                ("date echeance", invoice.due_date),
                ("debut periode", invoice.period_start),
                ("fin periode", invoice.period_end),
            )
            if value is None
        ]
        return _make_basic_control(
            anchor,
            control_type="invoice_timeline",
            status="blocked",
            severity="warning",
            message=f"Calendrier facture incomplet : {', '.join(missing)} manquante(s).",
        )
    if invoice.period_start > invoice.period_end:
        return _make_basic_control(
            anchor,
            control_type="invoice_timeline",
            status="error",
            severity="error",
            message="Calendrier facture incoherent : debut de periode posterieur a la fin.",
        )
    # Modele de facturation DALKIA : l'echeance est ancree sur une borne de la periode facturee
    # (debut OU fin de periode), et l'edition suit (acomptes edites juste apres la cloture) ou
    # arrive bien plus tard (regularisations). Une echeance anterieure a la date d'edition est
    # donc STRUCTURELLE chez DALKIA, pas une incoherence : la signaler en erreur produisait 275
    # faux positifs sur 473 factures (verifie sur l'export reel). On ne retient comme vraie
    # incoherence que l'echeance situee AVANT le debut de la periode facturee (impossible :
    # echeance d'un service pas encore commence).
    if invoice.due_date < invoice.period_start:
        return _make_basic_control(
            anchor,
            control_type="invoice_timeline",
            status="error",
            severity="error",
            message="Calendrier facture incoherent : echeance anterieure au debut de la periode facturee.",
        )
    metrics = _invoice_timeline_metrics(invoice)
    message = (
        f"Calendrier facture coherent : periode {invoice.period_start} au {invoice.period_end}, "
        f"edition {invoice.invoice_date}, echeance {invoice.due_date}."
    )
    if invoice.due_date < invoice.invoice_date:
        message += " Echeance sur une borne de periode, edition posterieure (acompte/regularisation) : coherent."
    if metrics["issue_delay_days"] is not None:
        message += f" Edition {metrics['issue_delay_days']} jour(s) apres la fin de periode."
    return _make_basic_control(
        anchor,
        control_type="invoice_timeline",
        status="ok",
        severity="info",
        message=message,
    )


def _find_contract_reference(
    db: Session,
    *,
    city_id: int | None,
    contract_code: str | None,
    reference_kind: str,
    year: int,
) -> CpeContractReference | None:
    if not contract_code:
        return None
    normalized_contract = contract_code.strip().upper()
    for ref_city_id in (city_id, None):
        query = select(CpeContractReference).where(
            CpeContractReference.contract_code == normalized_contract,
            CpeContractReference.reference_kind == reference_kind,
            CpeContractReference.year == year,
            CpeContractReference.active.is_(True),
        )
        if ref_city_id is None:
            query = query.where(CpeContractReference.city_id.is_(None))
        else:
            query = query.where(CpeContractReference.city_id == ref_city_id)
        reference = db.scalars(query.order_by(CpeContractReference.id)).first()
        if reference is not None:
            return reference
    return None


def _control_p1_gaz_acompte_against_dpgf(
    db: Session,
    invoice: CpeFinanceInvoice,
    lines: list[CpeFinanceLine],
) -> CpeFinanceControl | None:
    invoice_contract = (invoice.contract_code or "").strip().upper()
    p1_market_lines = [
        line
        for line in lines
        if (line.contract_code or invoice_contract or "").strip().upper() == invoice_contract
        and (line.market or "").strip().upper() == "P1"
        and (line.billed_item or "").strip().upper() in P1_GAZ_ACOMPTE_ITEMS
    ]
    if not p1_market_lines:
        return None

    anchor = p1_market_lines[0]
    if not invoice.period_start or not invoice.period_end:
        return _make_basic_control(
            anchor,
            control_type="p1_gaz_acompte_dpgf",
            status="blocked",
            severity="warning",
            message="Controle acompte P1 gaz impossible : periode facture absente ou incomplete.",
            formula="Acompte P1 gaz = 1/4 du P1 annuel DPGF revise",
        )

    reference = _find_contract_reference(
        db,
        city_id=invoice.city_id,
        contract_code=invoice.contract_code,
        reference_kind=P1_GAZ_ACOMPTE_KIND,
        year=invoice.period_end.year,
    )
    if reference is None:
        if _is_current_cpe_contract(
            db,
            invoice.contract_code,
            city_id=invoice.city_id,
            year=invoice.period_end.year,
        ):
            return _make_basic_control(
                anchor,
                control_type="p1_gaz_acompte_dpgf",
                status="blocked",
                severity="warning",
                message=(
                    "Controle acompte P1 gaz impossible : reference contractuelle absente "
                    f"pour {invoice.contract_code} / {invoice.period_end.year}."
                ),
                formula="Reference attendue : contrat, exercice, postes inclus, montant annuel et tolerance",
            )
        return None

    expected_months = _split_csv_tokens(reference.expected_period_months)
    expected_month_numbers = {int(month) for month in expected_months if month.isdigit()} or {3, 6, 9}
    period_start = invoice.period_start
    period_end = invoice.period_end
    # L'acompte P1 est contractuellement TRIMESTRIEL : il ne se compare a 1/4 de l'annuel que
    # si la periode de la facture couvre reellement un trimestre entier se terminant a une
    # echeance d'acompte (31/03, 30/06, 30/09). Une facture MENSUELLE finissant un dernier jour
    # de mois de fin de trimestre (ex. 01/03 -> 31/03) passait a tort le seul test de fin de
    # trimestre : le scope_query n'agregeait alors qu'un seul mois et le comparait au trimestre
    # complet -> ecart trompeur. On exige donc en plus que period_start soit le 1er jour du
    # trimestre (mois de fin - 2), ce qui isole les vraies factures trimestrielles.
    is_quarter_end = (
        period_end.month in expected_month_numbers
        and period_end.day == monthrange(period_end.year, period_end.month)[1]
    )
    quarter_start_month = period_end.month - 2
    covers_full_quarter = (
        quarter_start_month >= 1
        and period_start == date(period_end.year, quarter_start_month, 1)
    )
    is_expected_acompte_date = is_quarter_end and covers_full_quarter
    if not is_expected_acompte_date:
        period_days = (period_end - period_start).days + 1
        return _make_basic_control(
            anchor,
            control_type="p1_gaz_acompte_dpgf",
            status="ok",
            severity="info",
            message=(
                f"Controle acompte P1 gaz non applique : la periode {period_start} - {period_end} "
                f"({period_days} j) ne couvre pas un trimestre complet se terminant a une echeance "
                "d'acompte (31/03, 30/06, 30/09)."
            ),
            formula=reference.formula or "Acomptes P1 attendus aux 31/03, 30/06 et 30/09 ; decompte definitif au 15/02/N+1",
        )

    if reference.expected_amount_ht is not None:
        expected = round(reference.expected_amount_ht, 2)
    elif reference.annual_amount_ht is not None and reference.installment_count:
        expected = round(reference.annual_amount_ht / reference.installment_count, 2)
    else:
        return _make_basic_control(
            anchor,
            control_type="p1_gaz_acompte_dpgf",
            status="blocked",
            severity="warning",
            message=(
                "Controle acompte P1 gaz impossible : la reference contractuelle ne contient "
                "ni montant attendu, ni montant annuel avec nombre d'acomptes."
            ),
            formula=reference.formula,
        )

    included_items = _reference_included_items(reference)
    scope_query = select(CpeFinanceLine).where(
        CpeFinanceLine.batch_id == invoice.batch_id,
        CpeFinanceLine.city_id == invoice.city_id,
        CpeFinanceLine.contract_code == reference.contract_code,
        CpeFinanceLine.market == reference.market,
        CpeFinanceLine.period_start == invoice.period_start,
        CpeFinanceLine.period_end == invoice.period_end,
    )
    if included_items:
        scope_query = scope_query.where(func.upper(CpeFinanceLine.billed_item).in_(included_items))
    scope_lines = list(db.scalars(scope_query).all())
    if not scope_lines:
        return _make_basic_control(
            anchor,
            control_type="p1_gaz_acompte_dpgf",
            status="blocked",
            severity="warning",
            message="Controle acompte P1 gaz impossible : aucune ligne importee ne correspond aux postes inclus dans la reference.",
            formula=reference.formula,
        )

    # L'acompte P1 est contractuellement un montant trimestriel au niveau du LOT, mais DALKIA
    # le ventile sur de nombreuses factures (une par site). `scope_lines` agrege donc toutes les
    # lignes P1 du lot pour la periode, toutes factures confondues. Pour ne pas dupliquer le meme
    # ecart de lot sur chaque facture (et gonfler le compteur d'ecarts), on n'emet ce controle
    # qu'une seule fois : sur la facture de plus petit id parmi celles qui portent ces lignes.
    scope_invoice_ids = {line.invoice_id for line in scope_lines}
    owner_invoice_id = min(scope_invoice_ids)
    if invoice.id != owner_invoice_id:
        return None
    invoice_count = len(scope_invoice_ids)
    scope_label = (
        f"sur le lot importe ({invoice_count} factures P1)"
        if invoice_count > 1
        else "sur le lot importe"
    )

    actual = round(sum(line.amount_ht or 0.0 for line in scope_lines), 2)
    annual_revised = (
        round(reference.annual_amount_ht, 2)
        if reference.annual_amount_ht is not None
        else round(expected * (reference.installment_count or 4), 2)
    )
    delta = round(actual - expected, 2)
    delta_pct = round(delta / expected, 6) if expected else None
    ratio = (actual / expected) if expected else None

    # Garde-fou ZERO tolerance : un acompte trimestriel ne peut pas depasser le P1 annuel revise
    # complet. C'est la seule incoherence certaine au stade de l'acompte (provisionnel).
    if actual > annual_revised + 1.0:
        return CpeFinanceControl(
            city_id=anchor.city_id,
            batch_id=anchor.batch_id,
            invoice_id=anchor.invoice_id,
            line_id=anchor.id,
            control_type="p1_gaz_acompte_dpgf",
            status="error",
            severity="error",
            message=(
                f"Acompte P1 gaz incoherent {scope_label} : l'acompte du trimestre "
                f"{invoice.period_start} - {invoice.period_end} ({actual:.2f} EUR HT) depasse le P1 "
                f"annuel revise complet ({annual_revised:.2f} EUR HT). Impossible pour un acompte trimestriel."
            ),
            formula=reference.formula,
            expected_revised_price=annual_revised,
            actual_revised_price=actual,
            delta_abs=round(actual - annual_revised, 2),
        )

    # Sinon : controle INFORMATIF (jamais d'erreur sur l'ecart au quart). L'acompte trimestriel est
    # PROVISIONNEL et ventile par site selon la logique DALKIA (conso historique/saisonniere) : il
    # n'est PAS egal au P1 annuel / 4 (verifie sur donnees reelles : ratio facture/(annuel/4) de 0,42
    # a 1,38 selon site). Mettre une tolerance dessus reviendrait a masquer une base de comparaison
    # fausse. La verification EXACTE du P1 se fait (1) sur le prix unitaire via la formule de revision
    # (controle p1_gaz_pu_os3 : Pu = Pu0 x (a + b.PEG/PEG0 + c.TVD/TVD0 + d.CEE/CEE0 + e.TICGN/TICGN0))
    # et (2) au decompte definitif (conso reelle x prix indexes, 15/02/N+1).
    ratio_txt = f"{ratio:.0%}" if ratio is not None else "n/a"
    return CpeFinanceControl(
        city_id=anchor.city_id,
        batch_id=anchor.batch_id,
        invoice_id=anchor.invoice_id,
        line_id=anchor.id,
        control_type="p1_gaz_acompte_dpgf",
        status="ok",
        severity="info",
        message=(
            f"Acompte P1 gaz provisionnel {scope_label} pour {invoice.period_start} - "
            f"{invoice.period_end} : total facture {actual:.2f} EUR HT ({ratio_txt} du quart theorique "
            f"{expected:.2f}, P1 annuel revise {annual_revised:.2f}). Acompte ventile par site "
            f"(non egal a annuel/4) ; verification exacte par le prix unitaire (formule de revision) "
            f"et au decompte definitif."
        ),
        formula=reference.formula,
        expected_revised_price=expected,
        actual_revised_price=actual,
        delta_abs=delta,
        delta_pct=delta_pct,
    )


def _control_p2p3_base_against_dalkia(
    db: Session,
    line: CpeFinanceLine,
    invoice: CpeFinanceInvoice,
) -> CpeFinanceControl | None:
    """Verifie que la base (forfait) P2/P3 facturee correspond au forfait contractuel DALKIA.

    Les controles de revision verifient que `revised_price = base_price x facteur`, mais rien
    ne verifiait que `base_price` lui-meme est conforme au contrat. Ce controle comble ce trou :
    `base_price` (euros base, stable d'un trimestre a l'autre) doit egaler le forfait du
    referentiel DALKIA actif pour (site, annee, poste). Semantique validee sur donnees reelles.
    """
    if (line.market or "").upper() not in {"P2", "P3"}:
        return None
    year, _quarter = _line_index_period(line)
    if year is None:
        return None
    if not _is_current_cpe_contract(
        db,
        invoice.contract_code or line.contract_code,
        city_id=invoice.city_id,
        year=year,
    ):
        return None
    base = line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base")
    if base is None:
        return None
    code_site = (line.site_code_detected or "").strip()
    if not code_site:
        return None

    # Seuls les postes rattachables au referentiel sont controles ; les sous-postes
    # (P2-11, P2-2, P1, P1EAU...) n'ont pas de correspondance -> pas de controle (skip).
    poste = normalize_p2p3_poste(line.billed_item or line.market)
    if poste not in BPU_P2P3_POSTES:
        return None

    expected = resolve_dalkia_p2p3_forfait(
        db, code_site=code_site, year=year, billed_item=poste, city_id=invoice.city_id
    )
    formula = "Base P2/P3 facturee = forfait contractuel DALKIA (Annexe 3.1 P2 / Annexe 4 P3)"

    if expected is None:
        # poste controllable mais aucune ligne referentiel pour (site, annee) -> desalignement code
        return _make_basic_control(
            line,
            control_type="p2p3_base_dpgf",
            status="blocked",
            severity="warning",
            message=(
                f"Pas de forfait DALKIA pour {code_site} / {year} ({poste}) : verifier l'alignement "
                "du code site avec le referentiel importe (sinon controle base impossible)."
            ),
            formula=formula,
        )

    base_r = round(base, 2)
    exp_r = round(expected, 2)
    delta = round(base_r - exp_r, 2)
    delta_pct = round(delta / exp_r, 6) if exp_r else None
    tolerance = max(1.0, abs(exp_r) * 0.005)

    if abs(delta) <= tolerance:
        status, severity = "ok", "info"
        message = (
            f"Base {poste} conforme au contrat pour {code_site} {year} : "
            f"{base_r:.2f} EUR (forfait DALKIA {exp_r:.2f})."
        )
    else:
        status, severity = "error", "error"
        message = (
            f"Base {poste} non conforme pour {code_site} {year} : facturee {base_r:.2f} EUR "
            f"vs forfait contractuel DALKIA {exp_r:.2f} (ecart {delta:+.2f})."
        )

    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type="p2p3_base_dpgf",
        status=status,
        severity=severity,
        message=message,
        formula=formula,
        base_price=base_r,
        expected_revised_price=exp_r,
        actual_revised_price=base_r,
        delta_abs=delta,
        delta_pct=delta_pct,
    )


def _control_p1_gaz_pu_os3(
    db: Session,
    line: CpeFinanceLine,
    invoice: CpeFinanceInvoice,
) -> CpeFinanceControl | None:
    """Verifie que le prix unitaire gaz facture (base_price des lignes CHAUFFAGE) correspond
    au prix fixe OS N°3 du tarif du site sur 2026-2030.

    Le `base_price` des lignes P1 / service "CHAUFFAGE" porte le Pu gaz (EUR HT/MWhPCS) applique
    a la periode (verifie sur donnees reelles : CCAS 04 -> 70,78 en 2026 = OS N°3 T3). On ne
    contraint que les lignes dont le base_price est dans une plage de Pu plausible (les autres
    lignes CHAUFFAGE portent des montants/regularisations, controles par ailleurs).
    `cpe_prix_gaz` stocke le Pu en PCI -> conversion PCS = PCI / ratio PCS-PCI.
    """
    if (line.market or "").upper() != "P1":
        return None
    if (line.service_sold or "").strip().upper() != "CHAUFFAGE":
        return None
    year, _quarter = _line_index_period(line)
    if year is None or year < 2026 or year > 2030:
        return None
    if not _is_current_cpe_contract(
        db,
        invoice.contract_code or line.contract_code,
        city_id=invoice.city_id,
        year=year,
    ):
        return None
    base = line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base")
    if base is None or not (30.0 <= base <= 250.0):
        return None  # pas une ligne "prix unitaire" (montant/regularisation)
    year, _quarter = _line_index_period(line)
    if year is None or year < 2026 or year > 2030:
        return None  # hors fenetre de prix fixe OS N°3

    code_site = (line.site_code_detected or "").strip()
    formula = "Prix unitaire gaz facture = prix fixe OS N°3 du tarif (2026-2030), en EUR HT/MWhPCS"
    tarif = resolve_p1_gaz_tarif(db, code_site=code_site, city_id=invoice.city_id) if code_site else None
    if tarif is None:
        return _make_basic_control(
            line,
            control_type="p1_gaz_pu_os3",
            status="blocked",
            severity="warning",
            message=(
                f"Tarif gaz introuvable pour {code_site or 'site inconnu'} (Annexe 6 import actif) : "
                "controle du prix unitaire impossible."
            ),
            formula=formula,
        )
    prix = get_prix_gaz(db, year, tarif)
    if prix is None or not prix.pu_eur_mwh_pci:
        return _make_basic_control(
            line,
            control_type="p1_gaz_pu_os3",
            status="blocked",
            severity="warning",
            message=f"Prix OS N°3 absent pour tarif {tarif} / {year} : controle du prix unitaire impossible.",
            formula=formula,
        )

    expected_pcs = round(prix.pu_eur_mwh_pci / PCS_PCI_RATIO, 2)
    base_r = round(base, 2)
    delta = round(base_r - expected_pcs, 2)
    delta_pct = round(delta / expected_pcs, 6) if expected_pcs else None
    tolerance = max(0.3, abs(expected_pcs) * 0.005)

    if abs(delta) <= tolerance:
        status, severity = "ok", "info"
        message = (
            f"Prix unitaire gaz conforme OS N°3 pour {code_site} {year} (tarif {tarif}) : "
            f"{base_r:.2f} EUR/MWhPCS (attendu {expected_pcs:.2f})."
        )
    else:
        status, severity = "error", "error"
        message = (
            f"Prix unitaire gaz non conforme pour {code_site} {year} (tarif {tarif}) : facture "
            f"{base_r:.2f} EUR/MWhPCS vs OS N°3 {expected_pcs:.2f} (ecart {delta:+.2f})."
        )

    return CpeFinanceControl(
        city_id=line.city_id,
        batch_id=line.batch_id,
        invoice_id=line.invoice_id,
        line_id=line.id,
        control_type="p1_gaz_pu_os3",
        status=status,
        severity=severity,
        message=message,
        formula=formula,
        base_price=base_r,
        expected_revised_price=expected_pcs,
        actual_revised_price=base_r,
        delta_abs=delta,
        delta_pct=delta_pct,
    )


def recompute_finance_invoice_controls(
    db: Session,
    invoice: CpeFinanceInvoice,
) -> list[CpeFinanceControl]:
    reference_values = _reference_index_map(db, invoice.city_id)
    icht_ime_base = reference_values.get("ICHT_IME0", ICHT_IME_BASE)
    fsd2_base = reference_values.get("FSD20", FSD2_BASE)
    bt40_base = reference_values.get("BT400", BT40_BASE)
    lines = list_finance_lines(db, invoice.id, invoice.city_id)
    revision_lines = [line for line in lines if (line.market or "").upper() in {"P2", "P3"}]
    p2_4_lines = [line for line in lines if _is_p2_4_line(line)]
    db.execute(delete(CpeFinanceControl).where(CpeFinanceControl.invoice_id == invoice.id))
    if not lines:
        db.commit()
        return []

    years = {year for line in revision_lines for year, _quarter in [_line_index_period(line)] if year is not None}
    index_rows = []
    if years:
        index_rows = db.scalars(
            select(CpeRevisionIndex).where(
                CpeRevisionIndex.city_id == invoice.city_id,
                CpeRevisionIndex.year.in_(years),
            )
        ).all()
    indices = {(item.index_code, item.year, item.quarter): item.value for item in index_rows}
    controls = [
        _control_revision_p2(line, indices, icht_ime_base=icht_ime_base, fsd2_base=fsd2_base)
        if (line.market or "").upper() == "P2"
        else _control_revision_p3(line, indices, icht_ime_base=icht_ime_base, bt40_base=bt40_base)
        for line in revision_lines
    ]
    controls.extend(_control_p2_4_objectives(db, line) for line in p2_4_lines)
    controls.extend(_control_accounting_nature(line) for line in lines)
    controls.extend(_control_accounting_site(db, line, invoice) for line in lines)
    controls.append(_control_invoice_type(invoice, lines[0]))
    controls.append(_control_invoice_total(invoice, lines, lines[0]))
    controls.append(_control_invoice_period(invoice, lines, lines[0]))
    controls.append(_control_invoice_timeline(invoice, lines[0]))
    p1_gaz_control = _control_p1_gaz_acompte_against_dpgf(db, invoice, lines)
    if p1_gaz_control is not None:
        controls.append(p1_gaz_control)
    for line in revision_lines:
        p2p3_base_control = _control_p2p3_base_against_dalkia(db, line, invoice)
        if p2p3_base_control is not None:
            controls.append(p2p3_base_control)
    for line in lines:
        pu_os3_control = _control_p1_gaz_pu_os3(db, line, invoice)
        if pu_os3_control is not None:
            controls.append(pu_os3_control)
    db.add_all(controls)
    db.commit()
    for control in controls:
        db.refresh(control)
    return controls


def list_finance_controls(db: Session, invoice_id: int, city_id: int | None = None) -> list[CpeFinanceControl]:
    query = select(CpeFinanceControl).where(CpeFinanceControl.invoice_id == invoice_id)
    if city_id is not None:
        query = query.where(CpeFinanceControl.city_id == city_id)
    query = query.order_by(CpeFinanceControl.status.desc(), CpeFinanceControl.id)
    return list(db.scalars(query).all())


def _should_auto_validate_cpe(invoice_status: str, error_count: int, blocked_count: int) -> bool:
    """Auto-validation CPE : une facture entièrement propre passe en `valide`.

    Critère strict, symétrique de l'énergie (`control_status == valid`) : aucun
    contrôle en `error` ni en `blocked`. On ne valide que depuis `a_controler`,
    jamais une décision humaine déjà prise (`valide` / `refuse` / `conteste`).
    """
    return invoice_status == "a_controler" and error_count == 0 and blocked_count == 0


def build_finance_control_report(
    db: Session,
    city_id: int | None = None,
    *,
    recalculate: bool = False,
) -> dict[str, Any]:
    invoices = [
        invoice
        for invoice in list_finance_invoices(db, city_id=city_id)
        if _is_current_cpe_contract(
            db,
            invoice.contract_code,
            city_id=invoice.city_id,
            year=(invoice.period_end.year if invoice.period_end else None),
        )
    ]
    summaries: list[dict[str, Any]] = []
    type_counts: dict[str, dict[str, int]] = defaultdict(lambda: {"ok": 0, "error": 0, "blocked": 0})
    controls_ok = 0
    controls_error = 0
    controls_blocked = 0
    auto_validated = 0

    invoice_ids = [invoice.id for invoice in invoices]
    lines_by_invoice: dict[int, list[CpeFinanceLine]] = defaultdict(list)
    if invoice_ids:
        for line in db.scalars(
            select(CpeFinanceLine).where(CpeFinanceLine.invoice_id.in_(invoice_ids))
        ).all():
            lines_by_invoice[line.invoice_id].append(line)

    def _join_distinct(values: Iterable[str | None]) -> str | None:
        seen = [value for value in dict.fromkeys(v for v in values if v) ]
        return ", ".join(seen) if seen else None

    for invoice in invoices:
        invoice_lines = lines_by_invoice.get(invoice.id, [])
        recipient_ref = _join_distinct(_line_raw_str(line, "ref_destinataire_1") for line in invoice_lines)
        markets = _join_distinct(line.market for line in invoice_lines)
        billed_items = _join_distinct(line.billed_item for line in invoice_lines)
        controls = (
            recompute_finance_invoice_controls(db, invoice)
            if recalculate
            else list_finance_controls(db, invoice.id, city_id)
        )
        status_counts = {
            "ok": sum(1 for control in controls if control.status == "ok"),
            "error": sum(1 for control in controls if control.status == "error"),
            "blocked": sum(1 for control in controls if control.status == "blocked"),
        }
        controls_ok += status_counts["ok"]
        controls_error += status_counts["error"]
        controls_blocked += status_counts["blocked"]
        if recalculate and _should_auto_validate_cpe(invoice.status, status_counts["error"], status_counts["blocked"]):
            invoice.status = "valide"
            note = "Validée automatiquement : contrôle sans écart ni point bloquant."
            invoice.notes = f"{invoice.notes} | {note}" if invoice.notes else note
            auto_validated += 1
        for control in controls:
            if control.status in type_counts[control.control_type]:
                type_counts[control.control_type][control.status] += 1
        summaries.append(
            {
                "invoice_id": invoice.id,
                "invoice_number": invoice.invoice_number,
                "contract_code": invoice.contract_code,
                "contract_label": invoice.contract_label,
                "invoice_type": invoice.invoice_type,
                "recipient_ref": recipient_ref,
                "market": markets,
                "billed_items": billed_items,
                "total_ht": invoice.total_ht,
                "invoice_status": invoice.status,
                "finance_exported_at": invoice.finance_exported_at,
                "due_date": invoice.due_date,
                "due_in_days": _invoice_timeline_metrics(invoice)["due_in_days"],
                "deadline_status": _invoice_timeline_metrics(invoice)["deadline_status"],
                **status_counts,
                "controls_total": len(controls),
                "control_types": sorted({control.control_type for control in controls if control.status != "ok"}),
            }
        )

    if auto_validated:
        db.commit()

    summaries.sort(key=lambda item: (-item["error"], -item["blocked"], item["invoice_number"]))
    return {
        "generated_at": datetime.utcnow(),
        "scope": "Contrats CPE Ville actifs",
        "invoice_count": len(invoices),
        "total_ht": round(sum(invoice.total_ht or 0.0 for invoice in invoices), 2),
        "invoices_ok": sum(1 for item in summaries if item["error"] == 0 and item["blocked"] == 0),
        "invoices_with_errors": sum(1 for item in summaries if item["error"] > 0),
        "invoices_blocked": sum(1 for item in summaries if item["error"] == 0 and item["blocked"] > 0),
        "controls_ok": controls_ok,
        "controls_error": controls_error,
        "controls_blocked": controls_blocked,
        "control_types": [
            {"control_type": control_type, **counts, "total": sum(counts.values())}
            for control_type, counts in sorted(type_counts.items())
        ],
        "invoices": summaries,
    }


def build_finance_control_report_workbook(
    db: Session,
    city_id: int | None = None,
) -> bytes:
    report = build_finance_control_report(db, city_id, recalculate=False)
    invoice_ids = [item["invoice_id"] for item in report["invoices"]]
    controls_by_invoice: dict[int, list[CpeFinanceControl]] = defaultdict(list)
    if invoice_ids:
        for control in db.scalars(
            select(CpeFinanceControl)
            .where(CpeFinanceControl.invoice_id.in_(invoice_ids))
            .order_by(CpeFinanceControl.invoice_id, CpeFinanceControl.status.desc(), CpeFinanceControl.id)
        ).all():
            controls_by_invoice[control.invoice_id].append(control)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthese"
    ws["A1"] = "Rapport controle global factures CPE"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A3"] = "Genere le"
    ws["B3"] = report["generated_at"].strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Perimetre"
    ws["B4"] = report["scope"]
    ws["A5"] = "Factures analysees"
    ws["B5"] = report["invoice_count"]
    ws["A6"] = "Montant total HT controle"
    ws["B6"] = report["total_ht"]
    ws["B6"].number_format = '#,##0.00 "EUR"'
    ws["A7"] = "Factures conformes"
    ws["B7"] = report["invoices_ok"]
    ws["A8"] = "Factures avec ecarts"
    ws["B8"] = report["invoices_with_errors"]
    ws["A9"] = "Factures bloquees"
    ws["B9"] = report["invoices_blocked"]

    ws["A11"] = "Repartition des controles"
    ws["A11"].font = Font(bold=True)
    headers = ["Type de controle", "OK", "Ecarts", "Bloques", "Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=12, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row = 13
    for item in report["control_types"]:
        ws.cell(row=row, column=1, value=_control_type_label(item["control_type"]))
        ws.cell(row=row, column=2, value=item["ok"])
        ws.cell(row=row, column=3, value=item["error"])
        ws.cell(row=row, column=4, value=item["blocked"])
        ws.cell(row=row, column=5, value=item["total"])
        row += 1
    _set_widths(ws, [42, 12, 12, 12, 12])

    ws_queue = wb.create_sheet("File priorisee")
    queue_headers = [
        "Facture",
        "Contrat",
        "Type",
        "Destinataire",
        "Marche",
        "Postes factures",
        "HT",
        "Decision",
        "Echeance",
        "Jours avant echeance",
        "Statut echeance",
        "OK",
        "Ecarts",
        "Bloques",
        "Familles a traiter",
    ]
    for col, header in enumerate(queue_headers, start=1):
        cell = ws_queue.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row = 2
    for item in report["invoices"]:
        ws_queue.cell(row=row, column=1, value=item["invoice_number"])
        ws_queue.cell(row=row, column=2, value=item["contract_label"] or item["contract_code"])
        ws_queue.cell(row=row, column=3, value=item["invoice_type"])
        ws_queue.cell(row=row, column=4, value=item["recipient_ref"])
        ws_queue.cell(row=row, column=5, value=item["market"])
        ws_queue.cell(row=row, column=6, value=item["billed_items"])
        ws_queue.cell(row=row, column=7, value=item["total_ht"]).number_format = '#,##0.00 "EUR"'
        ws_queue.cell(row=row, column=8, value=item["invoice_status"])
        ws_queue.cell(row=row, column=9, value=str(item["due_date"]) if item["due_date"] else None)
        ws_queue.cell(row=row, column=10, value=item["due_in_days"])
        ws_queue.cell(row=row, column=11, value=item["deadline_status"])
        ws_queue.cell(row=row, column=12, value=item["ok"])
        ws_queue.cell(row=row, column=13, value=item["error"])
        ws_queue.cell(row=row, column=14, value=item["blocked"])
        ws_queue.cell(
            row=row,
            column=15,
            value=", ".join(_control_type_label(control_type) for control_type in item["control_types"]) or "Aucune anomalie",
        )
        row += 1
    _set_widths(ws_queue, [18, 34, 12, 22, 14, 28, 14, 14, 14, 18, 20, 8, 8, 8, 46])
    ws_queue.auto_filter.ref = f"A1:O{max(2, row - 1)}"

    ws_detail = wb.create_sheet("Detail controles")
    detail_headers = [
        "Facture",
        "Type controle",
        "Statut",
        "Message",
        "Cause probable",
        "Action recommandee",
        "Formule",
        "Trace calcul",
    ]
    for col, header in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row = 2
    for invoice in report["invoices"]:
        for control in controls_by_invoice.get(invoice["invoice_id"], []):
            ws_detail.cell(row=row, column=1, value=invoice["invoice_number"])
            ws_detail.cell(row=row, column=2, value=_control_type_label(control.control_type))
            ws_detail.cell(row=row, column=3, value=control.status)
            ws_detail.cell(row=row, column=4, value=control.message)
            ws_detail.cell(row=row, column=5, value=_control_probable_cause(control))
            ws_detail.cell(row=row, column=6, value=_control_recommended_action(control))
            ws_detail.cell(row=row, column=7, value=control.formula)
            ws_detail.cell(row=row, column=8, value=_control_calculation_trace(control))
            row += 1
    _set_widths(ws_detail, [18, 28, 12, 56, 52, 52, 44, 64])
    ws_detail.auto_filter.ref = f"A1:H{max(2, row - 1)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def list_finance_batches(db: Session, city_id: int | None = None) -> list[CpeFinanceImportBatch]:
    query = select(CpeFinanceImportBatch)
    if city_id is not None:
        query = query.where(CpeFinanceImportBatch.city_id == city_id)
    query = query.order_by(CpeFinanceImportBatch.created_at.desc())
    return list(db.scalars(query).all())


def list_finance_invoices(
    db: Session,
    city_id: int | None = None,
    batch_id: int | None = None,
) -> list[CpeFinanceInvoice]:
    query = select(CpeFinanceInvoice)
    if city_id is not None:
        query = query.where(CpeFinanceInvoice.city_id == city_id)
    if batch_id is not None:
        query = query.where(CpeFinanceInvoice.batch_id == batch_id)
    query = query.order_by(CpeFinanceInvoice.invoice_date.desc().nullslast(), CpeFinanceInvoice.invoice_number)
    return list(db.scalars(query).all())


def list_finance_invoices_enriched(
    db: Session,
    city_id: int | None = None,
    batch_id: int | None = None,
) -> list[dict[str, Any]]:
    invoices = list_finance_invoices(db, city_id=city_id, batch_id=batch_id)
    if not invoices:
        return []

    invoice_ids = [item.id for item in invoices]
    line_query = select(CpeFinanceLine).where(CpeFinanceLine.invoice_id.in_(invoice_ids))
    if city_id is not None:
        line_query = line_query.where(CpeFinanceLine.city_id == city_id)
    lines = list(db.scalars(line_query).all())

    markets_by_invoice: dict[int, set[str]] = defaultdict(set)
    billed_items_by_invoice: dict[int, set[str]] = defaultdict(set)
    dest_ref1_by_invoice: dict[int, set[str]] = defaultdict(set)
    prestation_sites_by_invoice: dict[int, set[str]] = defaultdict(set)
    prestation_detail_by_invoice: dict[int, set[str]] = defaultdict(set)
    evidence_by_invoice: dict[int, CpeInvoiceEvidence] = {}
    evidence_query = (
        select(CpeInvoiceEvidence)
        .where(CpeInvoiceEvidence.invoice_id.in_(invoice_ids))
        .order_by(CpeInvoiceEvidence.created_at.desc(), CpeInvoiceEvidence.id.desc())
    )
    for evidence in db.scalars(evidence_query).all():
        evidence_by_invoice.setdefault(evidence.invoice_id, evidence)

    for line in lines:
        if line.market:
            markets_by_invoice[line.invoice_id].add(line.market.strip().upper())
        if line.billed_item:
            billed_items_by_invoice[line.invoice_id].add(line.billed_item.strip().upper())
        if line.site_code_detected:
            prestation_sites_by_invoice[line.invoice_id].add(line.site_code_detected.strip())
        if line.detail:
            prestation_detail_by_invoice[line.invoice_id].add(line.detail.strip())
        for key in ("ref_destinataire_1", "ref_destinataire1", "reference_destinataire_1"):
            value = _line_raw_str(line, key)
            if value:
                dest_ref1_by_invoice[line.invoice_id].add(value)
                break

    rows: list[dict[str, Any]] = []
    for invoice in invoices:
        evidence = evidence_by_invoice.get(invoice.id)
        row = {
            "id": invoice.id,
            "batch_id": invoice.batch_id,
            "city_id": invoice.city_id,
            "invoice_number": invoice.invoice_number,
            "contract_code": invoice.contract_code,
            "contract_label": invoice.contract_label,
            "invoice_type": invoice.invoice_type,
            "supplier": invoice.supplier,
            "customer_code": invoice.customer_code,
            "customer_name": invoice.customer_name,
            "invoice_date": invoice.invoice_date,
            "due_date": invoice.due_date,
            "period_start": invoice.period_start,
            "period_end": invoice.period_end,
            "markets": ", ".join(sorted(markets_by_invoice.get(invoice.id, set()))) or None,
            "billed_items": ", ".join(sorted(billed_items_by_invoice.get(invoice.id, set()))) or None,
            "recipient_reference_1": ", ".join(sorted(dest_ref1_by_invoice.get(invoice.id, set()))) or None,
            "prestation_sites": ", ".join(sorted(prestation_sites_by_invoice.get(invoice.id, set()))) or None,
            "prestation_detail": " · ".join(sorted(prestation_detail_by_invoice.get(invoice.id, set()))[:3]) or None,
            "evidence_id": evidence.id if evidence else None,
            "evidence_status": evidence.validation_status if evidence else None,
            "evidence_revision_date": evidence.revision_date if evidence else None,
            "evidence_declared_factor": evidence.declared_factor if evidence else None,
            "evidence_declared_icht_ime": evidence.declared_icht_ime if evidence else None,
            "evidence_declared_fsd2": evidence.declared_fsd2 if evidence else None,
            "evidence_declared_bt40": evidence.declared_bt40 if evidence else None,
            "total_ht": invoice.total_ht,
            "status": invoice.status,
            "finance_exported_at": invoice.finance_exported_at,
            **_invoice_timeline_metrics(invoice),
            "notes": invoice.notes,
            "created_at": invoice.created_at,
            "updated_at": invoice.updated_at,
        }
        rows.append(row)
    return rows


def mark_finance_liaison_exported(db: Session, invoice: CpeFinanceInvoice) -> CpeFinanceInvoice:
    invoice.finance_exported_at = datetime.utcnow()
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice


def get_finance_invoice(db: Session, invoice_id: int, city_id: int | None = None) -> CpeFinanceInvoice | None:
    query = select(CpeFinanceInvoice).where(CpeFinanceInvoice.id == invoice_id)
    if city_id is not None:
        query = query.where(CpeFinanceInvoice.city_id == city_id)
    return db.scalars(query).first()


def get_invoice_evidence(db: Session, evidence_id: int, city_id: int | None = None) -> CpeInvoiceEvidence | None:
    query = select(CpeInvoiceEvidence).where(CpeInvoiceEvidence.id == evidence_id)
    if city_id is not None:
        query = query.where(CpeInvoiceEvidence.city_id == city_id)
    return db.scalars(query).first()


def list_finance_lines(db: Session, invoice_id: int, city_id: int | None = None) -> list[CpeFinanceLine]:
    query = select(CpeFinanceLine).where(CpeFinanceLine.invoice_id == invoice_id)
    if city_id is not None:
        query = query.where(CpeFinanceLine.city_id == city_id)
    query = query.order_by(CpeFinanceLine.row_number)
    return list(db.scalars(query).all())


def update_finance_invoice(
    db: Session,
    invoice: CpeFinanceInvoice,
    *,
    status: str | None = None,
    notes: str | None = None,
) -> CpeFinanceInvoice:
    if status is not None:
        invoice.status = status
    if notes is not None:
        invoice.notes = notes
    db.commit()
    db.refresh(invoice)
    return invoice


def _controls_status_label(controls: list[CpeFinanceControl]) -> str | None:
    if not controls:
        return None
    priority = {"error": 3, "blocked": 2, "ok": 1}
    return max((control.status for control in controls), key=lambda status: priority.get(status, 0))


def _control_type_label(control_type: str | None) -> str:
    mapping = {
        "revision_p2": "Revision P2",
        "revision_p3": "Revision P3/P3.4",
        "p2_4_objectives": "Objectif P2.4",
        "accounting_nature": "Codification nature comptable",
        "accounting_site": "Rattachement site finance",
        "invoice_type": "Qualification type de facture",
        "invoice_total_ht": "Coherence total HT",
        "invoice_period": "Coherence periode facture",
        "invoice_timeline": "Calendrier edition et echeance",
        "p1_gaz_acompte_dpgf": "Acompte P1 vs reference DPGF",
        "p2p3_base_dpgf": "Base P2/P3 vs forfait DALKIA",
        "p1_gaz_pu_os3": "Prix unitaire gaz vs OS N°3",
    }
    return mapping.get(control_type or "", control_type or "Controle")


def _fmt_number(value: float | None, digits: int = 4) -> str | None:
    if value is None:
        return None
    return f"{value:.{digits}f}"


def _fmt_percent(value: float | None, digits: int = 2) -> str | None:
    if value is None:
        return None
    return f"{value * 100:.{digits}f}%"


def _control_probable_cause(control: CpeFinanceControl) -> str:
    if control.status == "ok":
        return "Aucun ecart detecte sur ce controle."

    if control.control_type in {"revision_p2", "revision_p3"}:
        if control.index_year is None or control.index_quarter is None:
            return "Periode de ligne absente ou incomplete : impossible de choisir le trimestre d'indices."
        if control.base_price is None or control.actual_revised_price is None:
            return "Prix de base ou prix revise absent dans l'export facture."
        missing_indices: list[str] = []
        if control.icht_ime_value is None:
            missing_indices.append("ICHT-IME")
        if control.control_type == "revision_p2" and control.fsd2_value is None:
            missing_indices.append("FSD2")
        if control.control_type == "revision_p3" and control.bt40_value is None:
            missing_indices.append("BT40")
        if missing_indices:
            return f"Indice(s) manquant(s) pour la periode: {', '.join(missing_indices)}."
        if control.expected_revised_price is None:
            return "Prix revise attendu non calcule (donnees incompletes)."
        return "Ecart entre prix revise facture et prix revise calcule selon la formule contractuelle."

    if control.control_type == "p1_gaz_acompte_dpgf":
        if control.expected_revised_price is None:
            return "Reference DPGF incomplete ou non disponible pour calculer l'acompte attendu."
        if control.actual_revised_price is None:
            return "Montant facture introuvable sur les lignes incluses dans le controle."
        return "Ecart entre l'acompte facture et le montant attendu selon reference contractuelle."

    if control.control_type == "accounting_nature":
        return "La ligne n'a pas de nature comptable rattachee (matrice de codification a completer)."

    if control.control_type == "accounting_site":
        return "Le site de destination finance n'est pas rattache (mapping site a completer)."

    if control.control_type == "invoice_type":
        return "Type de facture absent ou inconnu (acompte/avoir/regularisation/definitive)."

    if control.control_type == "invoice_total_ht":
        return "Le total HT facture n'est pas egal a la somme des lignes importees."

    if control.control_type == "invoice_period":
        return "Periode facture/lignes incoherente ou incomplete."

    if control.control_type == "invoice_timeline":
        return "Dates d'edition, d'echeance ou de periode absentes ou incoherentes."

    return control.message or "Controle non conforme."


def _control_recommended_action(control: CpeFinanceControl) -> str:
    if control.status == "ok":
        return "Aucune action requise."

    if control.control_type in {"revision_p2", "revision_p3"}:
        if control.icht_ime_value is None or (
            control.control_type == "revision_p2" and control.fsd2_value is None
        ) or (control.control_type == "revision_p3" and control.bt40_value is None):
            return "Completer les indices du trimestre puis relancer le controle."
        if control.base_price is None or control.actual_revised_price is None:
            return "Verifier les colonnes prix de base/prix revise dans le fichier source."
        return "Verifier formule, base contractuelle et prix revise facture avec DALKIA."

    if control.control_type == "p1_gaz_acompte_dpgf":
        return "Verifier la reference contractuelle P1 (montant, postes inclus, tolerances)."

    if control.control_type == "accounting_nature":
        return "Mettre a jour la matrice contrat/poste -> nature comptable."

    if control.control_type == "accounting_site":
        return "Mettre a jour la matrice site pour rattacher la ligne au bon code finance."

    if control.control_type == "invoice_type":
        return "Verifier le code type facture transmis par DALKIA."

    if control.control_type == "invoice_total_ht":
        return "Comparer total facture et total des lignes importees."

    if control.control_type == "invoice_period":
        return "Verifier les dates debut/fin sur facture et lignes."

    if control.control_type == "invoice_timeline":
        return "Verifier le calendrier facture avant emission de la fiche de liaison finances."

    return "Verifier la ligne et les donnees source."


def _control_calculation_trace(control: CpeFinanceControl) -> str | None:
    parts: list[str] = []
    if control.formula:
        parts.append(f"Formule: {control.formula}")
    if control.index_year is not None and control.index_quarter is not None:
        parts.append(f"Periode indices: {control.index_year} T{control.index_quarter}")
    index_parts: list[str] = []
    if control.icht_ime_value is not None:
        index_parts.append(f"ICHT-IME={_fmt_number(control.icht_ime_value, 3)}")
    if control.bt40_value is not None:
        index_parts.append(f"BT40={_fmt_number(control.bt40_value, 3)}")
    if control.fsd2_value is not None:
        index_parts.append(f"FSD2={_fmt_number(control.fsd2_value, 3)}")
    if index_parts:
        parts.append("Indices: " + ", ".join(index_parts))
    if control.expected_factor is not None:
        parts.append(f"Facteur attendu={_fmt_number(control.expected_factor, 6)}")
    if control.base_price is not None:
        parts.append(f"Prix base={_fmt_number(control.base_price, 4)}")
    if control.expected_revised_price is not None:
        parts.append(f"Prix revise calcule={_fmt_number(control.expected_revised_price, 4)}")
    if control.actual_revised_price is not None:
        parts.append(f"Prix revise facture={_fmt_number(control.actual_revised_price, 4)}")
    if control.delta_abs is not None:
        parts.append(f"Ecart={_fmt_number(control.delta_abs, 4)}")
    if control.delta_pct is not None:
        parts.append(f"Ecart %={_fmt_percent(control.delta_pct, 2)}")
    return " | ".join(parts) or None


def _line_primary_control(controls: list[CpeFinanceControl]) -> CpeFinanceControl | None:
    if not controls:
        return None
    priority = {"error": 3, "blocked": 2, "ok": 1}
    return sorted(
        controls,
        key=lambda control: (
            priority.get(control.status, 0),
            1 if control.delta_abs not in (None, 0) else 0,
            control.id,
        ),
        reverse=True,
    )[0]


def _reference_base_value(reference_values: dict[str, float], code: str, fallback: float | None = None) -> float | None:
    value = reference_values.get(code)
    if value is not None:
        return value
    return fallback


def _p1_reference_snapshot(reference_values: dict[str, float]) -> str:
    tokens: list[str] = []
    for code in (
        "P1_CPB0_T1",
        "P1_CPB0_T2",
        "P1_CPB0_T3",
        "P1_TVD0_T1",
        "P1_TVD0_T2",
        "P1_TVD0_T3",
        "P1_CEE0",
        "P1_TICGN0",
        "P1_PEG0",
        "P1_PUGAZ0",
    ):
        if code in reference_values:
            tokens.append(f"{code}={_fmt_number(reference_values[code], 4)}")
    return " | ".join(tokens)


def _style_header(cells: Any, fill: str = "1F4E78") -> None:
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)


def _set_widths(ws: Any, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def build_finance_liaison_workbook(db: Session, invoice: CpeFinanceInvoice) -> bytes:
    """Construit une fiche de liaison finances XLSX pour une facture."""
    lines = list_finance_lines(db, invoice.id, invoice.city_id)
    controls_by_line: dict[int, list[CpeFinanceControl]] = defaultdict(list)
    for control in db.scalars(select(CpeFinanceControl).where(CpeFinanceControl.invoice_id == invoice.id)).all():
        controls_by_line[control.line_id].append(control)
    sites = {
        mapping.id: mapping
        for mapping in db.scalars(
            select(CpeAccountingSiteMapping).where(
                CpeAccountingSiteMapping.id.in_([line.accounting_site_id for line in lines if line.accounting_site_id])
            )
        ).all()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiche liaison"
    ws["A1"] = "Fiche de liaison finance DALKIA"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A3"] = "Facture"
    ws["B3"] = invoice.invoice_number
    ws["A4"] = "Contrat"
    ws["B4"] = invoice.contract_code
    ws["A5"] = "Période"
    ws["B5"] = f"{invoice.period_start or '-'} au {invoice.period_end or '-'}"
    ws["A6"] = "Type facture"
    ws["B6"] = f"{_invoice_type_label(invoice.invoice_type) or '-'} ({invoice.invoice_type or '-'})"
    ws["A7"] = "Statut"
    ws["B7"] = invoice.status
    ws["A8"] = "Total HT"
    ws["B8"] = invoice.total_ht
    ws["B8"].number_format = '#,##0.00 "€"'
    if invoice.notes:
        ws["A9"] = "Notes"
        ws["B9"] = invoice.notes

    headers = [
        "Ligne",
        "Type facture",
        "Code type",
        "Marché",
        "Service vendu",
        "Poste facturé",
        "Site détecté",
        "Nom site",
        "Service",
        "Fonction",
        "Antenne",
        "Opération",
        "Nature",
        "Libellé nature",
        "Prix base",
        "Prix révisé",
        "Contrôle",
        "Formule",
        "Message contrôle",
        "Montant HT",
        "Conso",
        "Unité",
        "Détail",
    ]
    start_row = 11
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for row_index, line in enumerate(lines, start=start_row + 1):
        site = sites.get(line.accounting_site_id or 0)
        line_controls = controls_by_line.get(line.id, [])
        values = [
            line.row_number,
            _invoice_type_label(invoice.invoice_type),
            invoice.invoice_type,
            line.market,
            line.service_sold,
            line.billed_item,
            line.site_code_detected,
            site.site_name if site else None,
            site.service_code if site else None,
            site.function_code if site else None,
            site.antenna_code if site else None,
            site.operation_code if site else None,
            line.accounting_nature,
            line.accounting_label,
            line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base"),
            line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise"),
            _controls_status_label(line_controls),
            " | ".join(control.formula for control in line_controls if control.formula) or None,
            " | ".join(control.message for control in line_controls) or None,
            line.amount_ht,
            line.consumption,
            line.unit,
            line.detail,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
        ws.cell(row=row_index, column=15).number_format = '#,##0.0000'
        ws.cell(row=row_index, column=16).number_format = '#,##0.0000'
        ws.cell(row=row_index, column=20).number_format = '#,##0.00 "€"'

    widths = [10, 22, 10, 12, 20, 18, 18, 32, 14, 14, 16, 16, 12, 24, 12, 12, 12, 34, 48, 14, 12, 10, 42]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    ws.freeze_panes = "A12"
    ws.auto_filter.ref = f"A{start_row}:W{max(start_row + 1, start_row + len(lines))}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_detailed_finance_liaison_workbook(db: Session, invoice: CpeFinanceInvoice) -> bytes:
    """Construit une fiche de liaison finances complete pour audit et transmission."""
    reference_values = _reference_index_map(db, invoice.city_id)
    icht_ime_base = _reference_base_value(reference_values, "ICHT_IME0", ICHT_IME_BASE)
    fsd2_base = _reference_base_value(reference_values, "FSD20", FSD2_BASE)
    bt40_base = _reference_base_value(reference_values, "BT400", BT40_BASE)
    p1_reference_snapshot = _p1_reference_snapshot(reference_values)
    lines = list_finance_lines(db, invoice.id, invoice.city_id)
    controls_by_line: dict[int, list[CpeFinanceControl]] = defaultdict(list)
    for control in db.scalars(select(CpeFinanceControl).where(CpeFinanceControl.invoice_id == invoice.id)).all():
        controls_by_line[control.line_id].append(control)
    controls = [control for group in controls_by_line.values() for control in group]
    sites = {
        mapping.id: mapping
        for mapping in db.scalars(
            select(CpeAccountingSiteMapping).where(
                CpeAccountingSiteMapping.id.in_([line.accounting_site_id for line in lines if line.accounting_site_id])
            )
        ).all()
    }

    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Synthese"
    summary["A1"] = "Fiche de liaison finance DALKIA"
    summary["A1"].font = Font(bold=True, size=15)
    summary_rows = [
        ("Facture", invoice.invoice_number),
        ("Contrat", invoice.contract_code),
        ("Libelle contrat", invoice.contract_label),
        ("Societe", invoice.supplier),
        ("Client", invoice.customer_name),
        ("Code client", invoice.customer_code),
        ("Date edition", invoice.invoice_date),
        ("Date echeance", invoice.due_date),
        ("Periode", f"{invoice.period_start or '-'} au {invoice.period_end or '-'}"),
        ("Type facture", f"{_invoice_type_label(invoice.invoice_type) or '-'} ({invoice.invoice_type or '-'})"),
        ("Statut decision", invoice.status),
        ("Total HT", invoice.total_ht),
        ("Lignes facture", len(lines)),
        ("Lignes avec nature comptable", sum(1 for line in lines if line.accounting_nature)),
        ("Lignes avec site rattache", sum(1 for line in lines if line.accounting_site_id)),
        ("Base ICHT-IME0", icht_ime_base),
        ("Base FSD20", fsd2_base),
        ("Base BT400", bt40_base),
        ("References P1 (OS3)", p1_reference_snapshot or None),
        ("Notes", invoice.notes),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row_index, column=2, value=value)
    summary["B14"].number_format = '#,##0.00 "EUR"'
    summary["B17"].number_format = "#,##0.0000"
    summary["B18"].number_format = "#,##0.0000"
    summary["B19"].number_format = "#,##0.0000"
    summary["A21"] = "Synthese des controles"
    summary["A21"].font = Font(bold=True, size=13)
    for offset, (label, value) in enumerate(
        [
            ("OK", sum(1 for control in controls if control.status == "ok")),
            ("Ecarts", sum(1 for control in controls if control.status == "error")),
            ("Bloques / a completer", sum(1 for control in controls if control.status == "blocked")),
        ],
        start=1,
    ):
        summary.cell(row=21 + offset, column=1, value=label)
        summary.cell(row=21 + offset, column=2, value=value)
    _set_widths(summary, [26, 48, 18, 18])

    detail = wb.create_sheet("Lignes finance")
    headers = [
        "Ligne",
        "Type facture",
        "Code type",
        "Contrat",
        "Marche",
        "Type marche",
        "Service vendu",
        "Poste facture",
        "Site detecte",
        "Nom site",
        "Famille",
        "Gestionnaire",
        "Service",
        "Libelle service",
        "Fonction",
        "Libelle fonction",
        "Antenne",
        "Libelle antenne",
        "Operation",
        "Libelle operation",
        "Nature",
        "Libelle nature",
        "TVA",
        "Prix base annuel",
        "Prix revise annuel",
        "Acompte hors revision",
        "Revision appliquee",
        "Controle",
        "Formule",
        "Message controle",
        "Diagnostic probable",
        "Action recommandee",
        "Lecture calcul",
        "Montant HT (avec revision)",
        "Conso",
        "Unite",
        "Detail",
        "Debut periode",
        "Fin periode",
    ]
    for col, header in enumerate(headers, start=1):
        detail.cell(row=1, column=col, value=header)
    _style_header(detail[1])

    for row_index, line in enumerate(lines, start=2):
        site = sites.get(line.accounting_site_id or 0)
        line_controls = controls_by_line.get(line.id, [])
        primary_control = _line_primary_control(line_controls)
        calculation_trace = " || ".join(
            trace
            for trace in [_control_calculation_trace(control) for control in line_controls[:3]]
            if trace
        ) or None
        base_share, revision_share = _line_revision_breakdown(line)
        values = [
            line.row_number,
            _invoice_type_label(invoice.invoice_type),
            invoice.invoice_type,
            line.contract_code or invoice.contract_code,
            line.market,
            line.market_type,
            line.service_sold,
            line.billed_item,
            line.site_code_detected,
            site.site_name if site else None,
            site.family if site else None,
            site.manager if site else None,
            site.service_code if site else None,
            site.service_label if site else None,
            site.function_code if site else None,
            site.function_label if site else None,
            site.antenna_code if site else None,
            site.antenna_label if site else None,
            site.operation_code if site else None,
            site.operation_label if site else None,
            line.accounting_nature,
            line.accounting_label,
            line.vat_rate,
            line.base_price if line.base_price is not None else _line_raw_float(line, "prix_de_base"),
            line.revised_price if line.revised_price is not None else _line_raw_float(line, "prix_ou_forfait_revise"),
            base_share,
            revision_share,
            _controls_status_label(line_controls),
            " | ".join(control.formula for control in line_controls if control.formula) or None,
            " | ".join(control.message for control in line_controls) or None,
            _control_probable_cause(primary_control) if primary_control else None,
            _control_recommended_action(primary_control) if primary_control else None,
            calculation_trace,
            line.amount_ht,
            line.consumption,
            line.unit,
            line.detail,
            line.period_start,
            line.period_end,
        ]
        for col, value in enumerate(values, start=1):
            detail.cell(row=row_index, column=col, value=value)
        detail.cell(row=row_index, column=23).number_format = '0.00"%"'
        detail.cell(row=row_index, column=24).number_format = '#,##0.0000'
        detail.cell(row=row_index, column=25).number_format = '#,##0.0000'
        detail.cell(row=row_index, column=26).number_format = '#,##0.00 "EUR"'
        detail.cell(row=row_index, column=27).number_format = '#,##0.00 "EUR"'
        detail.cell(row=row_index, column=34).number_format = '#,##0.00 "EUR"'
    _set_widths(
        detail,
        [
            9,
            22,
            10,
            14,
            12,
            16,
            20,
            18,
            16,
            32,
            14,
            20,
            14,
            28,
            14,
            28,
            14,
            28,
            14,
            28,
            12,
            28,
            10,
            14,
            16,
            18,
            18,
            12,
            34,
            55,
            36,
            36,
            58,
            18,
            12,
            10,
            48,
            14,
            14,
        ],
    )
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:AM{max(2, len(lines) + 1)}"

    if controls:
        controls_sheet = wb.create_sheet("Controles")
        control_headers = [
            "Ligne",
            "Type controle",
            "Libelle controle",
            "Statut",
            "Severite",
            "Message",
            "Cause probable",
            "Action recommandee",
            "Formule",
            "Lecture calcul",
            "ICHT-IME0 (base)",
            "FSD20 (base)",
            "BT400 (base)",
            "References P1 (OS3)",
            "Annee indice",
            "Trimestre indice",
            "ICHT-IME",
            "BT40",
            "FSD2",
            "Facteur attendu",
            "Prix base",
            "Prix revise attendu",
            "Prix revise facture",
            "Ecart",
            "Ecart %",
        ]
        for col, header in enumerate(control_headers, start=1):
            controls_sheet.cell(row=1, column=col, value=header)
        _style_header(controls_sheet[1], "7C2D12")
        line_by_id = {line.id: line for line in lines}
        for row_index, control in enumerate(controls, start=2):
            line = line_by_id.get(control.line_id)
            values = [
                line.row_number if line else None,
                control.control_type,
                _control_type_label(control.control_type),
                control.status,
                control.severity,
                control.message,
                _control_probable_cause(control),
                _control_recommended_action(control),
                control.formula,
                _control_calculation_trace(control),
                icht_ime_base if control.control_type in {"revision_p2", "revision_p3"} else None,
                fsd2_base if control.control_type == "revision_p2" else None,
                bt40_base if control.control_type == "revision_p3" else None,
                p1_reference_snapshot or None,
                control.index_year,
                control.index_quarter,
                control.icht_ime_value,
                control.bt40_value,
                control.fsd2_value,
                control.expected_factor,
                control.base_price,
                control.expected_revised_price,
                control.actual_revised_price,
                control.delta_abs,
                control.delta_pct,
            ]
            for col, value in enumerate(values, start=1):
                controls_sheet.cell(row=row_index, column=col, value=value)
            controls_sheet.cell(row=row_index, column=10).number_format = "#,##0.0000"
            controls_sheet.cell(row=row_index, column=11).number_format = "#,##0.0000"
            controls_sheet.cell(row=row_index, column=12).number_format = "#,##0.0000"
        _set_widths(
            controls_sheet,
            [
                9,
                18,
                24,
                12,
                12,
                55,
                44,
                40,
                64,
                16,
                16,
                16,
                48,
                12,
                12,
                12,
                12,
                12,
                14,
                12,
                16,
                16,
                12,
                12,
            ],
        )
        controls_sheet.freeze_panes = "A2"
        controls_sheet.auto_filter.ref = f"A1:X{max(2, len(controls) + 1)}"

    raw_sheet = wb.create_sheet("Donnees source")
    raw_keys = sorted(
        {
            key
            for line in lines
            for key in (json.loads(line.raw_json or "{}").keys() if line.raw_json else [])
        }
    )
    raw_headers = ["Ligne"] + raw_keys
    for col, header in enumerate(raw_headers, start=1):
        raw_sheet.cell(row=1, column=col, value=header)
    _style_header(raw_sheet[1], "374151")
    for row_index, line in enumerate(lines, start=2):
        raw_data = json.loads(line.raw_json or "{}") if line.raw_json else {}
        raw_sheet.cell(row=row_index, column=1, value=line.row_number)
        for col, key in enumerate(raw_keys, start=2):
            raw_sheet.cell(row=row_index, column=col, value=raw_data.get(key))
    _set_widths(raw_sheet, [9] + [22] * len(raw_keys))
    raw_sheet.freeze_panes = "A2"
    raw_sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(raw_headers))}{max(2, len(lines) + 1)}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def delete_finance_batch(db: Session, batch: CpeFinanceImportBatch) -> None:
    invoice_ids = list(db.scalars(select(CpeFinanceInvoice.id).where(CpeFinanceInvoice.batch_id == batch.id)).all())
    if invoice_ids:
        db.execute(delete(CpeInvoiceEvidenceLink).where(CpeInvoiceEvidenceLink.invoice_id.in_(invoice_ids)))
        db.execute(
            CpeInvoiceEvidence.__table__.update()
            .where(CpeInvoiceEvidence.invoice_id.in_(invoice_ids))
            .values(invoice_id=None)
        )
    db.execute(delete(CpeFinanceControl).where(CpeFinanceControl.batch_id == batch.id))
    db.execute(delete(CpeFinanceLine).where(CpeFinanceLine.batch_id == batch.id))
    db.execute(delete(CpeFinanceInvoice).where(CpeFinanceInvoice.batch_id == batch.id))
    db.delete(batch)
    db.commit()


def purge_duplicate_finance_invoices(db: Session, city_id: int | None = None) -> dict[str, int]:
    """Supprime les factures DALKIA en double (même numéro de facture).

    Conserve la facture la plus récente (id le plus élevé) par numéro et supprime
    les autres ainsi que leurs lignes et contrôles. Retour : {"removed": N, "kept": K}.
    """
    query = select(CpeFinanceInvoice.id, CpeFinanceInvoice.invoice_number)
    if city_id is not None:
        query = query.where(CpeFinanceInvoice.city_id == city_id)
    rows = list(db.execute(query.order_by(CpeFinanceInvoice.id.desc())).all())
    seen: set[str] = set()
    to_delete: list[int] = []
    for invoice_id, invoice_number in rows:
        if not invoice_number:
            continue
        if invoice_number in seen:
            to_delete.append(invoice_id)
        else:
            seen.add(invoice_number)
    if not to_delete:
        return {"removed": 0, "kept": len(seen)}
    db.execute(delete(CpeInvoiceEvidenceLink).where(CpeInvoiceEvidenceLink.invoice_id.in_(to_delete)))
    db.execute(
        CpeInvoiceEvidence.__table__.update()
        .where(CpeInvoiceEvidence.invoice_id.in_(to_delete))
        .values(invoice_id=None)
    )
    db.execute(delete(CpeFinanceControl).where(CpeFinanceControl.invoice_id.in_(to_delete)))
    db.execute(delete(CpeFinanceLine).where(CpeFinanceLine.invoice_id.in_(to_delete)))
    removed = db.execute(delete(CpeFinanceInvoice).where(CpeFinanceInvoice.id.in_(to_delete))).rowcount or 0
    db.commit()
    return {"removed": removed, "kept": len(seen)}


def delete_finance_history(db: Session, city_id: int | None = None) -> dict[str, int]:
    """Supprime l'historique des factures DALKIA sans toucher au referentiel."""
    batch_query = select(CpeFinanceImportBatch.id)
    if city_id is not None:
        batch_query = batch_query.where(CpeFinanceImportBatch.city_id == city_id)
    batch_ids = list(db.scalars(batch_query).all())
    if not batch_ids:
        return {"batches_deleted": 0, "invoices_deleted": 0, "lines_deleted": 0, "controls_deleted": 0}

    invoice_ids = list(db.scalars(select(CpeFinanceInvoice.id).where(CpeFinanceInvoice.batch_id.in_(batch_ids))).all())
    if invoice_ids:
        db.execute(delete(CpeInvoiceEvidenceLink).where(CpeInvoiceEvidenceLink.invoice_id.in_(invoice_ids)))
        db.execute(
            CpeInvoiceEvidence.__table__.update()
            .where(CpeInvoiceEvidence.invoice_id.in_(invoice_ids))
            .values(invoice_id=None)
        )
    controls_deleted = db.execute(delete(CpeFinanceControl).where(CpeFinanceControl.batch_id.in_(batch_ids))).rowcount or 0
    lines_deleted = db.execute(delete(CpeFinanceLine).where(CpeFinanceLine.batch_id.in_(batch_ids))).rowcount or 0
    invoices_deleted = db.execute(delete(CpeFinanceInvoice).where(CpeFinanceInvoice.batch_id.in_(batch_ids))).rowcount or 0
    batches_deleted = db.execute(delete(CpeFinanceImportBatch).where(CpeFinanceImportBatch.id.in_(batch_ids))).rowcount or 0
    db.commit()
    return {
        "batches_deleted": batches_deleted,
        "invoices_deleted": invoices_deleted,
        "lines_deleted": lines_deleted,
        "controls_deleted": controls_deleted,
    }
