"""
Réexport ASTECH — le **retour** de l'aller-retour (incrément 3).

Produit le classeur que la collectivité réinjecte dans ASTECH. Trois règles données par
l'utilisateur encadrent tout ce module :

1. **Le `CODE_BIEN` n'est jamais réécrit.** C'est la clé de mise à jour d'ASTECH.
2. **Les en-têtes sont recopiés à l'octet près** depuis `PatrimoineLegacyImport.headers_json`,
   jamais retapés ici. Un `COD_COMPTABLE` devenu `CODE_COMPTABLE` et l'import échoue. Le
   gabarit est donc *dérivé du fichier source*, pas une constante du code.
3. **On n'écrit jamais une valeur qu'on n'a pas su produire proprement.** Une adresse
   inanalysable, un numéro de plan hors format, un rattachement non validé : la ligne part
   en feuille « à vérifier » plutôt que dans le fichier de la collectivité.

Le classeur produit porte trois feuilles :

- la **feuille réinjectable**, au gabarit ASTECH (décision Q12, feuille réduite) ;
- la **traçabilité**, ancienne → nouvelle valeur, pour relecture avant réinjection ;
- les lignes **à vérifier**, avec le motif — c'est le filet de la règle 3.
"""
from __future__ import annotations

import io
import json
import re
import unicodedata
from datetime import date
from typing import Any

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.patrimoine_legacy import (
    STATUS_LINKED,
    STATUS_TO_CREATE,
    PatrimoineLegacyAsset,
    PatrimoineLegacyImport,
)
from app.services.patrimoine_legacy import NEW_ASSET_CODE_PREFIX

# Colonnes ASTECH que Po2 maîtrise et réécrit (décision Q12, §11.3). Toute autre colonne
# du fichier d'origine est hors de notre portée et n'est pas émise.
# Chaque champ exporte, avec les en-tetes qui peuvent le porter selon la generation du
# fichier. Le referent ASTECH a renomme ses colonnes (2026-08-20) : `CODE_BIEN` est
# devenu `Code`, `BISTER` est devenu `Complement`... L'en-tete REELLEMENT ecrit est
# toujours celui trouve dans le fichier importe, jamais un nom choisi ici.
EXPORTED_FIELDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("CODE_BIEN", ("CODE_BIEN", "CODEBIEN", "CODBAR", "CODE")),
    ("DESIGNATION", ("DESIGNATION",)),
    ("NOMCOURT", ("NOMCOURT", "NOM COURT")),
    ("NORUE", ("NORUE", "NUMERO(S)", "NUMEROS", "NUMERO")),
    ("BISTER", ("BISTER", "COMPLEMENT")),
    ("LIBELVOIE", ("LIBELVOIE", "ADRESSE")),
    ("CODPOST", ("CODPOST", "CODE POSTAL")),
    ("VILLE", ("VILLE",)),
    ("COMMUNE", ("COMMUNE",)),
    ("REFCAD", ("REFCAD", "REF CADASTRALE")),
    ("LATITUDE", ("LATITUDE",)),
    ("LONGITUDE", ("LONGITUDE",)),
)

# Colonnes que Po2 AJOUTE quand le fichier source ne les porte pas.
#
# Le nouvel export ASTECH n'a plus ni cadastre ni coordonnees : sans ces colonnes,
# l'enrichissement se limiterait a l'adresse et tout le travail d'attribution IGN
# serait perdu au retour. Decision Q22 du 2026-08-20 : on les ajoute, le referent
# retraitant le fichier avant reinjection.
#
# C'est la SEULE exception a la regle « en-tetes recopies du fichier source » — elle
# est explicite, limitee a trois colonnes absentes, et annoncee dans le resultat.
ADDED_COLUMNS = {
    "REFCAD": "Ref cadastrale",
    "LATITUDE": "Latitude",
    "LONGITUDE": "Longitude",
}

EXPORTED_COLUMNS = tuple(field for field, _aliases in EXPORTED_FIELDS)

# Statuts autorisés dans la feuille réinjectable. Un rattachement **proposé** par le
# moteur n'a été validé par personne : il n'a rien à faire dans le fichier de la
# collectivité (§17.3). Il ressort en feuille « à vérifier ».
EXPORTABLE_STATUSES = (STATUS_LINKED, STATUS_TO_CREATE)

# Expansion des types de voie vers la forme longue d'ASTECH (§11.2).
# La source DGFIP est elle-même incohérente — `BD`/`BOULEVARD`, `AV`/`Avenue`,
# `IMP`/`Impasse` cohabitent dans le même référentiel. Recopier tel quel importerait ce
# désordre dans le fichier de la collectivité : la normalisation n'est pas un confort.
STREET_TYPES = {
    "RUE": "RUE",
    "AV": "AVENUE", "AVE": "AVENUE", "AVENUE": "AVENUE",
    "BD": "BOULEVARD", "BLD": "BOULEVARD", "BOULEVARD": "BOULEVARD",
    "QUA": "QUAI", "QUAI": "QUAI",
    "CHE": "CHEMIN", "CHEM": "CHEMIN", "CHEMIN": "CHEMIN",
    "IMP": "IMPASSE", "IMPASSE": "IMPASSE",
    "PL": "PLACE", "PLACE": "PLACE",
    "PRO": "PROMENADE", "PROM": "PROMENADE", "PROMENADE": "PROMENADE",
    "RTE": "ROUTE", "ROUTE": "ROUTE",
    "ALL": "ALLEE", "ALLEE": "ALLEE",
    "COR": "CORNICHE", "CORNICHE": "CORNICHE",
    "TRA": "TRAVERSE", "TRAVERSE": "TRAVERSE",
    "PAS": "PASSAGE", "PASSAGE": "PASSAGE",
    "MTE": "MONTEE", "MONTEE": "MONTEE",
    "ESP": "ESPLANADE", "ESPLANADE": "ESPLANADE",
    "SQ": "SQUARE", "SQUARE": "SQUARE",
    "BRE": "BARRIERE", "PARC": "PARC",
}

# Indice de répétition DGFIP -> forme ASTECH. Décision Q9 : `BISTER` retrouve son usage
# légitime, il ne porte plus le type de voie.
REPETITION_INDEX = {"B": "BIS", "T": "TER", "Q": "QUATER", "C": "QUATER"}

# Code postal en fin de libelle : marque la frontiere entre la voie et la commune.
_POSTCODE_PATTERN = re.compile(r"\b(\d{5})\b")


def _ascii_upper(value: str) -> str:
    """Majuscules sans accent : la convention majoritaire d'ASTECH (425 lignes sur 615)."""
    folded = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return " ".join(folded.upper().split())


def split_house_number(raw: str | None) -> tuple[str | None, str | None, str | None]:
    """Sépare `'0005B'` en (numéro `'5'`, `BISTER` `'BIS'`, motif de rejet).

    Les zéros de tête viennent du référentiel DGFIP (`0005B`) et n'ont aucun sens dans
    ASTECH. Le suffixe, lui, est l'indice de répétition : `B` = bis, `T` = ter.

    Un suffixe inconnu n'est **pas** écrit : mieux vaut une case vide qu'une valeur
    inventée dans le référentiel de la collectivité.
    """
    text = (raw or "").strip().upper().replace(" ", "")
    if not text:
        return None, None, None
    digits = "".join(char for char in text if char.isdigit())
    suffix = "".join(char for char in text if char.isalpha())
    if not digits:
        return None, None, f"numéro de voirie non numérique : « {raw} »"
    number = digits.lstrip("0") or "0"
    if not suffix:
        return number, None, None
    expanded = REPETITION_INDEX.get(suffix)
    if expanded is None:
        # On garde le numéro, on renonce au suffixe, et on le signale.
        return number, None, f"indice de répétition inconnu : « {suffix} »"
    return number, expanded, None


def normalize_street(raw: str | None) -> tuple[str | None, str | None, str | None]:
    """Rend `LIBELVOIE` : type de voie **en toutes lettres** + nom (§9.1).

    Renvoie `(valeur, code_postal_trouvé, motif_de_rejet)`. Le motif est renseigné quand
    la voie n'est pas analysable — relevé en base : `4674` (un numéro là où le type
    devrait être) et `QUAIDU` (mots collés). Ces adresses ne partent pas dans ASTECH.

    ⚠️ Le géocodage inverse rend un libellé **complet** (« 10 Quai de la Résistance 34200
    Sète ») : après découpage, le code postal et la commune restent collés au nom de voie.
    Constaté sur les vraies données — `LIBELVOIE` serait parti dans ASTECH avec
    « QUAI DE LA RESISTANCE 34200 SETE ». On coupe au code postal, et on le récupère au
    passage puisqu'il est juste, plutôt que de le jeter.
    """
    text = _ascii_upper(raw or "")
    if not text:
        return None, None, None
    postcode = None
    match = _POSTCODE_PATTERN.search(text)
    if match:
        postcode = match.group(1)
        text = text[: match.start()].strip()
    if not text:
        return None, postcode, f"voie réduite à un code postal : « {raw} »"
    words = text.split(" ")
    head = words[0]
    if head.isdigit():
        return None, postcode, f"voie non analysable, commence par un nombre : « {raw} »"
    expanded = STREET_TYPES.get(head)
    if expanded is None:
        # Pas de type reconnu : ce n'est pas forcement une erreur (« LE BARROU »,
        # « CORNICHE DE NEUBURG » sans article). On recopie tel quel, en majuscules.
        return text, postcode, None
    rest = " ".join(words[1:]).strip()
    if not rest:
        # « RUE » tout seul ne designe aucune voie.
        return None, postcode, f"voie réduite à un type sans nom : « {raw} »"
    return f"{expanded} {rest}", postcode, None


def format_coordinate(value: float | None) -> str | None:
    """Coordonnée à **virgule décimale**, en texte — le format constaté dans le fichier
    (`'43,436176'`). Un point décimal serait relu comme du texte par ASTECH."""
    if value is None:
        return None
    return f"{value:.6f}".replace(".", ",")


def build_refcad(asset: PatrimoineLegacyAsset) -> tuple[str | None, str | None]:
    """`REFCAD` = section + numéro de plan sur 3 chiffres (`AK149`).

    Garde-fou : au-delà de 999, la référence ne rentre pas dans le format observé. On
    n'écrit alors **rien** plutôt qu'une référence tronquée, qui pointerait vers une
    autre parcelle.
    """
    if asset.resolved_refcad:
        return asset.resolved_refcad, None
    section = (asset.resolved_section or "").strip()
    plan = (asset.resolved_numero_plan or "").strip().lstrip("0")
    if not section or not plan:
        return None, None
    if len(plan) > 3:
        return None, f"numéro de plan hors format ASTECH (>999) : « {section}{plan} »"
    return f"{section}{plan.zfill(3)}", None


def build_export_row(asset: PatrimoineLegacyAsset) -> dict[str, Any]:
    """Construit les valeurs ASTECH d'un bien, et la liste de ce qu'on n'a pas su produire.

    Ne renvoie **que** ce qui a été produit proprement : une valeur absente reste absente,
    elle n'est pas remplacée par une approximation.
    """
    issues: list[str] = []
    values: dict[str, Any] = {}

    # Clé pivot : recopiée telle quelle. Les biens « à créer » sortent avec une clé VIDE,
    # c'est ASTECH qui leur en attribuera une (décision Q13).
    is_new = asset.code_bien.startswith(NEW_ASSET_CODE_PREFIX)
    values["CODE_BIEN"] = "" if is_new else asset.code_bien

    name = (asset.nomcourt or asset.designation or "").strip()
    if name:
        values["DESIGNATION"] = name[:255]
        values["NOMCOURT"] = name[:255]

    number, bister, number_issue = split_house_number(asset.resolved_housenumber)
    if number_issue:
        issues.append(number_issue)
    if number:
        values["NORUE"] = number
    if bister:
        values["BISTER"] = bister

    street, street_postcode, street_issue = normalize_street(asset.resolved_street)
    if street_issue:
        issues.append(street_issue)
    if street:
        values["LIBELVOIE"] = street

    # Le code postal du bâtiment est rarement renseigné (2 sur 81 en prod) ; celui
    # extrait du libellé géocodé est juste, autant s'en servir que le jeter.
    postcode = asset.resolved_postcode or street_postcode
    if postcode:
        values["CODPOST"] = postcode
    if asset.resolved_city:
        values["VILLE"] = _ascii_upper(asset.resolved_city)
    if asset.resolved_citycode:
        values["COMMUNE"] = asset.resolved_citycode

    refcad, refcad_issue = build_refcad(asset)
    if refcad_issue:
        issues.append(refcad_issue)
    if refcad:
        values["REFCAD"] = refcad

    latitude = format_coordinate(asset.latitude)
    longitude = format_coordinate(asset.longitude)
    if latitude and longitude:
        values["LATITUDE"] = latitude
        values["LONGITUDE"] = longitude

    # Une ligne qui n'apporte AUCUNE information nouvelle n'a pas à être réinjectée :
    # elle ferait courir un risque d'écrasement pour rien.
    enriched = [key for key in values if key not in ("CODE_BIEN", "DESIGNATION", "NOMCOURT")]
    if not enriched:
        issues.append("aucune donnée à renvoyer : ni adresse, ni cadastre, ni position")

    return {"values": values, "issues": issues, "is_new": is_new}


def _reason_not_exportable(asset: PatrimoineLegacyAsset) -> str | None:
    """Pourquoi ce bien ne peut pas partir dans le fichier de la collectivité."""
    if asset.status in EXPORTABLE_STATUSES:
        return None
    if asset.status == "propose":
        return (
            "rattachement proposé par le moteur, jamais confirmé — "
            "clique « Confirmer les rattachements proposés »"
        )
    if asset.status == "hors_perimetre":
        return "hors périmètre (bien situé hors de la commune)"
    if asset.status == "ignore":
        return "bien écarté du parcours"
    return "aucun rattachement à un bâtiment Po2"


def build_astech_workbook(db: Session, city_id: int | None) -> dict[str, Any]:
    """Assemble le classeur de retour. Renvoie les octets et le compte de chaque feuille."""
    import_statement = select(PatrimoineLegacyImport).order_by(PatrimoineLegacyImport.id.desc())
    asset_statement = select(PatrimoineLegacyAsset).order_by(PatrimoineLegacyAsset.code_bien.asc())
    if city_id is not None:
        import_statement = import_statement.where(PatrimoineLegacyImport.city_id == city_id)
        asset_statement = asset_statement.where(PatrimoineLegacyAsset.city_id == city_id)

    source_import = db.scalars(import_statement.limit(1)).first()
    if source_import is None:
        raise ValueError(
            "Aucun export ASTECH n'a été importé : le gabarit des en-têtes vient du "
            "fichier d'origine, il ne peut pas être inventé."
        )
    headers: list[str] = json.loads(source_import.headers_json)

    # Les en-têtes sont pris DANS le gabarit, par index, et écrits tels quels : c'est la
    # condition de réinjection posée par la collectivité (§13). La comparaison ignore
    # accents et casse — le nouvel export écrit « Désignation », « Nom court » — mais ce
    # qu'on ÉCRIT reste l'orthographe exacte du fichier.
    header_index = {_ascii_upper(name): position for position, name in enumerate(headers) if name}
    exported: list[tuple[str, str]] = []  # (champ logique, en-tête réellement écrit)
    added: list[str] = []
    for field, aliases in EXPORTED_FIELDS:
        position = next((header_index[alias] for alias in aliases if alias in header_index), None)
        if position is not None:
            exported.append((field, headers[position]))
        elif field in ADDED_COLUMNS:
            # Colonne absente du fichier source mais indispensable au retour (Q22) :
            # sans elle, tout le travail d'attribution IGN serait perdu.
            exported.append((field, ADDED_COLUMNS[field]))
            added.append(ADDED_COLUMNS[field])
    missing = [
        field
        for field, aliases in EXPORTED_FIELDS
        if not any(alias in header_index for alias in aliases) and field not in ADDED_COLUMNS
    ]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = source_import.sheet_name[:31] or "Feuil1"

    # Le fichier source place ses en-têtes en ligne 2 (ligne 1 vide) : on reproduit sa
    # mise en page, pas une mise en page à nous.
    for _ in range(source_import.header_row - 1):
        sheet.append([None] * len(exported))
    sheet.append([header for _field, header in exported])
    for cell in sheet[source_import.header_row]:
        cell.font = Font(bold=True)

    trace = workbook.create_sheet("Traçabilité")
    trace.append(["CODE_BIEN", "Bien", "Champ", "Valeur ASTECH d'origine", "Valeur renvoyée", "Origine", "Date"])
    for cell in trace[1]:
        cell.font = Font(bold=True)

    review = workbook.create_sheet("À vérifier")
    review.append(["CODE_BIEN", "Bien", "Motif"])
    for cell in review[1]:
        cell.font = Font(bold=True)

    today = date.today().isoformat()
    source_by_column = {
        "DESIGNATION": lambda a: a.designation,
        "NOMCOURT": lambda a: a.nomcourt,
        "NORUE": lambda a: a.source_norue,
        "BISTER": lambda a: a.source_bister,
        "LIBELVOIE": lambda a: a.source_libelvoie,
        "CODPOST": lambda a: a.source_codpost,
        "VILLE": lambda a: a.source_ville,
        "COMMUNE": lambda a: a.source_commune,
        "REFCAD": lambda a: a.source_refcad,
    }

    exported_rows = 0
    review_rows = 0
    for asset in db.scalars(asset_statement):
        label = asset.nomcourt or asset.designation or asset.code_bien
        blocked = _reason_not_exportable(asset)
        if blocked:
            review.append([asset.code_bien, label, blocked])
            review_rows += 1
            continue

        built = build_export_row(asset)
        if built["issues"]:
            review.append([asset.code_bien, label, " · ".join(built["issues"])])
            review_rows += 1
            # On écrit quand même ce qui a été produit proprement : le reste des colonnes
            # reste vide. Le bien apparaît dans les deux feuilles, ce qui est l'intention —
            # il part, mais la référente sait quoi relire.
            if not built["values"].get("LIBELVOIE") and not built["values"].get("REFCAD"):
                continue

        sheet.append([built["values"].get(field, "") for field, _header in exported])
        exported_rows += 1

        origin = "IGN/DGFIP" if asset.resolved_source == "building" else "point posé sur la carte"
        for column, _header in exported:
            if column == "CODE_BIEN":
                continue
            new_value = built["values"].get(column)
            if new_value in (None, ""):
                continue
            old_value = source_by_column.get(column, lambda _a: None)(asset)
            if (old_value or "") == str(new_value):
                continue
            trace.append([asset.code_bien, label, column, old_value or "", new_value, origin, today])

    buffer = io.BytesIO()
    workbook.save(buffer)
    filename = f"retour_astech_{date.today().strftime('%Y%m%d')}.xlsx"
    return {
        "content": buffer.getvalue(),
        "filename": filename,
        "exported_rows": exported_rows,
        "review_rows": review_rows,
        "columns": [header for _field, header in exported],
        "missing_columns": missing,
        "added_columns": added,
        "sheet_name": sheet.title,
        "header_row": source_import.header_row,
    }
