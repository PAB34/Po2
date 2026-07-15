"""Rapport de controle comptable multi-marches.

Increment 1-2 de la spec `demande-comptable-rapport-controle-spec.md` :
parser les worklists comptables, rapprocher par numero fournisseur strictement
trime, puis produire une feuille par marche sans synthese ni revision de prix.
"""
from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import BinaryIO, Literal

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.cpe import CpeAccountingSiteMapping, CpeFinanceControl, CpeFinanceInvoice, CpeFinanceLine
from app.models.gas_invoice import GasInvoice
from app.models.invoice import EnergyInvoiceImport
from app.services import cpe_accounting, energie_accounting
from app.services import accounting_contract_budget, engie_elec_budget_revise, gas_budget_revise

MarketKey = Literal["dalkia", "engie", "edf", "totalenergies"]

_SUPPLIER_INVOICE_RE = re.compile(r"^FAC\.\s*(\S+)\s+DU", re.IGNORECASE)
_TTC_TOLERANCE = 0.01
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class WorklistInvoice:
    row_number: int
    accounting_number: str | None
    supplier_invoice_number: str | None
    label: str
    total_ttc: float | None
    invoice_date: object | None
    arrival_date: object | None
    supplier_code: str | None
    supplier_name: str | None
    invoice_status: str | None
    liquidation_status: str | None
    market_code: str | None
    raw: dict[str, object | None]


@dataclass(frozen=True)
class WorklistParseResult:
    sheet_name: str
    rows: list[WorklistInvoice]


@dataclass(frozen=True)
class PlatformInvoice:
    id: int
    invoice_number: str
    total_ttc: float | None
    control_status: str | None
    decision_status: str | None
    problem_summary: str | None
    raw: object


@dataclass(frozen=True)
class MarketConfig:
    key: MarketKey
    title: str
    family: str


MARKETS: tuple[MarketConfig, ...] = (
    MarketConfig("dalkia", "DALKIA", "cpe"),
    MarketConfig("engie", "ENGIE", "energy"),
    MarketConfig("edf", "EDF", "energy"),
    MarketConfig("totalenergies", "TotalEnergies", "gas"),
)


def extract_supplier_invoice_number(label: str | None) -> str | None:
    """Extrait le numero fournisseur depuis `FAC. <num> DU`, sans autre normalisation."""
    if not label:
        return None
    match = _SUPPLIER_INVOICE_RE.match(str(label).strip())
    if not match:
        return None
    return match.group(1).strip()


def parse_comptable_worklist(source: bytes | str | Path | BinaryIO) -> WorklistParseResult:
    """Parse une worklist comptable XLSX au format `_ShowList-NNN`.

    Les lignes de total sans numero fournisseur sont ignorees. Les lignes avec
    libelle facture malforme restent dans le resultat avec un numero a `None`,
    afin d'apparaitre comme non rapprochables dans le rapport.
    """
    workbook = load_workbook(_as_stream(source), read_only=True, data_only=True)
    sheet_name = next((name for name in workbook.sheetnames if name.startswith("_ShowList-")), workbook.sheetnames[0])
    ws = workbook[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
    except StopIteration as exc:
        raise ValueError("Worklist comptable vide.") from exc

    header_index = {header: idx for idx, header in enumerate(headers) if header}
    if "Libellé" not in header_index and "Libelle" not in header_index:
        raise ValueError("Colonne Libellé absente de la worklist comptable.")

    parsed: list[WorklistInvoice] = []
    for excel_row_number, values in enumerate(rows_iter, start=2):
        row = {header: values[idx] if idx < len(values) else None for header, idx in header_index.items()}
        if not any(value not in (None, "") for value in row.values()):
            continue
        label = _text(_get(row, "Libellé", "Libelle")) or ""
        if label.strip().upper() == "TOTAL":
            continue
        supplier_number = extract_supplier_invoice_number(label)
        if not supplier_number and not label:
            continue
        parsed.append(
            WorklistInvoice(
                row_number=excel_row_number,
                accounting_number=_text(_get(row, "Numéro", "Numero")),
                supplier_invoice_number=supplier_number,
                label=label,
                total_ttc=_float(_get(row, "TTC")),
                invoice_date=_get(row, "Date facture"),
                arrival_date=_get(row, "Arrivée le", "Arrivee le"),
                supplier_code=_text(_get(row, "Tiers (code)")),
                supplier_name=_text(_get(row, "Tiers (Nom)")),
                invoice_status=_text(_get(row, "Etat facture", "État facture")),
                liquidation_status=_text(_get(row, "Etat liquidation", "État liquidation")),
                market_code=_text(_get(row, "Marché", "Marche")),
                raw=row,
            )
        )
    return WorklistParseResult(sheet_name=sheet_name, rows=parsed)


def build_comptable_control_workbook(
    db: Session,
    city_id: int,
    files_by_market: dict[MarketKey, bytes],
) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    parsed_by_market: dict[MarketKey, WorklistParseResult] = {}
    for config in MARKETS:
        content = files_by_market.get(config.key)
        if content:
            parsed_by_market[config.key] = parse_comptable_worklist(content)

    report_year = _infer_report_year(parsed_by_market) or date.today().year
    summary_ws = wb.create_sheet("Synthèse")
    _write_summary_sheet(db, city_id, summary_ws, report_year, parsed_by_market)

    for config in MARKETS:
        ws = wb.create_sheet(config.title)
        parsed = parsed_by_market.get(config.key)
        if parsed is None:
            ws["A1"] = "Aucune facture à analyser"
            ws["A1"].font = Font(bold=True)
            ws.column_dimensions["A"].width = 34
            continue
        platform = _platform_index(db, city_id, config, [row.supplier_invoice_number for row in parsed.rows])
        _write_market_sheet(db, city_id, ws, config, parsed, platform)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



def _write_summary_sheet(
    db: Session,
    city_id: int,
    ws,
    report_year: int,
    parsed_by_market: dict[MarketKey, WorklistParseResult],
) -> None:
    ws["A1"] = "Synthèse - rapport de contrôle comptable"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"Année de référence : {report_year}"
    ws["A2"].font = Font(italic=True)

    headers = ["Marché", "Fichier compta", "Factures worklist", "Réalisé à date", "Atterrissage", "Note"]
    _write_header(ws, 4, headers)
    row_cursor = 5
    total_realise = 0.0
    total_landing = 0.0
    for config in MARKETS:
        parsed = parsed_by_market.get(config.key)
        summary = _market_summary(db, city_id, config, report_year)
        realise = summary.get("realise")
        landing = summary.get("atterrissage")
        if isinstance(realise, (int, float)):
            total_realise += float(realise)
        if isinstance(landing, (int, float)):
            total_landing += float(landing)
        values = [
            config.title,
            "oui" if parsed else "non",
            len(parsed.rows) if parsed else 0,
            realise,
            landing,
            summary.get("note"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_cursor, column=col, value=value)
        for col in (4, 5):
            ws.cell(row=row_cursor, column=col).number_format = '#,##0.00 "EUR"'
        row_cursor += 1

    ws.cell(row=row_cursor, column=1, value="Total général").font = Font(bold=True)
    ws.cell(row=row_cursor, column=4, value=round(total_realise, 2)).number_format = '#,##0.00 "EUR"'
    ws.cell(row=row_cursor, column=5, value=round(total_landing, 2)).number_format = '#,##0.00 "EUR"'
    _set_widths(ws, [20, 14, 18, 18, 18, 90])


def _market_summary(db: Session, city_id: int, config: MarketConfig, report_year: int) -> dict[str, object]:
    try:
        if config.key == "dalkia":
            data = accounting_contract_budget.build_contract_budget_landing(db, city_id, year=report_year)
            totals = data.get("totals", {})
            return {
                "realise": totals.get("realise"),
                "atterrissage": totals.get("atterrissage"),
                "note": data.get("source_note"),
            }
        if config.key == "engie":
            data = engie_elec_budget_revise.build_engie_elec_budget_revise(db, city_id, year=report_year)
        elif config.key == "edf":
            data = engie_elec_budget_revise.build_edf_elec_budget_revise(db, city_id, year=report_year)
        else:
            data = gas_budget_revise.build_gas_budget_revise(db, city_id, year=report_year)
        totals = data.get("totals", {})
        return {
            "realise": totals.get("realise"),
            "atterrissage": totals.get("atterrissage"),
            "note": data.get("source_note"),
        }
    except Exception as exc:  # pragma: no cover - l'export reste utile même si un moteur est incomplet
        return {"realise": None, "atterrissage": None, "note": f"Synthèse indisponible : {exc}"}


def _infer_report_year(parsed_by_market: dict[MarketKey, WorklistParseResult]) -> int | None:
    counts: dict[int, int] = {}
    for parsed in parsed_by_market.values():
        for row in parsed.rows:
            year = _date_year(row.invoice_date) or _year_from_raw(row.raw.get("Exercice"))
            if year:
                counts[year] = counts.get(year, 0) + 1
    if not counts:
        return None
    return sorted(counts.items(), key=lambda item: (item[1], item[0]), reverse=True)[0][0]


def _year_from_raw(value: object | None) -> int | None:
    if value is None:
        return None
    try:
        year = int(str(value).strip()[:4])
    except ValueError:
        return None
    return year if 1990 <= year <= 2100 else None
def _write_market_sheet(
    db: Session,
    city_id: int,
    ws,
    config: MarketConfig,
    parsed: WorklistParseResult,
    platform: dict[str, PlatformInvoice],
) -> None:
    ws["A1"] = f"Rapport de controle comptable - {config.title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Source : feuille {parsed.sheet_name}"
    ws["A2"].font = Font(italic=True)

    headers = [
        "Rapprochement",
        "Numero fournisseur",
        "Numero compta",
        "Fournisseur",
        "TTC compta",
        "TTC plateforme",
        "Ecart TTC",
        "Date facture",
        "Etat facture",
        "Etat liquidation",
        "Controle plateforme",
        "Decision plateforme",
        "Motif / point à vérifier",
        "Libelle compta",
    ]
    header_row = 4
    _write_header(ws, header_row, headers)

    matched: list[tuple[WorklistInvoice, PlatformInvoice]] = []
    for row_index, item in enumerate(parsed.rows, start=header_row + 1):
        current = platform.get(item.supplier_invoice_number or "")
        status, delta = _reconciliation_status(item, current)
        if current is not None:
            matched.append((item, current))
        values = [
            status,
            item.supplier_invoice_number,
            item.accounting_number,
            item.supplier_name,
            item.total_ttc,
            current.total_ttc if current else None,
            delta,
            item.invoice_date,
            item.invoice_status,
            item.liquidation_status,
            _control_label(config.family, current.control_status if current else None),
            _decision_label(config.family, current.decision_status if current else None),
            _row_problem_summary(status, delta, current),
            item.label,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
        for col in (5, 6, 7):
            ws.cell(row=row_index, column=col).number_format = '#,##0.00 "EUR"'

    revision_start = header_row + len(parsed.rows) + 3
    if config.family == "cpe":
        revision_start = _write_cpe_accounting_summary(db, ws, revision_start, matched) + 2
    next_start = _write_revision_section(db, city_id, ws, revision_start, config, matched)

    detail_start = next_start + 2
    ws.cell(row=detail_start, column=1, value="Décomposition comptable").font = Font(bold=True, size=12)
    if not matched:
        ws.cell(row=detail_start + 1, column=1, value="Aucune facture rapprochée.")
    elif config.family == "cpe":
        _write_cpe_decomposition(db, ws, detail_start + 1, matched)
    elif config.family == "energy":
        _write_energy_decomposition(db, ws, detail_start + 1, matched)
    else:
        _write_gas_decomposition(ws, detail_start + 1, matched)

    _set_widths(ws, [18, 22, 16, 18, 14, 14, 14, 14, 24, 28, 20, 20, 48, 46, 20, 32, 28, 52])
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:N{max(header_row + 1, header_row + len(parsed.rows))}"




def _write_cpe_accounting_summary(db: Session, ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> int:
    ws.cell(row=start_row, column=1, value="Vue comptable par facture").font = Font(bold=True, size=12)
    headers = [
        "Numero fournisseur",
        "Contrat",
        "Marches / postes",
        "Lignes",
        "Sites",
        "Prix base annuel",
        "Prix revise annuel",
        "Revision annuelle",
        "Montant HT facture",
        "Ecart revision",
        "Indices",
        "Services",
        "Fonctions",
        "Antennes",
        "Operations",
        "Natures comptables",
        "Statut imputation",
        "Point a corriger",
    ]
    header_row = start_row + 1
    _write_header(ws, header_row, headers)
    if not matched:
        ws.cell(row=header_row + 1, column=1, value="Aucune facture rapprochee.")
        return header_row + 2

    invoice_ids = [current.id for _worklist, current in matched]
    lines = list(db.scalars(
        select(CpeFinanceLine)
        .where(CpeFinanceLine.invoice_id.in_(invoice_ids))
        .order_by(CpeFinanceLine.invoice_id, CpeFinanceLine.row_number)
    ).all())
    controls = list(db.scalars(
        select(CpeFinanceControl)
        .where(CpeFinanceControl.invoice_id.in_(invoice_ids))
        .where(CpeFinanceControl.control_type.in_(("revision_p2", "revision_p3")))
        .order_by(CpeFinanceControl.invoice_id, CpeFinanceControl.control_type, CpeFinanceControl.id)
    ).all())
    site_ids = sorted({line.accounting_site_id for line in lines if line.accounting_site_id})
    sites = {
        site.id: site
        for site in db.scalars(
            select(CpeAccountingSiteMapping).where(CpeAccountingSiteMapping.id.in_(site_ids))
        ).all()
    } if site_ids else {}

    lines_by_invoice: dict[int, list[CpeFinanceLine]] = {}
    for line in lines:
        lines_by_invoice.setdefault(line.invoice_id, []).append(line)
    controls_by_invoice: dict[int, list[CpeFinanceControl]] = {}
    for control in controls:
        controls_by_invoice.setdefault(control.invoice_id, []).append(control)

    row_cursor = header_row + 1
    for worklist_row, current in matched:
        invoice = current.raw
        invoice_lines = lines_by_invoice.get(current.id, [])
        invoice_controls = controls_by_invoice.get(current.id, [])
        site_values = [
            _site_summary_value(line, sites.get(line.accounting_site_id or 0))
            for line in invoice_lines
        ]
        services = [sites[line.accounting_site_id].service_code for line in invoice_lines if line.accounting_site_id in sites and sites[line.accounting_site_id].service_code]
        functions = [sites[line.accounting_site_id].function_code for line in invoice_lines if line.accounting_site_id in sites and sites[line.accounting_site_id].function_code]
        antennas = [sites[line.accounting_site_id].antenna_code for line in invoice_lines if line.accounting_site_id in sites and sites[line.accounting_site_id].antenna_code]
        operations = [sites[line.accounting_site_id].operation_code for line in invoice_lines if line.accounting_site_id in sites and sites[line.accounting_site_id].operation_code]
        natures = [
            f"{line.accounting_nature} - {line.accounting_label}" if line.accounting_label else line.accounting_nature
            for line in invoice_lines
            if line.accounting_nature
        ]
        base_total = _sum_present(line.base_price for line in invoice_lines)
        revised_total = _sum_present(line.revised_price for line in invoice_lines)
        revision_total = (
            round(revised_total - base_total, 2)
            if base_total is not None and revised_total is not None
            else None
        )
        revision_delta = _sum_present(control.delta_abs for control in invoice_controls)
        values = [
            worklist_row.supplier_invoice_number,
            _text(getattr(invoice, "contract_code", None)) if isinstance(invoice, CpeFinanceInvoice) else None,
            _join_unique(f"{line.market or '-'} / {line.billed_item or '-'}" for line in invoice_lines),
            len(invoice_lines) or None,
            _join_unique(value for value in site_values if value),
            base_total,
            revised_total,
            revision_total,
            getattr(invoice, "total_ht", None) if isinstance(invoice, CpeFinanceInvoice) else None,
            revision_delta,
            _join_unique(f"{control.index_year} T{control.index_quarter}" for control in invoice_controls if control.index_year and control.index_quarter),
            _join_unique(services),
            _join_unique(functions),
            _join_unique(antennas),
            _join_unique(operations),
            _join_unique(natures) or "A COMPLETER",
            _cpe_accounting_status(invoice_lines),
            current.problem_summary,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_cursor, column=col, value=value)
        for col in (6, 7, 8, 9, 10):
            ws.cell(row=row_cursor, column=col).number_format = '#,##0.00 "EUR"'
        row_cursor += 1
    return row_cursor

def _write_revision_section(
    db: Session,
    city_id: int,
    ws,
    start_row: int,
    config: MarketConfig,
    matched: list[tuple[WorklistInvoice, PlatformInvoice]],
) -> int:
    ws.cell(row=start_row, column=1, value="Révision de prix").font = Font(bold=True, size=12)
    if not matched:
        ws.cell(row=start_row + 1, column=1, value="Aucune facture rapprochée.")
        return start_row + 2
    if config.family == "cpe":
        return _write_cpe_revision_section(db, ws, start_row + 1, matched)
    if config.family == "energy":
        return _write_energy_revision_section(db, city_id, ws, start_row + 1, config, matched)
    return _write_gas_revision_section(db, city_id, ws, start_row + 1, matched)


def _write_cpe_revision_section(db: Session, ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> int:
    headers = [
        "Numero fournisseur", "Controle", "Statut", "Indice annee", "Indice trimestre",
        "ICHT-IME", "BT40", "FSD2", "Facteur", "Prix base", "Prix revise attendu",
        "Prix revise facture", "Ecart", "Message",
    ]
    _write_header(ws, start_row, headers)
    row_cursor = start_row + 1
    invoice_ids = [current.id for _worklist, current in matched]
    controls = db.scalars(
        select(CpeFinanceControl)
        .where(CpeFinanceControl.invoice_id.in_(invoice_ids))
        .where(CpeFinanceControl.control_type.in_(("revision_p2", "revision_p3", "p1_gaz_pu_os3", "p1_gaz_acompte_dpgf")))
        .order_by(CpeFinanceControl.invoice_id, CpeFinanceControl.control_type)
    ).all() if invoice_ids else []
    by_invoice: dict[int, list[CpeFinanceControl]] = {}
    for control in controls:
        by_invoice.setdefault(control.invoice_id, []).append(control)
    for worklist_row, current in matched:
        invoice_controls = by_invoice.get(current.id, [])
        if not invoice_controls:
            ws.cell(row=row_cursor, column=1, value=worklist_row.supplier_invoice_number)
            ws.cell(row=row_cursor, column=2, value="Pas de révision applicable ou contrôle non disponible")
            row_cursor += 1
            continue
        for control in invoice_controls:
            values = [
                worklist_row.supplier_invoice_number,
                control.control_type,
                control.status,
                control.index_year,
                control.index_quarter,
                control.icht_ime_value,
                control.bt40_value,
                control.fsd2_value,
                control.expected_factor,
                control.base_price,
                control.expected_revised_price,
                control.actual_revised_price,
                control.delta_abs,
                control.message,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_cursor, column=col, value=value)
            for col in (10, 11, 12, 13):
                ws.cell(row=row_cursor, column=col).number_format = '#,##0.00 "EUR"'
            row_cursor += 1
    return row_cursor


def _write_energy_revision_section(
    db: Session,
    city_id: int,
    ws,
    start_row: int,
    config: MarketConfig,
    matched: list[tuple[WorklistInvoice, PlatformInvoice]],
) -> int:
    headers = [
        "Numero fournisseur", "PRM", "Annee", "Ratio BPU", "Ratio TURPE", "BPU disponible",
        "Source reference", "Realise", "Atterrissage", "Note",
    ]
    _write_header(ws, start_row, headers)
    row_cursor = start_row + 1
    budget_cache: dict[int, dict] = {}
    for worklist_row, current in matched:
        invoice = current.raw
        if not isinstance(invoice, EnergyInvoiceImport):
            continue
        year = _invoice_year_for_revision(invoice, worklist_row)
        budget = budget_cache.get(year)
        if budget is None:
            budget = _safe_energy_budget_revise(db, city_id, config.key, year)
            budget_cache[year] = budget
        point_by_prm = {str(point.get("prm")): point for point in budget.get("points", []) if point.get("prm")}
        prms = _invoice_prms(invoice)
        if not prms:
            ws.cell(row=row_cursor, column=1, value=worklist_row.supplier_invoice_number)
            ws.cell(row=row_cursor, column=10, value="Aucun PRM normalisé disponible pour rattacher la révision.")
            row_cursor += 1
            continue
        for prm in prms:
            point = point_by_prm.get(prm)
            values = [
                worklist_row.supplier_invoice_number,
                prm,
                year,
                point.get("bpu_ratio") if point else None,
                point.get("turpe_ratio") if point else None,
                point.get("bpu_available") if point else None,
                point.get("reference_source") if point else None,
                point.get("realise") if point else None,
                point.get("atterrissage") if point else None,
                "BPU/TURPE appliqués au point de livraison" if point else "Point absent du moteur budget révisé",
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_cursor, column=col, value=value)
            for col in (8, 9):
                ws.cell(row=row_cursor, column=col).number_format = '#,##0.00 "EUR"'
            row_cursor += 1
    return row_cursor


def _write_gas_revision_section(db: Session, city_id: int, ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> int:
    headers = [
        "Numero fournisseur", "PCE", "Annee", "Ratio PEG", "Ratio climat", "PEG disponible",
        "Realise", "Atterrissage", "Note",
    ]
    _write_header(ws, start_row, headers)
    row_cursor = start_row + 1
    budget_cache: dict[int, dict] = {}
    for worklist_row, current in matched:
        invoice = current.raw
        if not isinstance(invoice, GasInvoice):
            continue
        year = _date_year(invoice.fin_conso) or _date_year(invoice.debut_conso) or _date_year(invoice.date_comptable) or _date_year(worklist_row.invoice_date) or date.today().year
        budget = budget_cache.get(year)
        if budget is None:
            budget = _safe_gas_budget_revise(db, city_id, year)
            budget_cache[year] = budget
        point_by_pce = {str(point.get("pce")): point for point in budget.get("points", []) if point.get("pce")}
        point = point_by_pce.get(invoice.pce)
        values = [
            worklist_row.supplier_invoice_number,
            invoice.pce,
            year,
            point.get("peg_ratio") if point else None,
            point.get("climate_ratio") if point else None,
            budget.get("peg_available"),
            point.get("realise") if point else None,
            point.get("atterrissage") if point else None,
            "PEG + climat appliqués au PCE" if point else "PCE absent du moteur budget révisé gaz",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_cursor, column=col, value=value)
        for col in (7, 8):
            ws.cell(row=row_cursor, column=col).number_format = '#,##0.00 "EUR"'
        row_cursor += 1
    return row_cursor


def _safe_energy_budget_revise(db: Session, city_id: int, market: MarketKey, year: int) -> dict:
    try:
        if market == "edf":
            return engie_elec_budget_revise.build_edf_elec_budget_revise(db, city_id, year=year)
        return engie_elec_budget_revise.build_engie_elec_budget_revise(db, city_id, year=year)
    except Exception as exc:  # pragma: no cover - rapport robuste si reference incomplete
        return {"points": [], "totals": {}, "error": str(exc)}


def _safe_gas_budget_revise(db: Session, city_id: int, year: int) -> dict:
    try:
        return gas_budget_revise.build_gas_budget_revise(db, city_id, year=year)
    except Exception as exc:  # pragma: no cover
        return {"points": [], "totals": {}, "error": str(exc)}
def _write_cpe_decomposition(db: Session, ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> None:
    headers_written = False
    row_cursor = start_row
    for worklist_row, current in matched:
        invoice = current.raw
        if not isinstance(invoice, CpeFinanceInvoice):
            continue
        try:
            content = cpe_accounting.build_detailed_finance_liaison_workbook(db, invoice)
            liaison_wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            detail = liaison_wb["Lignes finance"]
        except Exception as exc:  # pragma: no cover - garde-fou rapport operateur
            ws.cell(row=row_cursor, column=1, value=worklist_row.supplier_invoice_number)
            ws.cell(row=row_cursor, column=2, value=f"Décomposition DALKIA indisponible : {exc}")
            row_cursor += 1
            continue
        detail_rows = detail.iter_rows(values_only=True)
        try:
            liaison_headers = list(next(detail_rows))
        except StopIteration:
            continue
        if not headers_written:
            _write_header(ws, row_cursor, ["Numero fournisseur", *liaison_headers])
            headers_written = True
            row_cursor += 1
        for values in detail_rows:
            ws.cell(row=row_cursor, column=1, value=worklist_row.supplier_invoice_number)
            for col, value in enumerate(values, start=2):
                ws.cell(row=row_cursor, column=col, value=value)
            row_cursor += 1


def _write_energy_decomposition(db: Session, ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> None:
    headers = [
        "Numero fournisseur", "PRM", "Nom site", "Poste", "Libelle", "Montant HT",
        "Service", "Fonction", "Antenne", "Operation", "Nature", "Libelle nature", "Codification",
    ]
    _write_header(ws, start_row, headers)
    row_cursor = start_row + 1
    for worklist_row, current in matched:
        invoice = current.raw
        if not isinstance(invoice, EnergyInvoiceImport):
            continue
        rows = energie_accounting.resolve_invoice_codification(db, invoice)
        if not rows:
            ws.cell(row=row_cursor, column=1, value=worklist_row.supplier_invoice_number)
            ws.cell(row=row_cursor, column=4, value="Aucune ligne de codification disponible")
            row_cursor += 1
            continue
        for line in rows:
            values = [
                worklist_row.supplier_invoice_number,
                line.prm_id,
                line.site_name,
                line.poste,
                line.label,
                line.amount_ht,
                line.service_code,
                line.function_code,
                line.antenna_code,
                line.operation_code,
                line.accounting_nature,
                line.accounting_label,
                "OK" if line.status == "ok" else "A CODIFIER",
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=row_cursor, column=col, value=value)
            ws.cell(row=row_cursor, column=6).number_format = '#,##0.00 "EUR"'
            row_cursor += 1


def _write_gas_decomposition(ws, start_row: int, matched: list[tuple[WorklistInvoice, PlatformInvoice]]) -> None:
    headers = ["Numero fournisseur", "PCE", "Site", "Total HT", "Total TTC", "Controle", "Note"]
    _write_header(ws, start_row, headers)
    row_cursor = start_row + 1
    for worklist_row, current in matched:
        invoice = current.raw
        if not isinstance(invoice, GasInvoice):
            continue
        values = [
            worklist_row.supplier_invoice_number,
            invoice.pce,
            invoice.nom_site,
            invoice.total_hors_tva,
            invoice.total_ttc,
            invoice.control_status,
            "Décomposition comptable gaz non disponible dans l'incrément 2.",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_cursor, column=col, value=value)
        ws.cell(row=row_cursor, column=4).number_format = '#,##0.00 "EUR"'
        ws.cell(row=row_cursor, column=5).number_format = '#,##0.00 "EUR"'
        row_cursor += 1


def _platform_index(
    db: Session,
    city_id: int,
    config: MarketConfig,
    invoice_numbers: list[str | None],
) -> dict[str, PlatformInvoice]:
    keys = sorted({number.strip() for number in invoice_numbers if number and number.strip()})
    if not keys:
        return {}
    if config.family == "cpe":
        return _cpe_index(db, city_id, keys)
    if config.family == "gas":
        return _gas_index(db, city_id, keys)
    return _energy_index(db, city_id, keys, config.key)



_CONTROL_LABELS: dict[str, str] = {
    "valid": "Conforme",
    "ok": "Conforme",
    "warning": "Alerte",
    "error": "Écart",
    "blocked": "Bloqué",
    "not_checked": "Non contrôlé",
    "a_controler": "À contrôler",
    "valide": "Validée",
    "refuse": "Refusée",
    "conteste": "Contestée",
}

_DECISION_LABELS: dict[str, str] = {
    "approved": "Approuvée",
    "to_review": "À contrôler",
    "rejected": "Refusée",
    "dispute_sent": "Contestée",
    "valide": "Validée",
    "a_controler": "À contrôler",
    "refuse": "Refusée",
    "conteste": "Contestée",
}

_PROBLEM_CODE_LABELS: dict[str, str] = {
    "invoice_total_ht": "Total HT incohérent",
    "invoice_period": "Période de facture incohérente",
    "invoice_timeline": "Dates facture/échéance incohérentes",
    "p1_gaz_pu_os3": "Prix unitaire P1 gaz différent de l'OS3",
    "p1_gaz_acompte_dpgf": "Acompte P1 gaz différent du DPGF",
    "p2p3_base_dpgf": "Montant P2/P3 différent du DPGF",
    "revision_p2": "Indice de révision P2 à vérifier",
    "revision_p3": "Indice de révision P3 à vérifier",
    "p2_4_objectives": "Intéressement P2.4 à vérifier",
    "accounting_nature": "Imputation comptable absente",
    "accounting_site": "Site comptable non rattaché",
    "BPU_PRICE_MISMATCH": "Écart prix BPU",
    "BPU_TARIFF_POSTE_INCONSISTENCY": "Incohérence poste/tarif BPU",
    "TOTAL_TTC_MISMATCH": "Écart total TTC",
    "LINE_AMOUNT_MISMATCH": "Écart montant de ligne",
    "HT_TOTAL_MISMATCH": "Écart total HT",
    "MISSING_INVOICE_NUMBER": "Numéro de facture absent",
    "MISSING_TOTAL_TTC": "Total TTC absent",
    "SUPPLIER_UNKNOWN": "Fournisseur non reconnu",
    "NO_SITE_FOUND": "Aucun point de livraison détecté",
    "UNKNOWN_PRM": "PRM hors référentiel",
    "SUPPLIER_CONTRACT_MISMATCH": "PRM rattaché à un autre fournisseur",
    "MISSING_PRM": "PRM absent",
    "BPU_REFERENCE_MISSING": "Référence BPU absente",
    "BPU_PRICE_MISSING": "Prix BPU absent",
    "ENEDIS_CONSUMPTION_MISSING": "Données ENEDIS absentes",
    "POWER_REFERENCE_MISSING": "Référence de puissance absente",
    "TAX_TOTALS_MISSING": "Totaux de taxes incomplets",
}

_NON_APPROVED_DECISIONS = {"to_review", "rejected", "dispute_sent", "a_controler", "refuse", "conteste"}


def _control_label(family: str, status: str | None) -> str | None:
    if not status:
        return None
    if "(" in status:
        code, details = status.split("(", 1)
        return f"{_CONTROL_LABELS.get(code.strip(), code.strip())} ({details}"
    return _CONTROL_LABELS.get(status, status)


def _decision_label(family: str, status: str | None) -> str | None:
    if not status:
        return None
    return _DECISION_LABELS.get(status, status)


def _row_problem_summary(status: str, delta: float | None, current: PlatformInvoice | None) -> str | None:
    if current is None:
        return "Facture non trouvée dans la plateforme."
    if status == "Numero fournisseur introuvable":
        return "Numéro fournisseur non extrait du libellé comptable."
    if delta is not None and abs(delta) > _TTC_TOLERANCE:
        return f"Écart TTC plateforme - compta : {delta:.2f} EUR."
    if current.decision_status in _NON_APPROVED_DECISIONS:
        return current.problem_summary or "Décision non finalisée dans la plateforme."
    return current.problem_summary


def _problem_label(code: object | None, fallback: object | None = None) -> str:
    if code:
        text = str(code)
        return _PROBLEM_CODE_LABELS.get(text, text.replace("_", " ").lower())
    if fallback:
        return str(fallback)
    return "Point de contrôle à vérifier"


def _summarize_problem_counts(items: list[tuple[str | None, str | None]]) -> str | None:
    counts: dict[str, int] = {}
    for code, message in items:
        label = _problem_label(code, message)
        counts[label] = counts.get(label, 0) + 1
    if not counts:
        return None
    selected = sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:3]
    parts = [f"{label} ({count})" if count > 1 else label for label, count in selected]
    remaining = sum(counts.values()) - sum(count for _label, count in selected)
    suffix = f" ; +{remaining} autre(s)" if remaining > 0 else ""
    return " ; ".join(parts) + suffix


def _energy_problem_summary(invoice: EnergyInvoiceImport) -> str | None:
    if invoice.decision_comment:
        return invoice.decision_comment
    issues = [
        (issue.get("code"), issue.get("message"))
        for issue in invoice.control_issues
        if isinstance(issue, dict)
    ]
    summary = _summarize_problem_counts(issues)
    if summary:
        return summary
    if invoice.decision_status == "to_review" and invoice.control_status == "valid":
        return "Contrôles conformes ; décision plateforme restant à valider."
    return None


def _gas_problem_summary(invoice: GasInvoice) -> str | None:
    if invoice.decision_comment:
        return invoice.decision_comment
    issues: list[tuple[str | None, str | None]] = []
    if invoice.control_issues_json:
        try:
            raw = json.loads(invoice.control_issues_json)
        except json.JSONDecodeError:
            raw = []
        if isinstance(raw, list):
            for issue in raw:
                if isinstance(issue, dict):
                    issues.append((issue.get("code"), issue.get("message")))
                else:
                    issues.append((None, str(issue)))
    summary = _summarize_problem_counts(issues)
    if summary:
        return summary
    if invoice.decision_status == "to_review" and invoice.control_status == "valid":
        return "Contrôles conformes ; décision plateforme restant à valider."
    return None

def _energy_index(db: Session, city_id: int, keys: list[str], market: MarketKey) -> dict[str, PlatformInvoice]:
    stmt = (
        select(EnergyInvoiceImport)
        .where(EnergyInvoiceImport.city_id == city_id, EnergyInvoiceImport.invoice_number.in_(keys))
        .order_by(EnergyInvoiceImport.updated_at.desc(), EnergyInvoiceImport.id.desc())
    )
    out: dict[str, PlatformInvoice] = {}
    for invoice in db.scalars(stmt).all():
        number = (invoice.invoice_number or "").strip()
        if not number or number in out:
            continue
        if market == "engie" and not _looks_like_supplier(invoice, "ENGIE"):
            continue
        if market == "edf" and not _looks_like_supplier(invoice, "EDF"):
            continue
        out[number] = PlatformInvoice(
            id=invoice.id,
            invoice_number=number,
            total_ttc=invoice.total_ttc,
            control_status=_energy_control_label(invoice),
            decision_status=invoice.decision_status,
            problem_summary=_energy_problem_summary(invoice),
            raw=invoice,
        )
    return out


def _cpe_index(db: Session, city_id: int, keys: list[str]) -> dict[str, PlatformInvoice]:
    stmt = (
        select(CpeFinanceInvoice)
        .where(CpeFinanceInvoice.city_id == city_id, CpeFinanceInvoice.invoice_number.in_(keys))
        .order_by(CpeFinanceInvoice.updated_at.desc(), CpeFinanceInvoice.id.desc())
    )
    invoices = db.scalars(stmt).all()
    invoice_ids = [invoice.id for invoice in invoices]
    totals = _cpe_ttc_by_invoice_id(db, invoice_ids)
    control_summaries = _cpe_control_summaries(db, invoice_ids)
    out: dict[str, PlatformInvoice] = {}
    for invoice in invoices:
        number = invoice.invoice_number.strip()
        if number in out:
            continue
        summary = control_summaries.get(invoice.id, {})
        out[number] = PlatformInvoice(
            id=invoice.id,
            invoice_number=number,
            total_ttc=totals.get(invoice.id),
            control_status=summary.get("control_status") or invoice.status,
            decision_status=invoice.status,
            problem_summary=summary.get("problem_summary") or _cpe_decision_problem_summary(invoice),
            raw=invoice,
        )
    return out



def _cpe_control_summaries(db: Session, invoice_ids: list[int]) -> dict[int, dict[str, str | None]]:
    if not invoice_ids:
        return {}
    rows = db.execute(
        select(CpeFinanceControl.invoice_id, CpeFinanceControl.status, CpeFinanceControl.control_type, CpeFinanceControl.message)
        .where(CpeFinanceControl.invoice_id.in_(invoice_ids))
    ).all()
    by_invoice: dict[int, list[tuple[str | None, str | None, str | None]]] = {}
    for invoice_id, status, control_type, message in rows:
        by_invoice.setdefault(invoice_id, []).append((status, control_type, message))

    summaries: dict[int, dict[str, str | None]] = {}
    for invoice_id in invoice_ids:
        controls = by_invoice.get(invoice_id, [])
        errors = [(control_type, message) for status, control_type, message in controls if status == "error"]
        blocked = [(control_type, message) for status, control_type, message in controls if status == "blocked"]
        warnings = [(control_type, message) for status, control_type, message in controls if status == "warning"]
        if errors:
            summaries[invoice_id] = {
                "control_status": f"error ({len(errors)} écart(s), {len(blocked)} bloqué(s))",
                "problem_summary": _summarize_problem_counts(errors + blocked),
            }
        elif blocked:
            summaries[invoice_id] = {
                "control_status": f"blocked ({len(blocked)} bloqué(s))",
                "problem_summary": _summarize_problem_counts(blocked),
            }
        elif warnings:
            summaries[invoice_id] = {
                "control_status": f"warning ({len(warnings)} alerte(s))",
                "problem_summary": _summarize_problem_counts(warnings),
            }
        elif controls:
            summaries[invoice_id] = {
                "control_status": "valid",
                "problem_summary": None,
            }
        else:
            summaries[invoice_id] = {
                "control_status": "not_checked",
                "problem_summary": "Aucun contrôle CPE disponible pour cette facture.",
            }
    return summaries


def _cpe_decision_problem_summary(invoice: CpeFinanceInvoice) -> str | None:
    if invoice.notes:
        return invoice.notes
    if invoice.status == "a_controler":
        return "Contrôles conformes ou non bloquants ; décision DALKIA restant à valider."
    return None

def _gas_index(db: Session, city_id: int, keys: list[str]) -> dict[str, PlatformInvoice]:
    stmt = (
        select(GasInvoice)
        .where(GasInvoice.city_id == city_id, GasInvoice.num_facture.in_(keys))
        .order_by(GasInvoice.updated_at.desc(), GasInvoice.id.desc())
    )
    out: dict[str, PlatformInvoice] = {}
    for invoice in db.scalars(stmt).all():
        number = invoice.num_facture.strip()
        if number in out:
            continue
        out[number] = PlatformInvoice(
            id=invoice.id,
            invoice_number=number,
            total_ttc=invoice.total_ttc,
            control_status=invoice.control_status,
            decision_status=invoice.decision_status,
            problem_summary=_gas_problem_summary(invoice),
            raw=invoice,
        )
    return out


def _cpe_ttc_by_invoice_id(db: Session, invoice_ids: list[int]) -> dict[int, float]:
    if not invoice_ids:
        return {}
    totals: dict[int, float] = {invoice_id: 0.0 for invoice_id in invoice_ids}
    rows = db.execute(
        select(CpeFinanceLine.invoice_id, CpeFinanceLine.amount_ht, CpeFinanceLine.vat_rate)
        .where(CpeFinanceLine.invoice_id.in_(invoice_ids))
    ).all()
    for invoice_id, amount_ht, vat_rate in rows:
        rate = (vat_rate or 0.0) / 100.0
        totals[invoice_id] = totals.get(invoice_id, 0.0) + float(amount_ht or 0.0) * (1.0 + rate)
    return {invoice_id: round(total, 2) for invoice_id, total in totals.items()}


def _reconciliation_status(item: WorklistInvoice, current: PlatformInvoice | None) -> tuple[str, float | None]:
    if item.supplier_invoice_number is None:
        return "Numero fournisseur introuvable", None
    if current is None:
        return "Absente plateforme", None
    if item.total_ttc is None or current.total_ttc is None:
        return "Rapprochée", None
    delta = round(current.total_ttc - item.total_ttc, 2)
    if abs(delta) > _TTC_TOLERANCE:
        return "Écart TTC", delta
    return "Rapprochée", delta



def _invoice_year_for_revision(invoice: EnergyInvoiceImport, worklist_row: WorklistInvoice) -> int:
    return (
        _date_year(invoice.period_end)
        or _date_year(invoice.period_start)
        or _date_year(invoice.invoice_date)
        or _date_year(worklist_row.invoice_date)
        or date.today().year
    )


def _invoice_prms(invoice: EnergyInvoiceImport) -> list[str]:
    normalized = invoice.normalized_invoice
    if normalized is None:
        return []
    prms: list[str] = []
    for site in normalized.sites:
        prm = _text(site.prm_id)
        if prm and prm not in prms:
            prms.append(prm)
    return prms


def _date_year(value: object | None) -> int | None:
    if value is None:
        return None
    year = getattr(value, "year", None)
    if isinstance(year, int):
        return year
    text = str(value).strip()
    if len(text) >= 4 and text[:4].isdigit():
        return int(text[:4])
    if len(text) >= 10 and text[6:10].isdigit():
        return int(text[6:10])
    return None
def _energy_control_label(invoice: EnergyInvoiceImport) -> str:
    return f"{invoice.control_status} ({invoice.control_errors_count} erreur(s), {invoice.control_warnings_count} alerte(s))"


def _looks_like_supplier(invoice: EnergyInvoiceImport, supplier: str) -> bool:
    value = " ".join([invoice.supplier_guess or "", invoice.source or "", invoice.original_filename or ""]).upper()
    return supplier in value



def _join_unique(values, *, limit: int = 6) -> str | None:
    seen: list[str] = []
    for value in values:
        text = _text(value)
        if text and text not in seen:
            seen.append(text)
    if not seen:
        return None
    suffix = f" ; +{len(seen) - limit}" if len(seen) > limit else ""
    return " ; ".join(seen[:limit]) + suffix


def _sum_present(values) -> float | None:
    total = 0.0
    found = False
    for value in values:
        if value is None:
            continue
        total += float(value)
        found = True
    return round(total, 2) if found else None


def _site_summary_value(line: CpeFinanceLine, site: CpeAccountingSiteMapping | None) -> str | None:
    if site and site.site_name:
        return f"{line.site_code_detected or site.code_site} - {site.site_name}"
    return line.site_code_detected


def _cpe_accounting_status(lines: list[CpeFinanceLine]) -> str:
    if not lines:
        return "Aucune ligne"
    missing_nature = sum(1 for line in lines if not line.accounting_nature)
    missing_site = sum(1 for line in lines if not line.accounting_site_id)
    if missing_nature == 0 and missing_site == 0:
        return "OK"
    parts = []
    if missing_nature:
        parts.append(f"{missing_nature} nature(s) a completer")
    if missing_site:
        parts.append(f"{missing_site} site(s) a rattacher")
    return "A completer : " + ", ".join(parts)

def _write_header(ws, row: int, headers: list[object]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _set_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _as_stream(source: bytes | str | Path | BinaryIO) -> str | Path | BinaryIO:
    if isinstance(source, bytes):
        return io.BytesIO(source)
    return source


def _get(row: dict[str, object | None], *names: str) -> object | None:
    for name in names:
        if name in row:
            return row[name]
    return None


def _text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _float(value: object | None) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


__all__ = [
    "MarketKey",
    "WorklistInvoice",
    "WorklistParseResult",
    "XLSX_MEDIA_TYPE",
    "build_comptable_control_workbook",
    "extract_supplier_invoice_number",
    "parse_comptable_worklist",
]