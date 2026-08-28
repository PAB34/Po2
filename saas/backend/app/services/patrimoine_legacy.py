"""
Import et rapprochement du référentiel patrimoine historique (ASTECH).

Flux cible (aller-retour) : export ASTECH -> import Po2 -> rapprochement / attribution
IGN -> réexport réinjectable dans ASTECH. Ce module couvre l'**aller** : lecture du
classeur, filtrage du périmètre, persistance, et proposition de candidats.

Contraintes de réinjection (données par la collectivité, cf.
`docs/refonte-v1/patrimoine-fichier-historique-rapprochement-decisions.md` §13) :
- le `CODE_BIEN` ne doit jamais être modifié : c'est la clé de mise à jour d'ASTECH ;
- les en-têtes de colonnes ne doivent jamais être modifiés : ils sont donc conservés
  **tels quels** dans `PatrimoineLegacyImport.headers_json` et serviront de gabarit au
  réexport, plutôt que d'être réécrits depuis le code.

Le moteur de reconnaissance réutilise `_site_similarity` du module CVC, déjà en
production pour rapprocher les inventaires DALKIA/SPIE des bâtiments.
"""
from __future__ import annotations

import io
import json
import re
import unicodedata
from datetime import datetime, timezone
from typing import Any

import openpyxl
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.models.building import Building
from app.models.city import City
from app.models.local import Local
from app.models.patrimoine_legacy import (
    ORIGIN_AUTO,
    ORIGIN_MANUAL,
    STATUS_GONE,
    STATUS_IGNORED,
    STATUS_LINKED,
    STATUS_PROPOSED,
    TARGET_BUILDING,
    TARGET_LOCAL,
    STATUS_OUT_OF_SCOPE,
    STATUS_TO_CREATE,
    STATUS_TODO,
    PatrimoineLegacyAsset,
    PatrimoineLegacyImport,
)
from app.services.building_naming import lookup_free_address_candidates, reverse_geocode_point
from app.services.cvc import _site_similarity

# --- Périmètre par défaut ----------------------------------------------------
# Décision utilisateur (2026-08-19) : importer **tous les bâtiments de la feuille
# `BAT`**, c'est-à-dire les genres `BATI` et `SITE`, y compris les biens sortis du
# parc — ils font partie du référentiel à traiter et sont simplement signalés.
# Le filtre reste paramétrable (`genres`, `include_out_of_park`).
DEFAULT_GENRES = ("BATI", "SITE")
DEFAULT_INCLUDE_OUT_OF_PARK = True
# 'O' = sorti du parc. Le fichier historique conserve les biens désaffectés.
EXCLUDED_HORSPARC = "O"

# Q4 : les biens hors Sète ne sont pas traités. On n'écarte QUE ceux dont la commune
# est explicitement autre : 41 bâtiments sétois n'ont aucune commune renseignée et
# seraient perdus par une lecture littérale de la règle.
SETE_INSEE = "34301"
SETE_NAMES = {"sete", "cette"}

# Colonnes ASTECH lues (orthographe de `Feuil1`, gabarit retenu).
# Clé pivot ASTECH. `CODBAR` est un repli : sur la feuille `BAT`, la colonne
# `CODEBIEN` est vide alors que `CODBAR` porte le même code (vérifié : 863/866
# lignes identiques sur l'export réel).
_KEY_COLUMNS = ("CODE_BIEN", "CODEBIEN", "CODBAR", "CODE")

# Le referent ASTECH a renomme les colonnes de son export (fichier du 2026-08-20) :
# `CODE_BIEN` devient `Code`, `NOMCOURT` devient `Nom court`, `BISTER` devient
# `Complement`... On accepte donc les DEUX generations plutot que d'imposer un format,
# et le premier alias trouve dans le fichier gagne. La comparaison se fait sans accent
# ni casse : le nouveau fichier ecrit « Designation » avec un accent.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "designation": ("DESIGNATION",),
    "nomcourt": ("NOMCOURT", "NOM COURT"),
    "genre": ("GENRE",),
    "categ": ("CATEG",),
    "categ_des": ("CATEG_DES",),
    "souscat_des": ("SOUSCAT_DES",),
    "horsparc": ("HORSPARC",),
    "code_parent": ("CODE_PARENT",),
    "source_norue": ("NORUE", "NUMERO(S)", "NUMEROS", "NUMERO"),
    # `Complement` porte le type de voie, exactement comme `BISTER` avant lui :
    # verifie sur le nouveau fichier — RUE 91, BD 19, QUA 13, AVE 9, et BIS 3 fois.
    "source_bister": ("BISTER", "COMPLEMENT"),
    "source_libelvoie": ("LIBELVOIE", "ADRESSE"),
    "source_codpost": ("CODPOST", "CODE POSTAL"),
    "source_ville": ("VILLE",),
    "source_commune": ("COMMUNE",),
    "source_refcad": ("REFCAD", "REF CADASTRALE"),
}

# Colonnes dont la valeur est composite, au format `CODE / LIBELLE` : `BATI / BATIMENT`,
# `34301 / 34301 SETE`. On ne garde que le code — sinon le filtre de perimetre ne
# reconnaitrait aucun genre et `COMMUNE` sortirait avec le libelle colle.
_COMPOSITE_FIELDS = ("genre", "source_commune")
# Longueurs déclarées sur le modèle : on tronque proprement plutôt que de laisser la
# base rejeter une ligne (le fichier historique contient des libellés très longs).
_MAX_LENGTHS = {
    "designation": 255, "nomcourt": 255, "genre": 20, "categ": 20, "categ_des": 120,
    "souscat_des": 120, "horsparc": 2, "code_parent": 40, "source_norue": 40,
    "source_bister": 40, "source_libelvoie": 255, "source_codpost": 10,
    "source_ville": 120, "source_commune": 10, "source_refcad": 40,
}

# Seuil de rattachement automatique (Q5 : on rattache les évidences, modifiables ensuite).
AUTO_LINK_SCORE = 0.90
# En dessous, aucun candidat n'est proposé : mieux vaut une case vide qu'une piste fausse.
CANDIDATE_MIN_SCORE = 0.50
# Écart minimal entre le 1er et le 2e candidat pour qu'un rattachement soit « évident ».
# Sans ce garde-fou, « TENNIS » se rattache seul à un bâtiment parmi plusieurs voisins.
AMBIGUITY_GAP = 0.05
# Au-dessus, le nom est identique au bâtiment cible : la présence d'un second candidat
# proche ne doit pas empêcher le rattachement. Mesuré sur le fichier réel : sans cette
# exception, « ECOLE ELEMENTAIRE PAUL BERT » -> « ECOLE ELEMENTAIRE PAUL BERT » (score 1,0)
# était bloqué à cause d'une école voisine au nom ressemblant.
EXACT_MATCH_SCORE = 0.98

# Provenance de l'adresse resolue, pour la feuille de tracabilite du reexport.
RESOLVED_FROM_BUILDING = "building"
RESOLVED_FROM_REVERSE = "ign_reverse"


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _normalize_ascii(value: Any) -> str:
    if not value:
        return ""
    ascii_value = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_value.lower()).strip()


# ---------------------------------------------------------------------------
# Lecture du classeur ASTECH
# ---------------------------------------------------------------------------

def _find_header_row(worksheet, max_scan: int = 6) -> tuple[int, list[str]] | None:
    """Localise la ligne d'en-têtes.

    L'ancien export la plaçait en ligne 2 (`Feuil1`) ou 1 (`BAT`) ; le nouveau en
    ligne 1. La reconnaissance se fait sur le CONTENU — une colonne clé et une
    désignation — et non sur un numéro de ligne figé.

    Comparaison sans accent ni casse : le nouveau fichier écrit « Désignation ».
    """
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        values = [("" if cell is None else str(cell).strip()) for cell in row]
        keys = {_header_key(v) for v in values if v}
        if any(key in keys for key in _KEY_COLUMNS) and "DESIGNATION" in keys:
            return row_index, values
    return None


def parse_astech_workbook(raw_bytes: bytes) -> dict[str, Any]:
    """Retourne la feuille exploitable, ses en-têtes **à l'octet près**, et ses lignes.

    Le classeur contient plusieurs feuilles dont les en-têtes divergent (`Feuil1` :
    317 colonnes, clé `CODE_BIEN` renseignée ; `BAT` : 122 colonnes, clé `CODEBIEN`
    vidée). On retient la feuille dont la clé est **effectivement renseignée**, car
    c'est elle qui porte le gabarit natif ASTECH.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    best: dict[str, Any] | None = None
    for worksheet in workbook.worksheets:
        found = _find_header_row(worksheet)
        if found is None:
            continue
        header_row, headers = found
        normalized_headers = [_header_key(h) for h in headers]
        key_indexes = [
            normalized_headers.index(key) for key in _KEY_COLUMNS if key in normalized_headers
        ]
        if not key_indexes:
            continue
        rows = [
            row
            for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

        def _filled(index: int) -> int:
            return sum(1 for row in rows if index < len(row) and _text(row[index]) is not None)

        # On retient la colonne clé réellement **renseignée**, pas la première trouvée :
        # sur la feuille `BAT`, `CODEBIEN` est vide alors que `CODBAR` porte le code.
        key_index = max(key_indexes, key=_filled)
        filled_keys = _filled(key_index)
        candidate = {
            "sheet_name": worksheet.title,
            "header_row": header_row,
            "headers": headers,
            "key_index": key_index,
            "key_header": headers[key_index],
            "rows": rows,
            "filled_keys": filled_keys,
        }
        if best is None or filled_keys > best["filled_keys"]:
            best = candidate
    if best is None:
        raise ValueError(
            "Aucune feuille exploitable : il faut une ligne d'en-têtes contenant "
            "CODE_BIEN (ou CODEBIEN) et DESIGNATION."
        )
    if best["filled_keys"] == 0:
        raise ValueError(
            f"La colonne « {best['key_header']} » est vide dans toutes les feuilles du "
            "classeur. Le code bien est la clé de rapprochement : redemander un export "
            "ASTECH avec cette colonne renseignée."
        )
    return best


# ---------------------------------------------------------------------------
# Périmètre
# ---------------------------------------------------------------------------

def _is_out_of_scope_commune(commune: str | None, ville: str | None) -> bool:
    """Vrai seulement si la commune est explicitement autre que Sète.

    Une commune absente n'est PAS hors périmètre : 41 bâtiments manifestement sétois
    (WC publics, restaurants scolaires…) n'ont aucune commune renseignée.
    """
    code = _text(commune)
    name = _normalize_ascii(ville)
    if code and code != SETE_INSEE:
        return True
    if not code and name and name not in SETE_NAMES:
        return True
    return False


def _row_in_scope(values: dict[str, Any], genres: tuple[str, ...], include_out_of_park: bool) -> bool:
    genre = (_text(values.get("genre")) or "").upper()
    if genres and genre not in genres:
        return False
    if not include_out_of_park and (_text(values.get("horsparc")) or "").upper() == EXCLUDED_HORSPARC:
        return False
    return True


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def _header_key(value: Any) -> str:
    """Cle de comparaison d'un en-tete : majuscules, sans accent, espaces normalises.

    Le nouveau fichier ecrit « Designation » avec un accent et « Nom court » en deux
    mots : comparer les octets bruts ne reconnaitrait aucune colonne.
    """
    return _normalize_ascii(value).upper()


def _composite_code(value: str | None) -> str | None:
    """`BATI / BATIMENT` -> `BATI`, `34301 / 34301 SETE` -> `34301`.

    Le nouveau export ASTECH livre ces deux colonnes au format `CODE / LIBELLE`.
    """
    if not value or " / " not in value:
        return value
    return value.split(" / ", 1)[0].strip() or value


def _extract_values(row: tuple[Any, ...], header_index: dict[str, int]) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for field, aliases in _COLUMN_ALIASES.items():
        index = next(
            (header_index[alias] for alias in aliases if alias in header_index), None
        )
        text = _text(row[index]) if index is not None and index < len(row) else None
        if text is not None and field in _COMPOSITE_FIELDS:
            text = _composite_code(text)
        limit = _MAX_LENGTHS.get(field)
        if text is not None and limit is not None and len(text) > limit:
            text = text[:limit]
        values[field] = text
    return values


def import_astech_file(
    db: Session,
    *,
    city_id: int | None,
    filename: str,
    raw_bytes: bytes,
    genres: tuple[str, ...] = DEFAULT_GENRES,
    include_out_of_park: bool = DEFAULT_INCLUDE_OUT_OF_PARK,
    batch: str | None = None,
) -> dict[str, Any]:
    """Charge un export ASTECH. **Idempotent** : rejouer le même fichier met à jour les
    biens existants (clé `CODE_BIEN`) sans dupliquer ni perdre les rattachements validés.
    """
    parsed = parse_astech_workbook(raw_bytes)
    headers: list[str] = parsed["headers"]
    header_index = {_header_key(header): position for position, header in enumerate(headers) if header}
    key_index: int = parsed["key_index"]
    batch_name = batch or f"astech_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    # Combien de biens etaient deja la AVANT cet import : sert a detecter un changement
    # de codification (cf. `codes_disjoints` plus bas).
    existing_before = db.scalar(
        select(func.count(PatrimoineLegacyAsset.id)).where(
            PatrimoineLegacyAsset.city_id == city_id
        )
    ) or 0

    created = updated = skipped_scope = skipped_out_of_city = skipped_no_key = 0

    for row in parsed["rows"]:
        code_bien = _text(row[key_index]) if key_index < len(row) else None
        if not code_bien:
            skipped_no_key += 1
            continue
        values = _extract_values(row, header_index)
        if not _row_in_scope(values, genres, include_out_of_park):
            skipped_scope += 1
            continue

        out_of_scope = _is_out_of_scope_commune(values.get("source_commune"), values.get("source_ville"))
        if out_of_scope:
            skipped_out_of_city += 1

        asset = db.scalar(
            select(PatrimoineLegacyAsset).where(
                PatrimoineLegacyAsset.city_id == city_id,
                PatrimoineLegacyAsset.code_bien == code_bien,
            )
        )
        is_new = asset is None
        if asset is None:
            asset = PatrimoineLegacyAsset(city_id=city_id, code_bien=code_bien, status=STATUS_TODO)

        for field, value in values.items():
            setattr(asset, field, value)
        asset.import_batch = batch_name
        asset.source_row_number = None
        # Payload complet : indispensable pour réémettre le fichier au format ASTECH.
        asset.source_payload_json = json.dumps(
            {
                header: ("" if position >= len(row) or row[position] is None else str(row[position]))
                for position, header in enumerate(headers)
                if header
            },
            ensure_ascii=False,
        )
        # Le statut d'un bien déjà traité n'est jamais rétrogradé par un réimport.
        if out_of_scope and asset.status == STATUS_TODO:
            asset.status = STATUS_OUT_OF_SCOPE
        elif not out_of_scope and asset.status == STATUS_OUT_OF_SCOPE:
            asset.status = STATUS_TODO

        db.add(asset)
        created += 1 if is_new else 0
        updated += 0 if is_new else 1

    import_row = db.scalar(
        select(PatrimoineLegacyImport).where(
            PatrimoineLegacyImport.city_id == city_id,
            PatrimoineLegacyImport.batch == batch_name,
        )
    )
    if import_row is None:
        import_row = PatrimoineLegacyImport(city_id=city_id, batch=batch_name)
    import_row.filename = filename[:255]
    import_row.sheet_name = str(parsed["sheet_name"])[:120]
    import_row.header_row = int(parsed["header_row"])
    # En-têtes conservés tels quels : gabarit obligatoire pour la réinjection ASTECH.
    import_row.headers_json = json.dumps(headers, ensure_ascii=False)
    import_row.total_rows = len(parsed["rows"])
    import_row.imported_rows = created + updated
    import_row.skipped_rows = skipped_scope + skipped_no_key
    db.add(import_row)
    db.commit()

    return {
        "batch": batch_name,
        "sheet_name": parsed["sheet_name"],
        "header_row": parsed["header_row"],
        "columns": len([h for h in headers if h]),
        "total_rows": len(parsed["rows"]),
        "created": created,
        "updated": updated,
        "skipped_scope": skipped_scope,
        "skipped_no_key": skipped_no_key,
        "out_of_scope_commune": skipped_out_of_city,
        # Garde-fou : ce fichier n'a AUCUN code en commun avec les biens deja presents.
        # C'est le signe d'un changement de codification cote ASTECH — constate le
        # 2026-08-20, `ADMICIMET02` devenu `BATI00272`. Les deux jeux cohabitent alors
        # au lieu de se mettre a jour, et l'ecran doit le dire : l'utilisateur s'est
        # retrouve avec 824 biens (444 + 380) sans comprendre pourquoi.
        "codes_disjoints": bool(existing_before and created and not updated),
        "existing_before": existing_before,
    }


# ---------------------------------------------------------------------------
# Reconnaissance des noms
# ---------------------------------------------------------------------------

def _load_building_targets(db: Session, city_id: int | None) -> list[tuple[int, str, str | None]]:
    """Bâtiments cibles. Q3 : la cible d'un rattachement est un `Building`, jamais un `Site`."""
    statement = select(Building.id, Building.nom_batiment, Building.adresse_reconstituee)
    if city_id is not None:
        statement = statement.where(Building.city_id == city_id)
    return [(row[0], row[1] or "", row[2]) for row in db.execute(statement) if row[1]]


def _address_bonus(asset: PatrimoineLegacyAsset, address: str | None) -> float:
    """Départage entre candidats de score voisin — jamais une moyenne pondérée.

    Mélanger nom et adresse dans un score unique dégrade le résultat (mesuré : 92 -> 60
    rapprochements sûrs), parce qu'une ligne sur cinq n'a pas de voie et que les adresses
    Po2 sont écrites autrement. L'adresse ne sert donc qu'à départager.
    """
    voie = _normalize_ascii(asset.source_libelvoie)
    if not voie or not address:
        return 0.0
    target = _normalize_ascii(address).lstrip("0123456789 ")
    if not target:
        return 0.0
    tokens = {token for token in voie.split() if len(token) > 2}
    if not tokens:
        return 0.0
    hits = sum(1 for token in tokens if token in target)
    return 0.02 * (hits / len(tokens))


def _best_candidate(
    asset: PatrimoineLegacyAsset, targets: list[tuple[int, str, str | None]]
) -> dict[str, Any]:
    scored: list[tuple[float, int, str]] = []
    for building_id, name, address in targets:
        score = max(
            _site_similarity(asset.nomcourt, name),
            _site_similarity(asset.designation, name),
        )
        if score <= 0:
            continue
        scored.append((min(1.0, score + _address_bonus(asset, address)), building_id, name))
    if not scored:
        return {"id": None, "label": None, "score": 0.0, "reason": None, "ambiguous": False}

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_id, best_label = scored[0]
    runner_up = scored[1][0] if len(scored) > 1 else 0.0
    if best_score < CANDIDATE_MIN_SCORE:
        return {"id": None, "label": None, "score": round(best_score, 3), "reason": None, "ambiguous": False}
    # Deux bâtiments quasi équivalents : on propose, on ne tranche pas. Sauf si le nom
    # est identique à la cible — un second candidat proche ne rend pas l'exact douteux.
    ambiguous = best_score < EXACT_MATCH_SCORE and (best_score - runner_up) < AMBIGUITY_GAP
    return {
        "id": best_id,
        "label": best_label,
        "score": round(best_score, 3),
        "reason": (
            "plusieurs bâtiments proches"
            if ambiguous
            else "nom identique" if best_score >= 0.999 else "nom approchant"
        ),
        "ambiguous": ambiguous,
    }


def compute_candidates(
    db: Session, city_id: int | None, *, auto_link: bool = True
) -> dict[str, int]:
    """(Re)calcule les candidats des biens non traités, et rattache les évidences.

    Q5 : le rattachement automatique est validé, avec modification manuelle toujours
    possible ensuite. Un bien déjà rattaché à la main n'est jamais recalculé.
    """
    targets = _load_building_targets(db, city_id)

    # Auto-reparation : supprimer le patrimoine Po2 met `building_id` a NULL en cascade
    # (ON DELETE SET NULL). Les biens restaient alors affiches « rattache » ou
    # « a confirmer » alors qu'ils ne pointaient plus vers rien, et disparaissaient de
    # la carte. On les remet a traiter pour qu'ils soient reproposes.
    orphan_statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.building_id.is_(None),
        PatrimoineLegacyAsset.status.in_([STATUS_LINKED, STATUS_PROPOSED]),
    )
    if city_id is not None:
        orphan_statement = orphan_statement.where(PatrimoineLegacyAsset.city_id == city_id)
    repaired = 0
    for asset in db.scalars(orphan_statement):
        asset.status = STATUS_TODO
        asset.local_id = None
        asset.target_type = TARGET_BUILDING
        asset.link_origin = None
        _clear_resolved_address(asset)
        db.add(asset)
        repaired += 1
    # Candidats perimes : `candidate_building_id` n'a PAS de cle etrangere, il survit
    # donc a une purge du patrimoine. Constate en prod le 2026-08-20 : le referentiel
    # Po2 avait ete reimporte, les batiments avaient de nouveaux identifiants, et les
    # 294 candidats proposes pointaient tous vers des batiments disparus. « Valider ce
    # rattachement » echouait alors sur la contrainte d'integrite.
    known_ids = {target[0] for target in targets}
    stale_statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.candidate_building_id.is_not(None)
    )
    if city_id is not None:
        stale_statement = stale_statement.where(PatrimoineLegacyAsset.city_id == city_id)
    stale = 0
    for asset in db.scalars(stale_statement):
        if asset.candidate_building_id in known_ids:
            continue
        asset.candidate_building_id = None
        asset.candidate_label = None
        asset.candidate_score = None
        asset.candidate_reason = None
        db.add(asset)
        stale += 1
    if repaired or stale:
        db.commit()

    statement = select(PatrimoineLegacyAsset).where(PatrimoineLegacyAsset.status == STATUS_TODO)
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)

    scanned = proposed = 0
    eligible: list[tuple[PatrimoineLegacyAsset, dict[str, Any]]] = []
    for asset in db.scalars(statement):
        scanned += 1
        candidate = _best_candidate(asset, targets)
        asset.candidate_building_id = candidate["id"]
        asset.candidate_label = candidate["label"]
        asset.candidate_score = candidate["score"] or None
        asset.candidate_reason = candidate["reason"]
        if candidate["id"] is not None:
            proposed += 1
        if (
            auto_link
            and candidate["id"] is not None
            and candidate["score"] >= AUTO_LINK_SCORE
            and not candidate["ambiguous"]
        ):
            eligible.append((asset, candidate))
        db.add(asset)

    # Second garde-fou : si plusieurs biens revendiquent le même bâtiment, le
    # rattachement n'est plus une évidence (cas « PAUL LANGEVIN » : maternelle,
    # élémentaire et restaurant scolaire portent le même nom court). La relation
    # N codes bien -> 1 bâtiment reste permise, mais elle se valide à la main.
    claims: dict[int, int] = {}
    for _, candidate in eligible:
        claims[candidate["id"]] = claims.get(candidate["id"], 0) + 1

    linked = 0
    for asset, candidate in eligible:
        if claims.get(candidate["id"], 0) > 1:
            asset.candidate_reason = "plusieurs biens visent ce bâtiment"
            db.add(asset)
            continue
        asset.building_id = candidate["id"]
        asset.target_type = TARGET_BUILDING
        # Le moteur PROPOSE, il ne valide pas : la confirmation reste humaine.
        asset.status = STATUS_PROPOSED
        asset.link_origin = ORIGIN_AUTO
        # Le rattachement manuel fait hériter l'adresse, le nom et le cadastre du
        # bâtiment ; le rattachement automatique doit en faire autant. Sans cela les
        # biens proposés par le moteur restent sans aucun champ résolu, et le réexport
        # ASTECH n'a rien à écrire pour eux — même une fois confirmés.
        _refresh_inherited_address(db, asset)
        db.add(asset)
        linked += 1

    db.commit()
    return {"scanned": scanned, "proposed": proposed, "auto_linked": linked, "repaired": repaired}


# ---------------------------------------------------------------------------
# Consultation et décision
# ---------------------------------------------------------------------------

def list_assets(
    db: Session,
    city_id: int | None,
    *,
    status: str | None = None,
    genre: str | None = None,
    search: str | None = None,
    limit: int = 500,
    offset: int = 0,
) -> list[PatrimoineLegacyAsset]:
    statement = select(PatrimoineLegacyAsset)
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)
    if status:
        statement = statement.where(PatrimoineLegacyAsset.status == status)
    if genre:
        statement = statement.where(PatrimoineLegacyAsset.genre == genre)
    if search:
        pattern = f"%{search.strip()}%"
        statement = statement.where(
            PatrimoineLegacyAsset.designation.ilike(pattern)
            | PatrimoineLegacyAsset.nomcourt.ilike(pattern)
            | PatrimoineLegacyAsset.code_bien.ilike(pattern)
        )
    statement = statement.order_by(
        PatrimoineLegacyAsset.candidate_score.desc().nullslast(),
        PatrimoineLegacyAsset.code_bien.asc(),
    ).limit(limit).offset(offset)
    return list(db.scalars(statement))


def counts_by_status(db: Session, city_id: int | None) -> dict[str, int]:
    statement = select(PatrimoineLegacyAsset.status, func.count(PatrimoineLegacyAsset.id))
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)
    counts = {status: count for status, count in db.execute(statement.group_by(PatrimoineLegacyAsset.status))}
    counts["total"] = sum(counts.values())
    return counts


def get_asset_or_none(db: Session, city_id: int | None, asset_id: int) -> PatrimoineLegacyAsset | None:
    statement = select(PatrimoineLegacyAsset).where(PatrimoineLegacyAsset.id == asset_id)
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)
    return db.scalar(statement)


def update_asset(
    db: Session,
    asset: PatrimoineLegacyAsset,
    *,
    status: str | None = None,
    building_id: int | None = None,
    local_id: int | None = None,
    designation: str | None = None,
    latitude: float | None = None,
    longitude: float | None = None,
    notes: str | None = None,
    clear_building: bool = False,
    clear_candidate: bool = False,
) -> PatrimoineLegacyAsset:
    """Décision utilisateur. Le `code_bien` n'est jamais modifiable : c'est la clé de
    mise à jour d'ASTECH.

    Deux enrichissements automatiques, pour éviter une saisie manuelle :
    - rattachement à un bâtiment Po2 → le bien **hérite de son adresse et de sa position** ;
    - point déplacé sans rattachement → l'adresse est **résolue par géocodage inverse**.
    Dans les deux cas l'adresse d'origine (`source_*`) reste intacte.
    """
    if clear_building:
        asset.building_id = None
        asset.local_id = None
        asset.target_type = TARGET_BUILDING
        asset.link_origin = None
        _clear_resolved_address(asset)
        # Le bien redevient ce qu'ASTECH en dit : son nom d'origine lui revient (Q26).
        # Une correction manuelle passée dans le même appel reste prioritaire.
        if designation is None:
            _restore_source_name(asset)
    elif local_id is not None:
        # Cible « local » : le bâtiment porteur reste renseigné, c'est lui qui porte
        # l'adresse, le cadastre et la position — un local n'a rien de tout cela.
        local = db.get(Local, local_id)
        if local is None:
            raise ValueError("Local introuvable.")
        asset.local_id = local_id
        asset.building_id = local.building_id
        asset.target_type = TARGET_LOCAL
        asset.status = STATUS_LINKED
        asset.link_origin = ORIGIN_MANUAL
        building = db.get(Building, local.building_id)
        if building is not None:
            _inherit_building_address(asset, building)
            # Le local peut porter SA propre adresse : le fichier d'inventaire en
            # fournit une par ligne, et l'entrée d'un local diffère parfois de celle
            # du bâtiment. Elle prime alors sur celle du bâtiment porteur.
            _override_with_local_address(asset, local)
            latitude_source = local.latitude if local.latitude is not None else building.latitude
            longitude_source = local.longitude if local.longitude is not None else building.longitude
            if latitude_source is not None and longitude_source is not None:
                asset.latitude = latitude_source
                asset.longitude = longitude_source
                latitude = longitude = None
    elif building_id is not None:
        building = db.get(Building, building_id)
        # Le batiment vise peut avoir disparu : `candidate_building_id` n'a pas de cle
        # etrangere et survit donc a une purge du patrimoine. Constate en prod le
        # 2026-08-20 : le referentiel Po2 avait ete reimporte, les batiments avaient
        # de nouveaux identifiants, et les 294 candidats pointaient tous dans le vide.
        # Sans ce controle, la contrainte d'integrite renvoyait une 500 opaque et le
        # bouton « Valider ce rattachement » semblait ne rien faire.
        if building is None:
            raise ValueError(
                "Ce bâtiment Po2 n'existe plus — le patrimoine a été réimporté depuis. "
                "Relance « 2. Reconnaître les noms » pour recalculer les candidats."
            )
        asset.building_id = building_id
        asset.local_id = None
        asset.target_type = TARGET_BUILDING
        asset.status = STATUS_LINKED
        asset.link_origin = ORIGIN_MANUAL
        if building is not None:
            _adopt_target_name(asset, building.nom_batiment, designation)
            _inherit_building_address(asset, building)
            # Le point ASTECH rejoint le bâtiment : c'est le geste « je dépose le
            # point sur le bâtiment Po2 », il ne doit pas rester à côté.
            if building.latitude is not None and building.longitude is not None:
                asset.latitude = building.latitude
                asset.longitude = building.longitude
                latitude = longitude = None
    if status is not None:
        asset.status = status
    if latitude is not None:
        asset.latitude = latitude
    if longitude is not None:
        asset.longitude = longitude
    moved = (latitude is not None or longitude is not None) and asset.latitude is not None and asset.longitude is not None
    if moved and asset.building_id is None:
        _resolve_address_from_point(asset)
    elif moved:
        # Bien rattaché : le point ASTECH et le bâtiment Po2 ne font plus qu'un.
        # Déplacer ce point unique déplace donc AUSSI le bâtiment Po2, et l'adresse
        # est recalculée pour les deux — c'est le sens du geste demandé.
        _resolve_address_from_point(asset)
        linked = db.get(Building, asset.building_id)
        if linked is not None:
            linked.latitude = asset.latitude
            linked.longitude = asset.longitude
            if asset.resolved_label:
                linked.adresse_reconstituee = asset.resolved_label[:255]
            if asset.resolved_city:
                linked.nom_commune = asset.resolved_city[:255]
            if asset.resolved_postcode:
                linked.code_postal = asset.resolved_postcode[:10]
            db.add(linked)
    if clear_candidate:
        # Rejeter la proposition du moteur SANS ecarter le bien : il reste a traiter.
        # « Ecarter » signifiait jusqu'ici « ignorer le bien », ce qui le sortait du
        # parcours — pas du tout la meme intention que « cette suggestion est fausse ».
        # Le bien n'a alors plus de candidat : il se positionne a la main sur la carte,
        # puis se rattache par depot sur un batiment ou par le selecteur.
        asset.candidate_building_id = None
        asset.candidate_label = None
        asset.candidate_score = None
        asset.candidate_reason = None
    if designation is not None:
        # Le nom affiché vient de `nomcourt` sinon de `designation` : on écrit les deux,
        # sinon corriger le libellé ne changerait rien à l'écran. C'est aussi ce qui
        # repartira dans ASTECH — le `code_bien`, lui, reste intouchable.
        cleaned = designation.strip()[:255] or None
        asset.designation = cleaned
        asset.nomcourt = cleaned
    if notes is not None:
        asset.notes = notes.strip() or None
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def _adopt_target_name(
    asset: PatrimoineLegacyAsset, target_name: str | None, explicit_designation: str | None
) -> None:
    """Le bien ASTECH prend le nom de la cible Po2 à laquelle on vient de le rattacher.

    C'est la décision Q11, appliquée jusqu'au bout : le nom Po2/IGN gagne et sera
    réécrit dans ASTECH. Le rattachement fait converger les deux référentiels — c'est
    l'objet même du chantier. Le `code_bien` reste la clé, donc renommer ne casse rien
    au cycle suivant.

    Deux garde-fous :
    - une **modification manuelle** passée dans le même appel gagne : l'utilisateur qui
      corrige un nom n'a pas à se le faire écraser par le rattachement ;
    - une cible sans nom ne vide pas le libellé ASTECH, mieux vaut l'ancien que rien.
    """
    if explicit_designation is not None:
        return
    cleaned = (target_name or "").strip()[:255]
    if not cleaned:
        return
    asset.designation = cleaned
    asset.nomcourt = cleaned


def _source_names(asset: PatrimoineLegacyAsset) -> dict[str, str]:
    """Libellés ASTECH d'ORIGINE du bien, relus depuis sa ligne source.

    Les clés de `source_payload_json` sont les **en-têtes bruts** du fichier importé, et
    les deux générations d'export n'ont pas les mêmes (`NOMCOURT` contre `Nom court`) :
    on repasse donc par la table d'alias, exactement comme à la lecture. Un bien créé
    depuis Po2 n'a pas de payload et renvoie un dictionnaire vide.
    """
    if not asset.source_payload_json:
        return {}
    try:
        payload = json.loads(asset.source_payload_json)
    except (ValueError, TypeError):
        return {}
    if not isinstance(payload, dict):
        return {}
    by_header = {_header_key(key): value for key, value in payload.items()}
    found: dict[str, str] = {}
    for field in ("designation", "nomcourt"):
        original = _text(
            next((by_header[alias] for alias in _COLUMN_ALIASES[field] if alias in by_header), None)
        )
        if original:
            found[field] = original[:255]
    return found


def _restore_source_name(asset: PatrimoineLegacyAsset) -> None:
    """Rend au bien son libellé ASTECH d'origine (Q26).

    Le rattachement fait adopter le nom Po2 (Q11) : c'est voulu tant que le lien existe.
    Mais détacher devait rendre au bien ce qu'il était, comme cela lui rend déjà son
    absence de position. Sans quoi un rattachement fait par erreur — un point deplacé qui
    accroche le bâtiment d'à côté — laissait un nom Po2 collé pour de bon.

    La source est `source_payload_json`, dont les clés sont les **en-têtes bruts** du
    fichier importé. Les deux générations d'export n'ont pas les mêmes (`NOMCOURT` contre
    `Nom court`) : on repasse donc par la table d'alias, exactement comme à la lecture.
    Un bien créé depuis Po2 n'a pas de payload — on ne touche alors à rien.
    """
    # Une source vide ne vide pas le libellé : mieux vaut le nom Po2 que rien.
    for field, original in _source_names(asset).items():
        setattr(asset, field, original)


def _clear_resolved_address(asset: PatrimoineLegacyAsset) -> None:
    asset.resolved_housenumber = None
    asset.resolved_street = None
    asset.resolved_postcode = None
    asset.resolved_city = None
    asset.resolved_citycode = None
    asset.resolved_label = None
    asset.resolved_source = None
    asset.resolved_name = None
    asset.resolved_section = None
    asset.resolved_numero_plan = None
    asset.resolved_refcad = None


_ADDRESS_PATTERN = re.compile(r"^\s*(\d+\s*[A-Za-z]?)\s+(.+?)\s*$")
# INSEE(5) + prefixe(3) + section(2) + numero de plan(4) : '34301000AK0149'.
_REFERENCE_NORM_PATTERN = re.compile(r"^(\d{5})(\d{3})([0-9A-Z]{2})(\d{4})$")


def _split_reconstituted_address(address: str | None) -> tuple[str | None, str | None]:
    """Sépare '208 AV DU MARECHAL JUIN' en ('208', 'AV DU MARECHAL JUIN').

    Les bâtiments Po2 ne stockent PAS l'adresse découpée (vérifié en prod : les colonnes
    numero_voirie / nom_voie sont vides sur toutes les lignes). Le découpage est donc
    reconstitué ici — le format est régulier — pour pouvoir alimenter NORUE et LIBELVOIE.
    """
    if not address:
        return None, None
    match = _ADDRESS_PATTERN.match(address)
    if match is None:
        return None, _text(address)
    return _text(match.group(1).replace(" ", "")), _text(match.group(2))


def _split_cadastral_reference(reference: str | None) -> tuple[str | None, str | None, str | None]:
    """Extrait (section, numéro de plan, REFCAD) de '34301000AK0149'.

    Le REFCAD attendu par ASTECH fait 5 caractères : section (2) + plan sur 3 chiffres
    (constaté : 'AS023'). Au-delà de 999, la valeur ne rentre pas dans ce format : on ne
    produit alors PAS de REFCAD plutôt que d'écrire une référence tronquée dans le
    référentiel de la collectivité. Section et plan restent disponibles séparément.
    """
    normalized = _text(reference)
    if not normalized:
        return None, None, None
    match = _REFERENCE_NORM_PATTERN.match(normalized.upper())
    if match is None:
        return None, None, None
    section = match.group(3)
    plan = match.group(4)
    plan_number = plan.lstrip("0") or "0"
    refcad = f"{section}{plan_number.zfill(3)}" if len(plan_number) <= 3 else None
    return section, plan, refcad


def _inherit_building_address(asset: PatrimoineLegacyAsset, building: Building) -> None:
    """Le bien reprend **tout** ce que Po2 sait du bâtiment : nom, adresse, cadastre.

    Les champs structurés du bâtiment étant vides en base, on reconstitue le découpage
    depuis `adresse_reconstituee` et `dgfip_reference_norm`, qui eux sont renseignés.
    """
    street = " ".join(part for part in [building.nature_voie, building.nom_voie] if part) or None
    housenumber = building.numero_voirie or None
    if not housenumber or not street:
        parsed_number, parsed_street = _split_reconstituted_address(building.adresse_reconstituee)
        housenumber = housenumber or parsed_number
        street = street or parsed_street

    section, numero_plan, refcad = _split_cadastral_reference(building.dgfip_reference_norm)
    if not section:
        section, numero_plan = building.section or None, building.numero_plan or None
        if section and numero_plan:
            plan_number = numero_plan.lstrip("0") or "0"
            refcad = f"{section}{plan_number.zfill(3)}" if len(plan_number) <= 3 else None

    asset.resolved_name = building.nom_batiment or None
    asset.resolved_housenumber = housenumber
    asset.resolved_street = street
    asset.resolved_postcode = building.code_postal or None
    asset.resolved_city = building.nom_commune or None
    asset.resolved_citycode = (building.dgfip_reference_norm or "")[:5] or None
    asset.resolved_label = building.adresse_reconstituee or (
        " ".join(part for part in [housenumber, street, building.nom_commune] if part) or None
    )
    asset.resolved_section = section
    asset.resolved_numero_plan = numero_plan
    asset.resolved_refcad = refcad
    asset.resolved_source = RESOLVED_FROM_BUILDING


def _override_with_local_address(asset: PatrimoineLegacyAsset, local: Local) -> None:
    """Remplace l'héritage par l'adresse propre du local, quand elle existe.

    Champ par champ : un local peut n'avoir qu'une partie de l'information, et il ne
    faut pas effacer ce que le bâtiment fournissait déjà.
    """
    if local.adresse_reconstituee:
        housenumber, street = _split_reconstituted_address(local.adresse_reconstituee)
        asset.resolved_label = local.adresse_reconstituee[:255]
        if housenumber:
            asset.resolved_housenumber = housenumber
        if street:
            asset.resolved_street = street
    if local.code_postal:
        asset.resolved_postcode = local.code_postal
    if local.nom_commune:
        asset.resolved_city = local.nom_commune
    if local.dgfip_reference_norm:
        section, numero_plan, refcad = _split_cadastral_reference(local.dgfip_reference_norm)
        if section:
            asset.resolved_section = section
            asset.resolved_numero_plan = numero_plan
            asset.resolved_refcad = refcad
            asset.resolved_citycode = local.dgfip_reference_norm[:5]
    if local.nom_local:
        asset.resolved_name = local.nom_local[:255]


def refresh_assets_of_building(db: Session, building_id: int) -> int:
    """Réaligne tous les biens ASTECH visant ce bâtiment, après un changement du bâtiment.

    Appelée après une attribution IGN : c'est elle qui fait descendre le cadastre et
    l'adresse fraîchement obtenus jusqu'au bien ASTECH, donc jusqu'au fichier de retour.
    Sans elle, « Attribuer IGN » enrichissait le bâtiment Po2 et le bien ASTECH gardait
    ses anciennes valeurs — l'action semblait sans effet sur l'export.

    `force` : une attribution IGN est un geste explicite et fait autorité, elle prime
    donc sur une adresse issue d'un point posé à la main.
    """
    statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.building_id == building_id
    )
    refreshed = 0
    for asset in db.scalars(statement):
        _refresh_inherited_address(db, asset, force=True)
        db.add(asset)
        refreshed += 1
    if refreshed:
        db.commit()
    return refreshed


def _refresh_inherited_address(
    db: Session, asset: PatrimoineLegacyAsset, force: bool = False
) -> None:
    """Réaligne les champs résolus du bien sur son bâtiment porteur.

    Sûr à rejouer : quand `resolved_source` vaut `building`, les champs résolus ne sont
    qu'une **copie** du bâtiment, donc les rafraîchir ne peut rien perdre. On préserve
    en revanche `ign_reverse`, qui est le résultat d'un point posé à la main.

    Nécessaire parce que le bâtiment porteur bouge : renommé, réadressé, réattaché à
    l'IGN. Constaté en prod le 2026-08-19 — deux biens portaient encore le nom
    « École Élémentaire Anatole France » hérité d'un bâtiment renommé depuis, et c'est
    ce nom périmé qui serait reparti dans ASTECH au réexport.
    """
    if asset.building_id is None:
        return
    if asset.resolved_source == RESOLVED_FROM_REVERSE and not force:
        return
    building = db.get(Building, asset.building_id)
    if building is None:
        return
    _inherit_building_address(asset, building)
    if asset.local_id is not None:
        local = db.get(Local, asset.local_id)
        if local is not None:
            _override_with_local_address(asset, local)


def _resolve_address_from_point(asset: PatrimoineLegacyAsset) -> None:
    """Géocodage inverse du point posé. Best effort : jamais bloquant."""
    try:
        found = reverse_geocode_point(asset.latitude, asset.longitude)
    except Exception:
        return
    if not found.get("found"):
        return
    asset.resolved_housenumber = (found.get("housenumber") or None)
    asset.resolved_street = (found.get("street") or None)
    asset.resolved_postcode = (found.get("postcode") or None)
    asset.resolved_city = (found.get("city") or None)
    asset.resolved_citycode = (found.get("citycode") or None)
    asset.resolved_label = (found.get("label") or None)
    asset.resolved_source = RESOLVED_FROM_REVERSE


def resolve_city_id(db: Session, city_id: int | None) -> int | None:
    if city_id is not None:
        return city_id
    return db.scalar(select(City.id).order_by(City.id.asc()).limit(1))


# Préfixe des biens créés depuis Po2 : ils n'ont pas encore de code ASTECH, c'est le
# logiciel de la collectivité qui l'attribuera au réimport (décision Q13, « lignes à
# créer »). Le réexport devra donc émettre ces lignes avec un CODE_BIEN vide.
NEW_ASSET_CODE_PREFIX = "NOUVEAU_"


def create_asset_from_building(
    db: Session, city_id: int | None, building: Building
) -> PatrimoineLegacyAsset:
    """Ajoute un bâtiment Po2 à la liste ASTECH comme bien **à créer**.

    Cas visé : un bâtiment connu de Po2 mais absent du référentiel de la collectivité.
    Il doit remonter dans le réexport pour y être créé, sinon les deux référentiels ne
    convergeront jamais.

    Idempotent : rappeler la fonction pour le même bâtiment renvoie le bien existant.
    """
    existing = db.scalar(
        select(PatrimoineLegacyAsset).where(
            PatrimoineLegacyAsset.city_id == city_id,
            PatrimoineLegacyAsset.building_id == building.id,
        )
    )
    if existing is not None:
        return existing

    asset = PatrimoineLegacyAsset(
        city_id=city_id,
        code_bien=f"{NEW_ASSET_CODE_PREFIX}{building.id}",
        designation=(building.nom_batiment or f"Bâtiment {building.id}")[:255],
        nomcourt=(building.nom_batiment or f"Bâtiment {building.id}")[:255],
        genre="BATI",
        horsparc="N",
        building_id=building.id,
        status=STATUS_TO_CREATE,
        link_origin=ORIGIN_MANUAL,
        latitude=building.latitude,
        longitude=building.longitude,
    )
    _inherit_building_address(asset, building)
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def create_asset_at_point(
    db: Session, city_id: int | None, *, name: str, latitude: float, longitude: float
) -> PatrimoineLegacyAsset:
    """Crée un bien ASTECH **de toutes pièces**, à un point posé sur la carte.

    Cas visé : un bien que la collectivité n'a ni dans ASTECH ni dans Po2 — on le voit
    sur le terrain ou sur le fond de carte, et il faut qu'il remonte dans le réexport
    pour y être créé. Les deux entrées existantes partent d'une entité Po2 déjà connue
    (`create_asset_from_building`, `create_asset_from_local`) ; il manquait le cas où il
    n'y a rien de préexistant.

    Statut « à créer » : le `CODE_BIEN` sortira **vide** du réexport, c'est ASTECH qui
    l'attribuera (Q13). Le code porté ici n'est qu'une clé interne, préfixée pour être
    reconnaissable et ne jamais être confondue avec un code de la collectivité.

    L'adresse n'est pas devinée : elle se remplira au rattachement à un bâtiment Po2, ou
    par le géocodage inverse si l'utilisateur déplace le point.
    """
    label = (name or "").strip()[:255]
    if not label:
        raise ValueError("Un nom est nécessaire pour créer un bien.")
    asset = PatrimoineLegacyAsset(
        city_id=city_id,
        code_bien=f"{NEW_ASSET_CODE_PREFIX}TMP",
        designation=label,
        nomcourt=label,
        genre="BATI",
        horsparc="N",
        status=STATUS_TO_CREATE,
        link_origin=ORIGIN_MANUAL,
        latitude=latitude,
        longitude=longitude,
    )
    db.add(asset)
    db.flush()
    # La clé interne ne peut être formée qu'une fois l'identifiant attribué.
    asset.code_bien = f"{NEW_ASSET_CODE_PREFIX}P{asset.id}"
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def create_asset_from_local(
    db: Session, city_id: int | None, local: Local
) -> PatrimoineLegacyAsset:
    """Ajoute un **local** Po2 à la liste ASTECH comme bien à créer.

    Même besoin que pour un bâtiment (décision Q13) : une entité connue de Po2 mais
    absente du référentiel de la collectivité doit remonter dans le fichier de retour,
    sinon les deux référentiels ne convergeront jamais. Or un `CODE_BIEN` désigne très
    souvent un local — logement de fonction, salle, WC publics — et seuls les bâtiments
    pouvaient être ajoutés.

    Le préfixe `NOUVEAU_L` distingue ces lignes de celles nées d'un bâtiment
    (`NOUVEAU_`) : deux entités différentes ne doivent pas produire la même clé, la
    contrainte d'unicité portant sur `(city_id, code_bien)`. Le réexport les émet avec
    un `CODE_BIEN` **vide** dans les deux cas — c'est ASTECH qui l'attribuera.

    Idempotent : rappeler la fonction pour le même local renvoie le bien existant.
    """
    existing = db.scalar(
        select(PatrimoineLegacyAsset).where(
            PatrimoineLegacyAsset.city_id == city_id,
            PatrimoineLegacyAsset.local_id == local.id,
        )
    )
    if existing is not None:
        return existing

    building = db.get(Building, local.building_id)
    if building is None:
        raise ValueError("Le bâtiment porteur de ce local est introuvable.")

    name = (local.nom_local or f"Local {local.id}")[:255]
    asset = PatrimoineLegacyAsset(
        city_id=city_id,
        code_bien=f"{NEW_ASSET_CODE_PREFIX}L{local.id}",
        designation=name,
        nomcourt=name,
        genre="BATI",
        horsparc="N",
        building_id=building.id,
        local_id=local.id,
        target_type=TARGET_LOCAL,
        status=STATUS_TO_CREATE,
        link_origin=ORIGIN_MANUAL,
        # Un local sans position propre prend celle de son bâtiment : il est dedans.
        latitude=local.latitude if local.latitude is not None else building.latitude,
        longitude=local.longitude if local.longitude is not None else building.longitude,
    )
    _inherit_building_address(asset, building)
    _override_with_local_address(asset, local)
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def get_local_for_city(db: Session, city_id: int | None, local_id: int) -> Local | None:
    """Le local n'a pas de `city_id` : on passe par son bâtiment porteur."""
    local = db.get(Local, local_id)
    if local is None:
        return None
    if city_id is None:
        return local
    building = db.get(Building, local.building_id)
    return local if building is not None and building.city_id == city_id else None


def get_building_for_city(db: Session, city_id: int | None, building_id: int) -> Building | None:
    statement = select(Building).where(Building.id == building_id)
    if city_id is not None:
        statement = statement.where(Building.city_id == city_id)
    return db.scalar(statement)


def confirm_proposed(
    db: Session, city_id: int | None, asset_ids: list[int] | None = None
) -> dict[str, int]:
    """Valide les rattachements proposes par le moteur.

    Sans `asset_ids`, confirme tout ce qui est en attente : sur 78 propositions, les
    confirmer une par une n'apporterait rien une fois la liste relue.
    """
    statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.status == STATUS_PROPOSED
    )
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)
    if asset_ids:
        statement = statement.where(PatrimoineLegacyAsset.id.in_(asset_ids))

    confirmed = 0
    for asset in db.scalars(statement):
        if asset.building_id is None:
            continue
        asset.status = STATUS_LINKED
        # Le bien prend le nom de sa cible Po2 (Q11) — au moment de la CONFIRMATION, et
        # pas quand le moteur propose : une proposition ne doit pas réécrire le libellé
        # ASTECH avant qu'un humain l'ait validée.
        #
        # Il manquait ici. Un bien rattaché à la main adoptait le nom Po2, un bien
        # confirmé en bloc gardait le sien : constaté le 2026-08-24, 49 biens « liés »
        # portaient encore leur libellé ASTECH. Ce n'était pas cosmétique — c'est
        # `nomcourt`/`designation` que le réexport écrit dans DESIGNATION et NOMCOURT
        # (`patrimoine_legacy_export.py`), donc c'est le nom ASTECH qui serait reparti
        # dans le fichier de la collectivité, à rebours de Q11.
        target = db.get(Building, asset.building_id)
        if target is not None:
            _adopt_target_name(asset, target.nom_batiment, None)
        # Filet de sécurité : c'est le dernier passage avant que le bien ne devienne
        # exportable vers ASTECH. On réaligne l'adresse héritée sur le bâtiment tel
        # qu'il est AUJOURD'HUI, pour ne pas réinjecter un nom ou une adresse périmés.
        _refresh_inherited_address(db, asset)
        db.add(asset)
        confirmed += 1
    db.commit()
    return {"confirmed": confirmed}


# Type des locaux nes d'un bien ASTECH : trace leur origine dans le referentiel Po2,
# a cote de 'PRINCIPAL' (import DGFIP) et 'RECLASSEMENT' (reclassement manuel).
LOCAL_TYPE_FROM_ASTECH = "ASTECH"


def convert_asset_to_local(db: Session, asset: PatrimoineLegacyAsset) -> PatrimoineLegacyAsset:
    """Fait du bien ASTECH un **local** du bâtiment auquel il est rattaché.

    C'était le chaînon manquant : l'écran savait viser un local *déjà existant*, mais
    rien ne permettait d'en créer un. Mesuré en prod le 2026-08-20 : 0 bien sur 79
    rattaché au niveau local, alors que c'est le cas de figure normal dès que plusieurs
    biens ASTECH désignent un même bâtiment (le club et ses salles, l'école et son
    restaurant scolaire).

    Le bâtiment porteur **reste** renseigné : c'est lui qui porte l'adresse, le cadastre
    et la position, et donc ce que le bien renverra à ASTECH (décision Q1 du
    2026-08-20). Passer au niveau local précise la structure sans rien retirer.

    Idempotent : un bien déjà rattaché à un local est renvoyé tel quel.
    """
    if asset.building_id is None:
        raise ValueError(
            "Ce bien n'est rattaché à aucun bâtiment Po2 : rattache-le d'abord, "
            "le local sera créé dans ce bâtiment."
        )
    if asset.target_type == TARGET_LOCAL and asset.local_id is not None:
        return asset

    building = db.get(Building, asset.building_id)
    if building is None:
        raise ValueError("Le bâtiment porteur est introuvable.")

    # Le local prend le nom ASTECH D'ORIGINE du bien, pas son libellé courant.
    #
    # Piège constaté le 2026-08-27 : rattacher d'abord au bâtiment fait adopter le nom
    # du bâtiment (Q11). Créer le local ensuite lui donnait donc le nom du BÂTIMENT —
    # « ECOLE MATERNELLE LOUISE MICHEL » pour ce qui est en réalité le restaurant
    # scolaire. On fabriquait ainsi un local homonyme de son bâtiment, exactement ce que
    # le nettoyage des doublons passe son temps à supprimer, et l'identité du bien
    # disparaissait de l'écran.
    #
    # Le bien reprend ensuite le nom du local : c'est sa cible Po2 réelle, et c'est ce
    # nom-là qui repartira dans ASTECH.
    source = _source_names(asset)
    name = (
        source.get("designation")
        or source.get("nomcourt")
        or asset.nomcourt
        or asset.designation
        or asset.code_bien
    ).strip()[:255]
    # Un local du meme nom existe deja dans ce batiment : on le reutilise plutot que
    # d'en creer un doublon. C'est le cas du TENNIS CLUB DU BARROU, dont le local
    # « SALLE TENNIS CLUB DU BARROU » etait deja en base.
    existing = db.scalar(
        select(Local).where(
            Local.building_id == building.id,
            func.lower(Local.nom_local) == name.lower(),
        )
    )
    local = existing
    if local is None:
        local = Local(
            building_id=building.id,
            nom_local=name,
            type_local=LOCAL_TYPE_FROM_ASTECH,
            # Le local herite de l'adresse et de la position du batiment : un local n'a
            # pas d'adresse propre tant que personne ne lui en donne une, et le laisser
            # vide ferait perdre au bien ce qu'il avait en visant le batiment.
            adresse_reconstituee=building.adresse_reconstituee,
            code_postal=building.code_postal,
            nom_commune=building.nom_commune,
            latitude=building.latitude,
            longitude=building.longitude,
            dgfip_reference_norm=building.dgfip_reference_norm,
        )
        db.add(local)
        db.flush()

    asset.local_id = local.id
    asset.target_type = TARGET_LOCAL
    _adopt_target_name(asset, local.nom_local, None)
    asset.status = STATUS_LINKED
    asset.link_origin = ORIGIN_MANUAL
    _inherit_building_address(asset, building)
    _override_with_local_address(asset, local)
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def set_asset_gone(db: Session, asset: PatrimoineLegacyAsset, gone: bool) -> PatrimoineLegacyAsset:
    """Marque un bien comme **à supprimer de AS-TECH**, ou annule cette consigne (Q23).

    Le bien n'a plus lieu d'être dans le référentiel de la collectivité. L'effacer de Po2
    serait irréversible — revenir sur une suppression à tort imposerait de recharger tout
    le fichier, donc de perdre les rattachements validés. On conserve donc la ligne, avec
    un statut qui la sort du parcours et du réexport.

    Annuler la consigne rend le statut **déduit de l'état réel** du bien : `lie` s'il a
    un bâtiment porteur, `a_traiter` sinon. Le statut antérieur n'est pas mémorisé, ce
    serait une donnée de plus à tenir à jour sans usage.
    """
    if gone:
        asset.status = STATUS_GONE
    elif asset.status == STATUS_GONE:
        asset.status = STATUS_LINKED if asset.building_id is not None else STATUS_TODO
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def _drop_inherited_position(db: Session, asset: PatrimoineLegacyAsset) -> None:
    """Efface la position du bien si elle n'est qu'un **emprunt** au bâtiment porteur.

    Un bien ASTECH n'a pas de position à lui : le fichier de la collectivité n'en porte
    qu'une seule sur 444. Quand un bien s'affiche sur la carte, c'est soit parce qu'il
    est rattaché (il prend le point du bâtiment), soit parce qu'on l'a posé à la main.

    Détacher un bien devait donc lui rendre son absence de position. La version
    précédente faisait l'inverse : elle figeait le point du bâtiment dans le bien pour
    qu'il ne disparaisse pas de la carte. Résultat mesuré en prod le 2026-08-20 :
    **73 des 82 positions** étaient des faux points posés exactement sur d'anciens
    bâtiments, contre ~8 réellement placés à la main.

    On ne garde donc que ce qui a été **déplacé volontairement** : une position qui ne
    coïncide pas avec celle du bâtiment porteur.
    """
    if asset.latitude is None or asset.longitude is None or asset.building_id is None:
        return
    building = db.get(Building, asset.building_id)
    if building is None or building.latitude is None or building.longitude is None:
        return
    # ~1 m : en deçà, la position est celle du bâtiment, pas un choix de l'utilisateur.
    if (
        abs(building.latitude - asset.latitude) < 0.00001
        and abs(building.longitude - asset.longitude) < 0.00001
    ):
        asset.latitude = None
        asset.longitude = None


def reset_everything(db: Session, city_id: int | None) -> dict[str, int]:
    """Remise à zéro **totale** : l'écran revient à l'état juste après l'import.

    Plus fort que `reset_all_links` : on efface aussi les positions posées à la main et
    on annule les décisions « ignoré ». Aucun bien ASTECH ne reste alors sur la carte —
    c'est normal, ils n'ont pas de coordonnées propres — et « Reconnaître les noms » les
    y ramène en les rattachant.

    Ce qui **n'est pas** remis à zéro : `hors_perimetre`. Ce n'est pas une décision de
    l'utilisateur mais un constat de périmètre (bien hors Sète, décision Q4), recalculé
    à chaque import. L'annuler ferait remonter dans la file 27 biens qui n'ont rien à
    y faire.
    """
    statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.status != STATUS_OUT_OF_SCOPE
    )
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)

    reset = 0
    for asset in db.scalars(statement):
        asset.building_id = None
        asset.local_id = None
        asset.target_type = TARGET_BUILDING
        asset.link_origin = None
        asset.latitude = None
        asset.longitude = None
        asset.candidate_building_id = None
        asset.candidate_label = None
        asset.candidate_score = None
        asset.candidate_reason = None
        asset.status = STATUS_TODO
        _clear_resolved_address(asset)
        db.add(asset)
        reset += 1
    db.commit()
    return {"reset": reset}


def delete_all_imports(db: Session, city_id: int | None) -> dict[str, int]:
    """Efface **tout** le référentiel ASTECH importé : les biens et les imports.

    Sert à repartir d'un export neuf. À distinguer de `reset_all_links`, qui ne coupe
    que les rapprochements et garde les biens.

    ⚠️ Destructif : tout le travail de rapprochement part avec les biens — statuts,
    cibles, points posés à la main. C'est à l'écran d'annoncer le compte exact avant
    de le faire.

    Ce qui **reste** délibérément, parce que ce sont désormais des données Po2 et non
    des données ASTECH :
    - les **bâtiments** créés depuis un bien ASTECH (« Créer le bâtiment Po2 ») ;
    - les **locaux** créés depuis un bien ASTECH (« En faire un local »).

    Les uns et les autres seront retrouvés par le moteur de reconnaissance au
    réimport, puisqu'ils portent le nom du bien. Les supprimer ferait perdre du
    patrimoine réel pour effacer un fichier source.
    """
    asset_statement = select(PatrimoineLegacyAsset)
    import_statement = select(PatrimoineLegacyImport)
    if city_id is not None:
        asset_statement = asset_statement.where(PatrimoineLegacyAsset.city_id == city_id)
        import_statement = import_statement.where(PatrimoineLegacyImport.city_id == city_id)

    assets_deleted = 0
    for asset in db.scalars(asset_statement):
        db.delete(asset)
        assets_deleted += 1
    imports_deleted = 0
    for import_row in db.scalars(import_statement):
        db.delete(import_row)
        imports_deleted += 1
    db.commit()
    return {"assets_deleted": assets_deleted, "imports_deleted": imports_deleted}


def reset_all_links(db: Session, city_id: int | None) -> dict[str, int]:
    """Supprime **tous** les rapprochements ASTECH ↔ Po2 et remet les biens à traiter.

    Sert à repartir d'une feuille blanche quand le référentiel Po2 a beaucoup bougé
    (renommages, réimport du patrimoine) : plutôt que de détacher 400 biens un par un,
    on efface et on relance « Reconnaître les noms ».

    Trois choix délibérés :

    - Les biens **`a_creer`** ne sont pas touchés : ils n'existent QUE parce qu'un
      bâtiment Po2 les a créés (décision Q13). Couper leur lien en ferait des lignes
      orphelines sans code ASTECH ni contrepartie.
    - Les décisions **`ignore`** et **`hors_perimetre`** sont conservées : ce ne sont
      pas des rapprochements mais des choix de périmètre, qu'une purge de liens n'a
      aucune raison d'annuler.
    - La **position de travail** (lat/lon) est conservée : c'est souvent un point posé
      à la main, et l'effacer ferait disparaître le bien de la carte — exactement le
      symptôme corrigé en #117.

    L'adresse résolue, elle, est effacée : elle n'était qu'une copie du bâtiment
    porteur, elle n'a plus de source une fois le lien coupé.
    """
    statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.status != STATUS_TO_CREATE,
        or_(
            PatrimoineLegacyAsset.building_id.is_not(None),
            PatrimoineLegacyAsset.local_id.is_not(None),
        ),
    )
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)

    cleared = 0
    for asset in db.scalars(statement):
        _drop_inherited_position(db, asset)
        asset.building_id = None
        asset.local_id = None
        asset.target_type = TARGET_BUILDING
        asset.link_origin = None
        # Une decision de perimetre n'est pas un rapprochement : elle survit a la purge.
        if asset.status not in (STATUS_IGNORED, STATUS_OUT_OF_SCOPE):
            asset.status = STATUS_TODO
        _clear_resolved_address(asset)
        db.add(asset)
        cleared += 1
    db.commit()
    return {"cleared": cleared}


def list_locals_for_building(db: Session, building_id: int) -> list[Local]:
    return list(
        db.scalars(
            select(Local).where(Local.building_id == building_id).order_by(Local.nom_local.asc())
        )
    )


# ---------------------------------------------------------------------------
# Géocodage en masse des biens sans position (§25, décision Q35)
# ---------------------------------------------------------------------------

def _astech_address(asset: PatrimoineLegacyAsset) -> str | None:
    """L'adresse du fichier ASTECH, composée comme à l'écran.

    `NORUE` vaut `0` sur 285 lignes sur 378 : c'est un défaut connu du fichier, pas un
    numéro. On l'écarte plutôt que de géocoder « 0 RUE X », qui ne rend rien.
    """
    parts = [
        asset.source_norue if asset.source_norue and asset.source_norue != "0" else None,
        asset.source_libelvoie,
    ]
    line = " ".join(part for part in parts if part)
    if not line.strip():
        return None
    return f"{line}, {asset.source_ville}" if asset.source_ville else line


def _commune_bounds(db: Session, city_id: int | None) -> tuple[float, float, float, float] | None:
    """Cadre géographique de la commune, déduit des bâtiments Po2 déjà positionnés.

    Sert de garde-fou au géocodage : le géocodeur national retombe sur une adresse
    homonyme ailleurs en France quand il ne trouve rien sur place, **même avec le code
    commune**. Constaté le 2026-08-24 sur 10 des 283 biens posés : `PLACE STALINGRAD`
    atterrissait à Reims, `LE CHATEAU VERT` près d'Aix, et `Le Globe - Ancienne mosquée`
    à Mayotte. Une position fausse est pire qu'une absence de position : elle a l'air
    d'une donnée.

    La marge de 0,15° (~16 km) tient compte de l'étirement d'une commune littorale — Sète
    et son lido font une douzaine de kilomètres — sans laisser passer un département
    voisin. Renvoie `None` s'il n'y a pas de quoi calibrer : on ne refuse pas tout sur la
    foi d'un cadre qu'on n'a pas.
    """
    statement = select(Building.latitude, Building.longitude).where(
        Building.latitude.is_not(None), Building.longitude.is_not(None)
    )
    if city_id is not None:
        statement = statement.where(Building.city_id == city_id)
    points = [(lat, lon) for lat, lon in db.execute(statement)]
    if len(points) < 10:
        return None
    margin = 0.15
    lats = [lat for lat, _ in points]
    lons = [lon for _, lon in points]
    return (min(lats) - margin, max(lats) + margin, min(lons) - margin, max(lons) + margin)


def geocode_pending_assets(
    db: Session, city_id: int | None, city_name: str | None, *, limit: int = 25, offset: int = 0
) -> dict[str, Any]:
    """Pose sur leur adresse ASTECH les biens qui n'ont aucune position.

    Pourquoi en masse : mesure du 2026-08-21 — **294 biens à traiter, dont 2 seulement
    sur la carte**, alors que 292 portent une adresse exploitable. La carte montrait donc
    ce qui était fini et cachait ce qui restait à faire. Le geste existait déjà (« Sur
    l'adresse ASTECH »), mais un par un sur 292 biens.

    Même moteur que le bouton unitaire (`lookup_free_address_candidates`, BAN).

    On s'arrête à `limit` biens par appel : l'écran rappelle jusqu'à épuisement et peut
    afficher l'avancement, là où une seule requête de 292 géocodages expirerait.

    `offset` sert à **passer les échecs déjà constatés**. Un bien introuvable reste sans
    position, donc dans le lot à traiter : sans ce décalage, chaque appel rejouerait les
    mêmes échecs et la boucle ne finirait jamais. L'appelant y met le nombre d'échecs
    cumulés depuis le début.

    Le géocodage inverse du point n'est **pas** rejoué ici, contrairement au bouton
    unitaire : on vient de partir de l'adresse ASTECH, la redemander au point qu'elle a
    produit serait circulaire et doublerait les appels réseau. Les champs résolus se
    remplissent au rattachement, depuis le bâtiment porteur.
    """
    statement = select(PatrimoineLegacyAsset).where(
        PatrimoineLegacyAsset.latitude.is_(None),
        # Un bien DEJA RATTACHE n'a pas a etre geocode : sa position est celle de son
        # batiment, empruntee et non propre (§19, et « les positions ASTECH sont
        # empruntees » — 1 bien sur 444 en porte une dans le fichier). Lui en donner une
        # propre le decolle de son batiment : la carte s'est couverte de traits partant
        # dans tous les sens, un par bien eloigne de sa cible. Constate le 2026-08-24,
        # 53 biens rattaches deplaces par erreur.
        PatrimoineLegacyAsset.building_id.is_(None),
        PatrimoineLegacyAsset.source_libelvoie.is_not(None),
        PatrimoineLegacyAsset.source_libelvoie != "",
        PatrimoineLegacyAsset.status.notin_([STATUS_OUT_OF_SCOPE, STATUS_GONE]),
    )
    if city_id is not None:
        statement = statement.where(PatrimoineLegacyAsset.city_id == city_id)

    # Ordre stable : le decalage des echecs n'a de sens que si le lot ne se reordonne pas
    # d'un appel a l'autre.
    pending = list(db.scalars(statement.order_by(PatrimoineLegacyAsset.code_bien.asc())))
    batch = pending[offset : offset + limit]
    bounds = _commune_bounds(db, city_id)

    positioned = 0
    failures: list[dict[str, str]] = []
    for asset in batch:
        address = _astech_address(asset)
        if not address:
            failures.append({"code_bien": asset.code_bien, "motif": "aucune adresse exploitable"})
            continue
        try:
            lookup = lookup_free_address_candidates(
                address,
                city_name=city_name,
                citycode=asset.source_commune,
                skip_ign_buildings=True,
            )
        except Exception as error:  # le géocodeur est un service externe : jamais bloquant
            failures.append({"code_bien": asset.code_bien, "motif": str(error)[:180]})
            continue
        latitude, longitude = lookup.get("lat"), lookup.get("lon")
        if latitude is None or longitude is None:
            failures.append({"code_bien": asset.code_bien, "motif": f"introuvable : « {address} »"})
            continue
        if bounds is not None and not (
            bounds[0] <= latitude <= bounds[1] and bounds[2] <= longitude <= bounds[3]
        ):
            # Le géocodeur a trouvé une adresse homonyme AILLEURS. On n'écrit pas une
            # position qu'on n'a pas su produire proprement : le bien reste sans point,
            # il se posera à la main.
            failures.append({
                "code_bien": asset.code_bien,
                "motif": f"trouvé hors de la commune ({latitude:.4f}, {longitude:.4f}) : « {address} »",
            })
            continue
        asset.latitude = latitude
        asset.longitude = longitude
        db.add(asset)
        positioned += 1

    db.commit()
    return {
        "traites": len(batch),
        "positionnes": positioned,
        "echecs": failures,
        "restants": max(0, len(pending) - offset - len(batch)),
    }
