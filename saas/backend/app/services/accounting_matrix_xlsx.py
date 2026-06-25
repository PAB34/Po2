"""Import/export XLSX des matrices comptables versionnées (doc 38, doc 35 §5).

Principes :

- l'aller-retour repose sur `stable_rule_key` : cette colonne ne doit jamais
  changer entre un export et un réimport ;
- un import ne modifie jamais une version active : il produit d'abord un
  aperçu de différences (`preview_import`), puis crée une nouvelle version
  brouillon (`commit_import`) que la comptabilité activera explicitement ;
- aucune suppression implicite : une règle présente dans la version de
  référence mais absente du classeur est signalée, pas supprimée.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.accounting_matrix import AccountingMatrixRule, AccountingMatrixVersion
from app.services import accounting_matrix as svc

# Ordre canonique des colonnes (doc 35 §5 + doc 38). `stable_rule_key` en tête.
COLUMNS: list[str] = [
    "stable_rule_key",
    "contract_code",
    "supplier",
    "domain",
    "scope",
    "site_code",
    "building_id",
    "meter_id",
    "billed_item_pattern",
    "supplier_item_code",
    "accounting_service",
    "accounting_function",
    "accounting_antenna",
    "operation_number",
    "accounting_nature",
    "accounting_label",
    "allocation_percent",
    "priority",
    "is_active",
    "comment",
]

# Champs comparés pour détecter une modification (tout sauf la clé stable).
_COMPARABLE_FIELDS = [c for c in COLUMNS if c != "stable_rule_key"]


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
def export_version_xlsx(db: Session, city_id: int | None, version_id: int) -> tuple[bytes, str]:
    """Exporte les règles d'une version en XLSX lisible et réimportable."""
    version = svc._require_version(db, city_id, version_id)
    contract = version.contract

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrice"
    ws.append(COLUMNS)
    for rule in version.rules:
        ws.append([_rule_cell(rule, contract, col) for col in COLUMNS])

    _append_readme(wb, contract, version)

    buf = BytesIO()
    wb.save(buf)
    filename = f"matrice_{contract.supplier}_{contract.contract_code or 'contrat'}_{version.id}.xlsx"
    return buf.getvalue(), _safe_filename(filename)


def _rule_cell(rule: AccountingMatrixRule, contract, col: str):
    if col == "contract_code":
        return contract.contract_code
    if col == "supplier":
        return contract.supplier
    if col == "domain":
        return contract.domain
    if col == "is_active":
        return "oui" if rule.is_active else "non"
    return getattr(rule, col, None)


def _append_readme(wb: Workbook, contract, version: AccountingMatrixVersion) -> None:
    ws = wb.create_sheet("Lisez-moi")
    ws.append(["Matrice comptable versionnée — fichier d'échange"])
    ws.append([])
    ws.append(["Contrat", contract.contract_label or contract.contract_code or ""])
    ws.append(["Fournisseur", contract.supplier])
    ws.append(["Domaine", contract.domain])
    ws.append(["Version", version.version_label])
    ws.append(["Statut version", version.status])
    ws.append([])
    ws.append(["Règle d'or : ne jamais modifier la colonne stable_rule_key."])
    ws.append(["Un réimport ne modifie jamais une version active : il crée une nouvelle version brouillon."])
    ws.append([])
    ws.append(["Colonne", "Rôle"])
    for label, role in _COLUMN_HELP:
        ws.append([label, role])


_COLUMN_HELP: list[tuple[str, str]] = [
    ("stable_rule_key", "Identifiant stable de règle (obligatoire, ne pas changer)"),
    ("contract_code", "Code contrat / lot (informationnel, défini par le contrat)"),
    ("supplier", "Fournisseur / prestataire (défini par le contrat)"),
    ("domain", "fluides, cpe, maintenance, travaux..."),
    ("scope", "site, meter, billed_item, subscription, tax, p1, p2, p3, other"),
    ("site_code", "Code site comptable si applicable"),
    ("building_id", "Identifiant bâtiment patrimoine si connu"),
    ("meter_id", "PRM / PCE / compteur eau si applicable"),
    ("billed_item_pattern", "Libellé ligne facture ou motif normalisé"),
    ("supplier_item_code", "Code fournisseur de la ligne si disponible"),
    ("accounting_service", "Axe comptable : service"),
    ("accounting_function", "Axe comptable : fonction"),
    ("accounting_antenna", "Axe comptable : antenne"),
    ("operation_number", "Numéro d'opération budgétaire"),
    ("accounting_nature", "Nature comptable"),
    ("accounting_label", "Libellé de la nature"),
    ("allocation_percent", "Pourcentage de ventilation (0 à 100, 100 par défaut)"),
    ("priority", "Priorité d'arbitrage si plusieurs règles matchent"),
    ("is_active", "oui / non"),
    ("comment", "Commentaire comptabilité"),
]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def parse_xlsx(raw: bytes) -> tuple[list[dict], list[str]]:
    """Lit le classeur et retourne (lignes normalisées, erreurs structurelles)."""
    errors: list[str] = []
    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — message utilisateur
        return [], [f"Fichier illisible : {exc}"]

    ws = wb["Matrice"] if "Matrice" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Classeur vide."]

    header_map = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    missing = [c for c in ("stable_rule_key", "scope", "accounting_nature", "allocation_percent") if c not in header_map]
    if missing:
        errors.append(f"Colonnes obligatoires absentes : {', '.join(missing)}.")
        return [], errors

    rows: list[dict] = []
    for line_no, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        rows.append(_row_to_dict(raw_row, header_map, line_no))
    return rows, errors


def _row_to_dict(raw_row, header_map: dict[str, int], line_no: int) -> dict:
    def cell(col: str):
        idx = header_map.get(col)
        if idx is None or idx >= len(raw_row):
            return None
        value = raw_row[idx]
        if isinstance(value, str):
            value = value.strip()
        return value if value not in ("", None) else None

    return {
        "_line": line_no,
        "stable_rule_key": _as_str(cell("stable_rule_key")),
        "scope": _as_str(cell("scope")) or "billed_item",
        "site_code": _as_str(cell("site_code")),
        "building_id": _as_int(cell("building_id")),
        "meter_id": _as_str(cell("meter_id")),
        "billed_item_pattern": _as_str(cell("billed_item_pattern")),
        "supplier_item_code": _as_str(cell("supplier_item_code")),
        "accounting_service": _as_str(cell("accounting_service")),
        "accounting_function": _as_str(cell("accounting_function")),
        "accounting_antenna": _as_str(cell("accounting_antenna")),
        "operation_number": _as_str(cell("operation_number")),
        "accounting_nature": _as_str(cell("accounting_nature")),
        "accounting_label": _as_str(cell("accounting_label")),
        "allocation_percent": _as_float(cell("allocation_percent")),
        "priority": _as_int(cell("priority")) or 0,
        "is_active": _as_bool(cell("is_active")),
        "comment": _as_str(cell("comment")),
    }


# ---------------------------------------------------------------------------
# Aperçu de différences (sans écriture)
# ---------------------------------------------------------------------------
def preview_import(db: Session, city_id: int | None, contract_id: int, raw: bytes) -> dict:
    contract = svc._require_contract(db, city_id, contract_id)
    rows, structural_errors = parse_xlsx(raw)

    reference = _reference_version(contract)
    ref_rules = {r.stable_rule_key: r for r in (reference.rules if reference else [])}

    summary = {"ajout": 0, "modifie": 0, "inchange": 0, "absentes_du_fichier": 0, "erreurs": 0}
    diff_rows: list[dict] = []
    seen_keys: set[str] = set()

    for row in rows:
        status, message = _classify_row(row, ref_rules, seen_keys)
        summary[status] = summary.get(status, 0) + 1
        diff_rows.append({
            "line": row["_line"],
            "stable_rule_key": row["stable_rule_key"],
            "status": status,
            "message": message,
        })
        if row["stable_rule_key"]:
            seen_keys.add(row["stable_rule_key"])

    absentes = [k for k in ref_rules if k not in seen_keys]
    summary["absentes_du_fichier"] = len(absentes)

    return {
        "contract_id": contract.id,
        "reference_version_id": reference.id if reference else None,
        "reference_version_label": reference.version_label if reference else None,
        "structural_errors": structural_errors,
        "summary": summary,
        "rows": diff_rows,
        "absentes_du_fichier": absentes,
        "can_commit": not structural_errors and summary["erreurs"] == 0,
        "warnings": _ventilation_warnings(rows),
    }


def _classify_row(row: dict, ref_rules: dict, seen_keys: set[str]) -> tuple[str, str | None]:
    key = row["stable_rule_key"]
    if not key:
        return "erreurs", "stable_rule_key manquant."
    if key in seen_keys:
        return "erreurs", "stable_rule_key en doublon dans le fichier."
    pct = row["allocation_percent"]
    if pct is None or pct < 0 or pct > 100:
        return "erreurs", "allocation_percent doit être entre 0 et 100."
    if not row["accounting_nature"]:
        return "erreurs", "accounting_nature obligatoire."

    ref = ref_rules.get(key)
    if ref is None:
        return "ajout", None
    if _row_differs(row, ref):
        return "modifie", None
    return "inchange", None


def _row_differs(row: dict, ref: AccountingMatrixRule) -> bool:
    for field in _COMPARABLE_FIELDS:
        # supplier/domain/contract_code sont portés par le contrat, pas la règle.
        if field in ("supplier", "domain", "contract_code"):
            continue
        if _norm(row.get(field)) != _norm(getattr(ref, field, None)):
            return True
    return False


def _ventilation_warnings(rows: list[dict]) -> list[str]:
    """Signale les groupes dont la somme des ventilations diffère de 100 %.

    Groupe = (scope, site_code, meter_id, billed_item_pattern). Non bloquant
    ici ; le contrôle dur appartient à la phase d'application des snapshots.
    """
    groups: dict[tuple, float] = {}
    counts: dict[tuple, int] = {}
    for row in rows:
        if not row["stable_rule_key"] or row["allocation_percent"] is None:
            continue
        gkey = (row["scope"], row["site_code"], row["meter_id"], row["billed_item_pattern"])
        groups[gkey] = groups.get(gkey, 0.0) + row["allocation_percent"]
        counts[gkey] = counts.get(gkey, 0) + 1
    warnings: list[str] = []
    for gkey, total in groups.items():
        if counts[gkey] > 1 and abs(total - 100.0) > 0.01:
            warnings.append(f"Ventilation {gkey} = {round(total, 2)} % (≠ 100 %).")
    return warnings


# ---------------------------------------------------------------------------
# Commit : nouvelle version brouillon depuis le classeur
# ---------------------------------------------------------------------------
def commit_import(
    db: Session, city_id: int | None, contract_id: int, raw: bytes,
    *, version_label: str, user_id: int | None,
) -> dict:
    contract = svc._require_contract(db, city_id, contract_id)
    rows, structural_errors = parse_xlsx(raw)
    if structural_errors:
        raise ValueError(" ".join(structural_errors))

    reference = _reference_version(contract)
    ref_rules = {r.stable_rule_key: r for r in (reference.rules if reference else [])}
    seen: set[str] = set()
    for row in rows:
        status, message = _classify_row(row, ref_rules, seen)
        if status == "erreurs":
            raise ValueError(f"Ligne {row['_line']} : {message}")
        if row["stable_rule_key"]:
            seen.add(row["stable_rule_key"])

    version = AccountingMatrixVersion(
        matrix_contract_id=contract.id,
        version_label=version_label,
        status="draft",
        source="import_xlsx",
        created_by_user_id=user_id,
    )
    db.add(version)
    db.flush()

    for row in rows:
        db.add(AccountingMatrixRule(
            matrix_version_id=version.id,
            stable_rule_key=row["stable_rule_key"],
            scope=row["scope"],
            site_code=row["site_code"],
            building_id=row["building_id"],
            meter_id=row["meter_id"],
            billed_item_pattern=row["billed_item_pattern"],
            supplier_item_code=row["supplier_item_code"],
            accounting_service=row["accounting_service"],
            accounting_function=row["accounting_function"],
            accounting_antenna=row["accounting_antenna"],
            operation_number=row["operation_number"],
            accounting_nature=row["accounting_nature"],
            accounting_label=row["accounting_label"],
            allocation_percent=row["allocation_percent"],
            priority=row["priority"],
            is_active=row["is_active"],
            comment=row["comment"],
        ))

    db.commit()
    db.refresh(version)
    return svc._version_out(version)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _reference_version(contract) -> AccountingMatrixVersion | None:
    """Version de référence pour le diff : l'active, sinon la plus récente."""
    active = svc._active_version(contract)
    if active:
        return active
    return contract.versions[-1] if contract.versions else None


def _norm(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _as_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _as_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _as_bool(value) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    return text not in ("non", "no", "false", "0", "n", "")


def _safe_filename(name: str) -> str:
    keep = "-_."
    cleaned = "".join(c if (c.isalnum() or c in keep) else "_" for c in name)
    return cleaned or "matrice.xlsx"
