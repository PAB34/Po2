"""Matrice comptable ENGIE + fiche de liaison finances.

Calqué sur `services/cpe_accounting.py` (DALKIA) mais autonome :
- `EnergyAccountingSiteMapping` : PRM -> codes analytiques (service/fonction/antenne/opération).
- `EnergyAccountingNatureRule` : poste facturé -> nature comptable.

Fonctions :
- `import_codification_workbook` : importe un xlsx (mêmes onglets que le fichier DALKIA) en upsert.
- CRUD site-mappings / nature-rules.
- `bootstrap_site_mappings_from_invoices` : pré-remplit la matrice depuis les PRM des factures.
- `resolve_line_codification` : ligne facture -> (codes analytiques, nature).
- `build_energy_liaison_workbook` : fiche de liaison finances (xlsx) pour une facture.
"""
from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.invoice import (
    EnergyAccountingNatureRule,
    EnergyAccountingSiteMapping,
    EnergyInvoiceImport,
)


# ---------------------------------------------------------------------------
# Helpers de normalisation (repris du pattern cpe_accounting)
# ---------------------------------------------------------------------------


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _upper(value: Any) -> str | None:
    text = _clean(value)
    return text.upper() if text else None


def _norm_header(value: Any) -> str:
    raw = "" if value is None else str(value).strip()
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "_", ascii_value.lower()).strip("_")


def _rows_from_sheet(ws, header_row: int) -> list[dict[str, Any]]:
    headers = [_norm_header(cell.value) for cell in ws[header_row]]
    rows: list[dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in excel_row):
            continue
        rows.append({headers[index]: value for index, value in enumerate(excel_row) if index < len(headers)})
    return rows


def _first(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


# ---------------------------------------------------------------------------
# Import du classeur de codification
# ---------------------------------------------------------------------------


@dataclass
class CodificationImportResult:
    filename: str | None = None
    nature_rules_created: int = 0
    nature_rules_updated: int = 0
    site_mappings_created: int = 0
    site_mappings_updated: int = 0
    errors: list[str] = field(default_factory=list)


# Onglets reconnus (tolérant : on cherche par sous-chaîne normalisée du titre).
_SITE_SHEET_HINTS = ("sites_vers_codes", "sites", "prm")
_NATURE_SHEET_HINTS = ("poste_facture_vers_nature", "postes_x_contrat_x_nature", "nature")


def _find_sheet(wb, hints: tuple[str, ...]) -> Any | None:
    norm = {name: _norm_header(name) for name in wb.sheetnames}
    # match exact prioritaire
    for name, n in norm.items():
        if n in hints:
            return wb[name]
    for name, n in norm.items():
        if any(h in n for h in hints):
            return wb[name]
    return None


def import_codification_workbook(
    db: Session,
    raw_bytes: bytes,
    *,
    filename: str | None,
    city_id: int | None,
) -> CodificationImportResult:
    """Importe la matrice comptable ENGIE (xlsx) en upsert."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    result = CodificationImportResult(filename=filename)

    # --- Sites vers codes (clé = PRM) ---
    site_ws = _find_sheet(wb, _SITE_SHEET_HINTS)
    if site_ws is not None:
        # En-tête sur la ligne 1 par défaut ; certains fichiers ont 1 ligne de titre.
        header_row = 1
        for candidate in (1, 2, 3):
            headers = [_norm_header(c.value) for c in site_ws[candidate]]
            if any(h in headers for h in ("prm", "pdl", "point_de_livraison", "code_site")):
                header_row = candidate
                break
        for row in _rows_from_sheet(site_ws, header_row):
            prm = _clean(_first(row, "prm", "pdl", "point_de_livraison", "code_site", "code_prm"))
            if not prm:
                continue
            payload = {
                "city_id": city_id,
                "prm_id": prm,
                "site_name": _clean(_first(row, "nom_du_site", "nom_site", "site", "libelle_site")),
                "regroupement": _clean(_first(row, "regroupement", "groupe")),
                "family": _clean(_first(row, "famille")),
                "manager": _clean(_first(row, "gestionnaire")),
                "alternate_manager": _clean(_first(row, "gestionnaire_alternatif")),
                "service_code": _clean(_first(row, "service")),
                "service_label": _clean(_first(row, "libelle_service")),
                "function_code": _clean(_first(row, "fonction")),
                "function_label": _clean(_first(row, "libelle_fonction")),
                "antenna_code": _clean(_first(row, "antenne")),
                "antenna_label": _clean(_first(row, "libelle_antenne")),
                "operation_code": _clean(_first(row, "operation_si_travaux", "operation")),
                "operation_label": _clean(_first(row, "libelle_operation")),
                "active": True,
            }
            existing = db.scalars(
                select(EnergyAccountingSiteMapping).where(
                    EnergyAccountingSiteMapping.city_id == city_id,
                    EnergyAccountingSiteMapping.prm_id == prm,
                )
            ).first()
            if existing:
                for key, value in payload.items():
                    setattr(existing, key, value)
                result.site_mappings_updated += 1
            else:
                db.add(EnergyAccountingSiteMapping(**payload))
                result.site_mappings_created += 1
    else:
        result.errors.append("Onglet 'Sites vers codes' (clé PRM) introuvable.")

    # --- Poste facturé vers Nature ---
    nature_ws = _find_sheet(wb, _NATURE_SHEET_HINTS)
    if nature_ws is not None:
        header_row = 1
        for candidate in (1, 2, 3):
            headers = [_norm_header(c.value) for c in nature_ws[candidate]]
            if any(h in headers for h in ("poste_facture", "poste", "libelle", "nature_proposee", "nature")):
                header_row = candidate
                break
        for row in _rows_from_sheet(nature_ws, header_row):
            billed_item = _upper(_first(row, "poste_facture", "poste", "code_poste", "libelle"))
            nature = _clean(_first(row, "nature_proposee", "nature", "nature_ctpab"))
            if not billed_item or not nature:
                continue
            if _upsert_nature_rule(
                db,
                city_id=city_id,
                supplier="ENGIE",
                market=_clean(_first(row, "marche")),
                billed_item=billed_item,
                frequency=_clean(_first(row, "frequence", "nb_lignes")),
                accounting_nature=nature,
                accounting_label=_clean(_first(row, "libelle_nature", "signification")),
            ):
                result.nature_rules_created += 1
            else:
                result.nature_rules_updated += 1
    else:
        result.errors.append("Onglet 'Poste facturé vers Nature' introuvable.")

    db.commit()
    return result


def _upsert_nature_rule(
    db: Session,
    *,
    city_id: int | None,
    supplier: str,
    market: str | None,
    billed_item: str,
    frequency: str | None,
    accounting_nature: str,
    accounting_label: str | None,
) -> bool:
    """Retourne True si créé, False si mis à jour."""
    from sqlalchemy import func as safunc

    existing = db.scalars(
        select(EnergyAccountingNatureRule).where(
            EnergyAccountingNatureRule.city_id == city_id,
            EnergyAccountingNatureRule.supplier == supplier,
            safunc.coalesce(EnergyAccountingNatureRule.market, "") == (market or ""),
            EnergyAccountingNatureRule.billed_item == billed_item,
            safunc.coalesce(EnergyAccountingNatureRule.frequency, "") == (frequency or ""),
        )
    ).first()
    payload = {
        "city_id": city_id,
        "supplier": supplier,
        "market": market,
        "billed_item": billed_item,
        "frequency": frequency,
        "accounting_nature": accounting_nature,
        "accounting_label": accounting_label,
        "active": True,
    }
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        return False
    db.add(EnergyAccountingNatureRule(**payload))
    return True


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------


def list_site_mappings(db: Session, city_id: int | None) -> list[EnergyAccountingSiteMapping]:
    query = select(EnergyAccountingSiteMapping)
    if city_id is not None:
        query = query.where(EnergyAccountingSiteMapping.city_id == city_id)
    return list(db.scalars(query.order_by(EnergyAccountingSiteMapping.prm_id)).all())


def list_nature_rules(db: Session, city_id: int | None) -> list[EnergyAccountingNatureRule]:
    query = select(EnergyAccountingNatureRule)
    if city_id is not None:
        query = query.where(EnergyAccountingNatureRule.city_id == city_id)
    return list(db.scalars(query.order_by(EnergyAccountingNatureRule.billed_item)).all())


def create_site_mapping(db: Session, data: dict) -> EnergyAccountingSiteMapping:
    obj = EnergyAccountingSiteMapping(**data)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def update_site_mapping(db: Session, obj: EnergyAccountingSiteMapping, data: dict) -> EnergyAccountingSiteMapping:
    for key, value in data.items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj


def delete_site_mapping(db: Session, obj: EnergyAccountingSiteMapping) -> None:
    db.delete(obj)
    db.commit()


def create_nature_rule(db: Session, data: dict) -> EnergyAccountingNatureRule:
    obj = EnergyAccountingNatureRule(**data)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def update_nature_rule(db: Session, obj: EnergyAccountingNatureRule, data: dict) -> EnergyAccountingNatureRule:
    for key, value in data.items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj


def delete_nature_rule(db: Session, obj: EnergyAccountingNatureRule) -> None:
    db.delete(obj)
    db.commit()


def get_site_mapping(db: Session, mapping_id: int, city_id: int | None) -> EnergyAccountingSiteMapping | None:
    obj = db.get(EnergyAccountingSiteMapping, mapping_id)
    if obj is None or (city_id is not None and obj.city_id != city_id):
        return None
    return obj


def get_nature_rule(db: Session, rule_id: int, city_id: int | None) -> EnergyAccountingNatureRule | None:
    obj = db.get(EnergyAccountingNatureRule, rule_id)
    if obj is None or (city_id is not None and obj.city_id != city_id):
        return None
    return obj


# ---------------------------------------------------------------------------
# Pré-remplissage depuis les factures
# ---------------------------------------------------------------------------


def bootstrap_site_mappings_from_invoices(db: Session, city_id: int) -> dict[str, int]:
    """Crée une ligne de matrice (PRM) vide pour chaque PRM vu dans les factures ENGIE."""
    from app.models.invoice import EnergyInvoice, EnergyInvoiceSite

    existing = {
        m.prm_id
        for m in db.scalars(
            select(EnergyAccountingSiteMapping).where(EnergyAccountingSiteMapping.city_id == city_id)
        ).all()
    }
    seen: dict[str, dict] = {}
    sites = db.scalars(
        select(EnergyInvoiceSite)
        .join(EnergyInvoice, EnergyInvoiceSite.invoice_id == EnergyInvoice.id)
        .where(EnergyInvoice.city_id == city_id)
    ).all()
    for site in sites:
        prm = _clean(site.prm_id)
        if not prm or prm in existing or prm in seen:
            continue
        seen[prm] = {
            "city_id": city_id,
            "prm_id": prm,
            "site_name": _clean(site.site_name),
            "regroupement": _clean(site.regroupement),
            "active": True,
        }
    for payload in seen.values():
        db.add(EnergyAccountingSiteMapping(**payload))
    db.commit()
    return {"created": len(seen), "existing": len(existing)}


# ---------------------------------------------------------------------------
# Résolution de la codification d'une ligne + fiche de liaison
# ---------------------------------------------------------------------------


def _index_site_mappings(db: Session, city_id: int) -> dict[str, EnergyAccountingSiteMapping]:
    return {
        m.prm_id: m
        for m in db.scalars(
            select(EnergyAccountingSiteMapping).where(EnergyAccountingSiteMapping.city_id == city_id)
        ).all()
    }


def _index_nature_rules(db: Session, city_id: int) -> dict[str, EnergyAccountingNatureRule]:
    """Indexe les règles par billed_item normalisé (upper)."""
    index: dict[str, EnergyAccountingNatureRule] = {}
    for rule in db.scalars(
        select(EnergyAccountingNatureRule).where(
            EnergyAccountingNatureRule.city_id == city_id,
            EnergyAccountingNatureRule.active.is_(True),
        )
    ).all():
        index.setdefault((rule.billed_item or "").upper(), rule)
    return index


def _resolve_nature(rules_index: dict[str, EnergyAccountingNatureRule], line) -> EnergyAccountingNatureRule | None:
    for candidate in (line.normalized_code, line.poste, line.label, line.family):
        key = (candidate or "").upper().strip()
        if key and key in rules_index:
            return rules_index[key]
    return None


@dataclass
class LiaisonRow:
    prm_id: str | None
    site_name: str | None
    poste: str | None
    label: str | None
    quantity: float | None
    unit_price_ht: float | None
    amount_ht: float | None
    service_code: str | None
    function_code: str | None
    antenna_code: str | None
    operation_code: str | None
    accounting_nature: str | None
    accounting_label: str | None
    status: str  # ok | blocked (PRM ou poste non codifié)


def resolve_invoice_codification(db: Session, invoice_import: EnergyInvoiceImport) -> list[LiaisonRow]:
    """Construit la codification ligne à ligne d'une facture (sans écrire)."""
    city_id = invoice_import.city_id
    sites_index = _index_site_mappings(db, city_id)
    rules_index = _index_nature_rules(db, city_id)
    rows: list[LiaisonRow] = []

    invoice = invoice_import.normalized_invoice
    if invoice is None:
        return rows
    for site in invoice.sites:
        mapping = sites_index.get(_clean(site.prm_id) or "")
        for period in site.periods:
            for line in period.lines:
                rule = _resolve_nature(rules_index, line)
                blocked = mapping is None or rule is None
                rows.append(
                    LiaisonRow(
                        prm_id=site.prm_id,
                        site_name=(mapping.site_name if mapping else None) or site.site_name,
                        poste=line.poste or line.normalized_code,
                        label=line.label,
                        quantity=line.quantity,
                        unit_price_ht=line.unit_price_ht,
                        amount_ht=line.amount_ht,
                        service_code=mapping.service_code if mapping else None,
                        function_code=mapping.function_code if mapping else None,
                        antenna_code=mapping.antenna_code if mapping else None,
                        operation_code=mapping.operation_code if mapping else None,
                        accounting_nature=rule.accounting_nature if rule else None,
                        accounting_label=rule.accounting_label if rule else None,
                        status="blocked" if blocked else "ok",
                    )
                )
    return rows


def build_energy_liaison_workbook(db: Session, invoice_import: EnergyInvoiceImport) -> bytes:
    """Construit la fiche de liaison finances XLSX pour une facture ENGIE."""
    rows = resolve_invoice_codification(db, invoice_import)
    invoice = invoice_import.normalized_invoice

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiche liaison"
    ws["A1"] = "Fiche de liaison finance ENGIE"
    ws["A1"].font = Font(bold=True, size=15)
    meta = [
        ("Facture", invoice_import.invoice_number or (invoice.invoice_number if invoice else None)),
        ("Fournisseur", (invoice.supplier if invoice else None) or invoice_import.supplier_guess or "ENGIE"),
        ("Période", f"{invoice_import.period_start or '-'} au {invoice_import.period_end or '-'}"),
        ("Total HT", invoice.total_ht if invoice else None),
        ("Total TTC", invoice_import.total_ttc),
        ("Statut contrôle", invoice_import.control_status),
        ("Décision", invoice_import.decision_status),
        ("Commentaire décision", invoice_import.decision_comment),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    headers = [
        "PRM", "Nom site", "Poste", "Libellé", "Quantité", "PU HT", "Montant HT",
        "Service", "Fonction", "Antenne", "Opération", "Nature", "Libellé nature", "Codification",
    ]
    start_row = 13
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for idx, r in enumerate(rows, start=start_row + 1):
        values = [
            r.prm_id, r.site_name, r.poste, r.label, r.quantity, r.unit_price_ht, r.amount_ht,
            r.service_code, r.function_code, r.antenna_code, r.operation_code,
            r.accounting_nature, r.accounting_label,
            "OK" if r.status == "ok" else "À CODIFIER",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=idx, column=col, value=value)
        ws.cell(row=idx, column=7).number_format = '#,##0.00 "€"'

    widths = [16, 30, 16, 36, 12, 12, 14, 12, 12, 14, 14, 12, 24, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    ws.freeze_panes = f"A{start_row + 1}"
    if rows:
        ws.auto_filter.ref = f"A{start_row}:N{start_row + len(rows)}"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
