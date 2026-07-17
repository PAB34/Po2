"""Service layer for CPE DALKIA OS / avenant preparation dossiers."""
from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.cpe_contract_change import CpeContractChangeLine, CpeContractChangeRequest
from app.models.cpe_dalkia import (
    CpeDalkiaRefImport,
    CpeDalkiaRefP1Elec,
    CpeDalkiaRefP1Gaz,
    CpeDalkiaRefP2P3,
    CpeDalkiaRefSite,
)
from app.schemas.cpe_os_avenant import CpeOsAvenantRequestCreate, CpeOsAvenantRequestUpdate

CONTRACT_END_YEAR = 2033
CONTRACT_END_DATE = date(2033, 10, 12)
VALID_STATUSES = {
    "draft",
    "sent_to_dalkia",
    "dalkia_completed",
    "pending_collectivity_validation",
    "os_ready",
    "os_signed",
    "in_service",
    "included_in_avenant",
    "cancelled",
}


def _city_filter(query, model, city_id: int | None):
    if city_id is None:
        return query
    return query.where(model.city_id == city_id)


def _active_import_ids(db: Session, city_id: int | None, lot: int | None = None) -> list[int]:
    query = select(CpeDalkiaRefImport.id).where(CpeDalkiaRefImport.is_active.is_(True))
    query = _city_filter(query, CpeDalkiaRefImport, city_id)
    if lot is not None:
        query = query.where(CpeDalkiaRefImport.lot == lot)
    return list(db.scalars(query).all())


def _sum_for_site(db: Session, model, column_name: str, *, import_ids: list[int], code_site: str, year: int) -> float:
    if not import_ids:
        return 0.0
    value = db.scalar(
        select(func.coalesce(func.sum(getattr(model, column_name)), 0.0)).where(
            model.import_id.in_(import_ids),
            model.code_site == code_site,
            model.period_year == year,
        )
    )
    return float(value or 0.0)


def _has_site_year(db: Session, model, *, import_ids: list[int], code_site: str, year: int) -> bool:
    if not import_ids:
        return False
    return bool(
        db.scalar(
            select(func.count(model.id)).where(
                model.import_id.in_(import_ids),
                model.code_site == code_site,
                model.period_year == year,
            )
        )
    )


def _p1_gaz_lines_for_site(db: Session, *, import_ids: list[int], code_site: str, year: int) -> list[dict[str, Any]]:
    if not import_ids:
        return []
    rows = db.scalars(
        select(CpeDalkiaRefP1Gaz)
        .where(
            CpeDalkiaRefP1Gaz.import_id.in_(import_ids),
            CpeDalkiaRefP1Gaz.code_site == code_site,
            CpeDalkiaRefP1Gaz.period_year == year,
        )
        .order_by(CpeDalkiaRefP1Gaz.pce, CpeDalkiaRefP1Gaz.id)
    ).all()
    return [
        {
            "pce": row.pce,
            "type_tarif": row.type_tarif,
            "prix_unitaire_ht": row.prix_unitaire_ht,
            "atrd_ht": row.atrd_ht,
            "cta_ht": row.cta_ht,
            "p10_fixe_ht": row.p10_fixe_ht,
            "qt_mwhpcs": row.qt_mwhpcs,
            "p10_var_ht": row.p10_var_ht,
            "p10_total_ht": row.p10_total_ht,
        }
        for row in rows
    ]


def _p1_elec_lines_for_site(db: Session, *, import_ids: list[int], code_site: str, year: int) -> list[dict[str, Any]]:
    if not import_ids:
        return []
    rows = db.scalars(
        select(CpeDalkiaRefP1Elec)
        .where(
            CpeDalkiaRefP1Elec.import_id.in_(import_ids),
            CpeDalkiaRefP1Elec.code_site == code_site,
            CpeDalkiaRefP1Elec.period_year == year,
        )
        .order_by(CpeDalkiaRefP1Elec.pdl, CpeDalkiaRefP1Elec.id)
    ).all()
    return [
        {
            "pdl": row.pdl,
            "prix_unitaire_ht": row.prix_unitaire_ht,
            "qt_mwh": row.qt_mwh,
            "p10_var_ht": row.p10_var_ht,
            "p10_total_ht": row.p10_total_ht,
        }
        for row in rows
    ]


def _p2p3_detail_for_site(db: Session, *, import_ids: list[int], code_site: str, year: int) -> dict[str, float]:
    if not import_ids:
        return {}
    columns = (
        "p2_1_ht",
        "p2_2_ht",
        "p2_3_ht",
        "p2_4_ht",
        "p2_total_ht",
        "p3_1_ht",
        "p3_2_ht",
        "p3_3_ht",
        "p3_4_ht",
        "p3_total_ht",
    )
    query = select(*[func.coalesce(func.sum(getattr(CpeDalkiaRefP2P3, column)), 0.0) for column in columns]).where(
        CpeDalkiaRefP2P3.import_id.in_(import_ids),
        CpeDalkiaRefP2P3.code_site == code_site,
        CpeDalkiaRefP2P3.period_year == year,
    )
    values = db.execute(query).one()
    return {column: float(value or 0.0) for column, value in zip(columns, values)}


def reference_for_site(
    db: Session,
    city_id: int | None,
    *,
    code_site: str,
    year: int,
    lot: int | None = None,
) -> dict[str, Any]:
    import_ids = _active_import_ids(db, city_id, lot)
    site = None
    if import_ids:
        site = db.scalars(
            select(CpeDalkiaRefSite)
            .where(CpeDalkiaRefSite.import_id.in_(import_ids), CpeDalkiaRefSite.code_site == code_site)
            .order_by(CpeDalkiaRefSite.id.desc())
        ).first()
    gaz = db.scalars(
        select(CpeDalkiaRefP1Gaz)
        .where(
            CpeDalkiaRefP1Gaz.import_id.in_(import_ids) if import_ids else False,
            CpeDalkiaRefP1Gaz.code_site == code_site,
            CpeDalkiaRefP1Gaz.period_year == year,
        )
        .order_by(CpeDalkiaRefP1Gaz.id.desc())
    ).first() if import_ids else None
    return {
        "code_site": code_site,
        "site_name": site.nom_batiment if site else code_site,
        "lot": site.lot if site else lot,
        "source_year": year,
        "pce": gaz.pce if gaz else None,
        "tarif": gaz.type_tarif if gaz else None,
        "p1_gaz_annual_ht": _sum_for_site(db, CpeDalkiaRefP1Gaz, "p10_total_ht", import_ids=import_ids, code_site=code_site, year=year),
        "p1_elec_annual_ht": _sum_for_site(db, CpeDalkiaRefP1Elec, "p10_total_ht", import_ids=import_ids, code_site=code_site, year=year),
        "p2_annual_ht": _sum_for_site(db, CpeDalkiaRefP2P3, "p2_total_ht", import_ids=import_ids, code_site=code_site, year=year),
        "p3_annual_ht": _sum_for_site(db, CpeDalkiaRefP2P3, "p3_total_ht", import_ids=import_ids, code_site=code_site, year=year),
        "p1_gaz_lines": _p1_gaz_lines_for_site(db, import_ids=import_ids, code_site=code_site, year=year),
        "p1_elec_lines": _p1_elec_lines_for_site(db, import_ids=import_ids, code_site=code_site, year=year),
        "p2p3_detail": _p2p3_detail_for_site(db, import_ids=import_ids, code_site=code_site, year=year),
    }


def list_site_options(db: Session, city_id: int | None, *, year: int, lot: int | None = None) -> list[dict[str, Any]]:
    import_ids = _active_import_ids(db, city_id, lot)
    if not import_ids:
        return []
    query = select(CpeDalkiaRefSite).where(CpeDalkiaRefSite.import_id.in_(import_ids))
    if lot is not None:
        query = query.where(CpeDalkiaRefSite.lot == lot)
    sites = db.scalars(query.order_by(CpeDalkiaRefSite.lot, CpeDalkiaRefSite.code_site)).all()
    options = []
    seen: set[str] = set()
    for site in sites:
        if site.code_site in seen:
            continue
        seen.add(site.code_site)
        ref = reference_for_site(db, city_id, code_site=site.code_site, year=year, lot=site.lot)
        total = ref["p2_annual_ht"] + ref["p3_annual_ht"]
        options.append({**ref, "total_annual_ht": round(total, 2)})
    return options


def _delta(action: str, current: float | None, target: float | None) -> float:
    current_value = float(current or 0.0)
    target_value = float(target or 0.0)
    if action == "remove":
        return -current_value if current is not None else -target_value
    if action == "modify":
        return target_value - current_value
    return target_value


def _line_impact(line: CpeContractChangeLine) -> dict[str, float]:
    action = line.action
    # P1 is energy consumption context. The avenant impact keeps only P2/P3.
    p2 = _delta(action, line.current_p2_annual_ht, line.p2_annual_ht)
    p3 = _delta(action, line.current_p3_annual_ht, line.p3_annual_ht)
    return {"p1_gaz_annual_ht": 0.0, "p1_elec_annual_ht": 0.0, "p2_annual_ht": p2, "p3_annual_ht": p3}


def _is_leap(year: int) -> bool:
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def _exercise_ratio(year: int, effective_date: date | None) -> float:
    """Share of a civil exercise impacted by the change, exact to the day."""
    start = date(year, 1, 1)
    end_exclusive = date(year + 1, 1, 1)
    if effective_date is not None:
        start = max(start, effective_date)
    end_exclusive = min(end_exclusive, CONTRACT_END_DATE + timedelta(days=1))
    if start >= end_exclusive:
        return 0.0
    days_in_year = 366 if _is_leap(year) else 365
    return (end_exclusive - start).days / days_in_year


def _line_year_impact(
    db: Session,
    city_id: int | None,
    request: CpeContractChangeRequest,
    line: CpeContractChangeLine,
    year: int,
) -> dict[str, float]:
    if line.code_site and line.action in {"remove", "modify"}:
        lot = line.lot or request.lot
        import_ids = _active_import_ids(db, city_id, lot)
        ref = reference_for_site(db, city_id, code_site=line.code_site, year=year, lot=lot)
        has_p2p3 = _has_site_year(db, CpeDalkiaRefP2P3, import_ids=import_ids, code_site=line.code_site, year=year)
        current_p2 = ref["p2_annual_ht"] if has_p2p3 else line.current_p2_annual_ht
        current_p3 = ref["p3_annual_ht"] if has_p2p3 else line.current_p3_annual_ht
        p2 = _delta(line.action, current_p2, line.p2_annual_ht)
        p3 = _delta(line.action, current_p3, line.p3_annual_ht)
        return {"p1_gaz_ht": 0.0, "p1_elec_ht": 0.0, "p2_ht": p2, "p3_ht": p3}
    impact = _line_impact(line)
    return {
        "p1_gaz_ht": impact["p1_gaz_annual_ht"],
        "p1_elec_ht": impact["p1_elec_annual_ht"],
        "p2_ht": impact["p2_annual_ht"],
        "p3_ht": impact["p3_annual_ht"],
    }


def build_annual_impacts(
    db: Session,
    city_id: int | None,
    request: CpeContractChangeRequest,
    lines: list[CpeContractChangeLine],
) -> list[dict[str, Any]]:
    start_year = request.effective_date.year if request.effective_date else date.today().year
    rows: list[dict[str, Any]] = []
    for year in range(start_year, CONTRACT_END_YEAR + 1):
        ratio = _exercise_ratio(year, request.effective_date)
        if ratio <= 0:
            continue
        totals = {"p1_gaz_ht": 0.0, "p1_elec_ht": 0.0, "p2_ht": 0.0, "p3_ht": 0.0}
        for line in lines:
            impact = _line_year_impact(db, city_id, request, line, year)
            for key, value in impact.items():
                totals[key] += value * ratio
        p1 = totals["p1_gaz_ht"] + totals["p1_elec_ht"]
        total = totals["p2_ht"] + totals["p3_ht"]
        rows.append(
            {
                "year": year,
                "ratio": round(ratio, 6),
                "p1_gaz_ht": round(totals["p1_gaz_ht"], 2),
                "p1_elec_ht": round(totals["p1_elec_ht"], 2),
                "p1_ht": round(p1, 2),
                "p2_ht": round(totals["p2_ht"], 2),
                "p3_ht": round(totals["p3_ht"], 2),
                "total_ht": round(total, 2),
            }
        )
    return rows


def build_impact(
    request: CpeContractChangeRequest,
    lines: list[CpeContractChangeLine],
    annual_impacts: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    totals = {"p1_gaz_annual_ht": 0.0, "p1_elec_annual_ht": 0.0, "p2_annual_ht": 0.0, "p3_annual_ht": 0.0}
    for line in lines:
        impact = _line_impact(line)
        for key, value in impact.items():
            totals[key] += value
    p1 = totals["p1_gaz_annual_ht"] + totals["p1_elec_annual_ht"]
    annual = totals["p2_annual_ht"] + totals["p3_annual_ht"]
    year = request.effective_date.year if request.effective_date else None
    ratio = _exercise_ratio(year, request.effective_date) if year else 1.0
    yearly = annual_impacts or []
    first_year_prorata = next((row["total_ht"] for row in yearly if row["year"] == year), annual * ratio)
    remaining_market = (
        sum(row["total_ht"] for row in yearly)
        if yearly
        else annual * ((CONTRACT_END_YEAR - year + ratio) if year else 1.0)
    )
    return {
        **{key: round(value, 2) for key, value in totals.items()},
        "p1_annual_ht": round(p1, 2),
        "total_annual_ht": round(annual, 2),
        "first_year_prorata_ht": round(first_year_prorata, 2),
        "remaining_market_ht": round(remaining_market, 2),
        "effective_year": year,
        "first_year_ratio": round(ratio, 4),
        "annual_impacts": yearly,
    }


def _request_to_dict(db: Session, request: CpeContractChangeRequest) -> dict[str, Any]:
    lines = list(
        db.scalars(
            select(CpeContractChangeLine)
            .where(CpeContractChangeLine.request_id == request.id)
            .order_by(CpeContractChangeLine.id)
        ).all()
    )
    annual_impacts = build_annual_impacts(db, request.city_id, request, lines)
    return {**request.__dict__, "lines": lines, "impact": build_impact(request, lines, annual_impacts)}


def list_requests(db: Session, city_id: int | None) -> list[dict[str, Any]]:
    query = select(CpeContractChangeRequest)
    query = _city_filter(query, CpeContractChangeRequest, city_id)
    requests = db.scalars(query.order_by(CpeContractChangeRequest.created_at.desc(), CpeContractChangeRequest.id.desc())).all()
    return [_request_to_dict(db, request) for request in requests]


def get_request(db: Session, city_id: int | None, request_id: int) -> dict[str, Any] | None:
    query = select(CpeContractChangeRequest).where(CpeContractChangeRequest.id == request_id)
    query = _city_filter(query, CpeContractChangeRequest, city_id)
    request = db.scalars(query).first()
    return _request_to_dict(db, request) if request else None


def _hydrate_line(db: Session, city_id: int | None, request: CpeContractChangeRequest, raw: dict[str, Any]) -> dict[str, Any]:
    data = dict(raw)
    action = data["action"]
    code_site = data.get("code_site")
    year = request.effective_date.year if request.effective_date else date.today().year
    if code_site and action in {"remove", "modify"}:
        ref = reference_for_site(db, city_id, code_site=code_site, year=year, lot=data.get("lot") or request.lot)
        data.setdefault("site_name", ref["site_name"])
        data["site_name"] = data.get("site_name") or ref["site_name"]
        data["lot"] = data.get("lot") or ref["lot"]
        data["pce"] = data.get("pce") or ref["pce"]
        data["tarif"] = data.get("tarif") or ref["tarif"]
        data["current_p1_gaz_annual_ht"] = data.get("current_p1_gaz_annual_ht") if data.get("current_p1_gaz_annual_ht") is not None else ref["p1_gaz_annual_ht"]
        data["current_p1_elec_annual_ht"] = data.get("current_p1_elec_annual_ht") if data.get("current_p1_elec_annual_ht") is not None else ref["p1_elec_annual_ht"]
        data["current_p2_annual_ht"] = data.get("current_p2_annual_ht") if data.get("current_p2_annual_ht") is not None else ref["p2_annual_ht"]
        data["current_p3_annual_ht"] = data.get("current_p3_annual_ht") if data.get("current_p3_annual_ht") is not None else ref["p3_annual_ht"]
        if action == "remove":
            data["p1_gaz_annual_ht"] = data.get("p1_gaz_annual_ht") if data.get("p1_gaz_annual_ht") is not None else ref["p1_gaz_annual_ht"]
            data["p1_elec_annual_ht"] = data.get("p1_elec_annual_ht") if data.get("p1_elec_annual_ht") is not None else ref["p1_elec_annual_ht"]
            data["p2_annual_ht"] = data.get("p2_annual_ht") if data.get("p2_annual_ht") is not None else ref["p2_annual_ht"]
            data["p3_annual_ht"] = data.get("p3_annual_ht") if data.get("p3_annual_ht") is not None else ref["p3_annual_ht"]
    data["site_name"] = data.get("site_name") or data.get("code_site") or "Site a preciser"
    return data


def create_request(db: Session, city_id: int | None, user_id: int | None, payload: CpeOsAvenantRequestCreate) -> dict[str, Any]:
    request = CpeContractChangeRequest(
        city_id=city_id,
        created_by_user_id=user_id,
        title=payload.title,
        change_type=payload.change_type,
        lot=payload.lot,
        effective_date=payload.effective_date,
        reason=payload.reason,
        requester_name=payload.requester_name,
        dalkia_contact_email=payload.dalkia_contact_email,
        notes=payload.notes,
    )
    db.add(request)
    db.flush()
    for line_payload in payload.lines:
        data = _hydrate_line(db, city_id, request, line_payload.model_dump())
        db.add(CpeContractChangeLine(request_id=request.id, city_id=city_id, **data))
    db.commit()
    db.refresh(request)
    return _request_to_dict(db, request)


def update_request(
    db: Session,
    city_id: int | None,
    request_id: int,
    payload: CpeOsAvenantRequestUpdate,
) -> dict[str, Any] | None:
    query = select(CpeContractChangeRequest).where(CpeContractChangeRequest.id == request_id)
    query = _city_filter(query, CpeContractChangeRequest, city_id)
    request = db.scalars(query).first()
    if request is None:
        return None
    data = payload.model_dump(exclude_unset=True)
    if "status" in data and data["status"] not in VALID_STATUSES:
        raise ValueError("INVALID_STATUS")
    for key, value in data.items():
        setattr(request, key, value)
    db.commit()
    db.refresh(request)
    return _request_to_dict(db, request)


def _style_header(ws, row: int, columns: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(1, columns + 1):
        cell = ws.cell(row=row, column=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _money(cell) -> None:
    cell.number_format = '#,##0 "EUR"'


def _autofit(ws, max_width: int = 42) -> None:
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        width = 12
        for cell in ws[letter]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def build_impact_workbook(db: Session, city_id: int | None, request_id: int) -> tuple[bytes, str] | None:
    dossier = get_request(db, city_id, request_id)
    if dossier is None:
        return None
    lines: list[CpeContractChangeLine] = dossier["lines"]
    impact = dossier["impact"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthese"
    ws["A1"] = "Dossier OS / avenant CPE DALKIA"
    ws["A1"].font = Font(bold=True, size=14)
    summary_rows = [
        ("Dossier", dossier["title"]),
        ("Statut", dossier["status"]),
        ("Type", dossier["change_type"]),
        ("Lot", dossier["lot"]),
        ("Date effet", dossier["effective_date"]),
        ("OS", dossier["os_number"]),
        ("Avenant", dossier["avenant_number"]),
        ("Motif", dossier["reason"]),
    ]
    row = 3
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    ws.append(["Indicateur", "Montant HT"])
    _style_header(ws, row, 2)
    for label, key in [
        ("Impact annuel P2+P3", "total_annual_ht"),
        ("Impact annee de prise d'effet", "first_year_prorata_ht"),
        ("Projection fin de marche", "remaining_market_ht"),
        ("P1 annuel non retenu", "p1_annual_ht"),
        ("P2 annuel", "p2_annual_ht"),
        ("P3 annuel", "p3_annual_ht"),
    ]:
        row += 1
        ws.cell(row=row, column=1, value=label)
        amount = ws.cell(row=row, column=2, value=impact.get(key))
        _money(amount)
    _autofit(ws)

    ws_lines = wb.create_sheet("Lignes")
    line_headers = [
        "Action", "Code site", "Site", "Lot", "PCE/PDL", "Tarif",
        "P1 gaz actuel", "P1 elec actuel", "P2 actuel", "P3 actuel",
        "P1 gaz cible", "P1 elec cible", "P2 cible", "P3 cible",
        "Impact annuel P2+P3",
    ]
    ws_lines.append(line_headers)
    _style_header(ws_lines, 1, len(line_headers))
    for line in lines:
        line_impact = _line_impact(line)
        total = sum(line_impact.values())
        values = [
            line.action, line.code_site, line.site_name, line.lot, line.pce, line.tarif,
            line.current_p1_gaz_annual_ht, line.current_p1_elec_annual_ht, line.current_p2_annual_ht, line.current_p3_annual_ht,
            line.p1_gaz_annual_ht, line.p1_elec_annual_ht, line.p2_annual_ht, line.p3_annual_ht,
            round(total, 2),
        ]
        ws_lines.append(values)
    for row_cells in ws_lines.iter_rows(min_row=2, min_col=7, max_col=15):
        for cell in row_cells:
            _money(cell)
    _autofit(ws_lines)
    ws_lines.freeze_panes = "A2"

    ws_projection = wb.create_sheet("Projection")
    projection_headers = ["Exercice", "Part exercice", "P1 non retenu", "P1 elec non retenu", "P1", "P2", "P3", "Total HT P2+P3"]
    ws_projection.append(projection_headers)
    _style_header(ws_projection, 1, len(projection_headers))
    for annual in impact.get("annual_impacts", []):
        ws_projection.append([
            annual["year"], annual["ratio"], annual["p1_gaz_ht"], annual["p1_elec_ht"],
            annual["p1_ht"], annual["p2_ht"], annual["p3_ht"], annual["total_ht"],
        ])
    for row_cells in ws_projection.iter_rows(min_row=2, min_col=2, max_col=8):
        row_cells[0].number_format = "0.0%"
        for cell in row_cells[1:]:
            _money(cell)
    _autofit(ws_projection)
    ws_projection.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    filename = f"impact-os-avenant-{request_id}.xlsx"
    return output.getvalue(), filename
