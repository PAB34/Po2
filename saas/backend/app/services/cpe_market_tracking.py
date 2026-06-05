"""Suivi de facturation marché CPE : enveloppes prévues (DPGF) vs montants reçus.

Le "prévu" provient du référentiel contractuel DALKIA importé depuis les fichiers DPGF
Lot 1 / Lot 2 (acte d'engagement) :

- P2 / P2.4 / P3 / P3.4 : sommes annuelles de ``cpe_dalkia_ref_p2p3`` (imports actifs) ;
- P1 (gaz) : somme annuelle du détail Annexe 6 ``cpe_dalkia_ref_p1_gaz.p10_total_ht``
  (imports actifs Lot 1 + Lot 2), conformément au choix de la source.

Le "reçu" agrège les lignes de factures DALKIA (``cpe_finance_lines``) du périmètre CPE
Ville, classées par poste à partir de ``market`` (P1/P2/P3) et ``billed_item`` (P2.4/P3.4),
et ventilées par année de période.

La page n'a besoin d'aucun nouveau parser : tout est déjà en base.
"""

from __future__ import annotations

import io
import re
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.cpe import CpeContractReference, CpeFinanceLine, CpeSite
from app.models.cpe_dalkia import CpeDalkiaRefImport, CpeDalkiaRefP1Gaz, CpeDalkiaRefP2P3
from app.models.cpe_dpgf_p1 import DPGF_P1_LEVELS, DPGF_P1_LEVEL_LABELS
from app.services.cpe_accounting import (
    CPE_CONTRACT_SCOPE_KIND,
    _is_current_cpe_contract,
    list_finance_invoices,
)
from app.services.cpe_dalkia_db import normalize_p2p3_poste
from app.services.cpe_dpgf_p1 import get_dpgf_p1_levels

# Le numero de lot est encode dans le billed_item des references de perimetre
# (kind cpe_contract_scope), ex. "CPE_VILLE_LOT_1" / "CPE_VILLE_LOT_2".
_LOT_RE = re.compile(r"LOT[_\s-]*(\d+)", re.IGNORECASE)

# Ordre et libellés des postes affichés dans la matrice.
POSTE_ORDER = ["P1", "P2", "P2-4", "P3", "P3-4"]
POSTE_LABELS = {
    "P1": "P1 — Fourniture gaz",
    "P2": "P2 — Maintenance (hors P2.4)",
    "P2-4": "P2.4 — Intéressement / objectifs",
    "P3": "P3 — Gros entretien (hors P3.4)",
    "P3-4": "P3.4 — Travaux obligatoires",
}
# Poste fourre-tout pour les lignes reçues non rattachables à un poste contractuel.
POSTE_OTHER = "AUTRE"
POSTE_OTHER_LABEL = "Autre (reçu non rattaché)"

P1_SOURCE_LABEL = "Annexe 6 DPGF — somme p10_total_ht (imports actifs Lot 1 + Lot 2)"

# Cadence d'acompte du marche CPE : 4 echeances par an (acomptes trimestriels + regularisation).
INSTALLMENTS_PER_YEAR = 4

# DJU de reference contractuel (base 18°C, station Montpellier 1981-2010). Fallback si aucun
# dju_reference en base. Le DJU reel (CSV Open-Meteo) est aussi en base 18 -> comparable.
DJU_REFERENCE_DEFAULT = 1426.0
DJU_SOURCE_LABEL = "DJU chauffage base 18°C — Open-Meteo / COSTIC (DJU/dju_sete.csv)"


def _classify_received_poste(line: CpeFinanceLine) -> str | None:
    """Classe une ligne de facture dans un poste contractuel, ou None si hors postes."""
    market = (line.market or "").strip().upper()
    item = normalize_p2p3_poste(line.billed_item)
    if market == "P1":
        return "P1"
    if item == "P2-4":
        return "P2-4"
    if item == "P3-4":
        return "P3-4"
    if market == "P2":
        return "P2"
    if market == "P3":
        return "P3"
    return None


def _line_year(line: CpeFinanceLine) -> int | None:
    anchor = line.period_end or line.period_start
    return anchor.year if anchor else None


def _line_quarter(line: CpeFinanceLine) -> int | None:
    """Numero de trimestre (1..4) de la periode de la ligne, ou None."""
    anchor = line.period_end or line.period_start
    return (anchor.month - 1) // 3 + 1 if anchor else None


def _cell(prevu: float, recu: float) -> dict[str, Any]:
    prevu = round(prevu, 2)
    recu = round(recu, 2)
    ecart = round(recu - prevu, 2)
    return {
        "prevu": prevu,
        "recu": recu,
        "ecart": ecart,
        "ecart_pct": round(ecart / prevu, 4) if prevu else None,
        "taux": round(recu / prevu, 4) if prevu else None,
    }


def _poste_row(poste: str, label: str, prevu_by_year: dict[int, float], recu_by_year: dict[int, float], years: list[int]) -> dict[str, Any]:
    by_year = [{"year": y, **_cell(prevu_by_year.get(y, 0.0), recu_by_year.get(y, 0.0))} for y in years]
    total_prevu = round(sum(prevu_by_year.get(y, 0.0) for y in years), 2)
    total_recu = round(sum(recu_by_year.get(y, 0.0) for y in years), 2)
    return {"poste": poste, "label": label, "by_year": by_year, "total": _cell(total_prevu, total_recu)}


def _contract_lot_map(db: Session, city_id: int | None) -> dict[str, int]:
    """Associe chaque code contrat CPE a son numero de lot, lu depuis les references de
    perimetre editables (billed_item type 'CPE_VILLE_LOT_1'). Aucun code en dur."""
    stmt = select(CpeContractReference).where(
        CpeContractReference.reference_kind == CPE_CONTRACT_SCOPE_KIND,
        CpeContractReference.active.is_(True),
    )
    if city_id is not None:
        stmt = stmt.where(CpeContractReference.city_id == city_id)
    mapping: dict[str, int] = {}
    for ref in db.scalars(stmt).all():
        match = _LOT_RE.search(ref.billed_item or "")
        code = (ref.contract_code or "").strip().upper()
        if code and match:
            mapping[code] = int(match.group(1))
    return mapping


def _collect(
    db: Session,
    city_id: int | None,
    years: list[int],
    year_set: set[int],
    *,
    prevu_lot: int | None = None,
    recu_contracts: set[str] | None = None,
) -> tuple[dict[str, dict[int, float]], dict[str, dict[int, float]], dict[int, float], int, dict[int, set[int]]]:
    """Accumule prévu (référentiel) et reçu (factures), éventuellement filtré par lot DPGF
    (``prevu_lot``) et par codes contrat reçus (``recu_contracts``)."""
    prevu: dict[str, dict[int, float]] = {poste: {y: 0.0 for y in years} for poste in POSTE_ORDER}

    p2p3_stmt = (
        select(CpeDalkiaRefP2P3)
        .join(CpeDalkiaRefImport, CpeDalkiaRefP2P3.import_id == CpeDalkiaRefImport.id)
        .where(CpeDalkiaRefImport.is_active.is_(True))
    )
    p1_stmt = (
        select(CpeDalkiaRefP1Gaz)
        .join(CpeDalkiaRefImport, CpeDalkiaRefP1Gaz.import_id == CpeDalkiaRefImport.id)
        .where(CpeDalkiaRefImport.is_active.is_(True))
    )
    if city_id is not None:
        p2p3_stmt = p2p3_stmt.where(CpeDalkiaRefImport.city_id == city_id)
        p1_stmt = p1_stmt.where(CpeDalkiaRefImport.city_id == city_id)
    if prevu_lot is not None:
        p2p3_stmt = p2p3_stmt.where(CpeDalkiaRefImport.lot == prevu_lot)
        p1_stmt = p1_stmt.where(CpeDalkiaRefImport.lot == prevu_lot)

    reference_rows = 0
    for row in db.scalars(p2p3_stmt).all():
        if row.period_year not in year_set:
            continue
        reference_rows += 1
        p2_4 = row.p2_4_ht or 0.0
        p3_4 = row.p3_4_ht or 0.0
        prevu["P2"][row.period_year] += (row.p2_total_ht or 0.0) - p2_4
        prevu["P2-4"][row.period_year] += p2_4
        prevu["P3"][row.period_year] += (row.p3_total_ht or 0.0) - p3_4
        prevu["P3-4"][row.period_year] += p3_4
    for row in db.scalars(p1_stmt).all():
        if row.period_year not in year_set:
            continue
        reference_rows += 1
        prevu["P1"][row.period_year] += row.p10_total_ht or 0.0

    recu: dict[str, dict[int, float]] = {poste: {y: 0.0 for y in years} for poste in POSTE_ORDER}
    recu_other: dict[int, float] = {y: 0.0 for y in years}
    quarters_seen: dict[int, set[int]] = {y: set() for y in years}
    invoices = [
        invoice
        for invoice in list_finance_invoices(db, city_id=city_id)
        if _is_current_cpe_contract(
            db,
            invoice.contract_code,
            city_id=invoice.city_id,
            year=(invoice.period_end.year if invoice.period_end else None),
        )
        and (recu_contracts is None or (invoice.contract_code or "").strip().upper() in recu_contracts)
    ]
    invoice_ids = [invoice.id for invoice in invoices]
    if invoice_ids:
        for line in db.scalars(
            select(CpeFinanceLine).where(CpeFinanceLine.invoice_id.in_(invoice_ids))
        ).all():
            year = _line_year(line)
            if year not in year_set:
                continue
            quarter = _line_quarter(line)
            if quarter is not None:
                quarters_seen[year].add(quarter)
            poste = _classify_received_poste(line)
            amount = line.amount_ht or 0.0
            if poste is None:
                recu_other[year] += amount
            else:
                recu[poste][year] += amount
    return prevu, recu, recu_other, reference_rows, quarters_seen


def _assemble(
    prevu: dict[str, dict[int, float]],
    recu: dict[str, dict[int, float]],
    recu_other: dict[int, float],
    years: list[int],
) -> dict[str, Any]:
    postes = [_poste_row(poste, POSTE_LABELS[poste], prevu[poste], recu[poste], years) for poste in POSTE_ORDER]
    if any(value for value in recu_other.values()):
        postes.append(_poste_row(POSTE_OTHER, POSTE_OTHER_LABEL, {y: 0.0 for y in years}, recu_other, years))

    totals_by_year = []
    for y in years:
        prevu_y = sum(prevu[poste][y] for poste in POSTE_ORDER)
        recu_y = sum(recu[poste][y] for poste in POSTE_ORDER) + recu_other[y]
        totals_by_year.append({"year": y, **_cell(prevu_y, recu_y)})

    grand_total = _cell(
        round(sum(cell["prevu"] for cell in totals_by_year), 2),
        round(sum(cell["recu"] for cell in totals_by_year), 2),
    )
    return {"postes": postes, "totals_by_year": totals_by_year, "grand_total": grand_total}


def _quarters_block(quarters_seen: dict[int, set[int]], years: list[int]) -> list[dict[str, Any]]:
    """Nombre de trimestres factures par annee (sur INSTALLMENTS_PER_YEAR attendus)."""
    return [
        {"year": y, "billed": len(quarters_seen.get(y, set())), "expected": INSTALLMENTS_PER_YEAR}
        for y in years
    ]


def _dju_reference(db: Session, city_id: int | None) -> float:
    """DJU de reference contractuel : valeur la plus frequente dans cpe_sites, sinon defaut 1426."""
    stmt = select(CpeSite.dju_reference).where(CpeSite.dju_reference.is_not(None))
    if city_id is not None:
        stmt = stmt.where(CpeSite.city_id == city_id)
    values = [v for v in db.scalars(stmt) if v]
    if not values:
        return DJU_REFERENCE_DEFAULT
    # mode (valeur la plus frequente)
    return max(set(values), key=values.count)


def _dju_block(db: Session, city_id: int | None, years: list[int]) -> dict[str, Any]:
    """Bandeau informatif de rigueur climatique : DJU chauffage reel vs reference, par annee.

    Purement explicatif (n'entre PAS dans le calcul prevu/recu en euros). Permet de
    contextualiser un P1 recu eleve par un hiver plus rigoureux. Le DJU reel vient du CSV
    Open-Meteo (base 18°C, meme base que la reference contractuelle). Une annee incomplete
    (< 12 mois de donnees) est marquee ``complete=false`` et son ratio est indicatif.
    """
    try:
        from app.services.energie import get_dju_monthly  # noqa: PLC0415 (lazy : lit un CSV)
        monthly = get_dju_monthly()
    except Exception:  # noqa: BLE001 — pas de CSV / source indispo -> bandeau absent
        monthly = []

    if not monthly:
        return {"reference": _dju_reference(db, city_id), "source": DJU_SOURCE_LABEL,
                "base": 18, "by_year": [], "has_data": False}

    sums: dict[int, float] = {y: 0.0 for y in years}
    months: dict[int, int] = {y: 0 for y in years}
    year_set = set(years)
    for row in monthly:
        ym = row.get("month", "")
        if len(ym) < 7:
            continue
        try:
            y = int(ym[:4])
        except ValueError:
            continue
        if y not in year_set:
            continue
        sums[y] += row.get("dju_chauffe", 0.0) or 0.0
        months[y] += 1

    reference = _dju_reference(db, city_id)
    by_year = []
    has_data = False
    for y in years:
        m = months[y]
        if m == 0:
            by_year.append({"year": y, "dju_real": None, "months": 0, "complete": False, "ratio": None})
            continue
        has_data = True
        dju_real = round(sums[y], 1)
        complete = m >= 12
        ratio = round(dju_real / reference, 4) if reference else None
        by_year.append({"year": y, "dju_real": dju_real, "months": m, "complete": complete, "ratio": ratio})

    return {
        "reference": round(reference, 1),
        "source": DJU_SOURCE_LABEL,
        "base": 18,
        "by_year": by_year,
        "has_data": has_data,
    }


def _dpgf_p1_block(db: Session, city_id: int | None, years: list[int], *, lot: int | None) -> dict[str, Any]:
    """Bloc informatif des niveaux P1 revises (DPGF) par annee : contrat / Rev Temp / Rev T° & prix.

    Purement additif : n'entre PAS dans le calcul prevu/recu. Le ``prevu P1`` reste au niveau
    contrat (cf. decision : DPGF P1 expose en plus, sans creer d'ecart artificiel).
    """
    levels = get_dpgf_p1_levels(db, city_id, years, lot=lot)
    rows = []
    has_data = False
    for level in DPGF_P1_LEVELS:
        by_year = levels.get(level, {})
        amounts = [round(by_year.get(y, 0.0), 2) for y in years]
        total = round(sum(amounts), 2)
        if total:
            has_data = True
        rows.append({
            "level": level,
            "label": DPGF_P1_LEVEL_LABELS[level],
            "by_year": [{"year": y, "total": a} for y, a in zip(years, amounts)],
            "total": total,
        })
    return {"levels": rows, "has_data": has_data}


def build_market_tracking(
    db: Session,
    city_id: int | None = None,
    *,
    year_from: int = 2026,
    year_to: int = 2030,
) -> dict[str, Any]:
    """Construit la matrice poste × année (prévu DPGF vs reçu factures).

    Renvoie le suivi combiné (tous lots) plus, dans ``by_lot``, le même découpage par lot
    contractuel (Lot 1 / Lot 2) quand le périmètre les distingue.
    """
    if year_to < year_from:
        year_from, year_to = year_to, year_from
    years = list(range(year_from, year_to + 1))
    year_set = set(years)

    prevu, recu, recu_other, reference_rows, quarters_seen = _collect(db, city_id, years, year_set)
    result: dict[str, Any] = {
        "years": years,
        **_assemble(prevu, recu, recu_other, years),
        "p1_source": P1_SOURCE_LABEL,
        "has_reference": reference_rows > 0,
        "p1_dpgf": _dpgf_p1_block(db, city_id, years, lot=None),
        "quarters_billed": _quarters_block(quarters_seen, years),
        "installments_per_year": INSTALLMENTS_PER_YEAR,
        "dju": _dju_block(db, city_id, years),
    }

    # ── Découpage par lot ────────────────────────────────────────────────────
    contract_lot = _contract_lot_map(db, city_id)
    by_lot: list[dict[str, Any]] = []
    for lot in sorted(set(contract_lot.values())):
        contracts = {code for code, value in contract_lot.items() if value == lot}
        l_prevu, l_recu, l_recu_other, l_refrows, l_quarters = _collect(
            db, city_id, years, year_set, prevu_lot=lot, recu_contracts=contracts
        )
        by_lot.append(
            {
                "lot": lot,
                "label": f"Lot {lot}",
                "contract_codes": sorted(contracts),
                **_assemble(l_prevu, l_recu, l_recu_other, years),
                "has_reference": l_refrows > 0,
                "p1_dpgf": _dpgf_p1_block(db, city_id, years, lot=lot),
                "quarters_billed": _quarters_block(l_quarters, years),
            }
        )
    result["by_lot"] = by_lot
    return result


def build_market_tracking_workbook(
    db: Session,
    city_id: int | None = None,
    *,
    year_from: int = 2026,
    year_to: int = 2030,
) -> bytes:
    """Export XLSX de la matrice de suivi marché (prévu vs reçu, poste × année)."""
    report = build_market_tracking(db, city_id, year_from=year_from, year_to=year_to)
    years = report["years"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi marche"
    ws["A1"] = "Suivi facturation marché CPE — prévu (DPGF) vs reçu"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Source du prévu P1 : {report['p1_source']}"
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")

    header_row = 4
    headers = ["Poste"]
    for year in years:
        headers += [f"{year} Prévu", f"{year} Reçu", f"{year} Écart", f"{year} Taux"]
    headers += ["Total Prévu", "Total Reçu", "Total Écart", "Total Taux"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    def _write_cell_group(ws_row: int, start_col: int, cell: dict[str, Any]) -> None:
        ws.cell(row=ws_row, column=start_col, value=cell["prevu"]).number_format = '#,##0 "€"'
        ws.cell(row=ws_row, column=start_col + 1, value=cell["recu"]).number_format = '#,##0 "€"'
        ws.cell(row=ws_row, column=start_col + 2, value=cell["ecart"]).number_format = '#,##0 "€"'
        taux_cell = ws.cell(row=ws_row, column=start_col + 3, value=cell["taux"])
        taux_cell.number_format = "0.0%"

    row = header_row + 1
    for poste in report["postes"]:
        ws.cell(row=row, column=1, value=poste["label"]).font = Font(bold=True)
        col = 2
        for cell in poste["by_year"]:
            _write_cell_group(row, col, cell)
            col += 4
        _write_cell_group(row, col, poste["total"])
        row += 1

    # Ligne TOTAL marché
    ws.cell(row=row, column=1, value="TOTAL marché").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E5E7EB")
    col = 2
    for cell in report["totals_by_year"]:
        _write_cell_group(row, col, cell)
        col += 4
    _write_cell_group(row, col, report["grand_total"])
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="E5E7EB")
    row += 1

    # Ligne "Trimestres factures / N" (lecture annuel vs trimestriel)
    quarters = report.get("quarters_billed") or []
    expected = report.get("installments_per_year", 4)
    if quarters:
        q_label = ws.cell(row=row, column=1, value=f"Trimestres facturés (sur {expected})")
        q_label.font = Font(italic=True, color="6B7280")
        col = 2
        for q in quarters:
            cell = ws.cell(row=row, column=col, value=f"{q['billed']}/{q.get('expected', expected)}")
            cell.font = Font(italic=True, color="6B7280")
            col += 4
        row += 1

    # Ligne "Rigueur climatique" : DJU reel vs reference (informatif)
    dju = report.get("dju") or {}
    if dju.get("has_data"):
        ref = round(dju.get("reference") or 0)
        dlabel = ws.cell(row=row, column=1, value=f"Rigueur climatique — DJU réel / réf. {ref}")
        dlabel.font = Font(italic=True, color="6B7280")
        by_year_dju = {d["year"]: d for d in dju.get("by_year", [])}
        col = 2
        for y in years:
            d = by_year_dju.get(y)
            if d and d.get("dju_real") is not None:
                pct = f" ({round(d['ratio'] * 100)} %)" if d.get("ratio") is not None else ""
                suffix = "" if d.get("complete") else " partiel"
                txt = f"{round(d['dju_real'])}{pct}{suffix}"
            else:
                txt = "—"
            c = ws.cell(row=row, column=col, value=txt)
            c.font = Font(italic=True, color="6B7280")
            col += 4
        row += 1

    # ── Bloc informatif : P1 gaz revise (DPGF apres OS) ──────────────────────
    p1_dpgf = report.get("p1_dpgf") or {}
    if p1_dpgf.get("has_data"):
        row += 1  # ligne vide de separation
        title_cell = ws.cell(
            row=row, column=1,
            value="P1 gaz révisé (DPGF) — informatif (le prévu P1 ci-dessus reste au niveau contrat)",
        )
        title_cell.font = Font(bold=True, italic=True, color="6B7280")
        row += 1
        for lvl in p1_dpgf["levels"]:
            label = lvl["label"] + (" (= prévu P1)" if lvl["level"] == "contrat" else "")
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            col = 2
            for cell in lvl["by_year"]:
                amount_cell = ws.cell(row=row, column=col, value=round(cell["total"], 2))
                amount_cell.number_format = '#,##0 "€"'
                col += 4  # une seule valeur par annee (colonne "Prévu" du groupe)
            total_cell = ws.cell(row=row, column=col, value=round(lvl["total"], 2))
            total_cell.number_format = '#,##0 "€"'
            row += 1

    ws.column_dimensions["A"].width = 34
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 13
    ws.freeze_panes = "B5"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
