"""Gabarit finance COMBINE de codification (aller-retour export/import).

Un seul classeur pour le service finance couvrant les deux tiers :
- DALKIA (CPE)      : feuilles « DALKIA - Sites » et « DALKIA - Postes ».
- ENGIE / EDF       : feuilles « ENGIE-EDF - Points » et « ENGIE-EDF - Postes ».

Export = reflet de la matrice en vigueur (DALKIA scopé au marché Ville en cours).
Import = upsert par feuille (aucune suppression). Vocation : remplacer à terme le
fichier `MATRICE_DALKIA-COMPATBILITE V2.xlsx`.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.cpe import CpeAccountingNatureRule, CpeAccountingSiteMapping
from app.models.invoice import EnergyAccountingNatureRule, EnergyAccountingSiteMapping
from app.services import cpe_accounting, energie_accounting
from app.services.cpe_accounting import (
    _clean,
    _detect_header_row,
    _norm_header,
    _parse_actif,
    _rows_from_sheet,
    _set_widths,
    _style_header,
    _upper,
)

# --- Noms de feuilles (normalisés pour le dispatch import) ---
SHEET_DALKIA_SITES = "DALKIA - Sites"
SHEET_DALKIA_POSTES = "DALKIA - Postes"
SHEET_ENERGY_POINTS = "ENGIE-EDF - Points"
SHEET_ENERGY_POSTES = "ENGIE-EDF - Postes"

_N_DALKIA_SITES = _norm_header(SHEET_DALKIA_SITES)
_N_DALKIA_POSTES = _norm_header(SHEET_DALKIA_POSTES)
_N_ENERGY_POINTS = _norm_header(SHEET_ENERGY_POINTS)
_N_ENERGY_POSTES = _norm_header(SHEET_ENERGY_POSTES)
# Formats DALKIA hérités (V2, gabarit simple) délégués au parseur cpe_accounting.
_N_DALKIA_LEGACY = {"sites_vers_codes", "poste_facture_vers_nature_ctpab", "postes_x_contrat_x_nature", "sites", "postes"}

_SITE_HEADERS = [
    "Code site", "Désignation", "Famille", "Gestionnaire", "Gestionnaire suppléant",
    "Service (code)", "Service (libellé)", "Fonction (code)", "Fonction (libellé)",
    "Antenne (code)", "Antenne (libellé)", "Opération (code)", "Opération (libellé)",
    "Actif", "Notes",
]
_POSTE_HEADERS = [
    "Code contrat", "Marché", "Poste facturé", "Service vendu", "Fréquence",
    "Nature comptable", "Libellé nature", "Actif", "Notes",
]
_POINT_HEADERS = [
    "PRM", "Désignation", "Regroupement", "Gestionnaire",
    "Service (code)", "Service (libellé)", "Fonction (code)", "Fonction (libellé)",
    "Antenne (code)", "Antenne (libellé)", "Opération (code)", "Opération (libellé)",
    "Actif", "Notes",
]
_ENERGY_POSTE_HEADERS = [
    "Fournisseur", "Marché", "Poste facturé", "Fréquence",
    "Nature comptable", "Libellé nature", "Actif", "Notes",
]


@dataclass
class FinanceCodificationResult:
    filename: str | None = None
    dalkia_sites_created: int = 0
    dalkia_sites_updated: int = 0
    dalkia_rules_created: int = 0
    dalkia_rules_updated: int = 0
    energy_points_created: int = 0
    energy_points_updated: int = 0
    energy_rules_created: int = 0
    energy_rules_updated: int = 0
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _actif(value: bool) -> str:
    return "Oui" if value else "Non"


def build_finance_codification_workbook(db: Session, city_id: int | None) -> bytes:
    """Classeur combiné DALKIA + ENGIE/EDF reflétant la matrice en vigueur."""
    dalkia_sites = cpe_accounting.list_accounting_site_mappings(db, city_id)
    dalkia_rules = cpe_accounting.list_accounting_nature_rules(db, city_id, only_current_scope=True)
    energy_points = energie_accounting.list_site_mappings(db, city_id)
    energy_rules = energie_accounting.list_nature_rules(db, city_id)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = SHEET_DALKIA_SITES
    ws.append(_SITE_HEADERS)
    _style_header(ws[1])
    for s in dalkia_sites:
        ws.append([
            s.code_site, s.site_name, s.family, s.manager, s.alternate_manager,
            s.service_code, s.service_label, s.function_code, s.function_label,
            s.antenna_code, s.antenna_label, s.operation_code, s.operation_label,
            _actif(s.active), s.notes,
        ])
    _set_widths(ws, [16, 42, 14, 16, 18, 12, 22, 12, 22, 14, 20, 14, 22, 8, 30])
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet(SHEET_DALKIA_POSTES)
    ws2.append(_POSTE_HEADERS)
    _style_header(ws2[1])
    for r in dalkia_rules:
        ws2.append([
            r.contract_code, r.market, r.billed_item, r.service_sold, r.frequency,
            r.accounting_nature, r.accounting_label, _actif(r.active), r.notes,
        ])
    _set_widths(ws2, [16, 14, 16, 18, 14, 16, 40, 8, 30])
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet(SHEET_ENERGY_POINTS)
    ws3.append(_POINT_HEADERS)
    _style_header(ws3[1])
    for p in energy_points:
        ws3.append([
            p.prm_id, p.site_name, p.regroupement, p.manager,
            p.service_code, p.service_label, p.function_code, p.function_label,
            p.antenna_code, p.antenna_label, p.operation_code, p.operation_label,
            _actif(p.active), p.notes,
        ])
    _set_widths(ws3, [18, 42, 16, 16, 12, 22, 12, 22, 14, 20, 14, 22, 8, 30])
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet(SHEET_ENERGY_POSTES)
    ws4.append(_ENERGY_POSTE_HEADERS)
    _style_header(ws4[1])
    for r in energy_rules:
        ws4.append([
            r.supplier, r.market, r.billed_item, r.frequency,
            r.accounting_nature, r.accounting_label, _actif(r.active), r.notes,
        ])
    _set_widths(ws4, [14, 14, 18, 14, 16, 40, 8, 30])
    ws4.freeze_panes = "A2"

    ws5 = wb.create_sheet("Mode d'emploi")
    for line in _GUIDE_LINES:
        ws5.append([line])
    ws5.column_dimensions["A"].width = 95

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


_GUIDE_LINES = [
    "Codification comptable — gabarit d'échange service finance (DALKIA + ENGIE/EDF)",
    "",
    "Ce classeur reflète la matrice en vigueur dans la plateforme.",
    "Modifiez les valeurs puis renvoyez-le pour réimport (menu Importer sur /refonte-v1/matrices).",
    "",
    "Feuilles :",
    "• « DALKIA - Sites » : un site par ligne, clé = Code site.",
    "• « DALKIA - Postes » : marché Ville EN COURS uniquement (C00190116O / C00190155J).",
    "   clé = Code contrat + Marché + Service vendu + Poste facturé + Fréquence.",
    "• « ENGIE-EDF - Points » : un point (PRM) par ligne, clé = PRM.",
    "• « ENGIE-EDF - Postes » : clé = Fournisseur + Marché + Poste facturé + Fréquence.",
    "",
    "Règles :",
    "• Ne pas renommer les feuilles ni la ligne d'en-tête.",
    "• Colonne « Actif » : Oui / Non.",
    "• L'import fonctionne en MISE À JOUR (upsert) : une ligne supprimée dans Excel",
    "  n'est PAS supprimée dans la plateforme (à retirer via l'interface).",
    "• L'opération d'investissement (98xxx) ne s'applique qu'aux postes P3 / P3.4.",
]


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def _bool_or_none(value: object) -> bool:
    return _parse_actif(value)


def _upsert_cpe_site(db: Session, city_id: int | None, row: dict) -> bool | None:
    code_site = _upper(row.get("code_site"))
    site_name = _clean(row.get("designation"))
    if not code_site or not site_name:
        return None
    payload = {
        "city_id": city_id, "code_site": code_site, "site_name": site_name,
        "family": _clean(row.get("famille")), "manager": _clean(row.get("gestionnaire")),
        "alternate_manager": _clean(row.get("gestionnaire_suppleant")),
        "service_code": _clean(row.get("service_code")), "service_label": _clean(row.get("service_libelle")),
        "function_code": _clean(row.get("fonction_code")), "function_label": _clean(row.get("fonction_libelle")),
        "antenna_code": _clean(row.get("antenne_code")), "antenna_label": _clean(row.get("antenne_libelle")),
        "operation_code": _clean(row.get("operation_code")), "operation_label": _clean(row.get("operation_libelle")),
        "active": _parse_actif(row.get("actif")), "notes": _clean(row.get("notes")),
    }
    existing = db.scalars(
        select(CpeAccountingSiteMapping).where(
            CpeAccountingSiteMapping.city_id == city_id,
            CpeAccountingSiteMapping.code_site == code_site,
        )
    ).first()
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        return False
    db.add(CpeAccountingSiteMapping(**payload))
    return True


def _upsert_energy_point(db: Session, city_id: int | None, row: dict) -> bool | None:
    prm = _clean(row.get("prm"))
    if not prm:
        return None
    payload = {
        "city_id": city_id, "prm_id": prm, "site_name": _clean(row.get("designation")),
        "regroupement": _clean(row.get("regroupement")), "manager": _clean(row.get("gestionnaire")),
        "service_code": _clean(row.get("service_code")), "service_label": _clean(row.get("service_libelle")),
        "function_code": _clean(row.get("fonction_code")), "function_label": _clean(row.get("fonction_libelle")),
        "antenna_code": _clean(row.get("antenne_code")), "antenna_label": _clean(row.get("antenne_libelle")),
        "operation_code": _clean(row.get("operation_code")), "operation_label": _clean(row.get("operation_libelle")),
        "active": _parse_actif(row.get("actif")), "notes": _clean(row.get("notes")),
    }
    existing = db.scalars(
        select(EnergyAccountingSiteMapping).where(
            EnergyAccountingSiteMapping.city_id == city_id,
            EnergyAccountingSiteMapping.prm_id == prm,
        )
    ).first()
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        return False
    db.add(EnergyAccountingSiteMapping(**payload))
    return True


def _upsert_energy_rule(db: Session, city_id: int | None, row: dict) -> bool | None:
    from sqlalchemy import func as safunc

    billed_item = _upper(row.get("poste_facture"))
    nature = _clean(row.get("nature_comptable"))
    if not billed_item or not nature:
        return None
    supplier = _upper(row.get("fournisseur")) or "ENGIE"
    market = _clean(row.get("marche"))
    frequency = _clean(row.get("frequence"))
    payload = {
        "city_id": city_id, "supplier": supplier, "market": market, "billed_item": billed_item,
        "frequency": frequency, "accounting_nature": nature,
        "accounting_label": _clean(row.get("libelle_nature")),
        "active": _parse_actif(row.get("actif")), "notes": _clean(row.get("notes")),
    }
    existing = db.scalars(
        select(EnergyAccountingNatureRule).where(
            EnergyAccountingNatureRule.city_id == city_id,
            EnergyAccountingNatureRule.supplier == supplier,
            safunc.coalesce(EnergyAccountingNatureRule.market, "") == (market or ""),
            EnergyAccountingNatureRule.billed_item == billed_item,
            safunc.coalesce(EnergyAccountingNatureRule.frequency, "") == (frequency or ""),
        )
    ).first()
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        return False
    db.add(EnergyAccountingNatureRule(**payload))
    return True


def import_finance_codification_workbook(
    db: Session,
    raw_bytes: bytes,
    *,
    filename: str | None,
    city_id: int | None,
) -> FinanceCodificationResult:
    """Importe le gabarit combiné (dispatch par feuille). Upsert, aucune suppression."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    result = FinanceCodificationResult(filename=filename)
    sheets = {_norm_header(name): name for name in wb.sheetnames}

    # --- DALKIA ---
    if _N_DALKIA_SITES in sheets:
        ws = wb[sheets[_N_DALKIA_SITES]]
        header_row = _detect_header_row(ws, {"code_site", "designation"})
        if header_row is None:
            result.errors.append(f"« {SHEET_DALKIA_SITES} » : en-tête introuvable")
        else:
            for row in _rows_from_sheet(ws, header_row):
                created = _upsert_cpe_site(db, city_id, row)
                if created is True:
                    result.dalkia_sites_created += 1
                elif created is False:
                    result.dalkia_sites_updated += 1
    if _N_DALKIA_POSTES in sheets:
        ws = wb[sheets[_N_DALKIA_POSTES]]
        header_row = _detect_header_row(ws, {"poste_facture", "nature_comptable"})
        if header_row is None:
            result.errors.append(f"« {SHEET_DALKIA_POSTES} » : en-tête introuvable")
        else:
            for row in _rows_from_sheet(ws, header_row):
                billed_item = _upper(row.get("poste_facture"))
                nature = _clean(row.get("nature_comptable"))
                if not billed_item or not nature:
                    continue
                market = _upper(row.get("marche")) or cpe_accounting._market_from_billed_item(billed_item)
                if cpe_accounting._upsert_nature_rule(
                    db, city_id=city_id, contract_code=_upper(row.get("code_contrat")),
                    market=market, service_sold=_upper(row.get("service_vendu")),
                    billed_item=billed_item, frequency=_clean(row.get("frequence")),
                    accounting_nature=nature, accounting_label=_clean(row.get("libelle_nature")),
                    notes=_clean(row.get("notes")), active=_parse_actif(row.get("actif")),
                ):
                    result.dalkia_rules_created += 1
                else:
                    result.dalkia_rules_updated += 1
    # Rétro-compat : classeur DALKIA hérité (V2 / gabarit simple) sans feuilles combinées.
    if _N_DALKIA_SITES not in sheets and _N_DALKIA_POSTES not in sheets and (sheets.keys() & _N_DALKIA_LEGACY):
        legacy = cpe_accounting.import_codification_workbook(db, raw_bytes, filename=filename, city_id=city_id)
        result.dalkia_sites_created += legacy.site_mappings_created
        result.dalkia_sites_updated += legacy.site_mappings_updated
        result.dalkia_rules_created += legacy.nature_rules_created
        result.dalkia_rules_updated += legacy.nature_rules_updated
        result.errors.extend(e for e in legacy.errors if "absente" not in e.lower())

    # --- ENGIE / EDF ---
    if _N_ENERGY_POINTS in sheets:
        ws = wb[sheets[_N_ENERGY_POINTS]]
        header_row = _detect_header_row(ws, {"prm", "designation"})
        if header_row is None:
            result.errors.append(f"« {SHEET_ENERGY_POINTS} » : en-tête introuvable")
        else:
            for row in _rows_from_sheet(ws, header_row):
                created = _upsert_energy_point(db, city_id, row)
                if created is True:
                    result.energy_points_created += 1
                elif created is False:
                    result.energy_points_updated += 1
    if _N_ENERGY_POSTES in sheets:
        ws = wb[sheets[_N_ENERGY_POSTES]]
        header_row = _detect_header_row(ws, {"poste_facture", "nature_comptable"})
        if header_row is None:
            result.errors.append(f"« {SHEET_ENERGY_POSTES} » : en-tête introuvable")
        else:
            for row in _rows_from_sheet(ws, header_row):
                created = _upsert_energy_rule(db, city_id, row)
                if created is True:
                    result.energy_rules_created += 1
                elif created is False:
                    result.energy_rules_updated += 1

    db.commit()
    return result
