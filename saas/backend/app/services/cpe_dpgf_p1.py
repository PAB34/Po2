"""Import du DPGF P1 revise (livrable separe DALKIA) — parsing, upsert isole, lecture.

Quand DALKIA signe un OS impactant le prix gaz, il livre un fichier
``P1 - DPGF LOT x AAAAVn.xlsx`` separe du fichier maitre. Ce module :

1. **parse** les 3 feuilles ``Annexe 6`` (contrat / Rev Temp / Rev T° & prix) de facon
   generique (reperage par libelle d'en-tete, insensible aux offsets de colonnes propres a
   chaque feuille) ;
2. **persiste** dans une lignee d'import PROPRE (``cpe_dpgf_p1_*``) ; un nouvel import ne
   desactive QUE le DPGF P1 precedent du meme lot, et **ne touche jamais** le referentiel
   maitre (``cpe_dalkia_ref_*`` via ``persist_dalkia_import``) ni ``cpe_contract_references`` ;
3. **expose** les totaux P1 revises par lot x annee x niveau, consommes en mode informatif
   par le suivi marche (le ``prevu P1`` reste au niveau contrat — aucun changement de
   semantique).

Voir docs/energie/CPE-DALKIA/12-OS3-Prix-gaz.md et la memoire project_p1_dpgf_revisions.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.cpe_dpgf_p1 import (
    DPGF_P1_LEVELS,
    CpeDpgfP1Import,
    CpeDpgfP1Line,
)
from app.models.user import User
from app.services.cpe_dalkia_import import (
    PERIOD_YEARS,
    _is_site_row,
    _clean_str,
    _to_float,
)

N_PERIODS = len(PERIOD_YEARS)


# ─────────────────────────────────────────────────────────────────────────────
# Parsing
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class DpgfP1LineRow:
    level: str
    code_site: str
    pce: str | None
    type_tarif: str | None
    prix_unitaire_ht: float | None
    period_idx: int
    period_label: str
    period_year: int
    qt_mwhpcs: float | None
    p10_var_ht: float | None
    p10_total_ht: float | None


@dataclass
class DpgfP1ParseResult:
    lot: int
    filename: str
    lines: list[DpgfP1LineRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    # totaux[level][year] = somme p10_total_ht (pour preview + validation)
    totals: dict[str, dict[int, float]] = field(default_factory=dict)
    nb_sites: dict[str, int] = field(default_factory=dict)


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _sheet_level(sheet_name: str) -> str | None:
    """Associe un nom de feuille Annexe 6 a un niveau de revision, ou None.

    Insensible aux accents / espaces : ``Annexe 6 - P1GAZ Rev T° & prix`` -> rev_temp_prix.
    L'ordre des tests compte : "prix" prime sur "rev temp" (la feuille prix contient les deux).
    """
    norm = _strip_accents(sheet_name).lower()
    if "annexe 6" not in norm or "p1" not in norm or "gaz" not in norm:
        return None
    if "prix" in norm:
        return "rev_temp_prix"
    if "rev" in norm and "temp" in norm:
        return "rev_temp"
    if "contrat" in norm:
        return "contrat"
    return None


def _find_header_row(rows: list[tuple]) -> int | None:
    """Ligne d'en-tete du tableau par site (contient 'LOT' et 'PROG')."""
    for i, row in enumerate(rows):
        vals = [str(v).upper() for v in row if v is not None]
        if any("LOT" in v for v in vals) and any("PROG" in v for v in vals):
            return i
    return None


def _col_with(header: tuple, *needles: str) -> int | None:
    """Premier index de colonne dont l'en-tete (majuscule) contient l'un des ``needles``."""
    for j, c in enumerate(header):
        if c is None:
            continue
        up = str(c).upper()
        if all(n.upper() in up for n in needles):
            return j
    return None


def _period_year_label(period_label_row: tuple, total_col: int, period_idx: int) -> str:
    """Libelle de periode (annee) lu au-dessus de la colonne QT (= total_col - 2)."""
    for c in (total_col - 2, total_col):
        val = period_label_row[c] if 0 <= c < len(period_label_row) else None
        if val:
            years = re.findall(r"\d{4}", str(val))
            if years:
                return years[0] if len(set(years)) == 1 else f"{years[0]}-{years[-1]}"
    return str(PERIOD_YEARS[period_idx]) if period_idx < N_PERIODS else "?"


def _parse_annexe6_sheet(rows: list[tuple], level: str) -> tuple[list[DpgfP1LineRow], list[str]]:
    """Parse une feuille Annexe 6 (un niveau). Generic : repere les colonnes par en-tete.

    La structure est invariante au niveau pres d'un decalage de colonnes :
      - ``code_site`` = colonne 'N° PROG' ;
      - 9 colonnes 'P10 - TOTAL' (une par periode 2025..2033), chacune precedee de QT et P10 var.
    """
    warnings: list[str] = []
    hr = _find_header_row(rows)
    if hr is None:
        return [], [f"DPGF P1 [{level}] : ligne d'en-tete (LOT/PROG) introuvable."]

    header = rows[hr]
    code_col = _col_with(header, "PROG")
    if code_col is None:
        return [], [f"DPGF P1 [{level}] : colonne 'N° PROG' (code site) introuvable."]
    pce_col = _col_with(header, "PCE")
    tarif_col = _col_with(header, "TYPE DE TA")
    pu_col = _col_with(header, "PRIX UNITA")

    total_cols = [j for j, c in enumerate(header) if c is not None and str(c).upper().startswith("P10 - TOTA")]
    if len(total_cols) < N_PERIODS:
        warnings.append(
            f"DPGF P1 [{level}] : {len(total_cols)} colonnes 'P10 - TOTAL' detectees "
            f"(attendu {N_PERIODS})."
        )
    total_cols = total_cols[:N_PERIODS]

    period_label_row = rows[hr - 2] if hr >= 2 else ()
    out: list[DpgfP1LineRow] = []
    for row in rows[hr + 1:]:
        code_val = row[code_col] if len(row) > code_col else None
        if not _is_site_row(code_val):
            continue
        code_site = str(code_val).strip()
        pce = _clean_str(row[pce_col]) if pce_col is not None and len(row) > pce_col else None
        type_tarif = _clean_str(row[tarif_col]) if tarif_col is not None and len(row) > tarif_col else None
        if type_tarif in ("0", ""):
            type_tarif = None
        prix_unitaire = _to_float(row[pu_col]) if pu_col is not None and len(row) > pu_col else None

        for period_i, total_col in enumerate(total_cols):
            qt = _to_float(row[total_col - 2]) if total_col - 2 >= 0 and len(row) > total_col - 2 else None
            p10_var = _to_float(row[total_col - 1]) if total_col - 1 >= 0 and len(row) > total_col - 1 else None
            p10_tot = _to_float(row[total_col]) if len(row) > total_col else None
            out.append(DpgfP1LineRow(
                level=level,
                code_site=code_site,
                pce=pce,
                type_tarif=type_tarif,
                prix_unitaire_ht=prix_unitaire,
                period_idx=period_i + 1,
                period_label=_period_year_label(period_label_row, total_col, period_i),
                period_year=PERIOD_YEARS[period_i],
                qt_mwhpcs=qt,
                p10_var_ht=p10_var,
                p10_total_ht=p10_tot,
            ))
    return out, warnings


def parse_dpgf_p1_file(raw_bytes: bytes, filename: str, lot: int) -> DpgfP1ParseResult:
    """Parse un DPGF P1 revise (xlsx) : les 3 feuilles Annexe 6 (contrat / rev / rev+prix)."""
    import io

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — message lisible cote API
        raise ValueError(f"Fichier DPGF P1 illisible : {exc}") from exc

    result = DpgfP1ParseResult(lot=lot, filename=filename)
    seen_levels: set[str] = set()

    for sheet_name in wb.sheetnames:
        level = _sheet_level(sheet_name)
        if level is None or level in seen_levels:
            continue
        seen_levels.add(level)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        lines, warnings = _parse_annexe6_sheet(rows, level)
        result.lines.extend(lines)
        result.warnings.extend(warnings)

    missing = [lvl for lvl in DPGF_P1_LEVELS if lvl not in seen_levels]
    if missing:
        result.warnings.append(
            "Niveaux P1 absents du fichier : " + ", ".join(missing)
            + " (feuilles Annexe 6 non trouvees)."
        )
    if not result.lines:
        raise ValueError(
            "Aucune ligne P1 revisee trouvee : ce fichier ne ressemble pas a un DPGF P1 "
            "(feuilles attendues : 'Annexe 6 - P1 GAZ contrat / Rev Temp / Rev T° & prix')."
        )

    # Totaux et comptages par niveau (pour preview + validation lecture seule)
    totals: dict[str, dict[int, float]] = {}
    sites: dict[str, set[str]] = {}
    for ln in result.lines:
        totals.setdefault(ln.level, {})
        totals[ln.level][ln.period_year] = round(
            totals[ln.level].get(ln.period_year, 0.0) + (ln.p10_total_ht or 0.0), 2
        )
        sites.setdefault(ln.level, set()).add(ln.code_site)
    result.totals = totals
    result.nb_sites = {lvl: len(codes) for lvl, codes in sites.items()}
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Persistance (isolee — ne touche JAMAIS le referentiel maitre)
# ─────────────────────────────────────────────────────────────────────────────


def persist_dpgf_p1_import(
    db: Session,
    result: DpgfP1ParseResult,
    current_user: User,
    *,
    deactivate_previous: bool = True,
    acte_label: str | None = None,
    date_effet=None,
) -> CpeDpgfP1Import:
    """Persiste un import DPGF P1 dans sa lignee propre.

    ``deactivate_previous`` ne desactive QUE les imports DPGF P1 precedents du MEME lot
    (table ``cpe_dpgf_p1_imports``). Cette fonction n'appelle pas ``persist_dalkia_import``
    et ne modifie aucun ``cpe_dalkia_ref_import`` : P2/P3/APE/cibles/RECAP du maitre intacts.
    """
    city_id = current_user.city_id

    if deactivate_previous:
        prev_stmt = select(CpeDpgfP1Import).where(
            CpeDpgfP1Import.lot == result.lot,
            CpeDpgfP1Import.is_active.is_(True),
        )
        if city_id is not None:
            prev_stmt = prev_stmt.where(CpeDpgfP1Import.city_id == city_id)
        for prev in db.scalars(prev_stmt):
            prev.is_active = False
            db.add(prev)

    batch = CpeDpgfP1Import(
        city_id=city_id,
        lot=result.lot,
        filename=result.filename,
        nb_lines=len(result.lines),
        is_active=True,
        notes=f"Warnings: {len(result.warnings)}" if result.warnings else None,
        acte_type="dpgf",
        acte_label=acte_label,
        date_effet=date_effet,
    )
    db.add(batch)
    db.flush()  # batch.id

    for ln in result.lines:
        db.add(CpeDpgfP1Line(
            import_id=batch.id,
            city_id=city_id,
            lot=result.lot,
            level=ln.level,
            code_site=ln.code_site,
            pce=ln.pce,
            type_tarif=ln.type_tarif,
            prix_unitaire_ht=ln.prix_unitaire_ht,
            period_idx=ln.period_idx,
            period_label=ln.period_label,
            period_year=ln.period_year,
            qt_mwhpcs=ln.qt_mwhpcs,
            p10_var_ht=ln.p10_var_ht,
            p10_total_ht=ln.p10_total_ht,
        ))

    db.commit()
    db.refresh(batch)
    return batch


# ─────────────────────────────────────────────────────────────────────────────
# Lecture
# ─────────────────────────────────────────────────────────────────────────────


def update_dpgf_p1_acte(
    db: Session, import_id: int, current_user: User, *,
    acte_label: str | None, date_effet,
) -> CpeDpgfP1Import | None:
    """Qualifie un import DPGF P1 (libellé / date d'effet) — éditable a posteriori."""
    stmt = select(CpeDpgfP1Import).where(CpeDpgfP1Import.id == import_id)
    if current_user.city_id is not None:
        stmt = stmt.where(CpeDpgfP1Import.city_id == current_user.city_id)
    imp = db.scalars(stmt).first()
    if imp is None:
        return None
    imp.acte_label = acte_label
    imp.date_effet = date_effet
    db.add(imp)
    db.commit()
    db.refresh(imp)
    return imp


def get_active_dpgf_p1_imports(db: Session, current_user: User) -> list[CpeDpgfP1Import]:
    stmt = select(CpeDpgfP1Import).where(CpeDpgfP1Import.is_active.is_(True))
    if current_user.city_id is not None:
        stmt = stmt.where(CpeDpgfP1Import.city_id == current_user.city_id)
    return list(db.scalars(stmt.order_by(CpeDpgfP1Import.lot, CpeDpgfP1Import.import_date.desc())))


def get_all_dpgf_p1_imports(
    db: Session, current_user: User, *, lot: int | None = None
) -> list[CpeDpgfP1Import]:
    """Tous les imports DPGF P1 (actifs ET remplaces conserves) — pour le journal du marche."""
    stmt = select(CpeDpgfP1Import)
    if current_user.city_id is not None:
        stmt = stmt.where(CpeDpgfP1Import.city_id == current_user.city_id)
    if lot is not None:
        stmt = stmt.where(CpeDpgfP1Import.lot == lot)
    return list(db.scalars(stmt.order_by(CpeDpgfP1Import.import_date.desc())))


def get_dpgf_p1_levels(
    db: Session,
    city_id: int | None,
    years: list[int],
    *,
    lot: int | None = None,
) -> dict[str, dict[int, float]]:
    """Totaux P1 revises (somme p10_total_ht) par niveau x annee, des imports DPGF P1 actifs.

    Retourne ``{level: {year: total}}`` filtre sur ``years`` (et eventuellement ``lot``).
    Consomme en mode informatif par le suivi marche — ne change pas le ``prevu P1``.
    """
    year_set = set(years)
    stmt = (
        select(CpeDpgfP1Line)
        .join(CpeDpgfP1Import, CpeDpgfP1Line.import_id == CpeDpgfP1Import.id)
        .where(CpeDpgfP1Import.is_active.is_(True))
    )
    if city_id is not None:
        stmt = stmt.where(CpeDpgfP1Import.city_id == city_id)
    if lot is not None:
        stmt = stmt.where(CpeDpgfP1Import.lot == lot)

    out: dict[str, dict[int, float]] = {lvl: {y: 0.0 for y in years} for lvl in DPGF_P1_LEVELS}
    for ln in db.scalars(stmt):
        if ln.period_year not in year_set:
            continue
        out.setdefault(ln.level, {y: 0.0 for y in years})
        out[ln.level][ln.period_year] = round(
            out[ln.level].get(ln.period_year, 0.0) + (ln.p10_total_ht or 0.0), 2
        )
    return out
