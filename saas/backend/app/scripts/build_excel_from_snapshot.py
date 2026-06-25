"""Branche le snapshot football-data.org dans le classeur Excel du modele.

Ajoute, de facon ADDITIVE et non destructive, deux feuilles de reference
alimentees uniquement par les donnees FACTUELLES de l'API (provenance = API) :

- ``10_API_Equipes``  : 48 equipes (nom officiel, zone, coach, taille effectif,
  forme reelle EURO 2024 quand disponible).
- ``11_API_Effectifs``: pool complet des ~1244 joueurs (poste FR, nationalite,
  age) dans lequel piocher les joueurs retenus de ``03_Joueurs_Retenus``.

Les feuilles du modele existant (formules, parametres) ne sont PAS modifiees.
Les donnees absentes de l'API (Elo, FIFA, forme des non-Europeens) restent a
completer par recherche web -> elles ne sont volontairement pas ecrites ici.

Usage (depuis saas/backend/) :
    python -m app.scripts.build_excel_from_snapshot \
        --snapshot ../../outputs/pronostics_model_v0/football_data_snapshot.json \
        --in  ../../outputs/pronostics_model_v0/modele_pronostics_cdm2026_v0.xlsx \
        --out ../../outputs/pronostics_model_v0/modele_pronostics_cdm2026_v1.xlsx
"""

from __future__ import annotations

import argparse
import json
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.services.pronostics import _normalize_team

POSITION_FR = {
    "Goalkeeper": "Gardien",
    "Defence": "Defenseur",
    "Midfield": "Milieu",
    "Offence": "Attaquant",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _age(dob: str | None) -> int | None:
    if not dob:
        return None
    try:
        y, m, d = (int(x) for x in dob[:10].split("-"))
    except ValueError:
        return None
    today = date.today()
    return today.year - y - ((today.month, today.day) < (m, d))


def _reset_sheet(wb: openpyxl.Workbook, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--snapshot", required=True)
    parser.add_argument("--in", dest="src", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    feed = json.load(open(args.snapshot, encoding="utf-8"))
    releve = date.today().isoformat()

    wb = openpyxl.load_workbook(args.src)

    # Nom FR canonique du classeur, indexe par cle normalisee.
    equipes_ws = wb["02_Equipes"]
    excel_name_by_key: dict[str, str] = {}
    for row in range(2, equipes_ws.max_row + 1):
        name = equipes_ws.cell(row, 1).value
        if name:
            excel_name_by_key[_normalize_team(name)] = name

    def fr_name(team: dict) -> str:
        key = _normalize_team(team.get("official_name") or "")
        return excel_name_by_key.get(key, team.get("official_name") or "")

    # --- 10_API_Equipes ---
    eq = _reset_sheet(wb, "10_API_Equipes")
    _write_header(
        eq,
        [
            "Equipe", "Nom_API", "Zone", "Coach", "Coach_nationalite",
            "Nb_effectif", "EURO_J", "EURO_V", "EURO_N", "EURO_D",
            "EURO_BP", "EURO_BC", "EURO_CS", "BP/match", "BC/match",
            "Source", "Date_releve",
        ],
    )
    for team in sorted(feed["teams"], key=fr_name):
        coach = team.get("coach") or {}
        rf = team.get("recent_form") or {}
        played = rf.get("played_count") or 0
        eq.append([
            fr_name(team),
            team.get("official_name"),
            team.get("area_name"),
            coach.get("name"),
            coach.get("nationality"),
            team.get("squad_count"),
            played or None,
            rf.get("wins") if played else None,
            rf.get("draws") if played else None,
            rf.get("losses") if played else None,
            rf.get("goals_for") if played else None,
            rf.get("goals_against") if played else None,
            rf.get("clean_sheets") if played else None,
            rf.get("goals_for_per_match") if played else None,
            rf.get("goals_against_per_match") if played else None,
            "football-data.org",
            releve,
        ])

    # --- 11_API_Effectifs ---
    ef = _reset_sheet(wb, "11_API_Effectifs")
    _write_header(
        ef,
        ["Equipe", "Joueur", "Poste", "Poste_API", "Nationalite", "Naissance", "Age", "Source", "Date_releve"],
    )
    key_to_fr = {t["local_team_key"]: fr_name(t) for t in feed["teams"]}
    for player in sorted(feed["players"], key=lambda p: (key_to_fr.get(p["local_team_key"], ""), p.get("name") or "")):
        pos = player.get("position")
        ef.append([
            key_to_fr.get(player["local_team_key"], player.get("team_official_name")),
            player.get("name"),
            POSITION_FR.get(pos, pos),
            pos,
            player.get("nationality"),
            player.get("date_of_birth"),
            _age(player.get("date_of_birth")),
            "football-data.org",
            releve,
        ])

    wb.save(args.out)

    teams_with_form = sum(1 for t in feed["teams"] if (t.get("recent_form") or {}).get("played_count"))
    print(f"OK -> {args.out}")
    print(f"  10_API_Equipes : {len(feed['teams'])} equipes ({teams_with_form} avec forme EURO 2024)")
    print(f"  11_API_Effectifs : {len(feed['players'])} joueurs")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
