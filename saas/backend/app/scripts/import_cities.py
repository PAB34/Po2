from __future__ import annotations

import argparse
from pathlib import Path
import sys
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.db import SessionLocal
from app.models import City


def _normalize_city_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    return " ".join(normalized.strip().split()).upper()


def _iter_commune_names(workbook_path: Path) -> tuple[str, list[str]]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    source_name = workbook_path.name
    names: list[str] = []

    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=22, values_only=True):
        commune_name = row[0]
        category = row[18]

        if commune_name is None or category is None:
            continue

        commune_text = str(commune_name).strip()
        category_text = str(category).strip()

        if not commune_text or category_text != "4 - Commune":
            continue

        names.append(commune_text)

    workbook.close()
    return source_name, names


def import_cities_from_xlsx(workbook_path: Path, truncate: bool) -> tuple[int, int, int]:
    source_name, imported_names = _iter_commune_names(workbook_path)

    unique_names: dict[str, str] = {}
    for name in imported_names:
        normalized_name = _normalize_city_name(name)
        unique_names.setdefault(normalized_name, name.strip())

    if not unique_names:
        raise ValueError("Aucune commune correspondant au filtre '4 - Commune' n'a été trouvée.")

    with SessionLocal() as db:
        if truncate:
            db.query(City).delete()
            db.flush()

        existing_cities = {
            _normalize_city_name(city.nom_commune): city
            for city in db.scalars(select(City)).all()
        }

        created_count = 0
        updated_count = 0

        for normalized_name, display_name in sorted(unique_names.items()):
            city = existing_cities.get(normalized_name)
            if city is None:
                db.add(City(nom_commune=display_name, source_file=source_name))
                created_count += 1
                continue

            changed = False
            if city.nom_commune != display_name:
                city.nom_commune = display_name
                changed = True
            if city.source_file != source_name:
                city.source_file = source_name
                changed = True
            if changed:
                updated_count += 1

        db.commit()

    return len(unique_names), created_count, updated_count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook_path")
    parser.add_argument("--truncate", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook_path)
    if not workbook_path.exists():
        print(f"Fichier introuvable: {workbook_path}", file=sys.stderr)
        return 1

    try:
        total_count, created_count, updated_count = import_cities_from_xlsx(
            workbook_path=workbook_path,
            truncate=args.truncate,
        )
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1

    print(
        f"Communes filtrées: {total_count} | créées: {created_count} | mises à jour: {updated_count}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
