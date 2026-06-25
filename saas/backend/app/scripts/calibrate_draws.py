"""Calibre le seuil de nul sur les probabilites du modele lui-meme.

Lit les vraies probabilites de Poisson (06_Diagnostic cols S/T/U) DEJA
recalculees, puis fixe ``01_Parametres!Seuil_nul_prioritaire`` de sorte que le
NOMBRE de nuls predits = nombre de nuls ATTENDU par le modele (somme des P(nul)).

Le taux de nuls n'est donc pas un % choisi : il emerge des buts attendus (lambda).
A relancer apres chaque recalcul (les lambda changent avec les donnees).

Regle de prediction (formule 07) : nul si P(nul) >= max(P_A,P_B) - seuil,
c.-a-d. ecart = max(P_A,P_B) - P(nul) <= seuil. On choisit donc seuil = le
N-ieme plus petit ecart, ou N = round(somme des P(nul)).

Usage : python -m app.scripts.calibrate_draws --file <classeur_recalcule.xlsx>
"""

from __future__ import annotations

import argparse

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    vals = openpyxl.load_workbook(args.file, data_only=True)["06_Diagnostic_Match"]
    gaps, expected = [], 0.0
    for r in range(2, vals.max_row + 1):
        pa, pn, pb = vals.cell(r, 19).value, vals.cell(r, 20).value, vals.cell(r, 21).value
        if pa is None or pn is None or pb is None:
            continue
        gaps.append(max(pa, pb) - pn)
        expected += pn
    if not gaps:
        print("Aucune proba trouvee (recalcule d'abord le classeur).")
        return 1

    gaps.sort()
    n = len(gaps)
    expected_draws = max(0, min(n, round(expected)))
    if expected_draws == 0:
        seuil = 0.0
    else:
        # seuil = N-ieme plus petit ecart (+epsilon) -> ~expected_draws nuls predits
        seuil = round(gaps[expected_draws - 1] + 0.0005, 4)

    wb = openpyxl.load_workbook(args.file)
    par = wb["01_Parametres"]
    for r in range(1, par.max_row + 1):
        if par.cell(r, 1).value == "Seuil_nul_prioritaire":
            par.cell(r, 2).value = seuil
            par.cell(r, 4).value = (
                f"Auto-calibre: nuls predits = attendus (somme P(nul)={expected:.1f} "
                f"sur {n} matchs, soit {round(100*expected/n)}%)"
            )
            break
    wb.save(args.file)
    print(f"OK -> {args.file}")
    print(f"  nuls attendus (somme P(nul)) = {expected:.1f} sur {n} ({round(100*expected/n)}%)")
    print(f"  seuil auto-calibre = {seuil} -> ~{expected_draws} nuls predits")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
