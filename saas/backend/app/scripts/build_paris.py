"""Genere les paris Winamax (mode Coupe du Monde) depuis le classement du modele.

Lit 13_Classement_Poules + Elo (02_Equipes), produit 5 marches :
  1. Vainqueur de groupe
  2. Deux premiers (sans ordre)
  3. Deux premiers dans l'ordre
  4. Dernier de groupe
  5. Qualification au Tour de barrage (2 premiers + 8 meilleurs 3es)

Pour chaque pari : marge de points a la frontiere decisive + coherence Elo ->
niveau de confiance -> style SECURE (marge confortable + Elo coherent) ou
COMPLET (tous). Ecrit la feuille 15_Paris_Winamax (sans toucher au reste).

Usage : python -m app.scripts.build_paris --file <classeur.xlsx>
"""

from __future__ import annotations

import argparse
import itertools

import openpyxl
from openpyxl.styles import Font, PatternFill

HF = PatternFill("solid", fgColor="C00000"); HFONT = Font(color="FFFFFF", bold=True)
SUB = PatternFill("solid", fgColor="F2F2F2"); SUBF = Font(bold=True)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    cl = wb["13_Classement_Poules"]; eq = wb["02_Equipes"]
    elo = {eq.cell(r, 1).value: (eq.cell(r, 11).value or 0) for r in range(2, eq.max_row + 1) if eq.cell(r, 1).value}

    teams = []
    for r in range(2, cl.max_row + 1):
        if not cl.cell(r, 1).value:
            continue
        teams.append({"g": cl.cell(r, 1).value, "eq": cl.cell(r, 2).value, "pts": cl.cell(r, 10).value,
                      "diff": cl.cell(r, 9).value, "bp": cl.cell(r, 7).value, "rang": cl.cell(r, 11).value})
    groups = {g: sorted(rows, key=lambda x: x["rang"]) for g, rows in itertools.groupby(sorted(teams, key=lambda x: x["g"]), key=lambda x: x["g"])}

    def elo_ok(a, b, tol=30):  # a au moins ~aussi fort que b (cohérence Elo)
        return elo.get(a, 0) >= elo.get(b, 0) - tol

    rows_out = []  # (marche, groupe, pari, marge, elo_gap, secure)

    # 1. Vainqueur de groupe
    for g, gr in groups.items():
        m = gr[0]["pts"] - gr[1]["pts"]
        secure = m >= 3 and elo_ok(gr[0]["eq"], gr[1]["eq"])
        rows_out.append(("1. Vainqueur de groupe", g, gr[0]["eq"], m, elo.get(gr[0]["eq"], 0) - elo.get(gr[1]["eq"], 0), secure))

    # 2. Deux premiers (sans ordre)
    for g, gr in groups.items():
        m = gr[1]["pts"] - gr[2]["pts"]   # marge du 2e sur le 3e = solidite du duo
        secure = m >= 3
        rows_out.append(("2. Deux premiers", g, f"{gr[0]['eq']} + {gr[1]['eq']}", m, 0, secure))

    # 3. Deux premiers dans l'ordre
    for g, gr in groups.items():
        m1 = gr[0]["pts"] - gr[1]["pts"]; m2 = gr[1]["pts"] - gr[2]["pts"]
        secure = m1 >= 3 and m2 >= 3 and elo_ok(gr[0]["eq"], gr[1]["eq"])
        rows_out.append(("3. Deux premiers dans l'ordre", g, f"1er {gr[0]['eq']}, 2e {gr[1]['eq']}", min(m1, m2), 0, secure))

    # 4. Dernier de groupe
    for g, gr in groups.items():
        m = gr[2]["pts"] - gr[3]["pts"]   # ecart 3e -> 4e
        secure = gr[3]["pts"] <= 1 and m >= 2
        rows_out.append(("4. Dernier de groupe", g, gr[3]["eq"], m, 0, secure))

    # 5. Qualification (2 premiers + 8 meilleurs 3es)
    thirds = sorted([gr[2] for gr in groups.values()], key=lambda x: (x["pts"], x["diff"], x["bp"]), reverse=True)
    qualified_thirds = thirds[:8]
    cut_pts = qualified_thirds[-1]["pts"]
    elim_pts = thirds[8]["pts"] if len(thirds) > 8 else -1   # 1er éliminé = ligne de coupure
    next_pts = elim_pts
    # Confiance d'une qualification = marge de points AU-DESSUS du 1er éliminé
    # (pour "se qualifier", c'est l'écart à la ligne d'élimination qui compte,
    #  pas l'écart au 3e du groupe). Vaut pour 1ers, 2es ET 3es qualifiés.
    for g, gr in groups.items():
        for idx, label in ((0, "1er"), (1, "2e")):
            buf = gr[idx]["pts"] - elim_pts
            rows_out.append(("5. Qualification (barrage)", g, f"{gr[idx]['eq']} ({label})", buf, 0, buf >= 3))
    for t in qualified_thirds:
        buf = t["pts"] - elim_pts
        rows_out.append(("5. Qualification (barrage)", t["g"], f"{t['eq']} (3e qualifié)", buf, 0, buf >= 3))

    # --- ecriture feuille 15 (mise en page claire, par marche) ---
    from openpyxl.styles import Alignment, Border, Side
    GREEN = PatternFill("solid", fgColor="C6EFCE"); GREENF = Font(color="006100", bold=True)
    GREY = PatternFill("solid", fgColor="F2F2F2"); GREYF = Font(color="808080")
    TITLE = PatternFill("solid", fgColor="1F4E78"); TITLEF = Font(color="FFFFFF", bold=True, size=12)
    thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    def confiance(marge):
        return "Élevée" if marge >= 3 else ("Moyenne" if marge >= 1 else "Faible (toss-up)")

    if "15_Paris_Winamax" in wb.sheetnames:
        del wb["15_Paris_Winamax"]
    ws = wb.create_sheet("15_Paris_Winamax")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26

    def put(row, vals, fill=None, font=None, align=None):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i); c.value = v
            if fill: c.fill = fill
            if font: c.font = font
            if align: c.alignment = Alignment(horizontal=align)
        return row + 1

    r = 1
    r = put(r, ["PARIS WINAMAX — Coupe du Monde 2026 (issus du modèle)"], TITLE, TITLEF); ws.merge_cells(f"A{r-1}:D{r-1}")
    r = put(r, ["🟢 SÉCURE = à jouer en priorité (marge confortable + cohérent Elo)  |  🔵 reste = style COMPLET seulement (plus risqué)"])
    r = put(r, ["Règle : 12 groupes de 4 → 2 premiers de chaque groupe + 8 meilleurs 3es = 32 qualifiés (Tour de barrage)"])
    r += 1

    marches = ["1. Vainqueur de groupe", "2. Deux premiers", "3. Deux premiers dans l'ordre",
               "4. Dernier de groupe", "5. Qualification (barrage)"]
    for marche in marches:
        sub = [x for x in rows_out if x[0] == marche]
        n_sec = sum(1 for x in sub if x[5])
        r = put(r, [f"{marche}   —   {n_sec} pari(s) en SÉCURE / {len(sub)} en COMPLET"], HF, HFONT); ws.merge_cells(f"A{r-1}:D{r-1}")
        r = put(r, ["Groupe", "Pari", "Confiance", "À jouer en SÉCURE ?"], GREY, Font(bold=True))
        for _, g, pari, marge, elog, secure in sub:
            fill = GREEN if secure else None
            font = GREENF if secure else GREYF
            cell_secure = "🟢 OUI" if secure else "— (complet seulement)"
            rr = put(r, [g, pari, confiance(marge), cell_secure], fill, font)
            for c in range(1, 5):
                ws.cell(r, c).border = thin
            r = rr
        r += 1

    # rappel des 8 meilleurs 3es retenus
    r = put(r, ["Détail — 8 meilleurs 3es retenus (ordre : pts → diff → buts)"], HF, HFONT); ws.merge_cells(f"A{r-1}:D{r-1}")
    for i, t in enumerate(qualified_thirds, 1):
        r = put(r, [t["g"], f"{i}. {t['eq']}", f"{t['pts']} pts, diff {t['diff']:+d}", ""])
    r = put(r, ["", f"Ligne de coupure : {cut_pts} pts (1er éliminé à {next_pts} pts)", "", ""], None, GREYF)
    ws.freeze_panes = "A5"

    wb.save(args.file)

    # --- sortie console : les 2 styles ---
    def show(marche):
        sub = [r for r in rows_out if r[0] == marche]
        sec = [r for r in sub if r[5]]
        print(f"\n=== {marche} ===")
        print(f"  SÉCURE ({len(sec)}): " + " | ".join(f"{r[1]}:{r[2]}" for r in sec))
        print(f"  COMPLET ({len(sub)}): " + " | ".join(f"{r[1]}:{r[2]}" for r in sub))
    for marche in ["1. Vainqueur de groupe", "2. Deux premiers", "3. Deux premiers dans l'ordre",
                   "4. Dernier de groupe", "5. Qualification (barrage)"]:
        show(marche)
    print(f"\n8 meilleurs 3es (ligne de coupure à {cut_pts} pts, premier éliminé à {next_pts} pts) :")
    print("  " + " | ".join(f"{t['eq']}({t['pts']}pts,{t['diff']:+d})" for t in qualified_thirds))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
