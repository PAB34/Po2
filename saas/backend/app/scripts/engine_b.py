"""Moteur B (test) : Poisson BIVARIE (Karlis-Ntzoufras / esprit Dixon-Coles).

Probleme du moteur A (Poisson independant) : pour avoir des nuls realistes il
faut baisser les buts -> triangle buts/nuls/ecarts insoluble.

Moteur B : on garde les moyennes de buts du moteur A (mu_A, mu_B = lambda de
06_Diagnostic) mais on ajoute une COVARIANCE lambda3 entre les deux scores qui
gonfle les nuls SANS toucher les moyennes (E[A]=lambda1+lambda3=mu_A). On
calibre lambda3 (via un coefficient rho) pour viser un taux de nuls cible, tout
en preservant buts moyens et gros ecarts. MOTEUR PROMU : ecrit directement
07_Modele_Probable (Proba_A/Nul/B bivariees en G/H/I, Score en J/K/L). rho
auto-calibre pour viser 01_Parametres!Cible_taux_nul, ecrit dans
01_Parametres!Rho_nul_bivarie. Buts realistes (round(mu)) + gros ecarts gardes ;
seuls les nuls sont gonfles -> 25% nuls + ~2.6 buts + gros ecarts ensemble.

A lancer APRES un recalcul (lit mu=lambda dans 07 E/F). Pipeline :
... -> recalc -> engine_b -> recalc -> compute_standings -> recalc.

Usage : python -m app.scripts.engine_b --file <classeur_recalcule.xlsx> [--cible-nuls 0.25]
"""

from __future__ import annotations

import argparse
import math

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="7030A0")
HEADER_FONT = Font(color="FFFFFF", bold=True)
N = 9   # scores 0..8
CAP = 7


def joint_matrix(mu_a: float, mu_b: float, lam3: float):
    l1 = max(0.0, mu_a - lam3)
    l2 = max(0.0, mu_b - lam3)
    base = math.exp(-(l1 + l2 + lam3))
    P = [[0.0] * N for _ in range(N)]
    for x in range(N):
        for y in range(N):
            s = 0.0
            for k in range(min(x, y) + 1):
                s += (l1 ** (x - k) / math.factorial(x - k)) * (l2 ** (y - k) / math.factorial(y - k)) * (lam3 ** k / math.factorial(k))
            P[x][y] = base * s
    return P


def proba_bivar(mu_a, mu_b, rho):
    """P(victoire A / nul / victoire B) en Poisson bivarie (covariance rho*min(mu))."""
    lam3 = rho * min(mu_a, mu_b)
    P = joint_matrix(mu_a, mu_b, lam3)
    pa = sum(P[x][y] for x in range(N) for y in range(N) if x > y)
    pn = sum(P[i][i] for i in range(N))
    pb = sum(P[x][y] for x in range(N) for y in range(N) if x < y)
    return pa, pn, pb


def score_from_means(mu_a, mu_b, is_draw, fav_a, cap=CAP):
    """Score realiste depuis les moyennes (comme moteur A) : round(mu), favori strictement devant."""
    if is_draw:
        g = min(cap, round((mu_a + mu_b) / 2))
        return g, g
    ga, gb = min(cap, round(mu_a)), min(cap, round(mu_b))
    if fav_a and ga <= gb:
        ga = min(cap, gb + 1)
    if (not fav_a) and gb <= ga:
        gb = min(cap, ga + 1)
    return ga, gb


def _read_param(par, name, default):
    for r in range(1, par.max_row + 1):
        if par.cell(r, 1).value == name:
            v = par.cell(r, 2).value
            return v if v is not None else default
    return default


def _set_param(par, name, value, role="", comment=""):
    for r in range(1, par.max_row + 1):
        if par.cell(r, 1).value == name:
            par.cell(r, 2).value = value
            if comment:
                par.cell(r, 4).value = comment
            return
    r = par.max_row + 1
    par.cell(r, 1).value = name; par.cell(r, 2).value = value
    par.cell(r, 3).value = role; par.cell(r, 4).value = comment


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    ap.add_argument("--cible-nuls", type=float, default=None,
                    help="Sinon lu depuis 01_Parametres!Cible_taux_nul (defaut 0.25).")
    args = ap.parse_args()

    vals = openpyxl.load_workbook(args.file, data_only=True)
    mp = vals["07_Modele_Probable"]
    par_v = vals["01_Parametres"]
    cible = args.cible_nuls if args.cible_nuls is not None else float(_read_param(par_v, "Cible_taux_nul", 0.25))
    cap = int(_read_param(par_v, "Plafond_buts_modele", 7))

    matches = []  # (row, A, B, muA, muB)
    for r in range(2, mp.max_row + 1):
        a, b = mp.cell(r, 3).value, mp.cell(r, 4).value
        ma, mb = mp.cell(r, 5).value, mp.cell(r, 6).value
        if not a or ma is None:
            continue
        matches.append([r, a, b, ma, mb])

    # calibrer rho (covariance) pour que la MASSE de nuls bivariee = cible
    target_mass = cible * len(matches)

    def draw_mass(rho):
        return sum(proba_bivar(ma, mb, rho)[1] for _, _, _, ma, mb in matches)

    lo, hi, best_rho = 0.0, 0.95, 0.0
    if draw_mass(0.0) < target_mass:
        for _ in range(50):
            mid = (lo + hi) / 2
            if draw_mass(mid) < target_mass:
                lo = mid
            else:
                hi = mid
            best_rho = mid

    probs = {row: proba_bivar(ma, mb, best_rho) for row, a, b, ma, mb in matches}
    n_draws = round(sum(p[1] for p in probs.values()))
    order = sorted(matches, key=lambda m: max(probs[m[0]][0], probs[m[0]][2]) - probs[m[0]][1])
    draw_rows = {order[i][0] for i in range(n_draws)}

    # ECRIRE dans 07 (moteur B promu) : probas bivariees (G/H/I) + score (J/K/L)
    wb = openpyxl.load_workbook(args.file)
    if "14_Moteur_B" in wb.sheetnames:   # ancienne feuille de comparaison, obsolete
        del wb["14_Moteur_B"]
    mpw = wb["07_Modele_Probable"]
    from collections import Counter
    res = Counter(); goals = []; big = 0
    for row, a, b, ma, mb in matches:
        pa, pn, pb = probs[row]
        ga, gb = score_from_means(ma, mb, row in draw_rows, pa >= pb, cap)
        mpw.cell(row, 7).value = round(pa, 3)   # Proba_A (bivariee)
        mpw.cell(row, 8).value = round(pn, 3)   # Proba_Nul
        mpw.cell(row, 9).value = round(pb, 3)   # Proba_B
        mpw.cell(row, 10).value = ga            # Score_Probable_A
        mpw.cell(row, 11).value = gb            # Score_Probable_B
        mpw.cell(row, 12).value = f"{ga}-{gb}"  # Score_Probable
        res["N" if ga == gb else ("A" if ga > gb else "B")] += 1
        goals.append(ga + gb)
        if abs(ga - gb) >= 3:
            big += 1

    _set_param(wb["01_Parametres"], "Rho_nul_bivarie", round(best_rho, 3), "Moteur B",
               f"Covariance auto-calibree pour viser Cible_taux_nul ({int(cible*100)}%) - ne pas editer a la main")
    wb.save(args.file)

    n = len(matches)
    print(f"OK -> 07 (moteur B bivarie promu) | rho={round(best_rho,3)} cible={int(cible*100)}%")
    print(f"  nuls {res['N']} ({round(100*res['N']/n)}%) | buts/match {round(sum(goals)/n,2)} | ecarts>=3 {big} | plafond {cap}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
