"""Moteur Monte-Carlo : simule la phase de groupes N fois -> probabilites.

Resout le probleme des classements "pile/face" : au lieu d'un seul score modal
par match (qui cree des egalites parfaites tranchees a l'Elo), on TIRE chaque
match dans la loi de Poisson bivariee deja calibree (mu_A, mu_B = lambda de 07,
covariance rho = 01_Parametres!Rho_nul_bivarie), N fois, et on agrege.

Sortie par equipe : P(vainqueur de groupe), P(top 2), P(qualifie au barrage),
P(dernier). Ecrit la feuille 16_Probabilites + regenere 15_Paris_Winamax avec
confiance = probabilite reelle. Ne touche pas a 07/13 (scenario central conserve).

Echantillonnage bivarie (Karlis-Ntzoufras) : Z3~Pois(lam3), X=Pois(mu_A-lam3)+Z3,
Y=Pois(mu_B-lam3)+Z3, avec lam3=rho*min(mu_A,mu_B) -> E[X]=mu_A, covariance via Z3.

Usage : python -m app.scripts.monte_carlo --file <classeur.xlsx> [--n 10000]
"""

from __future__ import annotations

import argparse
import math

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

N_QUALIF_THIRDS = 8
# seuils "SÉCURE" par marché (probabilité)
SEUIL = {"vainqueur": 0.55, "top2": 0.60, "ordre": 0.45, "dernier": 0.50, "qualif": 0.75}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    ap.add_argument("--n", type=int, default=10000)
    args = ap.parse_args()
    N = args.n
    rng = np.random.default_rng(12345)

    vals = openpyxl.load_workbook(args.file, data_only=True)
    mp = vals["07_Modele_Probable"]; par = vals["01_Parametres"]
    rho = next((par.cell(r, 2).value for r in range(1, par.max_row + 1)
                if par.cell(r, 1).value == "Rho_nul_bivarie"), 0.0) or 0.0

    matches = []  # (group, A, B, muA, muB)
    for r in range(2, mp.max_row + 1):
        g, a, b = mp.cell(r, 2).value, mp.cell(r, 3).value, mp.cell(r, 4).value
        ma, mb = mp.cell(r, 5).value, mp.cell(r, 6).value
        if not a or ma is None:
            continue
        matches.append((g, a, b, float(ma), float(mb)))

    teams = sorted({t for _, a, b, _, _ in matches for t in (a, b)})
    idx = {t: i for i, t in enumerate(teams)}
    team_group = {}
    for g, a, b, _, _ in matches:
        team_group[a] = g; team_group[b] = g
    groups = sorted(set(team_group.values()))
    gteams = {g: [t for t in teams if team_group[t] == g] for g in groups}

    # --- echantillonnage vectorise de tous les matchs : X,Y de forme (nmatch, N) ---
    X = np.zeros((len(matches), N), dtype=np.int16)
    Y = np.zeros((len(matches), N), dtype=np.int16)
    for m, (_, _, _, ma, mb) in enumerate(matches):
        lam3 = rho * min(ma, mb)
        z3 = rng.poisson(lam3, N)
        X[m] = rng.poisson(max(0.0, ma - lam3), N) + z3
        Y[m] = rng.poisson(max(0.0, mb - lam3), N) + z3

    nt = len(teams)
    win = np.zeros(nt); top2 = np.zeros(nt); last = np.zeros(nt); qualif = np.zeros(nt)
    rank_sum = np.zeros(nt)
    from collections import Counter
    pair_ord = {g: Counter() for g in groups}   # (1er, 2e) exact
    pair_set = {g: Counter() for g in groups}    # {top2} sans ordre

    # index des matchs par groupe
    gmatch = {g: [m for m, (gg, *_) in enumerate(matches) if gg == g] for g in groups}

    for s in range(N):
        pts = np.zeros(nt); gf = np.zeros(nt); ga = np.zeros(nt)
        for m, (g, a, b, _, _) in enumerate(matches):
            x, y = X[m, s], Y[m, s]
            ia, ib = idx[a], idx[b]
            gf[ia] += x; ga[ia] += y; gf[ib] += y; ga[ib] += x
            if x > y: pts[ia] += 3
            elif x < y: pts[ib] += 3
            else: pts[ia] += 1; pts[ib] += 1
        gd = gf - ga
        thirds = []
        for g in groups:
            ts = gteams[g]
            # cle de tri : pts, diff, buts, aleatoire (depart neutre des egalites)
            order = sorted(ts, key=lambda t: (pts[idx[t]], gd[idx[t]], gf[idx[t]], rng.random()), reverse=True)
            for rank, t in enumerate(order):
                i = idx[t]; rank_sum[i] += rank + 1
                if rank == 0: win[i] += 1
                if rank <= 1: top2[i] += 1; qualif[i] += 1
                if rank == 3: last[i] += 1
            pair_ord[g][(order[0], order[1])] += 1          # ordre exact 1er-2e
            pair_set[g][frozenset((order[0], order[1]))] += 1  # duo top2 sans ordre
            thirds.append(order[2])
        # 8 meilleurs 3es qualifies
        thirds.sort(key=lambda t: (pts[idx[t]], gd[idx[t]], gf[idx[t]], rng.random()), reverse=True)
        for t in thirds[:N_QUALIF_THIRDS]:
            qualif[idx[t]] += 1

    P = {t: {"win": win[idx[t]] / N, "top2": top2[idx[t]] / N, "qualif": qualif[idx[t]] / N,
             "last": last[idx[t]] / N, "rank": rank_sum[idx[t]] / N} for t in teams}
    # combinaisons les plus probables + leur VRAIE proba jointe
    best_ord = {g: (lambda mc: (mc[0][0], mc[0][1] / N))(pair_ord[g].most_common(1)) for g in groups}
    best_set = {g: (lambda mc: (tuple(sorted(mc[0][0], key=lambda t: -P[t]["win"])), mc[0][1] / N))(pair_set[g].most_common(1)) for g in groups}

    _write(args.file, groups, gteams, P, best_ord, best_set)
    # apercu console
    print(f"Monte-Carlo OK : {N} simulations, rho={rho}")
    for g in groups[:3]:
        ts = sorted(gteams[g], key=lambda t: -P[t]["qualif"])
        print(f"  {g}: " + " | ".join(f"{t[:9]} 1er{P[t]['win']*100:.0f}% top2 {P[t]['top2']*100:.0f}% qualif {P[t]['qualif']*100:.0f}%" for t in ts))
    return 0


def _write(path, groups, gteams, P, best_ord, best_set):
    HF = PatternFill("solid", fgColor="2E7D32"); HFONT = Font(color="FFFFFF", bold=True)
    GREEN = PatternFill("solid", fgColor="C6EFCE"); GREENF = Font(color="006100", bold=True)
    GREYF = Font(color="808080"); TITLE = PatternFill("solid", fgColor="1F4E78"); TITLEF = Font(color="FFFFFF", bold=True, size=12)
    PURP = PatternFill("solid", fgColor="C00000")
    wb = openpyxl.load_workbook(path)

    # ---- 13_Classement_Poules : ré-aligné sur le Monte-Carlo (plus de pile/face Elo) ----
    if "13_Classement_Poules" in wb.sheetnames:
        cl = wb["13_Classement_Poules"]
        # lire les lignes existantes par (groupe, equipe)
        rowvals = {}
        for r in range(2, cl.max_row + 1):
            g = cl.cell(r, 1).value
            if g:
                rowvals[(g, cl.cell(r, 2).value)] = [cl.cell(r, c).value for c in range(1, 11)]  # A..J (sans Rang)
        # réécrire chaque groupe trié par rang moyen Monte-Carlo (croissant = meilleur)
        rr = 2
        for g in groups:
            for t in sorted(gteams[g], key=lambda t: P[t]["rank"]):
                base = rowvals.get((g, t), [g, t, "", "", "", "", "", "", "", ""])
                for c, v in enumerate(base, 1):
                    cl.cell(rr, c).value = v
                cl.cell(rr, 11).value = sorted(gteams[g], key=lambda t: P[t]["rank"]).index(t) + 1  # rang MC
                rr += 1
        # note d'entête
        cl.cell(1, 13).value = "Classement = rang le plus probable (Monte-Carlo), cohérent avec 15/16. Les égalités ne sont plus tranchées à l'Elo."

    # ---- 16_Probabilites ----
    if "16_Probabilites" in wb.sheetnames:
        del wb["16_Probabilites"]
    ws = wb.create_sheet("16_Probabilites")
    ws.append(["Groupe", "Equipe", "P_vainqueur_%", "P_top2_%", "P_qualifie_%", "P_dernier_%", "Rang_moyen"])
    for c in range(1, 8):
        ws.cell(1, c).fill = HF; ws.cell(1, c).font = HFONT
    ws.freeze_panes = "A2"
    for g in groups:
        for t in sorted(gteams[g], key=lambda t: -P[t]["qualif"]):
            p = P[t]
            ws.append([g, t, round(p["win"]*100, 1), round(p["top2"]*100, 1),
                       round(p["qualif"]*100, 1), round(p["last"]*100, 1), round(p["rank"], 2)])
    for col, w in zip("ABCDEFG", [8, 22, 14, 12, 13, 12, 11]):
        ws.column_dimensions[col].width = w

    # ---- 15_Paris_Winamax (confiance = probabilite) ----
    def conf(p):
        return "Élevée" if p >= 0.70 else ("Moyenne" if p >= 0.50 else "Faible")
    if "15_Paris_Winamax" in wb.sheetnames:
        del wb["15_Paris_Winamax"]
    ps = wb.create_sheet("15_Paris_Winamax")
    for col, w in zip("ABCD", [8, 46, 16, 26]):
        ps.column_dimensions[col].width = w
    r = 1
    def put(vals, fill=None, font=None):
        nonlocal r
        for i, v in enumerate(vals, 1):
            c = ps.cell(r, i); c.value = v
            if fill: c.fill = fill
            if font: c.font = font
        r += 1
    put(["PARIS WINAMAX — CDM 2026 — moteur MONTE-CARLO (probabilités sur 10 000 simulations)"], TITLE, TITLEF); ps.merge_cells(f"A{r-1}:D{r-1}")
    put(["🟢 SÉCURE = probabilité élevée (seuils : vainqueur≥55% · top2≥60% · dernier≥50% · qualif≥75%)  |  reste = COMPLET"])
    put(["Règle : 12 groupes → 2 premiers + 8 meilleurs 3es = 32 qualifiés. Les % remplacent l'ancienne 'marge de points'."])
    r += 1

    def section(title):
        nonlocal r
        put([title], PURP, HFONT); ps.merge_cells(f"A{r-1}:D{r-1}")
        put(["Groupe", "Pari", "Probabilité", "À jouer en SÉCURE ?"], None, Font(bold=True))
    def line(g, pari, p, secure):
        put([g, pari, f"{p*100:.0f}%", "🟢 OUI" if secure else "— (complet seulement)"],
            GREEN if secure else None, GREENF if secure else GREYF)

    # 1. Vainqueur
    section("1. Vainqueur de groupe")
    for g in groups:
        t = max(gteams[g], key=lambda t: P[t]["win"]); p = P[t]["win"]
        line(g, t, p, p >= SEUIL["vainqueur"])
    r += 1
    # 2. Deux premiers (proba JOINTE que ce duo finisse top2, sans ordre)
    section("2. Deux premiers (sans ordre) — proba que le duo finisse top 2")
    for g in groups:
        (a, b), p = best_set[g]
        line(g, f"{a} + {b}", p, p >= SEUIL["top2"])
    r += 1
    # 3. Deux premiers dans l'ordre (proba JOINTE de l'ordre exact 1er-2e)
    section("3. Deux premiers dans l'ordre — proba de l'ordre exact 1er/2e")
    for g in groups:
        (first, second), p = best_ord[g]
        line(g, f"1er {first}, 2e {second}", p, p >= SEUIL["ordre"])
    r += 1
    # 4. Dernier
    section("4. Dernier de groupe")
    for g in groups:
        t = max(gteams[g], key=lambda t: P[t]["last"]); p = P[t]["last"]
        line(g, t, p, p >= SEUIL["dernier"])
    r += 1
    # 5. Qualification
    section("5. Qualification (barrage) — équipes avec P ≥ 50%")
    allq = sorted(((g, t, P[t]["qualif"]) for g in groups for t in gteams[g]), key=lambda x: -x[2])
    for g, t, p in allq:
        if p >= 0.50:
            line(g, t, p, p >= SEUIL["qualif"])

    ps.freeze_panes = "A5"

    # ---- 17_Tickets : 5 combinés prêts à jouer, du plus sûr au plus fou ----
    def leg(g, mk):
        if mk == "Q":
            t = max(gteams[g], key=lambda t: P[t]["qualif"]); return (f"{t} qualifié", P[t]["qualif"])
        if mk == "V":
            t = max(gteams[g], key=lambda t: P[t]["win"]); return (f"{t} vainqueur groupe {g}", P[t]["win"])
        if mk == "D2":
            (a, b), p = best_set[g]; return (f"{a} + {b} (top 2)", p)
        if mk == "O":
            (f1, f2), p = best_ord[g]; return (f"1er {f1}, 2e {f2}", p)
        if mk == "L":
            t = max(gteams[g], key=lambda t: P[t]["last"]); return (f"{t} dernier", P[t]["last"])

    tickets = [
        ("T1 — ULTRA-SÛR", 15, [("C", "Q"), ("J", "Q"), ("H", "D2")]),
        ("T2 — SÛR", 12, [("J", "V"), ("I", "V"), ("H", "D2"), ("L", "D2"), ("E", "L")]),
        ("T3 — ÉQUILIBRÉ", 10, [("J", "V"), ("I", "V"), ("H", "V"), ("A", "V"), ("B", "V"), ("C", "V")]),
        ("T4 — AUDACIEUX", 8, [("H", "O"), ("L", "O"), ("G", "V"), ("F", "D2"), ("E", "V"), ("I", "L")]),
        ("T5 — FOU (loterie)", 5, [("H", "O"), ("L", "O"), ("C", "O"), ("I", "O"), ("J", "O"), ("K", "O"), ("B", "O")]),
    ]
    if "17_Tickets" in wb.sheetnames:
        del wb["17_Tickets"]
    ts = wb.create_sheet("17_Tickets")
    for col, w in zip("ABCDE", [10, 46, 12, 14, 14]):
        ts.column_dimensions[col].width = w
    rr = 1
    def tput(vals, fill=None, font=None):
        nonlocal rr
        for i, v in enumerate(vals, 1):
            c = ts.cell(rr, i); c.value = v
            if fill: c.fill = fill
            if font: c.font = font
        rr += 1
    tput(["5 TICKETS COMBINÉS — du plus sûr au plus fou (mises totales = 50 €)"], TITLE, TITLEF); ts.merge_cells(f"A{rr-1}:E{rr-1}")
    tput(["Proba = chance modèle que TOUT le ticket passe (optimiste vs cotes réelles). Cote : à lire sur Winamax."], None, GREYF)
    rr += 1
    for name, stake, legs in tickets:
        legres = [leg(g, mk) for g, mk in legs]
        P_all = math.prod(p for _, p in legres)
        tput([f"{name}", f"MISE : {stake} €", f"{len(legres)} jambes", f"proba ~{P_all*100:.0f}%", ""], PURP, HFONT)
        tput(["Groupe", "Pari", "Proba", "", ""], None, Font(bold=True))
        for (g, mk), (lab, p) in zip(legs, legres):
            tput([g, lab, f"{p*100:.0f}%", "", ""])
        rr += 1
    ts.freeze_panes = "A3"

    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
