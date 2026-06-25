"""Selection du score probable : conversion continue + calibrage des nuls.

Remplace l'ancienne logique a paliers de 07_Modele_Probable (qui plafonnait les
scores) par une conversion CONTINUE buts = round(lambda) (plafond
``Plafond_buts_modele``), tout en gardant les nuls CALIBRES sur le modele :
on declare nul le nombre de matchs les plus serres = nuls attendus (somme des
P(nul) de Poisson). Ecrit 07 J/K/L (Score_Probable_A/B et "A-B") en valeurs.

A lancer APRES un recalcul (lit lambda et P(1/N/2) recalcules dans 07).
Pipeline : ... -> recalc -> finalize_scores -> recalc (classement).

Usage : python -m app.scripts.finalize_scores --file <classeur_recalcule.xlsx>
"""

from __future__ import annotations

import argparse

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    vals = openpyxl.load_workbook(args.file, data_only=True)
    mp = vals["07_Modele_Probable"]
    par = vals["01_Parametres"]
    cap = 7
    for r in range(1, par.max_row + 1):
        if par.cell(r, 1).value == "Plafond_buts_modele":
            cap = int(par.cell(r, 2).value or 7)
            break

    matches = []  # (row, lambdaA, lambdaB, pA, pN, pB)
    expected = 0.0
    for r in range(2, mp.max_row + 1):
        la, lb = mp.cell(r, 5).value, mp.cell(r, 6).value
        pa, pn, pb = mp.cell(r, 7).value, mp.cell(r, 8).value, mp.cell(r, 9).value
        if la is None or pn is None:
            continue
        matches.append([r, la, lb, pa, pn, pb])
        expected += pn

    # Nuls = les N matchs les plus serres (plus petit ecart max(pA,pB)-pN), N = nuls attendus.
    n_draws = max(0, min(len(matches), round(expected)))
    order = sorted(range(len(matches)), key=lambda i: max(matches[i][3], matches[i][5]) - matches[i][4])
    draw_rows = {matches[i][0] for i in order[:n_draws]}

    out = {}
    for r, la, lb, pa, pn, pb in matches:
        if r in draw_rows:
            g = min(cap, round((la + lb) / 2))
            ga = gb = g
        else:
            ga, gb = min(cap, round(la)), min(cap, round(lb))
            if pa >= pb:                       # A favori -> doit gagner
                if ga <= gb:
                    ga = min(cap, gb + 1)
            else:                              # B favori
                if gb <= ga:
                    gb = min(cap, ga + 1)
        out[r] = (ga, gb)

    wb = openpyxl.load_workbook(args.file)
    mpw = wb["07_Modele_Probable"]
    for r, (ga, gb) in out.items():
        mpw.cell(r, 10).value = ga              # Score_Probable_A
        mpw.cell(r, 11).value = gb              # Score_Probable_B
        mpw.cell(r, 12).value = f"{ga}-{gb}"    # Score_Probable
    wb.save(args.file)

    draws = sum(1 for ga, gb in out.values() if ga == gb)
    blow = sum(1 for ga, gb in out.values() if abs(ga - gb) >= 3)
    mx = max((max(ga, gb) for ga, gb in out.values()), default=0)
    print(f"OK -> {args.file}")
    print(f"  {len(out)} matchs | nuls {draws} ({round(100*draws/max(1,len(out)))}%, attendus {expected:.1f}) "
          f"| ecarts>=3 buts: {blow} | plafond {cap} | max buts {mx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
