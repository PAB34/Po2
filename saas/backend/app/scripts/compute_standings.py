"""Classement par poule avec departage FIFA complet (valeurs, pas formules).

Lit les scores finaux de 07_Modele_Probable et reconstruit 13_Classement_Poules
avec un classement SANS ex-aequo, selon l'ordre de departage FIFA :
  1. Points
  2. Difference de buts generale
  3. Buts marques generaux
  4. Confrontation directe : points entre equipes a egalite
  5. Conf. directe : difference de buts puis buts marques entre elles
  6. Ultime recours deterministe : Elo (eloratings) decroissant

Ecrit J/V/N/D/BP/BC/Diff/Pts/Rang en VALEURS. A lancer APRES finalize_scores.

Usage : python -m app.scripts.compute_standings --file <classeur.xlsx>
"""

from __future__ import annotations

import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _stats(team, matches):
    """Pts/GD/GF/V/N/D/J/BP/BC d'une equipe sur un ensemble de matchs."""
    pts = gf = ga = w = d = l = j = 0
    for a, b, sa, sb in matches:
        if team == a:
            mine, opp = sa, sb
        elif team == b:
            mine, opp = sb, sa
        else:
            continue
        j += 1; gf += mine; ga += opp
        if mine > opp: pts += 3; w += 1
        elif mine == opp: pts += 1; d += 1
        else: l += 1
    return {"J": j, "V": w, "N": d, "D": l, "BP": gf, "BC": ga, "Diff": gf - ga, "Pts": pts}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    mp = wb["07_Modele_Probable"]
    eq = wb["02_Equipes"]

    elo = {}
    group_of = {}
    for r in range(2, eq.max_row + 1):
        name = eq.cell(r, 1).value
        if name:
            group_of[name] = eq.cell(r, 2).value
            elo[name] = eq.cell(r, 11).value or 0

    all_matches = []  # (group, A, B, scoreA, scoreB)
    for r in range(2, mp.max_row + 1):
        a, b = mp.cell(r, 3).value, mp.cell(r, 4).value
        sa, sb = mp.cell(r, 10).value, mp.cell(r, 11).value
        if not a or sa is None:
            continue
        all_matches.append((mp.cell(r, 2).value, a, b, int(sa), int(sb)))

    groups: dict[str, list[str]] = {}
    for team, g in group_of.items():
        groups.setdefault(g, []).append(team)

    def rank_group(teams, gmatches):
        overall = {t: _stats(t, [(a, b, sa, sb) for _, a, b, sa, sb in gmatches]) for t in teams}

        def sort_key(t):
            s = overall[t]
            return (s["Pts"], s["Diff"], s["BP"])

        teams_sorted = sorted(teams, key=sort_key, reverse=True)
        # departager les blocs a (Pts,Diff,BP) identiques par confrontation directe
        final = []
        i = 0
        while i < len(teams_sorted):
            j = i + 1
            while j < len(teams_sorted) and sort_key(teams_sorted[j]) == sort_key(teams_sorted[i]):
                j += 1
            block = teams_sorted[i:j]
            if len(block) > 1:
                bm = [(a, b, sa, sb) for _, a, b, sa, sb in gmatches if a in block and b in block]
                h2h = {t: _stats(t, bm) for t in block}
                block.sort(key=lambda t: (h2h[t]["Pts"], h2h[t]["Diff"], h2h[t]["BP"], elo.get(t, 0)), reverse=True)
            final.extend(block)
            i = j
        return final, overall

    ws = wb["13_Classement_Poules"] if "13_Classement_Poules" in wb.sheetnames else wb.create_sheet("13_Classement_Poules")
    ws.delete_rows(1, ws.max_row)
    headers = ["Poule", "Equipe", "J", "V", "N", "D", "BP", "BC", "Diff", "Pts", "Rang"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = HEADER_FONT
    ws.freeze_panes = "A2"

    for g in sorted(groups):
        gmatches = [m for m in all_matches if m[0] == g]
        ordered, overall = rank_group(groups[g], gmatches)
        for rank, t in enumerate(ordered, 1):
            s = overall[t]
            ws.append([g, t, s["J"], s["V"], s["N"], s["D"], s["BP"], s["BC"], s["Diff"], s["Pts"], rank])

    wb.save(args.file)
    print(f"OK -> {args.file} : 13_Classement_Poules recalcule avec departage confrontation directe")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
