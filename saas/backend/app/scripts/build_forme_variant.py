"""Construit la variante "forme joueurs" du classeur a partir d'exports FBref.

Lit tous les CSV FBref (Player Standard Stats) presents dans --csvdir et/ou le
fichier --csv, agrege par joueur (somme minutes/buts/passes si plusieurs clubs),
puis :
  - enrichit la feuille 11_API_Effectifs avec, par joueur : minutes, buts,
    passes reels (FACTUEL) + notes derivees Minutes_index / Forme_club /
    Score_forme (0-100), provenance WEB/FBref ;
  - cable 04_Synthese_Joueurs col L (Score_forme_effectif) sur la moyenne
    ponderee par les minutes des joueurs de l'equipe disposant de donnees
    (sinon 70 neutre) -> alimente le modele via 06_Diagnostic.

Notes derivees (transparentes) :
  Minutes_index = min(100, minutes / 25)                  # ~2500 min = 100
  G_A_90        = (buts + passes) / (minutes / 90)
  Forme_club    = min(100, 50 + G_A_90 * 40)              # implication offensive
  Score_forme   = 0.6 * Minutes_index + 0.4 * Forme_club  # temps de jeu dominant

Les joueurs sans donnee FBref (hors ligues fournies) restent en note neutre,
marques 'neutre (hors couverture)'.

Usage (depuis saas/backend/) :
    python -m app.scripts.build_forme_variant \
        --in   ../../outputs/pronostics_model_v0/modele_pronostics_cdm2026_neutre.xlsx \
        --snapshot ../../outputs/pronostics_model_v0/football_data_snapshot.json \
        --csv  ../../outputs/pronostics_model_v0/fbref.csv \
        --csvdir ../../outputs/pronostics_model_v0/fbref_csv \
        --out  ../../outputs/pronostics_model_v0/modele_pronostics_cdm2026_forme.xlsx
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
import unicodedata
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
# Lissage (shrinkage) de la forme effectif vers le neutre 70 : plus l'echantillon
# de joueurs couverts est petit, plus on tire vers 70 (w = n/(n+k)).
SHRINKAGE_K = 5

# Coefficient de difficulte par championnat (0-100). Indicatif/ajustable :
# pondere la forme offensive (un but en Saoudie vaut moins qu'en Premier League).
# Cle = libelle FBref normalise (minuscules, sans le code pays accole).
DEFAULT_DIFFICULTY = 65
# Regles ORDONNEES (mot-cle sans accent -> difficulte). Premiere correspondance
# gagne : ambiguites placees d'abord (Bresil "serie a" avant Italie ; Autriche
# "bundesliga" avant Allemagne). Marche pour les libelles Comp ("br Serie A")
# ET les noms de fichier mono-ligue ("Campeonato Brasileiro Serie A").
LEAGUE_RULES = [
    (("premier league",), 100),          # Angleterre (Saudi = "saudi pro league")
    (("la liga", "primera division espan"), 96),
    (("brasileiro", "brazil", "bresil"), 80),   # AVANT 'serie a' (Italie)
    (("serie a",), 94),                   # Italie
    (("austrian", "autrich"), 70),        # AVANT 'bundesliga' (Allemagne)
    (("bundesliga",), 93),                # Allemagne
    (("ligue 1",), 88),
    (("primeira",), 82),
    (("eredivisie",), 80),
    (("championship",), 76),
    (("super lig", "super lig"), 75),     # Turquie
    (("argentin", "liga profesional"), 75),
    (("belg",), 74),
    (("liga mx", "mexiqu", "mexic"), 72),
    (("saudi",), 70),
    (("swiss", "suisse"), 70),
    (("greece", "grec"), 70),
    (("major league soccer", "mls"), 68),
    (("scottish", "ecosse"), 66),
    (("j1", "japan", "japon"), 66),
    (("k league", "korea", "coree"), 64),
    (("qatar",), 55),
]
# table indicative exposee dans la feuille 12 (libelle -> difficulte)
LEAGUE_DIFFICULTY = {kw[0]: diff for kw, diff in LEAGUE_RULES}


def _deaccent(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def league_difficulty(comp: str) -> int:
    key = _deaccent(comp.replace("\xa0", " ").strip().lower())
    for keywords, diff in LEAGUE_RULES:
        if any(k in key for k in keywords):
            return diff
    return DEFAULT_DIFFICULTY


def norm_name(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s).lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z ]+", " ", s).strip()


def to_num(s) -> float | None:
    s = str(s).replace("\xa0", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_fbref_csv(path: str, default_comp: str = "") -> dict[str, dict]:
    rows = list(csv.reader(open(path, encoding="utf-8-sig"), delimiter=";"))
    # Trouver la ligne d'entete (celle contenant 'Player' et 'Min').
    hdr_idx = next((i for i, r in enumerate(rows) if "Player" in r and "Min" in r), None)
    if hdr_idx is None:
        return {}
    hdr = rows[hdr_idx]
    # FBref duplique 'Gls'/'Ast' (totaux Performance PUIS Per-90) : garder la
    # PREMIERE occurrence (les totaux), pas la derniere.
    ix: dict[str, int] = {}
    for i, h in enumerate(hdr):
        ix.setdefault(h, i)
    out: dict[str, dict] = {}
    for r in rows[hdr_idx + 1:]:
        if len(r) <= ix["Min"] or not r[ix["Player"]].strip():
            continue
        name = r[ix["Player"]]
        nm = norm_name(name)
        if not nm:
            continue
        minutes = to_num(r[ix["Min"]]) or 0
        nineties = to_num(r[ix["90s"]]) if "90s" in ix else (minutes / 90.0 if minutes else 0)
        gls_raw = to_num(r[ix["Gls"]]) if "Gls" in ix else 0
        ast_raw = to_num(r[ix["Ast"]]) if "Ast" in ix else 0
        gls, ast = _totals(gls_raw, ast_raw, nineties)
        comp = r[ix["Comp"]] if "Comp" in ix and ix["Comp"] < len(r) and r[ix["Comp"]] else default_comp
        rec = out.setdefault(nm, {"name": name, "min": 0.0, "gls": 0.0, "ast": 0.0, "comp": comp})
        rec["min"] += minutes
        rec["gls"] += gls
        rec["ast"] += ast
    return out


def _totals(gls_raw, ast_raw, nineties) -> tuple[float, float]:
    """Reconstitue les totaux buts/passes.

    FBref expose tantot des TOTAUX (entiers), tantot des valeurs PAR-90 (petits
    decimaux) dans les memes colonnes selon l'export. Heuristique : une valeur
    non entiere et < 5 est une cadence par-90 (a multiplier par les 90s) ;
    sinon c'est deja un total.
    """
    def conv(v):
        v = v or 0
        if nineties and v < 5 and v != int(v):
            return v * nineties
        return v
    return conv(gls_raw), conv(ast_raw)


def derive_form(minutes: float, gls: float, ast: float, difficulty: int = 100) -> tuple[float, float, float]:
    minutes_index = min(100.0, minutes / 25.0)
    ga90 = (gls + ast) / (minutes / 90.0) if minutes > 0 else 0.0
    # la composante offensive est ponderee par la difficulte du championnat
    forme_club = min(100.0, 50.0 + ga90 * 40.0 * (difficulty / 100.0))
    score = 0.6 * minutes_index + 0.4 * forme_club
    return round(minutes_index, 1), round(forme_club, 1), round(score, 1)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--in", dest="src", required=True)
    ap.add_argument("--snapshot", required=True)
    ap.add_argument("--csv", action="append", default=[])
    ap.add_argument("--csvdir")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    paths = list(args.csv)
    if args.csvdir:
        paths += sorted(glob.glob(os.path.join(args.csvdir, "*.csv")))
    paths = [p for p in paths if os.path.isfile(p)]

    fb: dict[str, dict] = {}
    for p in paths:
        part = parse_fbref_csv(p, default_comp=os.path.splitext(os.path.basename(p))[0])
        for nm, rec in part.items():
            cur = fb.get(nm)
            if cur:
                cur["min"] += rec["min"]; cur["gls"] += rec["gls"]; cur["ast"] += rec["ast"]
            else:
                fb[nm] = rec
    print(f"Sources CSV: {len(paths)} fichier(s), {len(fb)} joueurs FBref uniques")

    feed = json.load(open(args.snapshot, encoding="utf-8"))
    wb = openpyxl.load_workbook(args.src)
    ef = wb["11_API_Effectifs"]

    # En-tetes ajoutees (a partir de la colonne 10).
    new_cols = ["Comp_club", "Min", "Buts", "Passes", "Minutes_index", "Forme_club", "Buts_recents", "Score_forme", "Source_forme"]
    base = 10
    for j, name in enumerate(new_cols):
        c = ef.cell(1, base + j)
        c.value = name; c.fill = HEADER_FILL; c.font = HEADER_FONT

    # Index nom normalise -> ligne dans 11 (l'ordre de 11 suit le snapshot trie).
    matched = 0
    comps_seen: dict[str, int] = {}
    team_scores: dict[str, list[tuple[float, float]]] = {}
    for r in range(2, ef.max_row + 1):
        jname = ef.cell(r, 2).value
        team = ef.cell(r, 1).value
        if not jname:
            continue
        rec = fb.get(norm_name(jname))
        if rec and rec["min"] > 0:
            diff = league_difficulty(rec["comp"])
            comps_seen[rec["comp"].replace("\xa0", " ").strip()] = comps_seen.get(rec["comp"].replace("\xa0", " ").strip(), 0) + 1
            mi, fc, sc = derive_form(rec["min"], rec["gls"], rec["ast"], diff)
            ef.cell(r, base + 0).value = rec["comp"].replace("\xa0", " ").strip()
            ef.cell(r, base + 1).value = int(rec["min"])
            ef.cell(r, base + 2).value = int(rec["gls"])
            ef.cell(r, base + 3).value = int(rec["ast"])
            ef.cell(r, base + 4).value = mi
            ef.cell(r, base + 5).value = fc
            ef.cell(r, base + 6).value = int(rec["gls"])
            ef.cell(r, base + 7).value = sc
            ef.cell(r, base + 8).value = "WEB/FBref"
            matched += 1
            team_scores.setdefault(team, []).append((sc, rec["min"]))
        else:
            ef.cell(r, base + 8).value = "neutre (hors couverture)"

    # 04_Synthese col L = moyenne ponderee par minutes des Score_forme (sinon 70).
    syn = wb["04_Synthese_Joueurs"]
    for r in range(2, syn.max_row + 1):
        team = syn.cell(r, 1).value
        if not team:
            continue
        scores = team_scores.get(team)
        if scores:
            wsum = sum(s * m for s, m in scores)
            msum = sum(m for _, m in scores)
            raw = wsum / msum if msum else 70.0
            # #1 SHRINKAGE vers le neutre 70 selon la fiabilite de l'echantillon :
            # w = n_couverts / (n_couverts + k). Peu de joueurs couverts -> on
            # tire la forme vers 70 (evite de punir/gonfler sur 1-2 joueurs).
            n_cov = len(scores)
            k = SHRINKAGE_K
            w = n_cov / (n_cov + k)
            syn.cell(r, 12).value = round(w * raw + (1 - w) * 70.0, 1)
        else:
            syn.cell(r, 12).value = 70
        # #4 nettoyage : en variante forme, la col L (Score_forme_effectif) est une
        # valeur directe FBref -> les colonnes intermediaires C..K (calculees depuis
        # 03_Joueurs_Retenus vide) ne servent plus, on les vide pour eviter la confusion.
        for c in range(3, 12):
            syn.cell(r, c).value = None

    # Feuille 12_Championnats : difficulte appliquee + volume (editable/transparence).
    if "12_Championnats" in wb.sheetnames:
        del wb["12_Championnats"]
    ch = wb.create_sheet("12_Championnats")
    for j, name in enumerate(["Championnat", "Difficulte_0_100", "Nb_joueurs", "Reconnu"], 1):
        c = ch.cell(1, j); c.value = name; c.fill = HEADER_FILL; c.font = HEADER_FONT
    unknown = []
    for comp, n in sorted(comps_seen.items(), key=lambda x: -x[1]):
        diff = league_difficulty(comp)
        recognized = diff != DEFAULT_DIFFICULTY
        ch.append([comp, diff, n, "oui" if recognized else f"NON -> defaut {DEFAULT_DIFFICULTY}"])
        if not recognized:
            unknown.append(comp)

    wb.save(args.out)
    print(f"OK -> {args.out}")
    print(f"  joueurs avec forme FACTUELLE : {matched}/{ef.max_row - 1}")
    print(f"  equipes avec >=1 joueur couvert : {len(team_scores)}/48")
    print(f"  championnats: {len(comps_seen)}")
    if unknown:
        print(f"  /!\\ championnats SANS difficulte definie (defaut {DEFAULT_DIFFICULTY}) : {unknown}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
