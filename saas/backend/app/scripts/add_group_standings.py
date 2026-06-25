"""Ajoute le classement par poule + corrige le seuil de nul dans un classeur.

- Met a jour ``01_Parametres`` : Seuil_nul_prioritaire (sinon ~0 nul en poules).
- (Re)cree ``13_Classement_Poules`` : par equipe d'une poule, J/V/N/D/BP/BC/Diff/
  Pts/Rang, en FORMULES live calculees depuis ``07_Modele_Probable`` (donc se
  recalcule quand les pronostics changent). Classement intra-poule = points puis
  difference de buts puis buts pour.

Usage (depuis saas/backend/) :
    python -m app.scripts.add_group_standings --file <classeur.xlsx> [--draw-threshold 0.12]
"""

from __future__ import annotations

import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
P = "'07_Modele_Probable'!"
C = f"{P}$C$2:$C$73"   # Equipe_A
D = f"{P}$D$2:$D$73"   # Equipe_B
GA = f"{P}$J$2:$J$73"  # buts A
GB = f"{P}$K$2:$K$73"  # buts B


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    ap.add_argument("--draw-threshold", type=float, default=None)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)

    if args.draw_threshold is not None:
        par = wb["01_Parametres"]
        for r in range(1, par.max_row + 1):
            if par.cell(r, 1).value == "Seuil_nul_prioritaire":
                par.cell(r, 2).value = args.draw_threshold
                break

    # Liste equipes + poule depuis 02_Equipes, triee par poule puis equipe.
    eq = wb["02_Equipes"]
    teams = sorted(
        ((eq.cell(r, 2).value, eq.cell(r, 1).value) for r in range(2, eq.max_row + 1) if eq.cell(r, 1).value),
        key=lambda t: (str(t[0]), str(t[1])),
    )

    if "13_Classement_Poules" in wb.sheetnames:
        del wb["13_Classement_Poules"]
    ws = wb.create_sheet("13_Classement_Poules")
    headers = ["Poule", "Equipe", "J", "V", "N", "D", "BP", "BC", "Diff", "Pts", "Cle", "Rang"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = HEADER_FILL
        ws.cell(1, c).font = HEADER_FONT
    ws.freeze_panes = "A2"

    for grp, team in teams:
        ws.append([grp, team])
    last = ws.max_row

    for r in range(2, last + 1):
        b = f"$B{r}"
        # J joues
        ws.cell(r, 3).value = f"=COUNTIF({C},{b})+COUNTIF({D},{b})"
        # V victoires
        ws.cell(r, 4).value = (
            f"=SUMPRODUCT(({C}={b})*({GA}>{GB}))+SUMPRODUCT(({D}={b})*({GB}>{GA}))"
        )
        # N nuls
        ws.cell(r, 5).value = (
            f"=SUMPRODUCT(({C}={b})*({GA}={GB}))+SUMPRODUCT(({D}={b})*({GB}={GA}))"
        )
        # D defaites
        ws.cell(r, 6).value = f"=C{r}-D{r}-E{r}"
        # BP / BC
        ws.cell(r, 7).value = f"=SUMIF({C},{b},{GA})+SUMIF({D},{b},{GB})"
        ws.cell(r, 8).value = f"=SUMIF({C},{b},{GB})+SUMIF({D},{b},{GA})"
        # Diff, Pts
        ws.cell(r, 9).value = f"=G{r}-H{r}"
        ws.cell(r, 10).value = f"=D{r}*3+E{r}"
        # Cle de tri (Pts, puis Diff, puis BP), Rang intra-poule
        ws.cell(r, 11).value = f"=J{r}*10000+(I{r}+50)*100+G{r}"
        ws.cell(r, 12).value = f"=SUMPRODUCT(($A$2:$A${last}=$A{r})*($K$2:$K${last}>$K{r}))+1"

    wb.save(args.file)
    print(f"OK -> {args.file} : 13_Classement_Poules ({last-1} lignes)"
          + (f", Seuil_nul_prioritaire={args.draw_threshold}" if args.draw_threshold is not None else ""))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
