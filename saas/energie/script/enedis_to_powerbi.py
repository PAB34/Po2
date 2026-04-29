"""
enedis_to_powerbi.py
--------------------
Flux ENEDIS → Power BI via Microsoft Fabric.

Collecte synchrone journalière :
  - Consommation journalière (mesure_synchrone v2)
  - Courbe de charge 30 min (consumption_load_curve)

Variables d'environnement requises : voir README.md et .env.example.

Usage :
    python scripts/enedis_to_powerbi.py
    python scripts/enedis_to_powerbi.py --regenerate-prm-list
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from dateutil import tz


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


LOG = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------


def getenv(name: str, default: Optional[str] = None, required: bool = False) -> str:
    value = os.getenv(name) or default  # si vide OU absente → default
    if required and not value:
        raise RuntimeError(f"Variable d'environnement obligatoire manquante : {name}")
    return value or ""


@dataclass
class Settings:
    enedis_auth_url: str
    enedis_client_id: str
    enedis_client_secret: str
    csv_output_path: str               # chemin du fichier CSV de sortie
    prm_list_path: str
    prm_xlsx_path: str                # source maîtresse des PRM (LISTE PRM.xlsx)
    state_file_path: str
    timezone: str
    day_offset_start: int
    day_offset_end: int
    max_prms_per_request: int
    max_prms: int
    enedis_sync_url: str
    enedis_load_curve_url: str
    enedis_canal_id: str
    enedis_type_donnee: str
    history_days: int
    load_curve_history_days: int
    load_curve_max_days_per_run: int
    sync_max_days_per_run: int
    enedis_base_url: str
    enedis_perimeter_url: str
    alert_webhook_url: str
    alert_on_success: bool
    enable_load_curve: bool
    enable_daily_consumption: bool
    enable_max_power: bool
    enedis_max_power_url: str
    max_power_history_days: int
    max_power_max_days_per_run: int

    @staticmethod
    def load() -> "Settings":
        return Settings(
            enedis_auth_url=getenv("ENEDIS_AUTH_URL", required=True),
            enedis_client_id=getenv("ENEDIS_CLIENT_ID", required=True),
            enedis_client_secret=getenv("ENEDIS_CLIENT_SECRET", required=True),
            csv_output_path=getenv("CSV_OUTPUT_PATH", "output/enedis_data.csv"),
            prm_list_path=getenv("PRM_LIST_PATH", "config/prm_list.txt"),
            prm_xlsx_path=getenv("PRM_XLSX_PATH", "LISTE PRM.xlsx"),
            state_file_path=getenv("STATE_FILE_PATH", "state/processed_files.json"),
            timezone=getenv("TIMEZONE", "Europe/Paris"),
            day_offset_start=int(getenv("DAY_OFFSET_START", "1")),
            day_offset_end=int(getenv("DAY_OFFSET_END", "0")),
            max_prms_per_request=int(getenv("MAX_PRMS_PER_REQUEST", "50")),
            max_prms=int(getenv("MAX_PRMS", "0")),
            enedis_sync_url=getenv(
                "ENEDIS_SYNC_URL",
                default="https://gw.ext.prod.api.enedis.fr/mesures/v2/metering_data/daily_consumption",
            ),
            enedis_load_curve_url=getenv(
                "ENEDIS_LOAD_CURVE_URL",
                default="https://gw.ext.prod.api.enedis.fr/mesures/v2/metering_data/consumption_load_curve",
            ),
            enedis_canal_id=getenv("ENEDIS_CANAL_ID", default="506350699"),
            enedis_type_donnee=getenv("ENEDIS_TYPE_DONNEE", default="ENERGIE"),
            history_days=int(getenv("HISTORY_DAYS", "365")),
            load_curve_history_days=int(getenv("LOAD_CURVE_HISTORY_DAYS", "365")),
            load_curve_max_days_per_run=int(getenv("LOAD_CURVE_MAX_DAYS_PER_RUN", "28")),
            sync_max_days_per_run=int(getenv("SYNC_MAX_DAYS_PER_RUN", "0")),
            enedis_base_url=getenv("ENEDIS_BASE_URL", "https://gw.ext.prod.api.enedis.fr"),
            enedis_perimeter_url=getenv(
                "ENEDIS_PERIMETER_URL",
                default="https://gw.ext.prod.api.enedis.fr/usage_point_id_perimeter/v1/usage_point_id",
            ),
            alert_webhook_url=getenv("ALERT_WEBHOOK_URL", default=""),
            alert_on_success=getenv("ALERT_ON_SUCCESS", default="true").lower() == "true",
            enable_load_curve=getenv("ENABLE_LOAD_CURVE", default="true").lower() == "true",
            enable_daily_consumption=getenv("ENABLE_DAILY_CONSUMPTION", default="true").lower() == "true",
            enable_max_power=getenv("ENABLE_MAX_POWER", default="false").lower() == "true",
            enedis_max_power_url=getenv(
                "ENEDIS_MAX_POWER_URL",
                default="https://gw.ext.prod.api.enedis.fr/mesures/v2/metering_data/daily_consumption_max_power",
            ),
            max_power_history_days=int(getenv("MAX_POWER_HISTORY_DAYS", "1095")),
            max_power_max_days_per_run=int(getenv("MAX_POWER_MAX_DAYS_PER_RUN", "0")),
        )


# ---------------------------------------------------------------------------
# Alerting Teams
# ---------------------------------------------------------------------------


def send_teams_alert(webhook_url: str, title: str, message: str, color: str = "0076D7") -> None:
    """Envoie une carte adaptative Teams via webhook entrant."""
    if not webhook_url:
        return
    payload = {
        "@type": "MessageCard",
        "@context": "https://schema.org/extensions",
        "themeColor": color,
        "summary": title,
        "sections": [{"activityTitle": title, "activityText": message}],
    }
    try:
        resp = requests.post(webhook_url, json=payload, timeout=10)
        if resp.status_code not in (200, 202):
            LOG.warning("Teams webhook : réponse inattendue %s — %s", resp.status_code, resp.text[:200])
    except Exception as exc:  # noqa: BLE001
        LOG.warning("Teams webhook : échec d'envoi — %s", exc)


# ---------------------------------------------------------------------------
# Contrôle de complétude
# ---------------------------------------------------------------------------


def check_completeness(settings: Settings, state: Dict[str, Any]) -> List[str]:
    """Vérifie que les dates max dans les CSV correspondent à l'état sauvegardé.

    Retourne une liste de messages d'avertissement (vide = tout OK).
    """
    warnings: List[str] = []
    today = date.today().isoformat()

    # --- enedis_data ---
    if settings.enable_daily_consumption:
        last_sync = state.get("last_sync_date")
        if last_sync:
            csv_path = Path(settings.csv_output_path)
            if csv_path.exists():
                try:
                    with csv_path.open(encoding="utf-8-sig") as f:
                        reader = csv.DictReader(f)
                        dates = [row.get("date", "") for row in reader if row.get("date")]
                    max_date = max(dates) if dates else None
                    if max_date and max_date < last_sync:
                        warnings.append(
                            f"⚠️ enedis_data : CSV max_date={max_date} < last_sync_date={last_sync}"
                        )
                    else:
                        LOG.info("✅ enedis_data : complétude OK (max_date=%s)", max_date)
                except Exception as exc:  # noqa: BLE001
                    warnings.append(f"⚠️ enedis_data : lecture CSV impossible — {exc}")

    # --- enedis_load_curve ---
    if settings.enable_load_curve:
        lc_last = state.get("last_load_curve_date")
        if lc_last:
            lc_path = Path(settings.csv_output_path).parent / "enedis_load_curve.csv"
            if lc_path.exists():
                try:
                    with lc_path.open(encoding="utf-8-sig") as f:
                        reader = csv.DictReader(f)
                        dates = [row.get("datetime", "")[:10] for row in reader if row.get("datetime")]
                    max_date = max(dates) if dates else None
                    if max_date and max_date < lc_last:
                        warnings.append(
                            f"⚠️ enedis_load_curve : CSV max_date={max_date} < last_load_curve_date={lc_last}"
                        )
                    else:
                        LOG.info("✅ enedis_load_curve : complétude OK (max_date=%s)", max_date)
                except Exception as exc:  # noqa: BLE001
                    warnings.append(f"⚠️ enedis_load_curve : lecture CSV impossible — {exc}")

    return warnings


# ---------------------------------------------------------------------------
# Gestion de la liste PRM
# ---------------------------------------------------------------------------


def generate_prm_list_from_xlsx(xlsx_path: str, output_path: str) -> List[str]:
    """
    Génère config/prm_list.txt à partir de la colonne B ('Identifiant PRM')
    de la feuille dont l'en-tête de la colonne B vaut exactement 'Identifiant PRM'.

    Dans le fichier 'LISTE PRM.xlsx' fourni, il s'agit de la feuille 5.

    Règles d'extraction :
    - Ignorer la ligne d'en-tête (ligne 1).
    - Ignorer les lignes vides ou dont la valeur de la colonne B est None.
    - Ignorer les valeurs non numériques ou dont la longueur n'est pas 14.
    - Dédupliquer et trier.

    Lève RuntimeError si le fichier, la colonne ou les données sont invalides.
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise RuntimeError(
            "Le module openpyxl est requis pour lire LISTE PRM.xlsx. "
            "Installez-le : pip install openpyxl"
        ) from exc

    path = Path(xlsx_path)
    if not path.exists():
        raise RuntimeError(
            f"Fichier Excel introuvable : {xlsx_path}. "
            "Vérifiez que LISTE PRM.xlsx est bien présent à la racine du dépôt."
        )

    LOG.info("Lecture du fichier Excel : %s", xlsx_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    target_sheet = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=2, values_only=True):
            header_b = str(row[0] or "").strip()
            if header_b == "Identifiant PRM":
                target_sheet = ws
                LOG.info("Feuille source trouvée : '%s'", sheet_name)
                break
        if target_sheet is not None:
            break

    if target_sheet is None:
        available = ", ".join(f"'{s}'" for s in wb.sheetnames)
        raise RuntimeError(
            f"Aucune feuille dont la colonne B (ligne 1) vaut 'Identifiant PRM' "
            f"n'a été trouvée dans {xlsx_path}. "
            f"Feuilles disponibles : {available}. "
            "Vérifiez la structure du fichier."
        )

    prms: List[str] = []
    skipped = 0
    for row in target_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        raw = row[0]
        if raw is None:
            continue
        val = str(raw).strip()
        if not val:
            continue
        if not val.isdigit() or len(val) != 14:
            LOG.warning("Valeur ignorée (format invalide, attendu 14 chiffres) : %r", val)
            skipped += 1
            continue
        prms.append(val)

    wb.close()

    unique_prms = sorted(set(prms))
    LOG.info(
        "PRMs extraits : %d uniques (%d total, %d ignorés).",
        len(unique_prms),
        len(prms),
        skipped,
    )

    if not unique_prms:
        raise RuntimeError(
            f"Aucun PRM valide (14 chiffres) extrait de {xlsx_path}. "
            "Vérifiez que la colonne B contient bien des identifiants ENEDIS."
        )

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(unique_prms) + "\n", encoding="utf-8")
    LOG.info("Fichier PRM généré : %s (%d entrées).", output_path, len(unique_prms))
    return unique_prms


def load_prms(prm_list_path: str) -> List[str]:
    path = Path(prm_list_path)
    if not path.exists():
        raise RuntimeError(
            f"Fichier PRM introuvable : {prm_list_path}. "
            "Lancez le script avec --regenerate-prm-list pour le regénérer depuis LISTE PRM.xlsx, "
            "ou créez-le manuellement (un PRM de 14 chiffres par ligne)."
        )
    prms = [
        line.strip()
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip() and line.strip().isdigit() and len(line.strip()) == 14
    ]
    if not prms:
        raise RuntimeError(
            f"La liste PRM est vide ou ne contient aucun identifiant valide : {prm_list_path}"
        )
    LOG.info("Liste PRM chargée : %d PRM(s).", len(prms))
    return prms


def discover_prms_from_api(settings: "Settings", token: str) -> List[str]:
    """Récupère tous les PRMs du périmètre via l'API usage_point_id_perimeter (paginé, 3000/page).
    Retourne la liste triée. En cas d'erreur, retourne [] pour déclencher le fallback.
    """
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    all_prms: List[str] = []
    page = 1
    while True:
        try:
            resp = requests.post(
                settings.enedis_perimeter_url,
                headers=headers,
                json={"page_number": page},
                timeout=30,
            )
            if resp.status_code != 200:
                LOG.warning(
                    "discover_prms_from_api : HTTP %d page %d — %s",
                    resp.status_code, page, resp.text[:300],
                )
                return []
            body = resp.json()
            data = body.get("query_parameters", {})
            prms_page = body.get("usage_point_id", [])
            all_prms.extend(prms_page)
            total_pages = int(data.get("page_total_count", 1))
            LOG.info(
                "Périmètre PRM : page %d/%d, %d PRM(s) reçus.",
                page, total_pages, len(prms_page),
            )
            if page >= total_pages:
                break
            page += 1
        except Exception as exc:
            LOG.warning("discover_prms_from_api : erreur page %d — %s", page, exc)
            return []
    unique = sorted(set(all_prms))
    LOG.info("Périmètre total : %d PRM(s) découverts.", len(unique))
    return unique


# ---------------------------------------------------------------------------
# Temps / état
# ---------------------------------------------------------------------------


def now_local(timezone_name: str) -> datetime:
    local_tz = tz.gettz(timezone_name)
    return datetime.now(local_tz)


def compute_dates(settings: Settings) -> Tuple[str, str]:
    current = now_local(settings.timezone).date()
    start_date = current - timedelta(days=settings.day_offset_start)
    end_date = current - timedelta(days=settings.day_offset_end)
    LOG.info("Période de données : %s → %s", start_date.isoformat(), end_date.isoformat())
    return start_date.isoformat(), end_date.isoformat()


def compute_incremental_dates(
    settings: Settings, state: Dict[str, Any]
) -> Tuple[Optional[str], Optional[str]]:
    """Calcule la plage incrémentale à collecter.

    - Pas de last_sync_date → backfill depuis today - history_days
    - Sinon → depuis last_sync_date + 1j jusqu'à today - day_offset_end
    - Retourne (None, None) si déjà à jour
    """
    today = now_local(settings.timezone).date()
    end_date = today - timedelta(days=settings.day_offset_end)
    last = state.get("last_sync_date")
    if last:
        start_date = date.fromisoformat(last) + timedelta(days=1)
    else:
        start_date = today - timedelta(days=settings.history_days)
    if start_date >= end_date:
        LOG.info(
            "Données déjà à jour (last_sync_date=%s, end=%s) — rien à collecter.",
            last, end_date,
        )
        return None, None
    LOG.info(
        "Plage incrémentale : %s \u2192 %s (last_sync=%s, history_days=%d).",
        start_date, end_date, last or "jamais", settings.history_days,
    )
    return start_date.isoformat(), end_date.isoformat()


def load_state(path: str) -> Dict[str, Any]:
    state_path = Path(path)
    state_path.parent.mkdir(parents=True, exist_ok=True)
    if not state_path.exists():
        LOG.info("Fichier d'état absent, initialisation vide : %s", path)
        return {"processed_files": [], "pending_dossiers": []}
    content = state_path.read_text(encoding="utf-8-sig").strip()
    if not content:
        LOG.warning("Fichier d'état vide, réinitialisation : %s", path)
        return {"processed_files": [], "pending_dossiers": []}
    state = json.loads(content)
    if "pending_dossiers" not in state:
        state["pending_dossiers"] = []
    LOG.info(
        "État chargé : %d fichier(s) traité(s), %d dossier(s) en attente.",
        len(state.get("processed_files", [])),
        len(state.get("pending_dossiers", [])),
    )
    return state


def save_state(path: str, state: Dict[str, Any]) -> None:
    state_path = Path(path)
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def get_last_load_curve_date(csv_path: str) -> Optional[str]:
    """Lit la date maximale présente dans enedis_load_curve.csv (colonne datetime).

    Source de vérité réelle pour éviter les désynchros avec le fichier d'état.
    Utilise csv.reader en streaming pour ne pas charger le fichier en mémoire.
    """
    path = Path(csv_path)
    if not path.exists():
        return None
    max_d: Optional[date] = None
    try:
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                return None
            try:
                dt_idx = header.index("datetime")
            except ValueError:
                LOG.warning("get_last_load_curve_date : colonne 'datetime' introuvable dans %s", csv_path)
                return None
            for row in reader:
                if len(row) > dt_idx:
                    dt_str = row[dt_idx][:10]
                    try:
                        d = date.fromisoformat(dt_str)
                        if max_d is None or d > max_d:
                            max_d = d
                    except ValueError:
                        pass
    except Exception as exc:
        LOG.warning("get_last_load_curve_date : lecture impossible — %s", exc)
        return None
    return max_d.isoformat() if max_d else None


def get_last_sync_date(csv_path: str) -> Optional[str]:
    """Lit la date maximale présente dans enedis_data.csv (colonne date)."""
    path = Path(csv_path)
    if not path.exists():
        return None
    max_d: Optional[date] = None
    try:
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                return None
            try:
                date_idx = header.index("date")
            except ValueError:
                LOG.warning("get_last_sync_date : colonne 'date' introuvable dans %s", csv_path)
                return None
            for row in reader:
                if len(row) > date_idx:
                    date_str = row[date_idx][:10]
                    try:
                        d = date.fromisoformat(date_str)
                        if max_d is None or d > max_d:
                            max_d = d
                    except ValueError:
                        pass
    except Exception as exc:
        LOG.warning("get_last_sync_date : lecture impossible — %s", exc)
        return None
    return max_d.isoformat() if max_d else None


# ---------------------------------------------------------------------------
# ENEDIS — authentification et soumission asynchrone
# ---------------------------------------------------------------------------


def get_enedis_token(settings: Settings) -> str:
    LOG.info("Obtention du token ENEDIS (client_credentials)…")
    payload = {
        "grant_type": "client_credentials",
        "client_id": settings.enedis_client_id,
        "client_secret": settings.enedis_client_secret,
    }
    try:
        resp = requests.post(settings.enedis_auth_url, data=payload, timeout=60)
        resp.raise_for_status()
    except requests.HTTPError as exc:
        raise RuntimeError(
            f"Authentification ENEDIS échouée (HTTP {exc.response.status_code}) : "
            f"{exc.response.text}"
        ) from exc
    except requests.RequestException as exc:
        raise RuntimeError(f"Erreur réseau lors de l'authentification ENEDIS : {exc}") from exc

    data = resp.json()
    token = data.get("access_token")
    if not token:
        raise RuntimeError(
            f"La réponse ENEDIS ne contient pas 'access_token'. Réponse : {data}"
        )
    LOG.info("Token ENEDIS obtenu.")
    # Décoder le JWT (sans vérification signature) pour inspecter le scope
    try:
        import base64 as _b64
        parts = token.split(".")
        if len(parts) == 3:
            padded = parts[1] + "==" * (-len(parts[1]) % 4)
            claims = json.loads(_b64.urlsafe_b64decode(padded).decode("utf-8", errors="replace"))
            LOG.info(
                "Token ENEDIS — sub=%s scope=%s subscribedAPIs=%s",
                claims.get("sub"),
                claims.get("scope"),
                str(claims.get("subscribedAPIs", []))[:2000],
            )
    except Exception as _jwt_exc:
        LOG.debug("Impossible de décoder le JWT : %s", _jwt_exc)
    return token


def upsert_rows_to_csv(
    rows: List[Dict[str, Any]],
    output_path: str,
    key_cols: Tuple[str, ...] = ("usage_point_id", "date"),
) -> int:
    """Fusionne les nouvelles lignes dans le CSV existant (upsert).

    Clé de déduplication : key_cols (par défaut usage_point_id + date).
    Les nouvelles lignes écrasent les anciennes sur la même clé.
    Le CSV résultant est trié par usage_point_id puis date.
    Retourne le nombre de lignes nouvelles (non présentes avant).
    """
    if not rows:
        return 0

    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    existing: Dict[tuple, Dict[str, Any]] = {}
    existing_cols: List[str] = []
    if out_path.exists() and out_path.stat().st_size > 0:
        with out_path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            existing_cols = list(reader.fieldnames or [])
            for r in reader:
                key = tuple(r.get(k, "") for k in key_cols)
                existing[key] = dict(r)

    new_count = 0
    for row in rows:
        key = tuple(row.get(k, "") for k in key_cols)
        if key not in existing:
            new_count += 1
        existing[key] = row

    all_cols: List[str] = list(dict.fromkeys(
        existing_cols + [k for row in rows for k in row]
    ))
    sorted_rows = sorted(
        existing.values(),
        key=lambda r: (r.get("usage_point_id", ""), r.get("date", "")),
    )
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=all_cols, extrasaction="ignore")
        writer.writeheader()
        for r in sorted_rows:
            writer.writerow({k: r.get(k, "") for k in all_cols})

    LOG.info(
        "CSV upsert : %s — %d ligne(s) totales, %d nouvelle(s).",
        output_path, len(sorted_rows), new_count,
    )
    return new_count


# ---------------------------------------------------------------------------
# APIs client ENEDIS (contrat, adresse, connexion)
# ---------------------------------------------------------------------------


def _flatten_into(obj: Any, result: Dict[str, Any], prefix: str = "") -> None:
    """Aplatit récursivement un objet JSON en clés underscore."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{prefix}_{k}" if prefix else k
            _flatten_into(v, result, new_key)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            _flatten_into(v, result, f"{prefix}_{i}" if prefix else str(i))
    else:
        result[prefix] = obj


def fetch_enedis_customer_data(
    base_url: str,
    token: str,
    prms: List[str],
    api_context: str,
) -> List[Dict[str, Any]]:
    """Appelle `base_url + api_context/{prm}` pour chaque PRM (parallèle, 10 threads)."""
    from concurrent.futures import ThreadPoolExecutor
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    ingested_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_rows: List[Dict[str, Any]] = []
    ok_count = err_count = 0

    import time as _time
    _RETRY_DELAYS = [5, 10, 20]
    _RETRY_429    = [60, 120, 180]

    def _call_one(prm: str) -> Tuple[Optional[Dict[str, Any]], int, int]:
        url = f"{base_url}{api_context}/{prm}"
        resp = None
        for attempt, delay in enumerate([0] + _RETRY_DELAYS):
            if delay:
                LOG.info("PRM %s [%s] retry %d/%d dans %ds...",
                         prm, api_context, attempt, len(_RETRY_DELAYS), delay)
                _time.sleep(delay)
            try:
                resp = requests.get(url, headers=headers, timeout=30)
                if resp.status_code == 429 and attempt < len(_RETRY_429):
                    LOG.warning("PRM %s [%s] → 429 quota, attente %ds...",
                                prm, api_context, _RETRY_429[attempt])
                    _time.sleep(_RETRY_429[attempt])
                    continue
                if resp.status_code < 500:
                    break
            except Exception as exc:
                if attempt == len(_RETRY_DELAYS):
                    LOG.warning("PRM %s [%s] → erreur : %s", prm, api_context, exc)
                    return None, 0, 1
        if resp is None:
            return None, 0, 1
        if resp.status_code == 200:
            row: Dict[str, Any] = {"usage_point_id": prm, "_ingested_at_utc": ingested_at}
            _flatten_into(resp.json(), row)
            return row, 1, 0
        LOG.warning(
            "PRM %s [%s] → HTTP %d : %s",
            prm, api_context, resp.status_code, resp.text[:300],
        )
        return None, 0, 1

    workers = min(10, len(prms))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for row, ok, err in ex.map(_call_one, prms):
            ok_count += ok
            err_count += err
            if row:
                all_rows.append(row)

    LOG.info("%s : %d OK, %d erreur(s), %d ligne(s).", api_context, ok_count, err_count, len(all_rows))
    return all_rows


def collect_customer_supplementary(
    settings: Settings, token: str, prms: List[str]
) -> None:
    """Collecte contrat + adresse + connexion et enregistre dans des CSVs séparés."""
    base = settings.enedis_base_url
    limit = settings.max_prms if settings.max_prms > 0 else len(prms)
    batch = prms[:limit]
    apis = [
        ("/contract/v1",              "customer_contract",          "output/enedis_contracts.csv"),
        ("/address/v1",               "customer_address",           "output/enedis_addresses.csv"),
        ("/connection/v1",            "customer_connection",        "output/enedis_connections.csv"),
        ("/contract_summary/v1",      "customer_contract_summary",  "output/enedis_contract_summary.csv"),
        ("/alimentation_auto/v1",     "customer_alimentation",      "output/enedis_alimentation.csv"),
        ("/situation_contrat_auto/v1", "customer_situation_contrat", "output/enedis_situation_contrat.csv"),
    ]
    for context, label, csv_path in apis:
        LOG.info("--- Collecte %s ---", label)
        rows = fetch_enedis_customer_data(base, token, batch, context)
        if rows:
            upsert_rows_to_csv(rows, csv_path, key_cols=("usage_point_id",))
        else:
            LOG.warning("%s : aucune donnée collectée.", label)


# ---------------------------------------------------------------------------
# Collecte mesures synchrones (daily_consumption + consumption_load_curve)
# ---------------------------------------------------------------------------


def fetch_enedis_sync_data(
    settings: "Settings",
    token: str,
    prms: List[str],
    start_date: str,
    end_date: str,
) -> List[Dict[str, Any]]:
    """Collecte la consommation journalière pour chaque PRM.

    GET /mesures/v2/metering_data/daily_consumption?usage_point_id=…&start=…&end=…
    Un appel par PRM, parallèle (10 threads), avec retry sur 429.
    """
    from concurrent.futures import ThreadPoolExecutor
    import time as _time

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    ingested_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_rows: List[Dict[str, Any]] = []
    ok_count = err_count = 0
    _RETRY_429 = [20, 40, 60]

    total_sync_prms = len(prms)
    LOG.info("daily_consumption : fenêtre %s → %s, %d PRM(s).", start_date, end_date, total_sync_prms)

    def _call_one(prm: str) -> Tuple[List[Dict[str, Any]], int, int]:
        q429 = 0
        resp = None
        for attempt in range(4):
            try:
                resp = requests.get(
                    settings.enedis_sync_url,
                    headers=headers,
                    params={"usage_point_id": prm, "start": start_date, "end": end_date},
                    timeout=30,
                )
                if resp.status_code == 429 and q429 < len(_RETRY_429):
                    LOG.warning("PRM %s [daily_consumption] → 429 quota, attente %ds...",
                                prm, _RETRY_429[q429])
                    _time.sleep(_RETRY_429[q429])
                    q429 += 1
                    continue
                if resp.status_code >= 500:
                    _time.sleep(5 * (attempt + 1))
                    continue
                break
            except Exception as exc:
                if attempt == 3:
                    LOG.warning("PRM %s [daily_consumption] → erreur réseau : %s", prm, exc)
                    return [], 0, 1
                _time.sleep(5)
        if resp is None:
            return [], 0, 1
        if resp.status_code == 200:
            mr = resp.json().get("meter_reading", {})
            unit = mr.get("reading_type", {}).get("unit", "Wh")
            quality = mr.get("quality", "")
            flow_dir = mr.get("reading_type", {}).get("flow_direction", "")
            rows = []
            for ir in mr.get("interval_reading", []):
                raw_date = ir.get("date", "")
                val = ir.get("value", "")
                try:
                    date_str = raw_date[:10]
                    value_wh = float(val) if val not in (None, "") else None
                except (ValueError, TypeError):
                    continue
                rows.append({
                    "usage_point_id": prm,
                    "date": date_str,
                    "value_wh": value_wh,
                    "unit": unit,
                    "quality": quality,
                    "flow_direction": flow_dir,
                    "_ingested_at_utc": ingested_at,
                })
            return rows, 1, 0
        LOG.warning("PRM %s [daily_consumption] → HTTP %d : %s",
                    prm, resp.status_code, resp.text[:200])
        return [], 0, 1

    workers = min(4, len(prms)) if prms else 1
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for idx, (rows, ok, err) in enumerate(ex.map(_call_one, prms), start=1):
            ok_count += ok
            err_count += err
            all_rows.extend(rows)
            if idx == 1 or idx % 50 == 0 or idx == total_sync_prms:
                LOG.info("daily_consumption : progression %d/%d PRM(s) (%.1f%%) — %d ligne(s).",
                         idx, total_sync_prms, idx / total_sync_prms * 100, len(all_rows))
    LOG.info("daily_consumption : %d OK, %d erreur(s), %d ligne(s) collectée(s).",
             ok_count, err_count, len(all_rows))
    return all_rows


def fetch_max_power_data(
    settings: "Settings",
    token: str,
    prms: List[str],
    start_date: str,
    end_date: str,
) -> List[Dict[str, Any]]:
    """Collecte la puissance max journalière pour chaque PRM.

    GET /mesures/v2/metering_data/daily_consumption_max_power?usage_point_id=…&start=…&end=…&measuring_period=P1D&grandeurPhysique=PMA
    Un appel par PRM, parallèle (4 threads), avec retry sur 429.
    Historique max : 36 mois + 15 jours.
    """
    from concurrent.futures import ThreadPoolExecutor
    import time as _time

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    ingested_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_rows: List[Dict[str, Any]] = []
    ok_count = err_count = 0
    _RETRY_429 = [20, 40, 60]

    total_prms = len(prms)
    LOG.info("max_power : fenêtre %s → %s, %d PRM(s).", start_date, end_date, total_prms)

    def _call_one(prm: str) -> Tuple[List[Dict[str, Any]], int, int]:
        q429 = 0
        resp = None
        for attempt in range(4):
            try:
                resp = requests.get(
                    settings.enedis_max_power_url,
                    headers=headers,
                    params={
                        "usage_point_id": prm,
                        "start": start_date,
                        "end": end_date,
                        "measuring_period": "P1D",
                        "grandeurPhysique": "PMA",
                    },
                    timeout=30,
                )
                if resp.status_code == 429 and q429 < len(_RETRY_429):
                    LOG.warning("PRM %s [max_power] → 429 quota, attente %ds...",
                                prm, _RETRY_429[q429])
                    _time.sleep(_RETRY_429[q429])
                    q429 += 1
                    continue
                if resp.status_code >= 500:
                    _time.sleep(5 * (attempt + 1))
                    continue
                break
            except Exception as exc:
                if attempt == 3:
                    LOG.warning("PRM %s [max_power] → erreur réseau : %s", prm, exc)
                    return [], 0, 1
                _time.sleep(5)
        if resp is None:
            return [], 0, 1
        if resp.status_code == 200:
            mr = resp.json().get("meter_reading", {})
            unit = mr.get("reading_type", {}).get("unit", "VA")
            quality = mr.get("quality", "")
            flow_dir = mr.get("reading_type", {}).get("flow_direction", "")
            rows = []
            for ir in mr.get("interval_reading", []):
                raw_date = ir.get("date", "")
                val = ir.get("value", "")
                try:
                    date_str = raw_date[:10]
                    value_va = float(val) if val not in (None, "") else None
                except (ValueError, TypeError):
                    continue
                rows.append({
                    "usage_point_id": prm,
                    "date": date_str,
                    "value_va": value_va,
                    "unit": unit,
                    "quality": quality,
                    "flow_direction": flow_dir,
                    "_ingested_at_utc": ingested_at,
                })
            return rows, 1, 0
        LOG.warning("PRM %s [max_power] → HTTP %d : %s",
                    prm, resp.status_code, resp.text[:200])
        return [], 0, 1

    workers = min(4, len(prms)) if prms else 1
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for idx, (rows, ok, err) in enumerate(ex.map(_call_one, prms), start=1):
            ok_count += ok
            err_count += err
            all_rows.extend(rows)
            if idx == 1 or idx % 50 == 0 or idx == total_prms:
                LOG.info("max_power : progression %d/%d PRM(s) (%.1f%%) — %d ligne(s).",
                         idx, total_prms, idx / total_prms * 100, len(all_rows))
    LOG.info("max_power : %d OK, %d erreur(s), %d ligne(s) collectée(s).",
             ok_count, err_count, len(all_rows))
    return all_rows


def get_last_max_power_date(csv_path: str) -> Optional[str]:
    """Lit la date max dans le CSV max_power (colonne 'date')."""
    if not Path(csv_path).exists():
        return None
    try:
        import csv as _csv
        max_d = None
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            reader = _csv.reader(f)
            header = next(reader, None)
            if not header:
                return None
            try:
                idx = header.index("date")
            except ValueError:
                return None
            for row in reader:
                if len(row) > idx:
                    try:
                        d = date.fromisoformat(row[idx][:10])
                        if max_d is None or d > max_d:
                            max_d = d
                    except ValueError:
                        pass
        return max_d.isoformat() if max_d else None
    except Exception as exc:
        LOG.warning("get_last_max_power_date(%s) : %s", csv_path, exc)
        return None


def fetch_load_curve_data(
    settings: "Settings",
    token: str,
    prms: List[str],
    start_date: str,
    end_date: str,
) -> List[Dict[str, Any]]:
    """Collecte la courbe de charge 30 min pour chaque PRM.

    GET /mesures/v2/metering_data/consumption_load_curve?usage_point_id=…&start=…&end=…
    Limite API ENEDIS : 7 jours max par appel → découpage interne en sous-fenêtres de 7j.
    """
    from concurrent.futures import ThreadPoolExecutor
    import time as _time

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    ingested_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_rows: List[Dict[str, Any]] = []
    ok_count = err_count = 0
    historical_skip_count = 0
    _RETRY_429 = [20, 40, 60]
    LC_API_MAX_DAYS = 7

    chunks: List[Tuple[str, str]] = []
    cur = date.fromisoformat(start_date)
    end = date.fromisoformat(end_date)
    while cur <= end:
        chunk_end = min(cur + timedelta(days=LC_API_MAX_DAYS - 1), end)
        chunks.append((cur.isoformat(), (chunk_end + timedelta(days=1)).isoformat()))
        cur = chunk_end + timedelta(days=1)

    total_prms = len(prms)
    total_chunks = len(chunks)
    total_estimated_calls = total_prms * total_chunks
    progress_every = 10 if total_prms >= 50 else max(1, total_prms // 5) if total_prms else 1

    LOG.info(
        "consumption_load_curve : fenêtre %s → %s, %d PRM(s), %d sous-fenêtre(s)/PRM, ~%d appel(s) API.",
        start_date, end_date, total_prms, total_chunks, total_estimated_calls,
    )

    def _request_one(prm: str, cs: str, ce: str) -> Tuple[List[Dict[str, Any]], int, int, int]:
        q429 = 0
        resp = None
        for attempt in range(4):
            try:
                resp = requests.get(
                    settings.enedis_load_curve_url,
                    headers=headers,
                    params={"usage_point_id": prm, "start": cs, "end": ce},
                    timeout=30,
                )
                if resp.status_code == 429 and q429 < len(_RETRY_429):
                    LOG.warning("PRM %s [load_curve %s→%s] → 429 quota, attente %ds...",
                                prm, cs, ce, _RETRY_429[q429])
                    _time.sleep(_RETRY_429[q429])
                    q429 += 1
                    continue
                if resp.status_code >= 500:
                    _time.sleep(5 * (attempt + 1))
                    continue
                break
            except Exception as exc:
                if attempt == 3:
                    LOG.warning("PRM %s [load_curve] → erreur réseau : %s", prm, exc)
                    return [], 0, 1, 0
                _time.sleep(5)
        if resp is None:
            return [], 0, 1, 0
        if resp.status_code == 200:
            mr = resp.json().get("meter_reading", {})
            unit = mr.get("reading_type", {}).get("unit", "W")
            quality = mr.get("quality", "")
            rows = []
            for ir in mr.get("interval_reading", []):
                raw_dt = ir.get("date", "")
                val = ir.get("value", "")
                try:
                    value_w = float(val) if val not in (None, "") else None
                except (ValueError, TypeError):
                    continue
                rows.append({
                    "usage_point_id": prm,
                    "datetime": raw_dt,
                    "value_w": value_w,
                    "unit": unit,
                    "quality": quality,
                    "_ingested_at_utc": ingested_at,
                })
            return rows, 1, 0, 0
        if resp.status_code == 400 and "date limite d'historique" in resp.text.lower():
            return [], 0, 0, 1
        LOG.warning("PRM %s [load_curve %s→%s] → HTTP %d : %s",
                    prm, cs, ce, resp.status_code, resp.text[:200])
        return [], 0, 1, 0

    def _call_one(prm: str) -> Tuple[List[Dict[str, Any]], int, int, int]:
        prm_rows: List[Dict[str, Any]] = []
        prm_ok = 0
        prm_err = 0
        prm_historical_skip = 0
        for cs, ce in chunks:
            rows, ok, err, historical_skip = _request_one(prm, cs, ce)
            prm_rows.extend(rows)
            prm_ok += ok
            prm_err += err
            prm_historical_skip += historical_skip
        return prm_rows, int(prm_ok > 0), prm_err, prm_historical_skip

    import threading as _threading
    _hb_stop = _threading.Event()
    def _hb():
        _n = 0
        while not _hb_stop.wait(60):
            _n += 1
            LOG.info("consumption_load_curve : ♥ heartbeat %d min — %d/%d PRM(s), %d ligne(s).",
                     _n, ok_count + err_count, total_prms, len(all_rows))
    _hb_t = _threading.Thread(target=_hb, daemon=True)
    _hb_t.start()

    workers = min(4, len(prms)) if prms else 1
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for idx, (rows, ok, err, historical_skip) in enumerate(ex.map(_call_one, prms), start=1):
            ok_count += ok
            err_count += err
            historical_skip_count += historical_skip
            all_rows.extend(rows)

            if idx == 1 or idx % progress_every == 0 or idx == total_prms:
                pct = (idx / total_prms * 100.0) if total_prms else 100.0
                LOG.info(
                    "consumption_load_curve : progression %d/%d PRM(s) (%.1f%%) — %d PRM(s) avec données, %d erreur(s), %d sous-fenêtre(s) hors historique ignorée(s), %d ligne(s) collectée(s).",
                    idx, total_prms, pct, ok_count, err_count, historical_skip_count, len(all_rows)
                )

    _hb_stop.set()

    LOG.info(
        "consumption_load_curve : %d PRM(s) OK, %d erreur(s), %d sous-fenêtre(s) hors historique ignorée(s), %d ligne(s) collectée(s).",
        ok_count, err_count, historical_skip_count, len(all_rows)
    )
    return all_rows


# Orchestration
# ----------------------------------------------------------------　　 　


def submit_next_enedis_sync_request(settings: Settings, state: Dict[str, Any]) -> None:
    LOG.info("--- Étape 2 : collecte synchrone ENEDIS (mesure_synchrone v2) ---")

    token = get_enedis_token(settings)

    # --- Découverte automatique des PRMs ---
    discovered = discover_prms_from_api(settings, token)
    if discovered:
        prms = discovered
        LOG.info("PRMs issus de l'API périmètre : %d.", len(prms))
    else:
        LOG.warning("Découverte API périmètre indisponible — fallback sur %s.", settings.prm_list_path)
        prms = load_prms(settings.prm_list_path)

    if settings.max_prms > 0:
        LOG.info("MAX_PRMS=%d — limitation à %d PRM(s) sur %d.", settings.max_prms, min(settings.max_prms, len(prms)), len(prms))
        prms = prms[:settings.max_prms]

    # --- Détection des nouveaux PRMs ---
    known_prms: List[str] = state.get("known_prms", [])
    new_prms = [p for p in prms if p not in known_prms]
    if new_prms:
        LOG.info("%d nouveau(x) PRM(s) détecté(s) : %s", len(new_prms), new_prms)
    state["known_prms"] = sorted(set(prms))

    # --- Collecte customer (si CSV absent OU nouveaux PRMs) ---
    skip_customer = os.environ.get("SKIP_CUSTOMER_COLLECTION", "").lower() in ("1", "true", "yes")
    force_customer = os.environ.get("FORCE_CUSTOMER_COLLECTION", "").lower() in ("1", "true", "yes")
    if skip_customer:
        LOG.info("SKIP_CUSTOMER_COLLECTION activé — collecte customer ignorée (mode backfill).")
    elif force_customer:
        LOG.info("FORCE_CUSTOMER_COLLECTION activé — collecte customer forcée pour %d PRM(s).", len(prms))
        collect_customer_supplementary(settings, token, prms)
    else:
        customer_missing = not Path("output/enedis_contracts.csv").exists()
        if customer_missing:
            LOG.info("Fichiers customer absents — collecte contrat, adresse, connexion.")
            collect_customer_supplementary(settings, token, prms)
        elif new_prms:
            LOG.info("Nouveaux PRMs — collecte customer pour les %d nouveaux uniquement.", len(new_prms))
            collect_customer_supplementary(settings, token, new_prms)

    # --- Courbe de charge (dates indépendantes — toujours exécutée) ---
    today = now_local(settings.timezone).date()
    if settings.enable_load_curve:
        LOG.info("--- Collecte courbe de charge 30 min ---")
        token = get_enedis_token(settings)  # refresh token avant collecte longue
        lc_end = (today - timedelta(days=settings.day_offset_end)).isoformat()

        # Source de vérité : max(state, CSV réel) pour être auto-résilient si l'état est perdu
        lc_last_state = state.get("last_load_curve_date")
        lc_csv_path = str(Path(settings.csv_output_path).parent / "enedis_load_curve.csv")
        LOG.info("Courbe de charge : lecture date max CSV (%s)...", lc_csv_path)
        lc_last_csv = get_last_load_curve_date(lc_csv_path)
        if lc_last_state and lc_last_csv:
            lc_last = max(lc_last_state, lc_last_csv)
        else:
            lc_last = lc_last_csv or lc_last_state
        if lc_last != lc_last_state:
            LOG.info("Courbe de charge : date de reprise syncée depuis CSV — %s (état était : %s).",
                     lc_last, lc_last_state)

        if lc_last:
            lc_start = (date.fromisoformat(lc_last) + timedelta(days=1)).isoformat()
        else:
            lc_start = (today - timedelta(days=settings.load_curve_history_days)).isoformat()
        if lc_start <= lc_end:
            if settings.load_curve_max_days_per_run > 0:
                capped = (date.fromisoformat(lc_start) + timedelta(days=settings.load_curve_max_days_per_run - 1)).isoformat()
                if capped < lc_end:
                    lc_end = capped
                    LOG.info("Courbe de charge : plage limitée à %d j/run — fin ajustée à %s.", settings.load_curve_max_days_per_run, lc_end)
            LOG.info("Courbe de charge : %s → %s (%d PRM(s)).", lc_start, lc_end, len(prms))
            lc_rows = fetch_load_curve_data(settings, token, prms, lc_start, lc_end)
            if lc_rows:
                upsert_rows_to_csv(lc_rows, "output/enedis_load_curve.csv",
                                   key_cols=("usage_point_id", "datetime"))
            else:
                LOG.warning("Courbe de charge : aucun point collecté pour la période %s → %s.",
                            lc_start, lc_end)
            state["last_load_curve_date"] = lc_end
            save_state(settings.state_file_path, state)
        else:
            LOG.info("Courbe de charge déjà à jour (last_load_curve_date=%s).", lc_last)
    else:
        LOG.info("ENABLE_LOAD_CURVE désactivé — collecte courbe de charge ignorée.")

    # --- Consommation journalière (dates indépendantes — par tranches) ---
    if not settings.enable_daily_consumption:
        LOG.info("ENABLE_DAILY_CONSUMPTION désactivé — collecte consommation ignorée.")
    else:
        LOG.info("--- Collecte consommation journalière ---")
        token = get_enedis_token(settings)  # refresh token avant collecte longue
        sync_end = (today - timedelta(days=settings.day_offset_end)).isoformat()

        sync_last_state = state.get("last_sync_date")
        sync_csv_path = settings.csv_output_path
        LOG.info("Consommation : lecture date max CSV (%s)...", sync_csv_path)
        sync_last_csv = get_last_sync_date(sync_csv_path)
        if sync_last_state and sync_last_csv:
            sync_last = max(sync_last_state, sync_last_csv)
        else:
            sync_last = sync_last_csv or sync_last_state
        if sync_last != sync_last_state:
            LOG.info("Consommation : date de reprise syncée depuis CSV — %s (état était : %s).",
                     sync_last, sync_last_state)

        if sync_last:
            start_date = (date.fromisoformat(sync_last) + timedelta(days=1)).isoformat()
        else:
            start_date = (today - timedelta(days=settings.history_days)).isoformat()

        if start_date > sync_end:
            LOG.info("Consommation déjà à jour (last_sync_date=%s).", sync_last)
        else:
            if settings.sync_max_days_per_run > 0:
                capped = (date.fromisoformat(start_date) + timedelta(days=settings.sync_max_days_per_run - 1)).isoformat()
                if capped < sync_end:
                    sync_end = capped
                    LOG.info("Consommation : plage limitée à %d j/run — fin ajustée à %s.", settings.sync_max_days_per_run, sync_end)

            LOG.info("Consommation : %s → %s (%d PRM(s)).", start_date, sync_end, len(prms))

            rows = fetch_enedis_sync_data(settings, token, prms, start_date, sync_end)
            if rows:
                upsert_rows_to_csv(rows, settings.csv_output_path)
            else:
                LOG.warning(
                    "Aucune donnée trouvée pour la période %s → %s "
                    "(PRMs sans données ou hors historique disponible).",
                    start_date, sync_end,
                )
            state["last_sync_date"] = sync_end
            LOG.info("last_sync_date avancé → %s (%d ligne(s) collectée(s)).", sync_end, len(rows))

    # --- Puissance max journalière (dates indépendantes — par tranches) ---
    if not settings.enable_max_power:
        LOG.info("ENABLE_MAX_POWER désactivé — collecte puissance max ignorée.")
    else:
        LOG.info("--- Collecte puissance max journalière ---")
        token = get_enedis_token(settings)
        mp_end = (today - timedelta(days=settings.day_offset_end)).isoformat()

        mp_csv_path = str(Path(settings.csv_output_path).parent / "enedis_max_power.csv")
        mp_last_state = state.get("last_max_power_date")
        LOG.info("Max power : lecture date max CSV (%s)...", mp_csv_path)
        mp_last_csv = get_last_max_power_date(mp_csv_path)
        if mp_last_state and mp_last_csv:
            mp_last = max(mp_last_state, mp_last_csv)
        else:
            mp_last = mp_last_csv or mp_last_state
        if mp_last != mp_last_state:
            LOG.info("Max power : date de reprise syncée depuis CSV — %s (état était : %s).",
                     mp_last, mp_last_state)

        if mp_last:
            mp_start = (date.fromisoformat(mp_last) + timedelta(days=1)).isoformat()
        else:
            mp_start = (today - timedelta(days=settings.max_power_history_days)).isoformat()

        if mp_start > mp_end:
            LOG.info("Max power déjà à jour (last_max_power_date=%s).", mp_last)
        else:
            if settings.max_power_max_days_per_run > 0:
                capped = (date.fromisoformat(mp_start) + timedelta(days=settings.max_power_max_days_per_run - 1)).isoformat()
                if capped < mp_end:
                    mp_end = capped
                    LOG.info("Max power : plage limitée à %d j/run — fin ajustée à %s.",
                             settings.max_power_max_days_per_run, mp_end)

            LOG.info("Max power : %s → %s (%d PRM(s)).", mp_start, mp_end, len(prms))

            mp_rows = fetch_max_power_data(settings, token, prms, mp_start, mp_end)
            if mp_rows:
                upsert_rows_to_csv(mp_rows, mp_csv_path, key_cols=("usage_point_id", "date"))
            else:
                LOG.warning("Max power : aucune donnée pour la période %s → %s.", mp_start, mp_end)
            state["last_max_power_date"] = mp_end
            save_state(settings.state_file_path, state)
            LOG.info("last_max_power_date avancé → %s (%d ligne(s) collectée(s)).", mp_end, len(mp_rows))


# ----------------------------------------------------------------　　 　 　
# Point d'entrée
# ----------------------------------------------------------------　　 　 　


def main(args: Optional[List[str]] = None) -> int:
    setup_logging()

    parser = argparse.ArgumentParser(
        description="Flux ENEDIS → Power BI (GitHub Actions)"
    )
    parser.add_argument(
        "--regenerate-prm-list",
        action="store_true",
        help=(
            "Régénère config/prm_list.txt depuis LISTE PRM.xlsx "
            "avant le run principal."
        ),
    )
    parsed = parser.parse_args(args)

    settings = Settings.load()

    # --- Diagnostic : variables d'environnement chargées ---
    _SECRETS = {"ENEDIS_CLIENT_ID", "ENEDIS_CLIENT_SECRET"}
    for var in [
        "ENEDIS_AUTH_URL",
        "ENEDIS_CLIENT_ID", "ENEDIS_CLIENT_SECRET",
        "CSV_OUTPUT_PATH", "PRM_LIST_PATH",
        "STATE_FILE_PATH", "TIMEZONE",
    ]:
        raw = os.getenv(var, "")
        if not raw:
            LOG.warning("DIAGNOSTIC — %s : VIDE (secret non sauvegardé ?)", var)
        elif var in _SECRETS:
            LOG.info("DIAGNOSTIC — %s : *** (longueur=%d)", var, len(raw))
        else:
            LOG.info("DIAGNOSTIC — %s : %s", var, raw)
    LOG.info("DIAGNOSTIC — ENEDIS_SYNC_URL (effectif) : %s", settings.enedis_sync_url)
    LOG.info("DIAGNOSTIC — ENEDIS_CANAL_ID (effectif) : %s", settings.enedis_canal_id)
    LOG.info("DIAGNOSTIC — ENEDIS_TYPE_DONNEE (effectif) : %s", settings.enedis_type_donnee)
    LOG.info("DIAGNOSTIC — MAX_PRMS_PER_REQUEST (effectif) : %d", settings.max_prms_per_request)
    LOG.info("DIAGNOSTIC — MAX_PRMS (effectif) : %d", settings.max_prms)
    LOG.info("DIAGNOSTIC — HISTORY_DAYS (effectif) : %d", settings.history_days)
    # --- Fin diagnostic ---

    if parsed.regenerate_prm_list:
        LOG.info("Régénération de la liste PRM depuis %s…", settings.prm_xlsx_path)
        generate_prm_list_from_xlsx(settings.prm_xlsx_path, settings.prm_list_path)

    state = load_state(settings.state_file_path)

    submit_next_enedis_sync_request(settings, state)
    save_state(settings.state_file_path, state)

    completeness_warnings = check_completeness(settings, state)
    for w in completeness_warnings:
        LOG.warning(w)

    run_msg = f"last_sync_date : {state.get('last_sync_date', 'N/A')}\nlast_load_curve_date : {state.get('last_load_curve_date', 'N/A')}"
    if completeness_warnings:
        run_msg += "\n" + "\n".join(completeness_warnings)
        send_teams_alert(settings.alert_webhook_url, "⚠️ ENEDIS Sync — Complétude incomplète", run_msg, color="FFA500")
    elif settings.alert_on_success:
        send_teams_alert(settings.alert_webhook_url, "✅ ENEDIS Sync — Succès", run_msg, color="00C853")

    LOG.info("=== Run terminé ===")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        LOG.error("Erreur fatale : %s", exc, exc_info=True)
        send_teams_alert(
            os.getenv("ALERT_WEBHOOK_URL", ""),
            "❌ ENEDIS Sync — Erreur fatale",
            str(exc),
            color="FF0000",
        )
        raise SystemExit(1)
